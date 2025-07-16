# raporlar.py dosyasının içeriği
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate, Spacer
from reportlab.lib import colors
from reportlab.pdfgen import canvas as rp_canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import threading

# Fontları veritabani.py'den import ediyoruz
from veritabani import TURKISH_FONT_NORMAL, TURKISH_FONT_BOLD # Bu importlar korunacak
from reportlab.pdfbase.ttfonts import TTFont # Bu import'un da olduğundan emin olun
from reportlab.pdfbase import pdfmetrics # Bu import'un da olduğundan emin olun
import os # os modülünü de import ettiğinizden emin olun
from yardimcilar import sort_treeview_column, setup_date_entry, DatePickerDialog

# PENCERELER MODÜLÜNDEN GEREKENLER
from pencereler import BeklemePenceresi, CariHesapEkstresiPenceresi

try:
    # Sadece eğer fontlar daha önce kaydedilmediyse kaydetmek iyi bir pratiktir.
    # Ancak ReportLab genellikle zaten kaydedilmiş bir fontu tekrar kaydetmeye çalışırken hata vermez.
    dejavu_sans_normal_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'DejaVuSans.ttf')
    dejavu_sans_bold_path = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), 'data')), 'DejaVuSans-Bold.ttf')

    # Sadece dosya varsa kaydet
    if os.path.exists(dejavu_sans_normal_path):
        pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_sans_normal_path))
    else:
        print(f"UYARI (raporlar.py): {dejavu_sans_normal_path} bulunamadı. Varsayılan Helvetica kullanılacak.")

    if os.path.exists(dejavu_sans_bold_path):
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_sans_bold_path))
    else:
        print(f"UYARI (raporlar.py): {dejavu_sans_bold_path} bulunamadı. Varsayılan Helvetica-Bold kullanılacak.")

except Exception as e:
    print(f"KRİTİK FONT YÜKLEME HATASI (raporlar.py): {e} - PDF'lerde Türkçe karakter sorunu olabilir.")


# YARDIMCI MODÜLLERDEN GEREKENLER
from yardimcilar import sort_treeview_column, setup_date_entry, DatePickerDialog

# PENCERELER MODÜLÜNDEN GEREKENLER
from pencereler import BeklemePenceresi, CariHesapEkstresiPenceresi 
#from arayuz import setup_date_entry

class CariYaslandirmaRaporuPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.title("Cari Hesap Yaşlandırma Raporu")
        self.geometry("1200x700")
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text="Cari Hesap Yaşlandırma Raporu", font=("Segoe UI", 16, "bold")).pack(pady=10, anchor=tk.W, padx=10)

        # Filtreleme Çerçevesi
        filter_frame = ttk.Frame(self, padding="10")
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(filter_frame, text="Rapor Tarihi (YYYY-AA-GG):").pack(side=tk.LEFT, padx=(0,2))
        self.rapor_tarihi_entry = ttk.Entry(filter_frame, width=12)
        self.rapor_tarihi_entry.pack(side=tk.LEFT, padx=(0,10))
        self.rapor_tarihi_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        setup_date_entry(self.app, self.rapor_tarihi_entry)
        # Rapor tarihi için takvim butonu
        ttk.Button(filter_frame, text="🗓️", command=lambda: self._open_date_picker(self.rapor_tarihi_entry), width=3).pack(side=tk.LEFT, padx=2)


        ttk.Button(filter_frame, text="Raporla", command=self.raporu_guncelle, style="Accent.TButton").pack(side=tk.LEFT, padx=(10,0))
        ttk.Button(filter_frame, text="Excel'e Aktar", command=self.excel_aktar).pack(side=tk.RIGHT, padx=5)
        ttk.Button(filter_frame, text="PDF'e Aktar", command=self.pdf_aktar).pack(side=tk.RIGHT, padx=5)

        # Rapor Sonuçları Çerçevesi (Müşteri ve Tedarikçi için ikiye ayrılacak)
        results_main_frame = ttk.Frame(self, padding="10")
        results_main_frame.pack(expand=True, fill=tk.BOTH)

        # Müşteri Alacakları
        musteri_frame = ttk.LabelFrame(results_main_frame, text="Müşteri Alacakları (Bize Borçlu)", padding="10")
        musteri_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=(0,5))
        self.musteri_tree = self._create_yaslandirma_treeview(musteri_frame)

        # Tedarikçi Borçları
        tedarikci_frame = ttk.LabelFrame(results_main_frame, text="Tedarikçi Borçları (Biz Borçluyuz)", padding="10")
        tedarikci_frame.pack(side=tk.RIGHT, expand=True, fill=tk.BOTH, padx=(5,0))
        self.tedarikci_tree = self._create_yaslandirma_treeview(tedarikci_frame)

        # Toplamlar Alanı
        summary_frame = ttk.Frame(self, padding="10")
        summary_frame.pack(fill=tk.X, side=tk.BOTTOM)
        self.toplam_alacak_label = ttk.Label(summary_frame, text="Toplam Alacak: 0.00 TL", font=("Segoe UI", 10, "bold"))
        self.toplam_alacak_label.pack(side=tk.LEFT, padx=10)
        self.toplam_borc_label = ttk.Label(summary_frame, text="Toplam Borç: 0.00 TL", font=("Segoe UI", 10, "bold"))
        self.toplam_borc_label.pack(side=tk.LEFT, padx=10)
        self.net_bakiye_label = ttk.Label(summary_frame, text="Net Bakiye: 0.00 TL", font=("Segoe UI", 12, "bold"))
        self.net_bakiye_label.pack(side=tk.RIGHT, padx=10)

        self.raporu_guncelle() # İlk yüklemede raporu oluştur

    def _open_date_picker(self, target_entry):
        """Bir Entry widget'ı için tarih seçici penceresi açar."""
        DatePickerDialog(self.app, target_entry)

    def _create_yaslandirma_treeview(self, parent_frame):
        cols = ("Cari Adı", "Tutar", "Vadesi Geçen Gün")
        tree = ttk.Treeview(parent_frame, columns=cols, show='headings', selectmode="browse")

        col_defs = [
            ("Cari Adı", 200, tk.W, tk.YES),
            ("Tutar", 100, tk.E, tk.NO),
            ("Vadesi Geçen Gün", 120, tk.E, tk.NO)
        ]
        for cn,w,a,s in col_defs:
            tree.column(cn, width=w, anchor=a, stretch=s)
            tree.heading(cn, text=cn, command=lambda _c=cn: sort_treeview_column(tree, _c, False))

        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(expand=True, fill=tk.BOTH)
        return tree

    def raporu_guncelle(self):
        rapor_tarihi_str = self.rapor_tarihi_entry.get()

        try:
            datetime.strptime(rapor_tarihi_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Tarih Formatı Hatası", "Rapor Tarihi formatı (YYYY-AA-GG) olmalıdır.", parent=self)
            return

        yaslandirma_sonuclari = self.db.get_cari_yaslandirma_verileri(rapor_tarihi_str)

        self._populate_treeview(self.musteri_tree, yaslandirma_sonuclari['musteri_alacaklari'])
        self._populate_treeview(self.tedarikci_tree, yaslandirma_sonuclari['tedarikci_borclari'])

        # Toplamları hesapla
        toplam_alacak = sum(item[2] for group in yaslandirma_sonuclari['musteri_alacaklari'].values() for item in group)
        toplam_borc = sum(item[2] for group in yaslandirma_sonuclari['tedarikci_borclari'].values() for item in group)
        net_bakiye = toplam_alacak - toplam_borc

        self.toplam_alacak_label.config(text=f"Toplam Alacak: {self.db._format_currency(toplam_alacak)}")
        self.toplam_borc_label.config(text=f"Toplam Borç: {self.db._format_currency(toplam_borc)}")
        self.net_bakiye_label.config(text=f"Net Bakiye: {self.db._format_currency(net_bakiye)}")

        self.app.set_status(f"Cari Hesap Yaşlandırma Raporu güncellendi ({rapor_tarihi_str}).")


    def _populate_treeview(self, tree, data_dict):
        for i in tree.get_children():
            tree.delete(i)

        for period, items in data_dict.items():
            if items:
                # Kategori başlığını ekle
                tree.insert("", tk.END, iid=period, text=f"--- {period} Gün ---", open=True, tags=('header',))
                for item in items:
                    # item: (cari_id, cari_adi, tutar, vadesi_gecen_gun_sayisi)
                    tree.insert(period, tk.END, values=(
                        item[1], # Cari Adı
                        self.db._format_currency(item[2]), # Tutar
                        item[3] # Vadesi Geçen Gün
                    ))
            else: # Boş kategoriler için bilgi mesajı
                tree.insert("", tk.END, iid=period, values=("", "", "Bu Kategori Boş"), tags=('empty',))

        # Header ve boş satırlar için stil
        tree.tag_configure('header', font=('Segoe UI', 9, 'bold'), background='#E0E0E0')
        tree.tag_configure('empty', foreground='gray')


    def excel_aktar(self):
        dosya_yolu = filedialog.asksaveasfilename(
            initialfile=f"Cari_Yaslandirma_Raporu_{datetime.now().strftime('%Y%m%d')}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyaları", "*.xlsx")],
            title="Cari Hesap Yaşlandırma Raporunu Excel'e Kaydet",
            parent=self
        )
        if not dosya_yolu:
            self.app.set_status("Excel'e aktarma iptal edildi.")
            return

        rapor_tarihi_str = self.rapor_tarihi_entry.get()

        bekleme_penceresi = BeklemePenceresi(self, message="Rapor Excel'e aktarılıyor, lütfen bekleyiniz...")
        threading.Thread(target=lambda: self._generate_excel_report_threaded(
            rapor_tarihi_str, dosya_yolu, bekleme_penceresi
        )).start()

    def _generate_excel_report_threaded(self, rapor_tarihi_str, dosya_yolu, bekleme_penceresi):
        try:
            yaslandirma_sonuclari = self.db.get_cari_yaslandirma_verileri(rapor_tarihi_str)
            if not yaslandirma_sonuclari['musteri_alacaklari'] and not yaslandirma_sonuclari['tedarikci_borclari']:
                success = False
                message = "Excel'e aktarılacak cari yaşlandırma verisi bulunamadı."
            else:
                wb = openpyxl.Workbook()
                ws_musteri = wb.active
                ws_musteri.title = "Musteri_Alacaklari"
                self._write_excel_sheet(ws_musteri, "Müşteri Alacakları", yaslandirma_sonuclari['musteri_alacaklari'])

                ws_tedarikci = wb.create_sheet("Tedarikci_Borclari")
                self._write_excel_sheet(ws_tedarikci, "Tedarikçi Borçları", yaslandirma_sonuclari['tedarikci_borclari'])

                wb.save(dosya_yolu)
                success = True
                message = f"Cari Hesap Yaşlandırma Raporu başarıyla '{dosya_yolu}' adresine kaydedildi."
        except Exception as e:
            success = False
            message = f"Rapor Excel'e aktarılırken bir hata oluştu:\n{e}"
            traceback.print_exc()
        finally:
            self.app.after(0, bekleme_penceresi.kapat)
            self.app.after(0, lambda: messagebox.showinfo("Excel Aktarım", message, parent=self.app) if success else messagebox.showerror("Excel Aktarım Hatası", message, parent=self.app))
            self.app.after(0, lambda: self.app.set_status(message))

    def _write_excel_sheet(self, ws, title, data_dict):
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.append([]) # Boş satır

        headers = ["Cari Adı", "Tutar", "Vadesi Geçen Gün"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for period, items in data_dict.items():
            ws.append([f"--- {period} Gün ---"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.append(headers)
            for col_idx, header_text in enumerate(headers, 1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            if not items:
                ws.append(["Bu Kategori Boş", "", ""])
            else:
                for item in items:
                    ws.append([item[1], item[2], item[3]])
                    # Tutar sütununu formatla
                    ws.cell(row=ws.max_row, column=2).number_format = '#,##0.00₺'
            ws.append([]) # Boş satır
        
        # Sütun genişliklerini ayarla
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    def pdf_aktar(self):
        dosya_yolu = filedialog.asksaveasfilename(
            initialfile=f"Cari_Yaslandirma_Raporu_{datetime.now().strftime('%Y%m%d')}.pdf",
            defaultextension=".pdf",
            filetypes=[("PDF Dosyaları", "*.pdf")],
            title="Cari Hesap Yaşlandırma Raporunu PDF'e Kaydet",
            parent=self
        )
        if not dosya_yolu:
            self.app.set_status("PDF'e aktarma iptal edildi.")
            return

        rapor_tarihi_str = self.rapor_tarihi_entry.get()

        bekleme_penceresi = BeklemePenceresi(self, message="Rapor PDF'e aktarılıyor, lütfen bekleyiniz...")
        threading.Thread(target=lambda: self._generate_pdf_report_threaded(
            rapor_tarihi_str, dosya_yolu, bekleme_penceresi
        )).start()

    def _generate_pdf_report_threaded(self, rapor_tarihi_str, dosya_yolu, bekleme_penceresi):
        try:
            yaslandirma_sonuclari = self.db.get_cari_yaslandirma_verileri(rapor_tarihi_str)
            if not yaslandirma_sonuclari['musteri_alacaklari'] and not yaslandirma_sonuclari['tedarikci_borclari']:
                success = False
                message = "PDF'e aktarılacak cari yaşlandırma verisi bulunamadı."
            else:
                # PDF dokümanı oluştur (SimpleDocTemplate kullanılıyor)
                doc = SimpleDocTemplate(dosya_yolu, pagesize=landscape(A4),
                                        rightMargin=cm, leftMargin=cm,
                                        topMargin=2.5*cm, bottomMargin=2.5*cm) # Kenar boşlukları

                # Stil yönetimi
                styles = getSampleStyleSheet()
                styles.add(ParagraphStyle(name='TurkishNormal', fontName='DejaVuSans', fontSize=8, leading=10)) # TURKISH_FONT_NORMAL yerine 'DejaVuSans'
                styles.add(ParagraphStyle(name='TurkishBold', fontName='DejaVuSans-Bold', fontSize=8, leading=10)) # TURKISH_FONT_BOLD yerine 'DejaVuSans-Bold'
                styles.add(ParagraphStyle(name='SectionTitle', fontName='DejaVuSans-Bold', fontSize=10, leading=12)) # TURKISH_FONT_BOLD yerine 'DejaVuSans-Bold'
                styles.add(ParagraphStyle(name='ReportTitle', fontName='DejaVuSans-Bold', fontSize=14, alignment=1)) # TURKISH_FONT_BOLD yerine 'DejaVuSans-Bold'
                styles.add(ParagraphStyle(name='CompanyInfo', fontName='DejaVuSans', fontSize=9, alignment=1)) # TURKISH_FONT_NORMAL yerine 'DejaVuSans'
                styles.add(ParagraphStyle(name='TableValueRight', fontName='DejaVuSans', fontSize=8, alignment=2)) # TURKISH_FONT_NORMAL yerine 'DejaVuSans'


                elements = []

                # Rapor Başlığı
                elements.append(Paragraph(f"Cari Hesap Yaşlandırma Raporu ({rapor_tarihi_str})", styles['ReportTitle']))
                elements.append(Paragraph(self.db.sirket_bilgileri.get("sirket_adi", ""), styles['CompanyInfo']))
                elements.append(Spacer(0, 0.5*cm)) # Boşluk

                # Müşteri Alacakları bölümü
                elements.append(Paragraph("Müşteri Alacakları (Bize Borçlu)", styles['SectionTitle']))
                elements.append(Spacer(0, 0.2*cm))

                musteri_data = []
                musteri_data.append([Paragraph("Cari Adı", styles['TurkishBold']),
                                     Paragraph("Tutar", styles['TurkishBold']),
                                     Paragraph("Vadesi Geçen Gün", styles['TurkishBold'])])

                for period, items in yaslandirma_sonuclari['musteri_alacaklari'].items():
                    musteri_data.append([Paragraph(f"--- {period} Gün ---", styles['TurkishBold']), "", ""])
                    if not items:
                        musteri_data.append([Paragraph("Bu Kategori Boş", styles['TurkishNormal']), "", ""])
                    else:
                        for item in items:
                            musteri_data.append([
                                Paragraph(str(item[1]), styles['TurkishNormal']), # Cari Adı
                                Paragraph(self.db._format_currency(item[2]), styles['TableValueRight']), # Tutar
                                Paragraph(str(item[3]), styles['TableValueRight']) # Vadesi Geçen Gün
                            ])
                
                # Müşteri tablosunu oluştur
                musteri_table = Table(musteri_data, colWidths=[8*cm, 3*cm, 3*cm])
                musteri_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#D0D0D0")), # Sütun başlıkları
                    ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), TURKISH_FONT_BOLD),
                    ('FONTSIZE', (0,0), (-1,0), 8),
                    ('BOTTOMPADDING', (0,0), (-1,0), 6),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('BACKGROUND', (0,1), (-1,1), colors.white), # Kategori başlıkları (ilk satır)
                    # Diğer kategori başlıkları için özel stil
                    ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
                ]))
                elements.append(musteri_table)
                elements.append(Spacer(0, 1*cm)) # Bölümler arası boşluk

                # Tedarikçi Borçları bölümü
                elements.append(Paragraph("Tedarikçi Borçları (Biz Borçluyuz)", styles['SectionTitle']))
                elements.append(Spacer(0, 0.2*cm))

                tedarikci_data = []
                tedarikci_data.append([Paragraph("Cari Adı", styles['TurkishBold']),
                                        Paragraph("Tutar", styles['TurkishBold']),
                                        Paragraph("Vadesi Geçen Gün", styles['TurkishBold'])])

                for period, items in yaslandirma_sonuclari['tedarikci_borclari'].items():
                    tedarikci_data.append([Paragraph(f"--- {period} Gün ---", styles['TurkishBold']), "", ""])
                    if not items:
                        tedarikci_data.append([Paragraph("Bu Kategori Boş", styles['TurkishNormal']), "", ""])
                    else:
                        for item in items:
                            tedarikci_data.append([
                                Paragraph(str(item[1]), styles['TurkishNormal']), # Cari Adı
                                Paragraph(self.db._format_currency(item[2]), styles['TableValueRight']), # Tutar
                                Paragraph(str(item[3]), styles['TableValueRight']) # Vadesi Geçen Gün
                            ])

                # Tedarikçi tablosunu oluştur
                tedarikci_table = Table(tedarikci_data, colWidths=[8*cm, 3*cm, 3*cm])
                tedarikci_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#D0D0D0")), # Sütun başlıkları
                    ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), TURKISH_FONT_BOLD),
                    ('FONTSIZE', (0,0), (-1,0), 8),
                    ('BOTTOMPADDING', (0,0), (-1,0), 6),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('BACKGROUND', (0,1), (-1,1), colors.white), # Kategori başlıkları (ilk satır)
                    ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
                ]))
                elements.append(tedarikci_table)
                
                # Toplamlar (son sayfaya veya yeni sayfaya sığdır)
                toplam_alacak_pdf = sum(item[2] for group in yaslandirma_sonuclari['musteri_alacaklari'].values() for item in group)
                toplam_borc_pdf = sum(item[2] for group in yaslandirma_sonuclari['tedarikci_borclari'].values() for item in group)
                net_bakiye_pdf = toplam_alacak_pdf - toplam_borc_pdf

                elements.append(Spacer(0, 1*cm))
                elements.append(Paragraph(f"Toplam Alacak: {self.db._format_currency(toplam_alacak_pdf)}", styles['TurkishBold']))
                elements.append(Paragraph(f"Toplam Borç: {self.db._format_currency(toplam_borc_pdf)}", styles['TurkishBold']))
                elements.append(Paragraph(f"Net Bakiye: {self.db._format_currency(net_bakiye_pdf)}", styles['TurkishBold']))

                # PDF'i oluştur
                doc.build(elements)

                success = True
                message = f"Cari Hesap Yaşlandırma Raporu başarıyla '{dosya_yolu}' adresine kaydedildi."
        except Exception as e:
            success = False
            message = f"Rapor PDF'e aktarılırken bir hata oluştu:\n{e}"
            traceback.print_exc()
        finally:
            self.app.after(0, bekleme_penceresi.kapat)
            self.app.after(0, lambda: messagebox.showinfo("PDF Aktarım", message, parent=self.app) if success else messagebox.showerror("PDF Aktarım Hatası", message, parent=self.app))
            self.app.after(0, lambda: self.app.set_status(message))


    def _draw_pdf_section(self, canvas, y_start, page_width, section_title, data_dict, styles, styleN, styleH, styleRight, format_currency_func):
        # Fontların ReportLab'e kaydedildiğinden emin olun (raporlar.py içinde).
        # Bu blok, veritabani.py'deki font kaydı bloğunun bir kopyasıdır.
        # Eğer fontların raporlar.py içinde erken kaydedilmesi gerekiyorsa bu blok kalmalı,
        # aksi takdirde kaldırılabilir (ve veritabani.py'den import edilmelidir).
        # Daha önceki konuşmalarda bunu raporlar.py'ye taşımıştık.
        try:
            if 'DejaVuSans' not in pdfmetrics.getRegisteredFontNames():
                dejavu_sans_normal_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'DejaVuSans.ttf')
                dejavu_sans_bold_path = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), 'data')), 'DejaVuSans-Bold.ttf')

                if os.path.exists(dejavu_sans_normal_path):
                    pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_sans_normal_path))
                else:
                    print(f"UYARI (raporlar.py): {dejavu_sans_normal_path} bulunamadı. Varsayılan Helvetica kullanılacak.")

                if os.path.exists(dejavu_sans_bold_path):
                    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_sans_bold_path))
                else:
                    print(f"UYARI (raporlar.py): {dejavu_sans_bold_path} bulunamadı. Varsayılan Helvetica-Bold kullanılacak.")
        except Exception as e:
            print(f"KRİTİK FONT YÜKLEME HATASI (raporlar.py - _draw_pdf_section): {e}")


        canvas.setFont('DejaVuSans-Bold' if 'DejaVuSans-Bold' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold', 10)
        canvas.drawString(40, y_start, section_title)
        y_start -= 15

        headers = ["Cari Adı", "Tutar", "Vadesi Geçen Gün"]
        col_widths = [8*cm, 3*cm, 3*cm]
        table_data = []

        # Başlık stilleri için Paragraph objelerini oluştur
        style_header_paragraph = ParagraphStyle(name='TableHeadingStyle', fontName='DejaVuSans-Bold' if 'DejaVuSans-Bold' in pdfmetrics.getRegisteredFontNames() else 'Helvetica-Bold', fontSize=8, alignment=1) # Center
        style_normal_paragraph = ParagraphStyle(name='TableNormalStyle', fontName='DejaVuSans' if 'DejaVuSans' in pdfmetrics.getRegisteredFontNames() else 'Helvetica', fontSize=8) # Left
        style_right_paragraph = ParagraphStyle(name='TableRightStyle', fontName='DejaVuSans' if 'DejaVuSans' in pdfmetrics.getRegisteredFontNames() else 'Helvetica', fontSize=8, alignment=2) # Right

        # Headers için Paragraph objeleri oluştur
        header_row_paragraphs = [
            Paragraph("Cari Adı", style_header_paragraph),
            Paragraph("Tutar", style_header_paragraph),
            Paragraph("Vadesi Geçen Gün", style_header_paragraph)
        ]

        for period, items in data_dict.items():
            # Kategori başlığı
            table_data.append([Paragraph(f"--- {period} Gün ---", style_header_paragraph), "", ""])
            table_data.append(header_row_paragraphs) # Sütun başlıkları

            if not items:
                table_data.append([Paragraph("Bu Kategori Boş", style_normal_paragraph), "", ""])
            else:
                for item in items:
                    table_data.append([
                        Paragraph(str(item[1]), style_normal_paragraph), # Cari Adı
                        Paragraph(format_currency_func(item[2]), style_right_paragraph), # Tutar
                        Paragraph(str(item[3]), style_right_paragraph) # Vadesi Geçen Gün
                    ])

        # TableStyle'ı güncelleyin - HER BİR ELEMANIN SONUNDA VİRGÜL OLDUĞUNDAN EMİN OLUN!
        table_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#D0D0D0")), # İlk satir (Kategori başlığı) için örnek renk
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'), # Varsayılan hizalama
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),

            # Kategori başlıkları için özel stil (her "--- X Gün ---" satırı)
            # Bu, Table objesine eklendikten sonra dinamik olarak ayarlanmalıdır.
            # Ancak ReportLab TableStyle'ı statik olarak tanımlandığı için,
            # her kategori başlığından sonra ve sütun başlıklarından sonra
            # arka plan rengini manuel olarak belirtmeliyiz.

            # Dinamik stil için daha karmaşık bir yapı gerekir, şimdilik sabit stil
            # Her alt-tablo için (kategori başlığı + header + data) ayrı stil uygulayacağız

        ])

        # TableStyle'ın her öğesi bir tuple'dır ve aralarında virgül olmalıdır.
        # Bu bölüm, Table objesi oluşturulurken çağrılmalıdır.
        # Tek bir büyük tablo oluşturma ve içinde dinamik stiller uygulama biraz karmaşık.
        # En basit yol, her "period" için ayrı bir Table oluşturmak ve bunları peş peşe çizmek.
        # Veya, tüm veriyi tek bir listeye toplayıp tek bir büyük Table objesi oluşturmak
        # ve TableStyle'ı bu yapıya göre dinamik olarak belirlemek.

        # Önceki kodunuzdaki gibi tek bir büyük tablo yapısını koruyarak devam edelim.
        # Style nesnesi tanımları, dışarıdan gelmeliydi, veya burada ParagraphStyle olarak tanımlanmalıydı.

        # Her bir kategori bloğu için ayrı bir Table objesi oluşturalım
        current_y_for_section = y_start
        for period, items in data_dict.items():
            section_table_data = []
            # Kategori başlığı
            section_table_data.append([Paragraph(f"--- {period} Gün ---", style_header_paragraph), "", ""])
            # Sütun başlıkları
            section_table_data.append(header_row_paragraphs)

            if not items:
                section_table_data.append([Paragraph("Bu Kategori Boş", style_normal_paragraph), "", ""])
            else:
                for item in items:
                    section_table_data.append([
                        Paragraph(str(item[1]), style_normal_paragraph),
                        Paragraph(format_currency_func(item[2]), style_right_paragraph),
                        Paragraph(str(item[3]), style_right_paragraph)
                    ])

            section_table = Table(section_table_data, colWidths=col_widths)
            table_style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E0E0E0")), # Kategori başlığı
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), TURKISH_FONT_BOLD), # Doğrudan global değişkeni kullanın
                ('FONTSIZE', (0,0), (-1,0), 8),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('ALIGN', (0,1), (-1,1), 'CENTER'), # Sütun başlıkları
                ('FONTNAME', (0,1), (-1,1), TURKISH_FONT_BOLD), # Doğrudan global değişkeni kullanın
                ('FONTSIZE', (0,1), (-1,1), 8),
                ('BACKGROUND', (0,1), (-1,1), colors.HexColor("#4F81BD")), # Sütun başlıkları arka plan
                ('TEXTCOLOR', (0,1), (-1,1), colors.whitesmoke),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTNAME', (0,2), (-1,-1), TURKISH_FONT_NORMAL), # Veri satırları font - Doğrudan global değişkeni kullanın
                ('ALIGN', (1,2), (1,-1), 'RIGHT'), # Tutar sağa hizalı
                ('ALIGN', (2,2), (2,-1), 'RIGHT'), # Vadesi Geçen Gün sağa hizalı
            ])
            section_table.setStyle(table_style)

            table_height = section_table.wrapOn(canvas, page_width - 80, current_y_for_section)[1]
            if current_y_for_section - table_height < 50: # Sayfa sonuna çok yakınsa
                canvas.showPage()
                current_y_for_section = page_width - 40 # Yeni sayfanın başı (yatay sayfa için)
                canvas.setFont(TURKISH_FONT_BOLD, 14) # Doğrudan global değişkeni kullanın
                canvas.drawCentredString(page_width/2, current_y_for_section, f"{section_title} (Devam)")
                current_y_for_section -= 20
                canvas.setFont(TURKISH_FONT_NORMAL, 9) # Doğrudan global değişkeni kullanın
                canvas.drawCentredString(page_width/2, current_y_for_section, self.db.sirket_bilgileri.get("sirket_adi", ""))
                current_y_for_section -= 30
                table_height = section_table.wrapOn(canvas, page_width - 80, current_y_for_section)[1]

            section_table.drawOn(canvas, 40, current_y_for_section - table_height)
            current_y_for_section -= table_height + 10 # Tablo sonrası boşluk

        return current_y_for_section

class CriticalStockWarningPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.title("Kritik Stok Uyarısı ve Sipariş Önerisi")
        self.geometry("800x500")
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text="Kritik Stoktaki Ürünler", font=("Segoe UI", 16, "bold")).pack(pady=10, anchor=tk.W, padx=10)

        info_frame = ttk.Frame(self, padding="10")
        info_frame.pack(fill=tk.X, padx=10)
        ttk.Label(info_frame, text="Minimum stok seviyesinin altında olan ürünler listelenmiştir. İstenilen stok seviyesine ulaşmak için önerilen miktarları sipariş edebilirsiniz.").pack(anchor=tk.W)

        # Kritik Stok Listesi (Treeview)
        tree_frame = ttk.Frame(self, padding="10")
        tree_frame.pack(expand=True, fill=tk.BOTH)

        cols = ("Ürün Kodu", "Ürün Adı", "Mevcut Stok", "Min. Stok", "Fark", "Önerilen Sipariş Mik.")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode="none") # Seçim olmasın

        col_defs = [
            ("Ürün Kodu", 100, tk.W, tk.NO),
            ("Ürün Adı", 250, tk.W, tk.YES),
            ("Mevcut Stok", 100, tk.E, tk.NO),
            ("Min. Stok", 100, tk.E, tk.NO),
            ("Fark", 80, tk.E, tk.NO),
            ("Önerilen Sipariş Mik.", 150, tk.E, tk.NO)
        ]
        for cn,w,a,s in col_defs:
            self.tree.column(cn, width=w, anchor=a, stretch=s)
            self.tree.heading(cn, text=cn, command=lambda _c=cn: sort_treeview_column(self.tree, _c, False)) # Sıralama eklendi

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(expand=True, fill=tk.BOTH)
        
        self.load_critical_stock()

        # Butonlar
        button_frame = ttk.Frame(self, padding="10")
        button_frame.pack(fill=tk.X)
        ttk.Button(button_frame, text="Yenile", command=self.load_critical_stock, style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Sipariş Oluştur", command=self._siparis_olustur_critical_stock, style="Accent.TButton").pack(side=tk.RIGHT, padx=5) # state=tk.DISABLED kaldırıldı
        ttk.Button(button_frame, text="Kapat", command=self.destroy).pack(side=tk.RIGHT)

    def load_critical_stock(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        critical_items = self.db.get_critical_stock_items()
        if not critical_items:
            self.tree.insert("", tk.END, values=("", "", "", "", "", "Kritik Stokta Ürün Bulunmuyor."))
            self.app.set_status("Kritik stokta ürün bulunmuyor.")
            return

        for item in critical_items:
            # item: (id, urun_kodu, urun_adi, stok_miktari, alis_fiyati, satis_fiyati, kdv_orani, min_stok_seviyesi)
            urun_kodu = item[1]
            urun_adi = item[2]
            mevcut_stok = item[3]
            min_stok = item[7]
            fark = min_stok - mevcut_stok
            onerilen_siparis = fark # Basitçe fark kadar öner

            self.tree.insert("", tk.END, values=(
                urun_kodu,
                urun_adi,
                f"{mevcut_stok:.2f}".rstrip('0').rstrip('.'),
                f"{min_stok:.2f}".rstrip('0').rstrip('.'),
                f"{fark:.2f}".rstrip('0').rstrip('.'),
                f"{onerilen_siparis:.2f}".rstrip('0').rstrip('.')
            ))
        self.app.set_status(f"{len(critical_items)} ürün kritik stok seviyesinin altında.")

    def create_purchase_order_placeholder(self):
        messagebox.showinfo("Bilgi", "Bu özellik henüz geliştirilmedi.\nSeçili ürünler için otomatik sipariş oluşturma fonksiyonu gelecekte eklenecektir.", parent=self)

    def _siparis_olustur_critical_stock(self):
        """
        Kritik stoktaki ürünleri toplar ve tedarikçi seçimi sonrası alış faturası oluşturma akışını başlatır.
        """
        from arayuz import TedarikciSecimDialog
        urunler_for_siparis = []
        all_critical_items_db = self.db.get_critical_stock_items() # Güncel listeyi al
        for item_db in all_critical_items_db:
            # item_db: (id, urun_kodu, urun_adi, stok_miktari, alis_fiyati, satis_fiyati, kdv_orani, min_stok_seviyesi, alis_fiyati_kdv_dahil)
            urun_id = item_db[0]
            urun_kodu_db = item_db[1]
            urun_adi_db = item_db[2]
            onerilen_miktar = item_db[7] - item_db[3] # Min. Stok - Mevcut Stok = Önerilen Miktar
            
            if onerilen_miktar > 0:
                urunler_for_siparis.append({
                    "id": urun_id,
                    "kodu": urun_kodu_db,
                    "adi": urun_adi_db,
                    "miktar": onerilen_miktar, # Miktar olarak gönderiyoruz
                    "alis_fiyati_kdv_haric": item_db[4], # KDV hariç alış fiyatı
                    "kdv_orani": item_db[6],   # KDV oranı
                    "alis_fiyati_kdv_dahil": item_db[8] # KDV dahil alış fiyatı
                })

        if not urunler_for_siparis:
            messagebox.showinfo("Bilgi", "Sipariş oluşturmak için kritik stokta ürün bulunmuyor.", parent=self)
            return

        # Tedarikçi Seçim Diyaloğunu aç ve callback'i _tedarikci_secildi_ve_siparis_olustur olarak ayarla
        from arayuz import TedarikciSecimDialog # Lokal import
        TedarikciSecimDialog(self, self.db, 
                             lambda selected_tedarikci_id, selected_tedarikci_ad: 
                             self._tedarikci_secildi_ve_siparis_olustur(selected_tedarikci_id, selected_tedarikci_ad, urunler_for_siparis))

    def _tedarikci_secildi_ve_siparis_olustur(self, tedarikci_id, tedarikci_ad, urunler_for_siparis):
        """
        Tedarikçi seçildikten sonra çağrılır. Alış siparişi oluşturma sayfasını başlatır.
        """
        if tedarikci_id:
            # Artık alış faturası değil, tedarikçi siparişi modülünü çağırıyoruz
            self.app.tedarikci_siparisi_goster(initial_cari_id=tedarikci_id, initial_urunler=urunler_for_siparis)
            self.app.set_status(f"'{tedarikci_ad}' için tedarikçi siparişi oluşturma ekranı açıldı.")
            self.destroy() # Kritik stok uyarısı penceresini kapat
        else:
            self.app.set_status("Tedarikçi seçimi iptal edildi. Sipariş oluşturulmadı.")
            messagebox.showwarning("İptal Edildi", "Tedarikçi seçimi yapılmadığı için sipariş oluşturma işlemi iptal edildi.", parent=self)



class NotificationDetailsPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager, notifications_data):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.notifications_data = notifications_data # Bildirim verileri
        self.title("Aktif Bildirim Detayları")
        self.geometry("900x600")
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text="Aktif Bildirim Detayları", font=("Segoe UI", 16, "bold")).pack(pady=10, anchor=tk.W, padx=10)

        self.notebook_details = ttk.Notebook(self)
        self.notebook_details.pack(expand=True, fill=tk.BOTH, padx=10, pady=5)

        # Kritik Stok Sekmesi
        if 'critical_stock' in self.notifications_data:
            critical_stock_frame = ttk.Frame(self.notebook_details, padding="10")
            self.notebook_details.add(critical_stock_frame, text="📦 Kritik Stok")
            self._create_critical_stock_tab(critical_stock_frame, self.notifications_data['critical_stock'])

        # Vadesi Geçmiş Alacaklar Sekmesi
        if 'overdue_receivables' in self.notifications_data:
            overdue_receivables_frame = ttk.Frame(self.notebook_details, padding="10")
            self.notebook_details.add(overdue_receivables_frame, text="💰 Vadesi Geçmiş Alacaklar")
            self._create_overdue_receivables_tab(overdue_receivables_frame, self.notifications_data['overdue_receivables'])

        # Vadesi Geçmiş Borçlar Sekmesi
        if 'overdue_payables' in self.notifications_data:
            overdue_payables_frame = ttk.Frame(self.notebook_details, padding="10")
            self.notebook_details.add(overdue_payables_frame, text="จ่าย Vadesi Geçmiş Borçlar")
            self._create_overdue_payables_tab(overdue_payables_frame, self.notifications_data['overdue_payables'])

        ttk.Button(self, text="Kapat", command=self.destroy).pack(pady=10)

    def _create_critical_stock_tab(self, parent_frame, data):
        cols = ("Ürün Kodu", "Ürün Adı", "Mevcut Stok", "Min. Stok", "Fark", "Önerilen Sipariş Mik.")
        tree = ttk.Treeview(parent_frame, columns=cols, show='headings', selectmode="none")

        col_defs = [
            ("Ürün Kodu", 100, tk.W, tk.NO),
            ("Ürün Adı", 250, tk.W, tk.YES),
            ("Mevcut Stok", 100, tk.E, tk.NO),
            ("Min. Stok", 100, tk.E, tk.NO),
            ("Fark", 80, tk.E, tk.NO),
            ("Önerilen Sipariş Mik.", 150, tk.E, tk.NO)
        ]
        for cn,w,a,s in col_defs:
            tree.column(cn, width=w, anchor=a, stretch=s)
            tree.heading(cn, text=cn)

        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(expand=True, fill=tk.BOTH)

        for item in data:
            # item: (id, urun_kodu, urun_adi, stok_miktari, alis_fiyati, satis_fiyati, kdv_orani, min_stok_seviyesi)
            urun_kodu = item[1]
            urun_adi = item[2]
            mevcut_stok = item[3]
            min_stok = item[7]
            fark = min_stok - mevcut_stok
            onerilen_siparis = fark
            tree.insert("", tk.END, values=(
                urun_kodu, urun_adi,
                f"{mevcut_stok:.2f}".rstrip('0').rstrip('.'),
                f"{min_stok:.2f}".rstrip('0').rstrip('.'),
                f"{fark:.2f}".rstrip('0').rstrip('.'),
                f"{onerilen_siparis:.2f}".rstrip('0').rstrip('.')
            ))

    def _create_overdue_receivables_tab(self, parent_frame, data):
        cols = ("Müşteri Adı", "Net Borç", "Vadesi Geçen Gün")
        tree = ttk.Treeview(parent_frame, columns=cols, show='headings', selectmode="browse")

        col_defs = [
            ("Müşteri Adı", 250, tk.W, tk.YES),
            ("Net Borç", 120, tk.E, tk.NO),
            ("Vadesi Geçen Gün", 120, tk.E, tk.NO)
        ]
        for cn,w,a,s in col_defs:
            tree.column(cn, width=w, anchor=a, stretch=s)
            tree.heading(cn, text=cn)

        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(expand=True, fill=tk.BOTH)

        for item in data:
            # item: (cari_id, cari_adi, net_borc, vadesi_gecen_gun)
            tree.insert("", tk.END, values=(
                item[1], self.db._format_currency(item[2]), item[3]
            ))
        tree.bind("<Double-1>", lambda event: self._open_cari_ekstresi_from_notification(event, tree, 'MUSTERI'))

    def _create_overdue_payables_tab(self, parent_frame, data):
        cols = ("Tedarikçi Adı", "Net Borç", "Vadesi Geçen Gün")
        tree = ttk.Treeview(parent_frame, columns=cols, show='headings', selectmode="browse")

        col_defs = [
            ("Tedarikçi Adı", 250, tk.W, tk.YES),
            ("Net Borç", 120, tk.E, tk.NO),
            ("Vadesi Geçen Gün", 120, tk.E, tk.NO)
        ]
        for cn,w,a,s in col_defs:
            tree.column(cn, width=w, anchor=a, stretch=s)
            tree.heading(cn, text=cn)

        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(expand=True, fill=tk.BOTH)

        for item in data:
            # item: (cari_id, cari_adi, net_borc, vadesi_gecen_gun)
            tree.insert("", tk.END, values=(
                item[1], self.db._format_currency(item[2]), item[3]
            ))
        tree.bind("<Double-1>", lambda event: self._open_cari_ekstresi_from_notification(event, tree, 'TEDARIKCI'))

    def _open_cari_ekstresi_from_notification(self, event, tree, cari_tip):
        selected_item = tree.focus()
        if not selected_item:
            return
        
        item_values = tree.item(selected_item, 'values')
        cari_adi = item_values[0] # İlk sütun cari adı
        
        # Bildirim verilerinden ilgili cari ID'yi bul
        cari_id = None
        if cari_tip == 'MUSTERI':
            for item in self.notifications_data.get('overdue_receivables', []):
                if item[1] == cari_adi:
                    cari_id = item[0]
                    break
        elif cari_tip == 'TEDARIKCI':
            for item in self.notifications_data.get('overdue_payables', []):
                if item[1] == cari_adi:
                    cari_id = item[0]
                    break
        
        if cari_id:
            CariHesapEkstresiPenceresi(self.app, self.db, cari_id, cari_tip, cari_adi)
        else:
            messagebox.showwarning("Hata", "Cari ID bulunamadı.", parent=self)

class NakitAkisRaporuPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.title("Nakit Akış Raporu")
        self.geometry("1000x700")
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text="Nakit Akış Raporu", font=("Segoe UI", 16, "bold")).pack(pady=10, anchor=tk.W, padx=10)

        # Filtreleme Çerçevesi
        filter_frame = ttk.Frame(self, padding="10")
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(filter_frame, text="Başlangıç Tarihi:").pack(side=tk.LEFT, padx=(0,2))
        self.bas_tarih_entry = ttk.Entry(filter_frame, width=12)
        self.bas_tarih_entry.pack(side=tk.LEFT, padx=(0,5))
        self.bas_tarih_entry.insert(0, (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        setup_date_entry(self.app, self.bas_tarih_entry)
        # Başlangıç tarihi için takvim butonu
        ttk.Button(filter_frame, text="🗓️", command=lambda: self._open_date_picker(self.bas_tarih_entry), width=3).pack(side=tk.LEFT, padx=2)

        ttk.Label(filter_frame, text="Bitiş Tarihi:").pack(side=tk.LEFT, padx=(0,2))
        self.bit_tarih_entry = ttk.Entry(filter_frame, width=12)
        self.bit_tarih_entry.pack(side=tk.LEFT, padx=(0,10))
        self.bit_tarih_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        setup_date_entry(self.app, self.bit_tarih_entry)
        # Bitiş tarihi için takvim butonu
        ttk.Button(filter_frame, text="🗓️", command=lambda: self._open_date_picker(self.bit_tarih_entry), width=3).pack(side=tk.LEFT, padx=2)

        ttk.Button(filter_frame, text="Filtrele/Yenile", command=self.nakit_akis_listesini_yukle, style="Accent.TButton").pack(side=tk.LEFT, padx=(10,0))
        ttk.Button(filter_frame, text="Excel'e Aktar", command=self.excel_aktar).pack(side=tk.RIGHT, padx=5)
        ttk.Button(filter_frame, text="PDF'e Aktar", command=self.pdf_aktar).pack(side=tk.RIGHT, padx=5)


        # Nakit Akış Listesi (Treeview)
        tree_frame = ttk.Frame(self, padding="10")
        tree_frame.pack(expand=True, fill=tk.BOTH)

        cols = ("Tarih", "Tip", "Tutar", "Açıklama", "Hesap Adı", "Hesap Tipi", "Kaynak")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings', selectmode="browse")

        col_defs = [
            ("Tarih", 90, tk.CENTER, tk.NO),
            ("Tip", 70, tk.CENTER, tk.NO), # GELİR/GİDER
            ("Tutar", 120, tk.E, tk.NO),
            ("Açıklama", 350, tk.W, tk.YES),
            ("Hesap Adı", 150, tk.W, tk.NO),
            ("Hesap Tipi", 80, tk.CENTER, tk.NO), # KASA/BANKA
            ("Kaynak", 100, tk.W, tk.NO) # FATURA, TAHSILAT, ODEME, MANUEL
        ]
        for cn,w,a,s in col_defs:
            self.tree.column(cn, width=w, anchor=a, stretch=s)
            self.tree.heading(cn, text=cn, command=lambda _c=cn: sort_treeview_column(self.tree, _c, False))

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(expand=True, fill=tk.BOTH)

        # Toplamlar Alanı
        summary_frame = ttk.Frame(self, padding="10")
        summary_frame.pack(fill=tk.X, side=tk.BOTTOM)
        self.toplam_gelir_label = ttk.Label(summary_frame, text="Toplam Gelir: 0.00 TL", font=("Segoe UI", 10, "bold"))
        self.toplam_gelir_label.pack(side=tk.LEFT, padx=10)
        self.toplam_gider_label = ttk.Label(summary_frame, text="Toplam Gider: 0.00 TL", font=("Segoe UI", 10, "bold"))
        self.toplam_gider_label.pack(side=tk.LEFT, padx=10)
        self.net_akis_label = ttk.Label(summary_frame, text="Net Nakit Akışı: 0.00 TL", font=("Segoe UI", 12, "bold"))
        self.net_akis_label.pack(side=tk.RIGHT, padx=10)

        # Kasa/Banka Bakiyeleri (Tek tanım)
        self.kasa_banka_bakiye_frame = ttk.LabelFrame(self, text="Kasa/Banka Güncel Bakiyeleri", padding="10")
        self.kasa_banka_bakiye_frame.pack(fill=tk.X, padx=10, pady=(0, 10), side=tk.BOTTOM)

        self.kasa_banka_bakiyeleri_labels = {} 
        ttk.Label(self.kasa_banka_bakiye_frame, text="Yükleniyor...", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)

        self.kayit_sayisi_per_sayfa = 20 # Her sayfada kaç kayıt gösterileceği (örnek değer)
        self.mevcut_sayfa = 1 # Başlangıç sayfası
        self.toplam_kayit_sayisi = 0 # Toplam kayıt sayısını tutacak

        pagination_frame = ttk.Frame(self, padding="10")
        pagination_frame.pack(fill=tk.X, padx=10, pady=5) # Treeview'in hemen altında

        ttk.Button(pagination_frame, text="Önceki Sayfa", command=self.onceki_sayfa).pack(side=tk.LEFT, padx=5)
        self.sayfa_bilgisi_label = ttk.Label(pagination_frame, text="Sayfa 1 / 1")
        self.sayfa_bilgisi_label.pack(side=tk.LEFT, padx=10)
        ttk.Button(pagination_frame, text="Sonraki Sayfa", command=self.sonraki_sayfa).pack(side=tk.LEFT, padx=5)

        self.nakit_akis_listesini_yukle() 
        self.guncelle_kasa_banka_bakiyeleri()

    def _open_date_picker(self, target_entry):
        """Bir Entry widget'ı için tarih seçici penceresi açar."""
        DatePickerDialog(self.app, target_entry)        

    def nakit_akis_listesini_yukle(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        bas_t = self.bas_tarih_entry.get()
        bit_t = self.bit_tarih_entry.get()

        try:
            if bas_t: datetime.strptime(bas_t, '%Y-%m-%d')
            if bit_t: datetime.strptime(bit_t, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Tarih Formatı Hatası", "Tarih formatı YYYY-AA-GG olmalıdır.", parent=self)
            return

        # Sayfalama mantığı
        self.toplam_kayit_sayisi = self.db.get_nakit_akis_count(bas_t, bit_t)
        toplam_sayfa = (self.toplam_kayit_sayisi + self.kayit_sayisi_per_sayfa - 1) // self.kayit_sayisi_per_sayfa
        if toplam_sayfa == 0:
            toplam_sayfa = 1

        if self.mevcut_sayfa > toplam_sayfa:
            self.mevcut_sayfa = toplam_sayfa
        
        offset = (self.mevcut_sayfa - 1) * self.kayit_sayisi_per_sayfa
        limit = self.kayit_sayisi_per_sayfa

        nakit_akis_verileri = self.db.get_nakit_akis_verileri(
            bas_t, bit_t, limit=limit, offset=offset
        )
        
        toplam_gelir = 0.0
        toplam_gider = 0.0

        for item in nakit_akis_verileri:
            # item: (tarih, tip, tutar, aciklama, hesap_adi, hesap_tipi, kaynak, kaynak_id)
            tarih_formatted = datetime.strptime(item[0], '%Y-%m-%d').strftime('%d.%m.%Y')
            tutar_formatted = self.db._format_currency(item[2])

            self.tree.insert("", tk.END, values=(
                tarih_formatted,
                item[1], # Tip (GELİR/GİDER)
                tutar_formatted,
                item[3], # Açıklama
                item[4], # Hesap Adı
                item[5], # Hesap Tipi
                item[6] # Kaynak
            ))
            
            if item[1] == 'GELİR':
                toplam_gelir += item[2]
            elif item[1] == 'GİDER':
                toplam_gider += item[2]

        self.toplam_gelir_label.config(text=f"Toplam Gelir: {self.db._format_currency(toplam_gelir)}")
        self.toplam_gider_label.config(text=f"Toplam Gider: {self.db._format_currency(toplam_gider)}")
        self.net_akis_label.config(text=f"Net Nakit Akışı: {self.db._format_currency(toplam_gelir - toplam_gider)}")
        
        self.app.set_status(f"Nakit Akış Raporu güncellendi ({len(nakit_akis_verileri)} kayıt). Toplam {self.toplam_kayit_sayisi} kayıt.") # Durum çubuğu
        self.sayfa_bilgisi_label.config(text=f"Sayfa {self.mevcut_sayfa} / {toplam_sayfa}") # Sayfa bilgisi

    def onceki_sayfa(self):
        if self.mevcut_sayfa > 1:
            self.mevcut_sayfa -= 1
            self.nakit_akis_listesini_yukle()

    def sonraki_sayfa(self):
        toplam_sayfa = (self.toplam_kayit_sayisi + self.kayit_sayisi_per_sayfa - 1) // self.kayit_sayisi_per_sayfa
        if toplam_sayfa == 0:
            toplam_sayfa = 1

        if self.mevcut_sayfa < toplam_sayfa:
            self.mevcut_sayfa += 1
            self.nakit_akis_listesini_yukle()

    def guncelle_kasa_banka_bakiyeleri(self):
        # Önceki bakiyeleri temizle
        for widget in self.kasa_banka_bakiye_frame.winfo_children():
            widget.destroy()

        hesaplar = self.db.get_tum_kasa_banka_bakiyeleri()
        if not hesaplar:
            ttk.Label(self.kasa_banka_bakiye_frame, text="Kasa/Banka Hesabı Bulunamadı.", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
            return

        for h_id, h_adi, bakiye, h_tip in hesaplar:
            bakiye_text = f"{h_adi} ({h_tip}): {self.db._format_currency(bakiye)}"
            ttk.Label(self.kasa_banka_bakiye_frame, text=bakiye_text, font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=10)

    def excel_aktar(self):
        dosya_yolu = filedialog.asksaveasfilename(
            initialfile=f"Nakit_Akis_Raporu_{datetime.now().strftime('%Y%m%d')}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyaları", "*.xlsx")],
            title="Nakit Akış Raporunu Excel'e Kaydet",
            parent=self
        )
        if dosya_yolu:
            bas_t = self.bas_tarih_entry.get()
            bit_t = self.bit_tarih_entry.get()

            bekleme_penceresi = BeklemePenceresi(self, message="Rapor Excel'e aktarılıyor, lütfen bekleyiniz...")
            threading.Thread(target=lambda: self._generate_excel_report_threaded(
                bas_t, bit_t, dosya_yolu, bekleme_penceresi
            )).start()
        else:
            self.app.set_status("Nakit Akış Raporu Excel'e aktarma iptal edildi.")

    def _generate_excel_report_threaded(self, bas_t, bit_t, dosya_yolu, bekleme_penceresi):
        success = False
        message = ""
        try:
            nakit_akis_verileri = self.db.get_nakit_akis_verileri(bas_t, bit_t)

            if not nakit_akis_verileri:
                message = "Excel'e aktarılacak nakit akış verisi bulunamadı."
                success = False
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Nakit_Akis_Raporu"

                headers = ["Tarih", "Tip", "Tutar", "Açıklama", "Hesap Adı", "Hesap Tipi", "Kaynak"]
                ws.append(headers)

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                for col_idx, header_text in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header_text) + 5

                for item in nakit_akis_verileri:
                    row_data = list(item)
                    try: row_data[0] = datetime.strptime(row_data[0], '%Y-%m-%d').strftime('%d.%m.%Y')
                    except: pass
                    ws.append(row_data[:-1])

                for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    if row_cells[2].value is not None: 
                        try:
                            val_float = float(row_cells[2].value)
                            row_cells[2].number_format = '#,##0.00₺'
                        except ValueError:
                            pass

                wb.save(dosya_yolu)
                success = True
                message = f"Nakit Akış Raporu başarıyla '{dosya_yolu}' adresine kaydedildi."
        except Exception as e:
            success = False
            message = f"Nakit Akış Raporu Excel'e aktarılırken bir hata oluştu:\n{e}"
            traceback.print_exc()
        finally:
            self.app.after(0, bekleme_penceresi.kapat)
            self.app.after(0, lambda: messagebox.showinfo("Excel Raporu", message, parent=self.app) if success else messagebox.showerror("Excel Raporu Hatası", message, parent=self.app))
            self.app.after(0, lambda: self.app.set_status(message))

    def pdf_aktar(self):
        dosya_yolu = filedialog.asksaveasfilename(
            initialfile=f"Nakit_Akis_Raporu_{datetime.now().strftime('%Y%m%d')}.pdf",
            defaultextension=".pdf",
            filetypes=[("PDF Dosyaları", "*.pdf")],
            title="Nakit Akış Raporunu PDF'e Kaydet",
            parent=self
        )
        if dosya_yolu:
            bas_t = self.bas_tarih_entry.get()
            bit_t = self.bit_tarih_entry.get()

            bekleme_penceresi = BeklemePenceresi(self, message="Rapor PDF'e aktarılıyor, lütfen bekleyiniz...")
            threading.Thread(target=lambda: self._generate_pdf_report_threaded(
                bas_t, bit_t, dosya_yolu, bekleme_penceresi
            )).start()
        else:
            self.app.set_status("Nakit Akış Raporu PDF'e aktarma iptal edildi.")

    def _generate_pdf_report_threaded(self, bas_t, bit_t, dosya_yolu, bekleme_penceresi):
        success = False
        message = ""
        try:
            nakit_akis_verileri = self.db.get_nakit_akis_verileri(bas_t, bit_t)

            if not nakit_akis_verileri:
                message = "PDF'e aktarılacak nakit akış verisi bulunamadı."
                success = False
            else:
                c = rp_canvas.Canvas(dosya_yolu, pagesize=landscape(A4))
                width, height = landscape(A4)

                styles = getSampleStyleSheet()
                styleN = styles['Normal']
                styleN.fontName = TURKISH_FONT_NORMAL # Doğrudan global değişkeni kullanın
                styleN.fontSize = 7
                styleH = styles['Normal']
                styleH.fontName = TURKISH_FONT_BOLD # Doğrudan global değişkeni kullanın
                styleH.fontSize = 7
                styleH.alignment = 1 # TA_CENTER
                styleRight = styles['Normal']
                styleRight.fontName = TURKISH_FONT_NORMAL # Doğrudan global değişkeni kullanın
                styleRight.fontSize = 7
                styleRight.alignment = 2 # TA_RIGHT

                # Başlık
                c.setFont(TURKISH_FONT_BOLD, 14) # Doğrudan global değişkeni kullanın
                c.drawCentredString(width/2, height - 40, f"Nakit Akış Raporu ({bas_t} - {bit_t})")
                c.setFont(TURKISH_FONT_NORMAL, 9) # Doğrudan global değişkeni kullanın
                c.drawCentredString(width/2, height - 55, self.db.sirket_bilgileri.get("sirket_adi", ""))
                y_pos = height - 80

                data = [
                    [Paragraph(h, styleH) for h in ["Tarih", "Tip", "Tutar", "Açıklama", "Hesap Adı", "Hesap Tipi", "Kaynak"]]
                ]

                toplam_gelir_rapor = 0
                toplam_gider_rapor = 0

                for item in nakit_akis_verileri:
                    tarih_f = datetime.strptime(item[0], '%Y-%m-%d').strftime('%d.%m.%y') if item[0] else '-'

                    data.append([
                        Paragraph(tarih_f, styleN),
                        Paragraph(str(item[1]), styleN),
                        Paragraph(self.db._format_currency(item[2]), styleRight),
                        Paragraph(str(item[3])[:50], styleN),
                        Paragraph(str(item[4]), styleN),
                        Paragraph(str(item[5]), styleN),
                        Paragraph(str(item[6]), styleN)
                    ])

                    if item[1] == 'GELİR':
                        toplam_gelir_rapor += item[2]
                    elif item[1] == 'GİDER':
                        toplam_gider_rapor += item[2]

                col_widths = [1.8*cm, 1.5*cm, 2.5*cm, 7*cm, 3*cm, 1.8*cm, 2*cm]

                rows_per_page = 30
                num_pages = (len(data) -1 + rows_per_page - 1) // rows_per_page
                if num_pages == 0: num_pages = 1

                for page_num in range(num_pages):
                    start_row = page_num * rows_per_page + (1 if page_num > 0 else 0)
                    end_row = min((page_num + 1) * rows_per_page, len(data) -1 )

                    page_data = [data[0]] + data[start_row+1 : end_row+1]
                    if not page_data[1:]:
                        if page_num > 0 : break
                        elif not data[1:]:
                                 page_data.append([Paragraph("Veri Yok", styleN)]*len(col_widths))


                    table = Table(page_data, colWidths=col_widths)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E78")),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                        ('ALIGN', (3,1), (3,-1), 'LEFT'),
                        ('ALIGN', (2,1), (2,-1), 'RIGHT'),
                        ('FONTNAME', (0,0), (-1,-1), TURKISH_FONT_NORMAL), # Doğrudan global değişkeni kullanın
                        ('FONTNAME', (0,0), (-1,0), TURKISH_FONT_BOLD), # Doğrudan global değişkeni kullanın
                        ('FONTSIZE', (0,0), (-1,-1), 7),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ]))

                    table.wrapOn(c, width - 80, height - 100)
                    table_h = table._height
                    table.drawOn(c, 40, y_pos - table_h)

                    c.setFont(TURKISH_FONT_NORMAL, 8) # Doğrudan global değişkeni kullanın
                    c.drawRightString(width - 40, 30, f"Sayfa {page_num + 1} / {num_pages}")

                    if page_num < num_pages - 1:
                        c.showPage()
                        c.setFont(TURKISH_FONT_BOLD, 14) # Doğrudan global değişkeni kullanın
                        c.drawCentredString(width/2, height - 40, f"Nakit Akış Raporu ({bas_t} - {bit_t}) - Devam")
                        c.setFont(TURKISH_FONT_NORMAL, 9) # Doğrudan global değişkeni kullanın
                        c.drawCentredString(width/2, height - 55, self.db.sirket_bilgileri.get("sirket_adi", ""))
                        y_pos = height - 80

                y_pos_summary = y_pos - table_h - 20
                if y_pos_summary < 80 :
                    c.showPage()
                    c.setFont(TURKISH_FONT_BOLD, 14) # Doğrudan global değişkeni kullanın
                    c.drawCentredString(width/2, height - 40, f"Nakit Akış Raporu ({bas_t} - {bit_t}) - Toplamlar")
                    y_pos_summary = height - 70
                    c.setFont(TURKISH_FONT_NORMAL, 8) # Doğrudan global değişkeni kullanın
                    c.drawRightString(width - 40, 30, f"Sayfa {num_pages} / {num_pages}")


                c.setFont(TURKISH_FONT_BOLD, 9) # Doğrudan global değişkeni kullanın
                c.drawRightString(width - 50, y_pos_summary, f"Toplam Gelir: {self.db._format_currency(toplam_gelir_rapor)}")
                y_pos_summary -= 15
                c.drawRightString(width - 50, y_pos_summary, f"Toplam Gider: {self.db._format_currency(toplam_gider_rapor)}")
                y_pos_summary -= 15
                c.setFont(TURKISH_FONT_BOLD, 10) # Doğrudan global değişkeni kullanın
                c.drawRightString(width - 50, y_pos_summary, f"Net Nakit Akışı: {self.db._format_currency(toplam_gelir_rapor - toplam_gider_rapor)}")

                c.save()
                success = True
                message = f"Nakit Akış Raporu PDF olarak '{dosya_yolu}' adresine kaydedildi."
        except Exception as e:
            success = False
            message = f"Nakit Akış Raporu PDF'e aktarılırken hata: {e}"
            traceback.print_exc()
        finally:
            self.app.after(0, bekleme_penceresi.kapat)
            self.app.after(0, lambda: messagebox.showinfo("PDF Raporu", message, parent=self.app) if success else messagebox.showerror("PDF Raporu Hatası", message, parent=self.app))
            self.app.after(0, lambda: self.app.set_status(message))

class KarZararRaporuPenceresi(tk.Toplevel):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.title("Kâr/Zarar Raporu")
        self.geometry("500x300")
        self.transient(parent_app)
        self.grab_set()

        ttk.Label(self, text="Kâr/Zarar Raporu", font=("Segoe UI", 16, "bold")).pack(pady=10, anchor=tk.W, padx=10)

        # Filtreleme Çerçevesi
        filter_frame = ttk.Frame(self, padding="10")
        filter_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(filter_frame, text="Başlangıç Tarihi:").pack(side=tk.LEFT, padx=(0,2))
        self.bas_tarih_entry = ttk.Entry(filter_frame, width=12)
        self.bas_tarih_entry.pack(side=tk.LEFT, padx=(0,5))
        self.bas_tarih_entry.insert(0, (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d'))
        setup_date_entry(self.app, self.bas_tarih_entry) 
        # Başlangıç tarihi için takvim butonu
        ttk.Button(filter_frame, text="🗓️", command=lambda: self._open_date_picker(self.bas_tarih_entry), width=3).pack(side=tk.LEFT, padx=2)

        ttk.Label(filter_frame, text="Bitiş Tarihi:").pack(side=tk.LEFT, padx=(0,2))
        self.bit_tarih_entry = ttk.Entry(filter_frame, width=12)
        self.bit_tarih_entry.pack(side=tk.LEFT, padx=(0,10))
        self.bit_tarih_entry.insert(0, datetime.now().strftime('%Y-%m-%d'))
        setup_date_entry(self.app, self.bit_tarih_entry)
        # Bitiş tarihi için takvim butonu
        ttk.Button(filter_frame, text="🗓️", command=lambda: self._open_date_picker(self.bit_tarih_entry), width=3).pack(side=tk.LEFT, padx=2)

        ttk.Button(filter_frame, text="Raporla", command=self.raporu_guncelle, style="Accent.TButton").pack(side=tk.LEFT, padx=(10,0))

        # Rapor Sonuçları Alanı
        results_frame = ttk.Frame(self, padding="10")
        results_frame.pack(expand=True, fill=tk.BOTH)

        self.toplam_gelir_label = ttk.Label(results_frame, text="Toplam Gelir: 0.00 TL", font=("Segoe UI", 12, "bold"))
        self.toplam_gelir_label.pack(pady=5, anchor=tk.W)

        self.toplam_gider_label = ttk.Label(results_frame, text="Toplam Gider: 0.00 TL", font=("Segoe UI", 12, "bold"))
        self.toplam_gider_label.pack(pady=5, anchor=tk.W)

        ttk.Separator(results_frame, orient='horizontal').pack(fill='x', pady=10)

        self.net_kar_zarar_label = ttk.Label(results_frame, text="Net Kâr/Zarar: 0.00 TL", font=("Segoe UI", 14, "bold"))
        self.net_kar_zarar_label.pack(pady=5, anchor=tk.W)

        self.raporu_guncelle() # İlk yüklemede raporu oluştur

    def _open_date_picker(self, target_entry):
        """Bir Entry widget'ı için tarih seçici penceresi açar."""
        DatePickerDialog(self.app, target_entry)

    def raporu_guncelle(self):
        bas_t = self.bas_tarih_entry.get()
        bit_t = self.bit_tarih_entry.get()

        try:
            if bas_t: datetime.strptime(bas_t, '%Y-%m-%d')
            if bit_t: datetime.strptime(bit_t, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Tarih Formatı Hatası", "Tarih formatı YYYY-AA-GG olmalıdır.", parent=self)
            return

        toplam_gelir, toplam_gider = self.db.get_kar_zarar_verileri(bas_t, bit_t)
        net_kar_zarar = toplam_gelir - toplam_gider

        self.toplam_gelir_label.config(text=f"Toplam Gelir: {self.db._format_currency(toplam_gelir)}")
        self.toplam_gider_label.config(text=f"Toplam Gider: {self.db._format_currency(toplam_gider)}")
        self.net_kar_zarar_label.config(text=f"Net Kâr/Zarar: {self.db._format_currency(net_kar_zarar)}")

        self.app.set_status(f"Kâr/Zarar Raporu güncellendi ({bas_t} - {bit_t}).")
