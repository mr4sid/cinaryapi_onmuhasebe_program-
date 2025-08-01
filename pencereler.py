# pencereler.py Dosyasının. Tamamım.

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QMessageBox, QFileDialog,
    QWidget, QMenuBar, QStatusBar, QDialog, QPushButton, QVBoxLayout,
    QHBoxLayout, QGridLayout, QLabel, QLineEdit, QComboBox,
    QTreeWidget, QTreeWidgetItem, QAbstractItemView, QHeaderView, QTextEdit,
    QCheckBox, QFrame, QTableWidget, QTableWidgetItem, QGroupBox,
    QMenu, QTabWidget,QSizePolicy, QProgressBar, QListWidget, QListWidgetItem )
from PySide6.QtGui import QFont, QPixmap, QImage, QDoubleValidator, QIntValidator, QBrush, QColor
from PySide6.QtCore import Qt, QDate, QTimer, Signal, Slot
import requests # Bu import, bazı eski direct request'ler için kalmış olabilir, ama kullanılmamalı.
from datetime import datetime, date, timedelta
import os
import shutil
import traceback
import threading
import calendar
import multiprocessing
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill 
from veritabani import OnMuhasebe
# AŞAMA 1 - Adım 1.2 için eklendi: CariService
# AŞAMA 1 - Adım 1.1 için güncellendi: format_and_validate_numeric_input
from hizmetler import FaturaService, TopluIslemService, CariService # <-- BU SATIR GÜNCELLENDİ
from yardimcilar import DatePickerDialog, normalize_turkish_chars, setup_locale, format_and_validate_numeric_input # <-- BU SATIR GÜNCELLENDİ
from config import API_BASE_URL # Bu UI tarafında doğrudan kullanılmamalı, OnMuhasebe sınıfı kullanmalı
# Logger kurulumu
logger = logging.getLogger(__name__)
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)

def setup_numeric_entry(parent_app, entry_widget, allow_negative=False, decimal_places=2, max_value=None): # parent_app parametresi eklendi
    validator = QDoubleValidator()
    validator.setBottom(0.0 if not allow_negative else -999999999.0)
    validator.setTop(999999999.0 if max_value is None else float(max_value))
    validator.setDecimals(decimal_places)
    validator.setNotation(QDoubleValidator.StandardNotation)
    entry_widget.setValidator(validator)
    # textChanged sinyaline bağlanırken lambda içinde line_edit ve decimals argümanlarını geçirin
    # decimal_places parametresi artık format_and_validate_numeric_input içinde kullanılmayacak.
    # Formatlama locale ayarına göre 2 ondalık basamakla yapılacak.
    entry_widget.textChanged.connect(lambda: format_and_validate_numeric_input(entry_widget, parent_app))
    entry_widget.editingFinished.connect(lambda: format_and_validate_numeric_input(entry_widget, parent_app))    
def setup_date_entry(parent_app, entry_widget):
    pass

class SiparisPenceresi(QDialog):
    def __init__(self, parent, db_manager, app_ref, siparis_tipi, siparis_id_duzenle=None, yenile_callback=None, initial_cari_id=None, initial_urunler=None, initial_data=None):
        super().__init__(parent)
        self.app = app_ref
        self.db = db_manager
        
        self.siparis_tipi = siparis_tipi
        self.siparis_id_duzenle = siparis_id_duzenle
        self.yenile_callback = yenile_callback
        self.initial_cari_id = initial_cari_id
        self.initial_urunler = initial_urunler
        self.initial_data = initial_data

        title = "Yeni Sipariş"
        if siparis_id_duzenle:
            try:
                siparis_info = self.db.siparis_getir_by_id(siparis_id_duzenle)
                siparis_no_display = siparis_info.get('siparis_no', 'Bilinmiyor')
                title = f"Sipariş Güncelleme: {siparis_no_display}"
            except Exception as e:
                logging.error(f"Sipariş bilgisi çekilirken hata: {e}")
                QMessageBox.critical(self, "Hata", "Sipariş bilgisi yüklenirken hata oluştu.")
                title = "Sipariş Güncelleme: Hata"
        else:
            title = "Yeni Müşteri Siparişi" if siparis_tipi == "SATIŞ_SIPARIS" else "Yeni Tedarikçi Siparişi"

        self.setWindowTitle(title)
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)

        dialog_layout = QVBoxLayout(self)

        from arayuz import SiparisOlusturmaSayfasi
        self.siparis_form = SiparisOlusturmaSayfasi(
            self,
            self.db,
            self.app,
            self.siparis_tipi,
            duzenleme_id=self.siparis_id_duzenle,
            yenile_callback=self.yenile_callback,
            initial_cari_id=self.initial_cari_id,
            initial_urunler=self.initial_urunler,
            initial_data=self.initial_data
        )
        dialog_layout.addWidget(self.siparis_form)

        self.siparis_form.saved_successfully.connect(self.accept)
        self.siparis_form.cancelled_successfully.connect(self.reject)

        self.finished.connect(self.on_dialog_finished)

    def on_dialog_finished(self, result):
        if result == QDialog.Rejected and self.siparis_id_duzenle is None:
            self.siparis_form._save_current_form_data_to_temp()
        
        if self.yenile_callback:
            self.yenile_callback()

# pencereler.py dosyasındaki CariHesapEkstresiPenceresi sınıfının TAMAMI

class CariHesapEkstresiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, cari_id, cari_tip, pencere_basligi, parent_list_refresh_func=None):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.cari_id = cari_id
        self.cari_tip = cari_tip
        self.cari_ad_gosterim = pencere_basligi
        self.parent_list_refresh_func = parent_list_refresh_func
        self.hareket_detay_map = {}

        self.setWindowTitle(f"Cari Hesap Ekstresi: {self.cari_ad_gosterim}")
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)

        main_container = QWidget(self)
        self.setLayout(QVBoxLayout(main_container))
        
        self.ozet_ve_bilgi_frame = QGroupBox("Cari Özet Bilgileri", self)
        self.layout().addWidget(self.ozet_ve_bilgi_frame)
        self._create_ozet_bilgi_alani()

        self.notebook = QTabWidget(self)
        self.layout().addWidget(self.notebook)
        self.notebook.currentChanged.connect(self._on_tab_change)

        self.hesap_hareketleri_tab = QWidget(self.notebook)
        self.notebook.addTab(self.hesap_hareketleri_tab, "Hesap Hareketleri")
        self._create_hesap_hareketleri_tab()

        self.siparisler_tab = QWidget(self.notebook)
        self.notebook.addTab(self.siparisler_tab, "Siparişler")
        self._create_siparisler_tab()

        self.hizli_islemler_ana_frame = QFrame(self)
        self.layout().addWidget(self.hizli_islemler_ana_frame)
        self._create_hizli_islem_alanlari()

        today = date.today()
        start_date = today - timedelta(days=3 * 365) if self.cari_tip == "TEDARIKCI" else today - timedelta(days=6 * 30)
        self.bas_tarih_entry.setText(start_date.strftime('%Y-%m-%d'))
        self.bit_tarih_entry.setText(today.strftime('%Y-%m-%d'))

        self._yukle_ozet_bilgileri()
        self.ekstreyi_yukle()

        self.finished.connect(self.on_dialog_finished)
        self.app.register_cari_ekstre_window(self)

    def on_dialog_finished(self, result):
        self.app.unregister_cari_ekstre_window(self)
        if self.parent_list_refresh_func:
            self.parent_list_refresh_func()

    def _on_tab_change(self, index):
        selected_tab_text = self.notebook.tabText(index)
        if selected_tab_text == "Siparişler":
            self._siparisleri_yukle()
        elif selected_tab_text == "Hesap Hareketleri":
            self.ekstreyi_yukle()

    def _yukle_cari_bilgileri(self):
        try:
            cari_adi = "Bilinmiyor"
            cari_telefon = ""
            
            if self.cari_tip == "MUSTERI":
                cari_data = self.db.musteri_getir_by_id(self.cari_id)
                if cari_data:
                    cari_adi = cari_data.get("ad", "Bilinmeyen Müşteri")
                    cari_telefon = cari_data.get("telefon", "")
                
            elif self.cari_tip == "TEDARIKCI":
                cari_data = self.db.tedarikci_getir_by_id(self.cari_id)
                if cari_data:
                    cari_adi = cari_data.get("ad", "Bilinmeyen Tedarikçi")
                    cari_telefon = cari_data.get("telefon", "")
            
            self.setWindowTitle(f"{cari_adi} - Cari Hesap Ekstresi")
        except Exception as e:
            logger.error(f"Cari bilgileri yüklenirken hata oluştu: {e}")
            QMessageBox.warning(self, "Hata", f"Cari bilgileri yüklenirken bir hata oluştu: {e}")

    def _create_hesap_hareketleri_tab(self):
        parent_frame = self.hesap_hareketleri_tab
        parent_frame.setLayout(QVBoxLayout(parent_frame))
        
        filter_frame = QFrame(parent_frame)
        parent_frame.layout().addWidget(filter_frame)
        self._create_filter_alani(filter_frame)

        tree_frame = QFrame(parent_frame)
        parent_frame.layout().addWidget(tree_frame)
        self._create_treeview_alani(tree_frame)

    def _create_siparisler_tab(self):
        parent_frame = self.siparisler_tab
        parent_frame.setLayout(QVBoxLayout(parent_frame))
        
        cols = ("ID", "Sipariş No", "Tarih", "Teslimat Tarihi", "Toplam Tutar", "Durum", "Fatura No")
        self.siparisler_tree = QTreeWidget(parent_frame)
        self.siparisler_tree.setHeaderLabels(cols)
        self.siparisler_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.siparisler_tree.setSortingEnabled(True)

        col_defs = [
            ("ID", 40, Qt.AlignCenter), ("Sipariş No", 150, Qt.AlignCenter), ("Tarih", 100, Qt.AlignCenter),
            ("Teslimat Tarihi", 100, Qt.AlignCenter), ("Toplam Tutar", 120, Qt.AlignRight), ("Durum", 120, Qt.AlignCenter),
            ("Fatura No", 150, Qt.AlignCenter)
        ]
        for i, (col_id, w, a) in enumerate(col_defs):
            self.siparisler_tree.setColumnWidth(i, w)
            self.siparisler_tree.headerItem().setTextAlignment(i, a)
            self.siparisler_tree.headerItem().setFont(i, QFont("Segoe UI", 9, QFont.Bold))

        self.siparisler_tree.header().setStretchLastSection(False)
        self.siparisler_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        
        parent_frame.layout().addWidget(self.siparisler_tree)
        self.siparisler_tree.itemDoubleClicked.connect(self._on_siparis_double_click)

    def _siparisleri_yukle(self):
        self.siparisler_tree.clear()
        
        try:
            params = {
                'cari_id': self.cari_id,
                'cari_turu': self.cari_tip
            }
            siparisler_data_response = self.db.siparis_listesi_al(**params) 
            siparisler_data = siparisler_data_response.get("items", []) 

            for siparis in siparisler_data:
                item_qt = QTreeWidgetItem(self.siparisler_tree)
                item_qt.setData(0, Qt.UserRole, siparis.get('id', -1))

                tarih_obj = datetime.strptime(str(siparis.get('tarih')), '%Y-%m-%d').date() if siparis.get('tarih') else None
                teslimat_tarihi_obj = datetime.strptime(str(siparis.get('teslimat_tarihi')), '%Y-%m-%d').date() if siparis.get('teslimat_tarihi') else None
                
                formatted_tarih = tarih_obj.strftime('%d.%m.%Y') if isinstance(tarih_obj, date) else '-'
                formatted_teslimat_tarihi = teslimat_tarihi_obj.strftime('%d.%m.%Y') if isinstance(teslimat_tarihi_obj, date) else '-'

                item_qt.setText(0, str(siparis.get('id', '')))
                item_qt.setText(1, siparis.get('siparis_no', ''))
                item_qt.setText(2, formatted_tarih)
                item_qt.setText(3, formatted_teslimat_tarihi)
                item_qt.setText(4, self.db._format_currency(siparis.get('toplam_tutar', 0.0)))
                item_qt.setText(5, siparis.get('durum', ''))
                
                fatura_no_text = "-"
                if siparis.get('fatura_id'):
                    try:
                        fatura_data = self.db.fatura_getir_by_id(siparis.get('fatura_id'))
                        fatura_no_text = fatura_data.get('fatura_no', '-')
                    except Exception:
                        fatura_no_text = "Hata"
                item_qt.setText(6, fatura_no_text)

                if siparis.get('durum') == "TAMAMLANDI":
                    for col_idx in range(self.siparisler_tree.columnCount()):
                        item_qt.setBackground(col_idx, QBrush(QColor("lightgreen")))
                elif siparis.get('durum') == "İPTAL_EDİLDİ":
                    for col_idx in range(self.siparisler_tree.columnCount()):
                        item_qt.setBackground(col_idx, QBrush(QColor("lightgray")))
                        item_qt.setForeground(col_idx, QBrush(QColor("gray")))
                        font = item_qt.font(col_idx)
                        font.setStrikeOut(True)
                        item_qt.setFont(col_idx, font)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Siparişler yüklenirken hata: {e}")
            logging.error(f"Cari Hesap Ekstresi - Siparişler yükleme hatası: {e}", exc_info=True)
        self.app.set_status_message(f"{self.cari_ad_gosterim} için {len(siparisler_data)} sipariş listelendi.", "blue")

    def _on_siparis_double_click(self, item, column):
        siparis_id = item.data(0, Qt.UserRole)
        if siparis_id:
            from pencereler import SiparisDetayPenceresi
            SiparisDetayPenceresi(self.app, self.db, siparis_id).exec()

    def _create_ozet_bilgi_alani(self):
        frame = self.ozet_ve_bilgi_frame
        frame.setLayout(QGridLayout(frame))

        label_font_buyuk = QFont("Segoe UI", 12, QFont.Bold)
        deger_font_buyuk = QFont("Segoe UI", 12)
        label_font_kucuk = QFont("Segoe UI", 9, QFont.Bold)
        deger_font_kucuk = QFont("Segoe UI", 9)

        finans_ozet_cerceve = QGroupBox("Finansal Özet", frame)
        finans_ozet_cerceve.setLayout(QGridLayout(finans_ozet_cerceve))
        frame.layout().addWidget(finans_ozet_cerceve, 0, 0)

        row_idx_finans = 0
        finans_ozet_cerceve.layout().addWidget(QLabel("Dönem Başı Bakiye:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_donem_basi_bakiye = QLabel("0,00 TL", font=deger_font_kucuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_donem_basi_bakiye, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Toplam Borç Hareketi:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_toplam_borc_hareketi = QLabel("0,00 TL", font=deger_font_kucuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_toplam_borc_hareketi, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Toplam Alacak Hareketi:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_toplam_alacak_hareketi = QLabel("0,00 TL", font=deger_font_kucuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_toplam_alacak_hareketi, row_idx_finans, 1)
        row_idx_finans += 1
        
        finans_ozet_cerceve.layout().addWidget(QLabel("Toplam Tahsilat/Ödeme:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_toplam_tahsilat_odeme = QLabel("0,00 TL", font=deger_font_kucuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_toplam_tahsilat_odeme, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Vadesi Gelmiş Borç/Alacak:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_vadesi_gelmis = QLabel("0,00 TL", font=deger_font_kucuk, styleSheet="color: red;")
        finans_ozet_cerceve.layout().addWidget(self.lbl_vadesi_gelmis, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Vadesi Gelecek Borç/Alacak:", font=label_font_kucuk), row_idx_finans, 0)
        self.lbl_vadesi_gelecek = QLabel("0,00 TL", font=deger_font_kucuk, styleSheet="color: blue;")
        finans_ozet_cerceve.layout().addWidget(self.lbl_vadesi_gelecek, row_idx_finans, 1)
        row_idx_finans += 1

        finans_ozet_cerceve.layout().addWidget(QLabel("Dönem Sonu Bakiye:", font=label_font_buyuk), row_idx_finans, 0)
        self.lbl_ozet_net_bakiye = QLabel("0,00 TL", font=deger_font_buyuk)
        finans_ozet_cerceve.layout().addWidget(self.lbl_ozet_net_bakiye, row_idx_finans, 1)

        cari_detay_cerceve = QGroupBox("Cari Detay Bilgileri", frame)
        cari_detay_cerceve.setLayout(QGridLayout(cari_detay_cerceve))
        frame.layout().addWidget(cari_detay_cerceve, 0, 1)

        row_idx_cari = 0
        cari_detay_cerceve.layout().addWidget(QLabel("Cari Adı:", font=label_font_kucuk), row_idx_cari, 0)
        self.lbl_cari_detay_ad = QLabel("-", font=deger_font_kucuk)
        cari_detay_cerceve.layout().addWidget(self.lbl_cari_detay_ad, row_idx_cari, 1)
        row_idx_cari += 1

        cari_detay_cerceve.layout().addWidget(QLabel("Telefon:", font=label_font_kucuk), row_idx_cari, 0)
        self.lbl_cari_detay_tel = QLabel("-", font=deger_font_kucuk)
        cari_detay_cerceve.layout().addWidget(self.lbl_cari_detay_tel, row_idx_cari, 1)
        row_idx_cari += 1

        cari_detay_cerceve.layout().addWidget(QLabel("Adres:", font=label_font_kucuk), row_idx_cari, 0, Qt.AlignTop)
        self.lbl_cari_detay_adres = QLabel("-", font=deger_font_kucuk, wordWrap=True)
        cari_detay_cerceve.layout().addWidget(self.lbl_cari_detay_adres, row_idx_cari, 1)
        row_idx_cari += 1

        cari_detay_cerceve.layout().addWidget(QLabel("Vergi No:", font=label_font_kucuk), row_idx_cari, 0)
        self.lbl_cari_detay_vergi = QLabel("-", font=deger_font_kucuk)
        cari_detay_cerceve.layout().addWidget(self.lbl_cari_detay_vergi, row_idx_cari, 1)
        row_idx_cari += 1

        export_buttons_frame = QFrame(frame)
        export_buttons_frame.setLayout(QVBoxLayout(export_buttons_frame))
        frame.layout().addWidget(export_buttons_frame, 0, 2, Qt.AlignTop)

        btn_pdf = QPushButton("PDF'e Aktar")
        btn_pdf.clicked.connect(self.pdf_aktar)
        export_buttons_frame.layout().addWidget(btn_pdf)

        btn_excel = QPushButton("Excel'e Aktar")
        btn_excel.clicked.connect(self.excel_aktar)
        export_buttons_frame.layout().addWidget(btn_excel)
        
        btn_update_cari = QPushButton("Cari Bilgilerini Güncelle")
        btn_update_cari.clicked.connect(self._cari_bilgileri_guncelle)
        cari_detay_cerceve.layout().addWidget(btn_update_cari, row_idx_cari, 0, 1, 2)

    def _create_filter_alani(self, filter_frame):
        filter_frame.setLayout(QHBoxLayout(filter_frame))
        
        filter_frame.layout().addWidget(QLabel("Başlangıç Tarihi:"))
        self.bas_tarih_entry = QLineEdit()
        filter_frame.layout().addWidget(self.bas_tarih_entry)
        
        btn_date_start = QPushButton("🗓️")
        btn_date_start.setFixedWidth(30)
        btn_date_start.clicked.connect(lambda: DatePickerDialog(self.app, self.bas_tarih_entry))
        filter_frame.layout().addWidget(btn_date_start)

        filter_frame.layout().addWidget(QLabel("Bitiş Tarihi:"))
        self.bit_tarih_entry = QLineEdit()
        filter_frame.layout().addWidget(self.bit_tarih_entry)
        
        btn_date_end = QPushButton("🗓️")
        btn_date_end.setFixedWidth(30)
        btn_date_end.clicked.connect(lambda: DatePickerDialog(self.app, self.bit_tarih_entry))
        filter_frame.layout().addWidget(btn_date_end)

        btn_filter = QPushButton("Filtrele")
        btn_filter.clicked.connect(self.ekstreyi_yukle)
        filter_frame.layout().addWidget(btn_filter)
        
    def _create_treeview_alani(self, tree_frame):
        tree_frame.setLayout(QVBoxLayout(tree_frame))
        
        cols = ("ID", "Tarih", "Saat", "İşlem Tipi", "Referans", "Ödeme Türü", "Açıklama/Detay", "Borç", "Alacak", "Bakiye", "Vade Tarihi")
        self.ekstre_tree = QTreeWidget(tree_frame)
        self.ekstre_tree.setHeaderLabels(cols)
        self.ekstre_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.ekstre_tree.setSortingEnabled(True)

        col_defs = [
            ("ID", 40, Qt.AlignCenter), ("Tarih", 80, Qt.AlignCenter),
            ("Saat", 60, Qt.AlignCenter), ("İşlem Tipi", 120, Qt.AlignCenter),
            ("Referans", 120, Qt.AlignCenter), ("Ödeme Türü", 100, Qt.AlignCenter),
            ("Açıklama/Detay", 300, Qt.AlignLeft),
            ("Borç", 100, Qt.AlignRight),
            ("Alacak", 100, Qt.AlignRight),
            ("Bakiye", 120, Qt.AlignRight),
            ("Vade Tarihi", 90, Qt.AlignCenter)
        ]
        for i, (col_name, width, alignment) in enumerate(col_defs):
            self.ekstre_tree.setColumnWidth(i, width)
            self.ekstre_tree.headerItem().setTextAlignment(i, alignment)
            self.ekstre_tree.headerItem().setFont(i, QFont("Segoe UI", 9, QFont.Bold))
        
        self.ekstre_tree.header().setStretchLastSection(False)
        self.ekstre_tree.header().setSectionResizeMode(6, QHeaderView.Stretch)

        tree_frame.layout().addWidget(self.ekstre_tree)
        
        self.ekstre_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ekstre_tree.customContextMenuRequested.connect(self._show_context_menu)
        self.ekstre_tree.itemDoubleClicked.connect(self.on_double_click_hareket_detay)

    def _create_hizli_islem_alanlari(self):
        self.hizli_islemler_ana_frame.setLayout(QHBoxLayout(self.hizli_islemler_ana_frame))

        ot_frame_text = "Ödeme Ekle" if self.cari_tip == "TEDARIKCI" else "Tahsilat Ekle"
        odeme_tahsilat_frame = QGroupBox(ot_frame_text, self.hizli_islemler_ana_frame)
        odeme_tahsilat_frame.setLayout(QGridLayout(odeme_tahsilat_frame))
        self.hizli_islemler_ana_frame.layout().addWidget(odeme_tahsilat_frame)

        odeme_tahsilat_frame.layout().addWidget(QLabel("Ödeme Tipi:"), 0, 0)
        self.ot_odeme_tipi_combo = QComboBox()
        self.ot_odeme_tipi_combo.addItems([self.db.ODEME_TURU_NAKIT, self.db.ODEME_TURU_KART, 
                                            self.db.ODEME_TURU_EFT_HAVALE, self.db.ODEME_TURU_CEK, 
                                            self.db.ODEME_TURU_SENET])
        self.ot_odeme_tipi_combo.setCurrentText(self.db.ODEME_TURU_NAKIT)
        self.ot_odeme_tipi_combo.currentIndexChanged.connect(self._ot_odeme_tipi_degisince)
        odeme_tahsilat_frame.layout().addWidget(self.ot_odeme_tipi_combo, 0, 1)

        odeme_tahsilat_frame.layout().addWidget(QLabel("Tutar:"), 1, 0)
        self.ot_tutar_entry = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.ot_tutar_entry, decimal_places=2)
        odeme_tahsilat_frame.layout().addWidget(self.ot_tutar_entry, 1, 1)

        odeme_tahsilat_frame.layout().addWidget(QLabel("Kasa/Banka:"), 2, 0)
        self.ot_kasa_banka_combo = QComboBox()
        self.ot_kasa_banka_combo.setEnabled(False)
        odeme_tahsilat_frame.layout().addWidget(self.ot_kasa_banka_combo, 2, 1)

        odeme_tahsilat_frame.layout().addWidget(QLabel("Not:"), 3, 0)
        self.ot_not_entry = QLineEdit()
        odeme_tahsilat_frame.layout().addWidget(self.ot_not_entry, 3, 1)

        btn_ot_save = QPushButton(ot_frame_text)
        btn_ot_save.clicked.connect(self._hizli_odeme_tahsilat_kaydet)
        odeme_tahsilat_frame.layout().addWidget(btn_ot_save, 4, 0, 1, 2)

        borc_frame = QGroupBox("Veresiye Borç Ekle", self.hizli_islemler_ana_frame)
        borc_frame.setLayout(QGridLayout(borc_frame))
        self.hizli_islemler_ana_frame.layout().addWidget(borc_frame)

        borc_frame.layout().addWidget(QLabel("Türü Seçiniz:"), 0, 0)
        self.borc_tur_combo = QComboBox()
        self.borc_tur_combo.addItems(["Diğer Borç", "Satış Faturası"])
        borc_frame.layout().addWidget(self.borc_tur_combo, 0, 1)

        borc_frame.layout().addWidget(QLabel("Tutar:"), 1, 0)
        self.borc_tutar_entry = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.borc_tutar_entry, decimal_places=2)
        borc_frame.layout().addWidget(self.borc_tutar_entry, 1, 1)

        borc_frame.layout().addWidget(QLabel("Not:"), 2, 0)
        self.borc_not_entry = QLineEdit()
        borc_frame.layout().addWidget(self.borc_not_entry, 2, 1)

        btn_borc_save = QPushButton("Veresiye Ekle")
        btn_borc_save.clicked.connect(self._hizli_veresiye_borc_kaydet)
        borc_frame.layout().addWidget(btn_borc_save, 3, 0, 1, 2)

        alacak_frame = QGroupBox("Alacak Ekleme", self.hizli_islemler_ana_frame)
        alacak_frame.setLayout(QGridLayout(alacak_frame))
        self.hizli_islemler_ana_frame.layout().addWidget(alacak_frame)

        alacak_frame.layout().addWidget(QLabel("Türü Seçiniz:"), 0, 0)
        self.alacak_tur_combo = QComboBox()
        self.alacak_tur_combo.addItems(["Diğer Alacak", "İade Faturası"])
        alacak_frame.layout().addWidget(self.alacak_tur_combo, 0, 1)

        alacak_frame.layout().addWidget(QLabel("Tutar:"), 1, 0)
        self.alacak_tutar_entry = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.alacak_tutar_entry, decimal_places=2)
        alacak_frame.layout().addWidget(self.alacak_tutar_entry, 1, 1)

        alacak_frame.layout().addWidget(QLabel("Not:"), 2, 0)
        self.alacak_not_entry = QLineEdit()
        alacak_frame.layout().addWidget(self.alacak_not_entry, 2, 1)

        btn_alacak_save = QPushButton("Alacak Kaydet")
        btn_alacak_save.clicked.connect(self._hizli_alacak_kaydet)
        alacak_frame.layout().addWidget(btn_alacak_save, 3, 0, 1, 2)

    def _yukle_kasa_banka_hesaplarini_hizli_islem_formu(self):
        self.ot_kasa_banka_combo.clear()
        self.kasa_banka_map.clear()
        
        try:
            hesaplar_response = self.db.kasa_banka_listesi_al()
            if isinstance(hesaplar_response, dict) and "items" in hesaplar_response:
                hesaplar = hesaplar_response["items"]
            elif isinstance(hesaplar_response, list):
                hesaplar = hesaplar_response
                self.app.set_status_message("Uyarı: Kasa/Banka listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                hesaplar = []
                self.app.set_status_message("Hata: Kasa/Banka listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Kasa/Banka listesi API'den beklenen formatta gelmedi: {type(hesaplar_response)} - {hesaplar_response}", exc_info=True)
                self.ot_kasa_banka_combo.addItem("Hesap Yok", None)
                self.ot_kasa_banka_combo.setEnabled(False)
                return

            if hesaplar:
                for h in hesaplar:
                    display_text = f"{h.get('hesap_adi')} ({h.get('tip')}) - Bakiye: {self.db._format_currency(h.get('bakiye', 0.0))}"
                    if h.get('tip') == "BANKA" and h.get('banka_adi'):
                        display_text += f" ({h.get('banka_adi')})"
                    self.kasa_banka_map[display_text] = h.get('id')
                    self.ot_kasa_banka_combo.addItem(display_text, h.get('id'))
                self.ot_kasa_banka_combo.setCurrentIndex(0)
                self.ot_kasa_banka_combo.setEnabled(True)
            else:
                self.ot_kasa_banka_combo.clear()
                self.ot_kasa_banka_combo.addItem("Hesap Yok", None)
                self.ot_kasa_banka_combo.setEnabled(False)

            self.app.set_status_message(f"{len(hesaplar)} kasa/banka hesabı API'den yüklendi.", "blue")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kasa/Banka hesapları yüklenirken hata: {e}")
            logging.error(f"Kasa/Banka yükleme hatası: {e}", exc_info=True)
            self.ot_kasa_banka_combo.addItem("Hesap Yok", None)
            self.ot_kasa_banka_combo.setEnabled(False)

    def _ot_odeme_tipi_degisince(self):
        selected_odeme_sekli = self.ot_odeme_tipi_combo.currentText()
        
        self.ot_kasa_banka_combo.blockSignals(True)
        
        default_account_id = None
        try:
            default_account_data = self.db.get_varsayilan_kasa_banka(selected_odeme_sekli)
            if default_account_data:
                default_account_id = default_account_data.get('id')
        except Exception as e:
            logging.warning(f"Default kasa/banka for {selected_odeme_sekli} could not be fetched: {e}")

        if default_account_id:
            for i in range(self.ot_kasa_banka_combo.count()):
                if self.ot_kasa_banka_combo.itemData(i) == default_account_id:
                    self.ot_kasa_banka_combo.setCurrentIndex(i)
                    break
            else:
                if self.ot_kasa_banka_combo.count() > 0 and self.ot_kasa_banka_combo.itemData(0) is not None:
                    self.ot_kasa_banka_combo.setCurrentIndex(0)
                else:
                    self.ot_kasa_banka_combo.setCurrentText("")
        elif self.ot_kasa_banka_combo.count() > 0 and self.ot_kasa_banka_combo.itemData(0) is not None:
            self.ot_kasa_banka_combo.setCurrentIndex(0)
        else:
            self.ot_kasa_banka_combo.setCurrentText("")

        self.ot_kasa_banka_combo.blockSignals(False)

    def _yukle_ozet_bilgileri(self):
        try:
            cari_detail = None
            if self.cari_tip == "MUSTERI":
                cari_detail = self.db.musteri_getir_by_id(self.cari_id)
            else:
                cari_detail = self.db.tedarikci_getir_by_id(self.cari_id)

            if not cari_detail:
                self.app.set_status_message(f"Hata: Cari bilgiler yüklenemedi. ID {self.cari_id} bulunamadı.", "red")
                return

            self.lbl_cari_detay_ad.setText(cari_detail.get('ad', '-'))
            self.lbl_cari_detay_tel.setText(cari_detail.get('telefon', '-'))
            self.lbl_cari_detay_adres.setText(cari_detail.get('adres', '-'))
            vergi_info = f"{cari_detail.get('vergi_dairesi', '-')} / {cari_detail.get('vergi_no', '-')}"
            self.lbl_cari_detay_vergi.setText(vergi_info)

            net_bakiye = cari_detail.get("net_bakiye", 0.0)

            bakiye_metni = self.db._format_currency(net_bakiye)
            if net_bakiye > 0:
                bakiye_metni = f"<b style='color: green;'>{bakiye_metni} ALACAKLI</b>"
            elif net_bakiye < 0:
                bakiye_metni = f"<b style='color: red;'>{bakiye_metni} BORÇLU</b>"
            else:
                bakiye_metni = f"<b style='color: blue;'>{bakiye_metni}</b>"
            self.lbl_ozet_net_bakiye.setText(bakiye_metni)

            self.lbl_donem_basi_bakiye.setText(self.db._format_currency(0.0))
            self.lbl_toplam_borc_hareketi.setText(self.db._format_currency(0.0))
            self.lbl_toplam_alacak_hareketi.setText(self.db._format_currency(0.0))
            self.lbl_toplam_tahsilat_odeme.setText(self.db._format_currency(0.0))
            self.lbl_vadesi_gelmis.setText(self.db._format_currency(0.0))
            self.lbl_vadesi_gelecek.setText(self.db._format_currency(0.0))

            self.app.set_status_message("Cari özet bilgileri güncellendi.", "green")

        except Exception as e:
            logger.error(f"Cari özet bilgileri yüklenirken hata oluştu: {e}", exc_info=True)
            self.app.set_status_message(f"Hata: Cari özet bilgileri yüklenemedi. Detay: {e}", "red")

    def _cari_bilgileri_guncelle(self):
        try:
            cari_data = None
            if self.cari_tip == "MUSTERI":
                cari_data = self.db.musteri_getir_by_id(self.cari_id)
                if cari_data:
                    from pencereler import YeniMusteriEklePenceresi
                    dialog = YeniMusteriEklePenceresi(self, self.db, self._ozet_ve_liste_yenile, musteri_duzenle=cari_data, app_ref=self.app)
                    dialog.exec()
                else:
                    self.app.set_status_message(f"Hata: Müşteri bilgileri yüklenemedi. ID {self.cari_id} bulunamadı.", "red")
                    return
            elif self.cari_tip == "TEDARIKCI":
                cari_data = self.db.tedarikci_getir_by_id(self.cari_id)
                if cari_data:
                    from pencereler import YeniTedarikciEklePenceresi
                    dialog = YeniTedarikciEklePenceresi(self, self.db, self._ozet_ve_liste_yenile, tedarikci_duzenle=cari_data, app_ref=self.app)
                    dialog.exec()
                else:
                    self.app.set_status_message(f"Hata: Tedarikçi bilgileri yüklenemedi. ID {self.cari_id} bulunamadı.", "red")
                    return

            self.app.set_status_message(f"{self.cari_tip} kartı açıldı.", "blue")

        except Exception as e:
            logger.error(f"Cari bilgiler güncellenmek üzere yüklenirken hata oluştu: {e}", exc_info=True)
            self.app.set_status_message(f"Hata: Cari bilgiler yüklenemedi. Detay: {e}", "red")

    def _ozet_ve_liste_yenile(self):
        self._yukle_ozet_bilgileri()
        self.ekstreyi_yukle()

    def _hizli_odeme_tahsilat_kaydet(self):
        islem_turu = self.sender().text()
        islem_turu_enum = "GIDER" if islem_turu == "Ödeme Yap" else "GELIR"

        tutar_str = self.ot_tutar_entry.text().replace(".", "").replace(",", ".")
        try:
            tutar = float(tutar_str)
            if tutar <= 0:
                self.app.set_status_message("Tutar sıfırdan büyük olmalıdır.", "orange")
                return
        except ValueError:
            self.app.set_status_message("Geçerli bir tutar girin.", "orange")
            return

        aciklama = self.ot_not_entry.text().strip()
        if not aciklama:
            self.app.set_status_message("Açıklama alanı boş bırakılamaz.", "orange")
            return

        selected_hesap_idx = self.ot_kasa_banka_combo.currentIndex()
        if selected_hesap_idx < 0:
            self.app.set_status_message("Lütfen bir Kasa/Banka hesabı seçin.", "orange")
            return
        
        kasa_banka_id = self.ot_kasa_banka_combo.currentData()

        gelir_gider_data = {
            "tarih": date.today().strftime('%Y-%m-%d'),
            "tip": islem_turu_enum,
            "tutar": tutar,
            "aciklama": aciklama,
            "kasa_banka_id": kasa_banka_id,
            "cari_id": self.cari_id,
            "cari_tip": self.cari_tip,
        }

        try:
            success = self.db.gelir_gider_ekle(gelir_gider_data)
            if success:
                self.app.set_status_message(f"Hızlı {islem_turu.lower()} kaydı başarıyla oluşturuldu.", "green")
                self.ot_tutar_entry.clear()
                self.ot_not_entry.clear()
                self.ot_odeme_tipi_combo.setCurrentText(self.db.ODEME_TURU_NAKIT)
                self._ot_odeme_tipi_degisince()
                self._ozet_ve_liste_yenile()
            else:
                self.app.set_status_message(f"Hızlı {islem_turu.lower()} kaydı oluşturulamadı.", "red")
        except Exception as e:
            logger.error(f"Hızlı {islem_turu.lower()} kaydı oluşturulurken hata oluştu: {e}", exc_info=True)
            self.app.set_status_message(f"Hata: Hızlı {islem_turu.lower()} kaydı oluşturulamadı. Detay: {e}", "red")

    def _hizli_veresiye_borc_kaydet(self):
        borc_tur = self.borc_tur_combo.currentText()
        tutar_str = self.borc_tutar_entry.text().replace(',', '.')
        not_str = self.borc_not_entry.text()

        if not tutar_str or float(tutar_str) <= 0:
            QMessageBox.warning(self, "Eksik Bilgi", "Lütfen geçerli bir tutar giriniz.")
            return

        if borc_tur == "Satış Faturası":
            QMessageBox.information(self, "Yönlendirme", "Fatura oluşturmak için lütfen ana menüden 'Yeni Satış Faturası' ekranını kullanın.")
        else:
            try:
                tutar_f = float(tutar_str)
                data = {
                    "cari_id": self.cari_id,
                    "cari_turu": self.cari_tip,
                    "tarih": date.today().strftime('%Y-%m-%d'),
                    "tutar": tutar_f,
                    "aciklama": not_str,
                    "islem_turu": "VERESİYE_BORÇ",
                    "islem_yone": "BORC",
                    "kaynak": self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL
                }
                success = self.db.cari_hareket_ekle_manuel(data)

                if success:
                    QMessageBox.information(self, "Başarılı", "Veresiye borç başarıyla eklendi.")
                    self.borc_tutar_entry.clear()
                    self.borc_not_entry.clear()
                    self._ozet_ve_liste_yenile()
                else:
                    QMessageBox.critical(self, "Hata", "Veresiye borç eklenirken hata.")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Veresiye borç eklenirken hata: {e}")
                logging.error(f"Hızlı veresiye borç kaydetme hatası: {e}", exc_info=True)

    def _hizli_alacak_kaydet(self):
        QMessageBox.information(self, "Geliştirme Aşamasında", "Alacak ekleme özelliği henüz tamamlanmamıştır.")

    def excel_aktar(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Cari Hesap Ekstresini Excel'e Kaydet", 
                                                 f"Cari_Ekstresi_{self.cari_ad_gosterim.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx", 
                                                 "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)")
        if file_path:
            bekleme_penceresi = BeklemePenceresi(self, message="Ekstre Excel'e aktarılıyor, lütfen bekleyiniz...")
            threading.Thread(target=lambda: self._generate_ekstre_excel_threaded(
                self.cari_tip, self.cari_id, self.bas_tarih_entry.text(), self.bit_tarih_entry.text(),
                file_path, bekleme_penceresi
            )).start()
        else:
            self.app.set_status_message("Excel'e aktarma iptal edildi.", "blue")

    def pdf_aktar(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Cari Hesap Ekstresini PDF'e Kaydet", 
                                                 f"Cari_Ekstresi_{self.cari_ad_gosterim.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf", 
                                                 "PDF Dosyaları (*.pdf);;Tüm Dosyalar (*)")
        if file_path:
            bekleme_penceresi = BeklemePenceresi(self, message="Ekstre PDF'e aktarılıyor, lütfen bekleyiniz...")
            
            result_queue = multiprocessing.Queue()
            pdf_process = multiprocessing.Process(target=self.db.cari_ekstresi_pdf_olustur, args=(
                self.db.data_dir,
                self.cari_tip,
                self.cari_id,
                self.bas_tarih_entry.text(),
                self.bit_tarih_entry.text(),
                file_path,
                result_queue
            ))
            pdf_process.start()

            self.app.process_queue_timer = QTimer(self.app)
            self.app.process_queue_timer.timeout.connect(lambda: self._check_pdf_process_completion(result_queue, pdf_process, bekleme_penceresi))
            self.app.process_queue_timer.start(100)
        else:
            self.app.set_status_message("PDF'e aktarma iptal edildi.", "blue")

    def _check_pdf_process_completion(self, result_queue, pdf_process, bekleme_penceresi):
        if not result_queue.empty():
            success, message = result_queue.get()
            bekleme_penceresi.close()
            self.app.process_queue_timer.stop()

            if success:
                QMessageBox.information(self, "Başarılı", message)
                self.app.set_status_message(message, "green")
            else:
                QMessageBox.critical(self, "Hata", message)
                self.app.set_status_message(f"Ekstre PDF'e aktarılırken hata: {message}", "red")
            pdf_process.join()
            
        elif not pdf_process.is_alive():
            bekleme_penceresi.close()
            self.app.process_queue_timer.stop()
            QMessageBox.critical(self, "Hata", "PDF işlemi beklenmedik şekilde sonlandı.")
            self.app.set_status_message("PDF işlemi beklenmedik şekilde sonlandı.", "red")
            pdf_process.join()

    def _generate_ekstre_excel_threaded(self, cari_tip, cari_id, bas_t, bit_t, dosya_yolu, bekleme_penceresi):
        local_db_manager = self.db.__class__(api_base_url=self.db.api_base_url, app_ref=self.app)
        
        success = False
        message = ""
        try:
            hareketler_listesi, devreden_bakiye, success_db, message_db = local_db_manager.cari_hesap_ekstresi_al(
                cari_id, cari_tip, bas_t, bit_t
            )
            
            if not success_db:
                message = f"Ekstre verisi alınırken hata: {message_db}"
            elif not hareketler_listesi and devreden_bakiye == 0:
                message = "Excel'e aktarılacak cari ekstre verisi bulunamadı."
            else:
                success, message = local_db_manager.tarihsel_satis_raporu_excel_olustur(
                    rapor_verileri=hareketler_listesi,
                    dosya_yolu=dosya_yolu,
                    bas_t=bas_t,
                    bit_t=bit_t
                )
                if not success: message = f"Excel oluşturulurken hata: {message}"

        except Exception as e:
            message = f"Rapor Excel'e aktarılırken bir hata oluştu:\n{e}"
            logging.error(f"Excel export thread error: {e}", exc_info=True)
        finally:
            self.app.set_status_message(message, "blue" if success else "red")
            self.app.after(0, bekleme_penceresi.kapat)
            self.app.after(0, lambda: QMessageBox.information(self, "Excel Aktarım", message) if success else QMessageBox.critical(self, "Excel Aktarım Hatası", message))

    def ekstreyi_yukle(self):
        self.ekstre_tree.clear()
        self.hareket_detay_map.clear()

        bas_tarih_str = self.bas_tarih_entry.text()
        bitis_tarih_str = self.bit_tarih_entry.text()

        try:
            datetime.strptime(bas_tarih_str, '%Y-%m-%d')
            datetime.strptime(bitis_tarih_str, '%Y-%m-%d')
        except ValueError:
            QMessageBox.critical(self, "Hata", "Tarih formatı 'YYYY-AA-GG' şeklinde olmalıdır.")
            return
        
        hareketler_listesi, devreden_bakiye, success_db, message_db = self.db.cari_hesap_ekstresi_al(
            self.cari_id, self.cari_tip, bas_tarih_str, bitis_tarih_str
        )

        if not success_db:
            QMessageBox.critical(self, "Hata", f"Ekstre verisi alınırken hata: {message_db}")
            self.app.set_status_message(f"{self.cari_ad_gosterim} için ekstre yüklenemedi: {message_db}", "red")
            return
        
        devir_item = QTreeWidgetItem(self.ekstre_tree)
        devir_item.setText(0, "")
        devir_item.setText(1, bas_tarih_str)
        devir_item.setText(2, "")
        devir_item.setText(3, "DEVİR")
        devir_item.setText(4, "")
        devir_item.setText(5, "Devreden Bakiye")
        devir_item.setText(6, "")
        devir_item.setText(7, self.db._format_currency(devreden_bakiye) if devreden_bakiye > 0 else "")
        devir_item.setText(8, self.db._format_currency(abs(devreden_bakiye)) if devreden_bakiye < 0 else "")
        devir_item.setText(9, self.db._format_currency(devreden_bakiye))
        devir_item.setText(10, "")
        
        for col_idx in range(self.ekstre_tree.columnCount()):
            devir_item.setBackground(col_idx, QBrush(QColor("#EFEFEF")))
            devir_item.setFont(col_idx, QFont("Segoe UI", 9, QFont.Bold))


        current_bakiye = devreden_bakiye
        
        for hareket in hareketler_listesi:
            item_qt = QTreeWidgetItem(self.ekstre_tree)
            
            tarih_formatted = hareket['tarih'].strftime('%d.%m.%Y') if isinstance(hareket['tarih'], date) else str(hareket['tarih'])
            vade_tarihi_formatted = hareket['vade_tarihi'].strftime('%d.%m.%Y') if isinstance(hareket['vade_tarihi'], date) else (str(hareket['vade_tarihi']) if hareket['vade_tarihi'] else '-')
            
            borc_val = ""
            alacak_val = ""
            
            if self.cari_tip == 'MUSTERI':
                if hareket['islem_yone'] == 'ALACAK':
                    alacak_val = self.db._format_currency(hareket['tutar'])
                    current_bakiye += hareket['tutar']
                elif hareket['islem_yone'] == 'BORC':
                    borc_val = self.db._format_currency(hareket['tutar'])
                    current_bakiye -= hareket['tutar']
            elif self.cari_tip == 'TEDARIKCI':
                if hareket['islem_yone'] == 'ALACAK':
                    alacak_val = self.db._format_currency(hareket['tutar'])
                    current_bakiye += hareket['tutar']
                elif hareket['islem_yone'] == 'BORC':
                    borc_val = self.db._format_currency(hareket['tutar'])
                    current_bakiye -= hareket['tutar']
            
            display_islem_tipi = hareket['islem_turu']
            display_ref_gosterim = hareket['fatura_no'] if hareket.get('fatura_no') else (hareket.get('kaynak') or '-')

            if hareket.get('kaynak') in (self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA):
                if hareket.get('fatura_turu') == self.db.FATURA_TIP_SATIS:
                    display_islem_tipi = "Satış Faturası"
                elif hareket.get('fatura_turu') == self.db.FATURA_TIP_ALIS:
                    display_islem_tipi = "Alış Faturası"
                elif hareket.get('fatura_turu') == self.db.FATURA_TIP_SATIS_IADE:
                    display_islem_tipi = "Satış İade Faturası"
                elif hareket.get('fatura_turu') == self.db.FATURA_TIP_ALIS_IADE:
                    display_islem_tipi = "Alış İade Faturası"
                display_ref_gosterim = hareket['fatura_no']
            elif hareket.get('kaynak') in (self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME):
                display_islem_tipi = "Tahsilat" if hareket.get('islem_turu') == "GELIR" else "Ödeme"
                display_ref_gosterim = hareket.get('kaynak')

            item_qt.setText(0, str(hareket['id']))
            item_qt.setText(1, tarih_formatted)
            item_qt.setText(2, hareket.get('islem_saati') or '')
            item_qt.setText(3, display_islem_tipi)
            item_qt.setText(4, display_ref_gosterim)
            item_qt.setText(5, hareket.get('odeme_turu') or '-')
            item_qt.setText(6, hareket.get('aciklama') or '-')
            item_qt.setText(7, borc_val)
            item_qt.setText(8, alacak_val)
            item_qt.setText(9, self.db._format_currency(current_bakiye))
            item_qt.setText(10, vade_tarihi_formatted)

            self.hareket_detay_map[hareket['id']] = hareket

            if hareket.get('kaynak') in (self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA):
                if hareket.get('odeme_turu') in self.db.pesin_odeme_turleri:
                    for col_idx in range(self.ekstre_tree.columnCount()):
                        item_qt.setBackground(col_idx, QBrush(QColor("lightgray")))
                        item_qt.setForeground(col_idx, QBrush(QColor("darkgray")))
                else:
                    for col_idx in range(self.ekstre_tree.columnCount()):
                        item_qt.setForeground(col_idx, QBrush(QColor("red")))
                if "İADE" in hareket.get('fatura_turu', ''):
                    for col_idx in range(self.ekstre_tree.columnCount()):
                        item_qt.setBackground(col_idx, QBrush(QColor("#FFF2CC")))
                        item_qt.setForeground(col_idx, QBrush(QColor("#A67400")))
            elif hareket.get('kaynak') in (self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME, self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL):
                for col_idx in range(self.ekstre_tree.columnCount()):
                    item_qt.setForeground(col_idx, QBrush(QColor("green")))

        self.app.set_status_message(f"{self.cari_ad_gosterim} için {len(hareketler_listesi)} hareket yüklendi.", "blue")

    def _show_context_menu(self, pos):
        item = self.ekstre_tree.itemAt(pos)
        if not item: return

        item_id = int(item.text(0))
        if item.text(3) == "DEVİR": return

        hareket_detayi = self.hareket_detay_map.get(item_id)
        if not hareket_detayi: return

        ref_tip = hareket_detayi.get('kaynak')

        context_menu = QMenu(self)
        
        if ref_tip in [self.db.KAYNAK_TIP_MANUEL, self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME, self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL, self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            context_menu.addAction("İşlemi Sil").triggered.connect(self.secili_islemi_sil)
        
        if ref_tip in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            context_menu.addAction("Faturayı Güncelle").triggered.connect(self.secili_islemi_guncelle)
        
        if context_menu.actions():
            context_menu.exec(self.ekstre_tree.mapToGlobal(pos))

    def secili_islemi_sil(self):
        selected_items = self.ekstre_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek için bir işlem seçin.")
            return

        item_qt = selected_items[0]
        hareket_id = int(item_qt.text(0))

        hareket_detayi = self.hareket_detay_map.get(hareket_id)
        if not hareket_detayi:
            QMessageBox.critical(self, "Hata", "İşlem detayları bulunamadı.")
            return
        
        ref_id = hareket_detayi.get('kaynak_id')
        ref_tip = hareket_detayi.get('kaynak')
        aciklama_text = hareket_detayi.get('aciklama')
        fatura_no = hareket_detayi.get('fatura_no')
        
        confirm_msg = f"'{aciklama_text}' açıklamalı işlemi silmek istediğinizden emin misiniz?\nBu işlem geri alınamaz."
        if ref_tip in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            confirm_msg = f"'{fatura_no}' numaralı FATURA ve ilişkili tüm hareketlerini silmek istediğinizden emin misiniz?\nBu işlem geri alınamaz."

        reply = QMessageBox.question(self, "Silme Onayı", confirm_msg, QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            success = False
            message = "Bilinmeyen işlem tipi."
            try:
                if ref_tip in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
                    success = self.db.fatura_sil(ref_id)
                    message = f"Fatura {fatura_no} başarıyla silindi."
                else:
                    if ref_tip in [self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME]:
                        success = self.db.gelir_gider_sil(hareket_id)
                    elif ref_tip == self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL:
                        success = self.db.cari_hareket_sil_manuel(hareket_id)
                        if not success:
                            QMessageBox.critical(self, "Hata", f"{ref_tip} türündeki hareket silinemiyor. API desteği gerekli olabilir.")
                    else:
                         QMessageBox.critical(self, "Hata", f"İşlem tipi ({ref_tip}) silinemiyor. API desteği gerekli.")
                         return

                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self._ozet_ve_liste_yenile()
                    
                    if hasattr(self.app, 'fatura_listesi_sayfasi'):
                        if hasattr(self.app.fatura_listesi_sayfasi, 'satis_fatura_frame'):
                            self.app.fatura_listesi_sayfasi.satis_fatura_frame.fatura_listesini_yukle()
                        if hasattr(self.app.fatura_listesi_sayfasi, 'alis_fatura_frame'):
                            self.app.fatura_listesi_sayfasi.alis_fatura_frame.fatura_listesini_yukle()
                    if hasattr(self.app, 'gelir_gider_sayfasi'):
                        if hasattr(self.app.gelir_gider_sayfasi, 'gelir_listesi_frame'):
                            self.app.gelir_gider_sayfasi.gelir_listesi_frame.gg_listesini_yukle()
                        if hasattr(self.app.gelir_gider_sayfasi, 'gider_listesi_frame'):
                            self.app.gelir_gider_sayfasi.gider_listesi_frame.gg_listesini_yukle()
                    if hasattr(self.app, 'kasa_banka_yonetimi_sayfasi'):
                        self.app.kasa_banka_yonetimi_sayfasi.hesap_listesini_yenile()
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                error_detail = str(e)
                QMessageBox.critical(self, "Hata", f"Silinirken hata: {error_detail}")
                logging.error(f"Cari Ekstresi silme hatası: {error_detail}", exc_info=True)
        else:
            self.app.set_status_message("Silme işlemi iptal edildi.", "blue")

    def secili_islemi_guncelle(self):
        selected_items = self.ekstre_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen güncellemek için bir fatura işlemi seçin.")
            return

        item_qt = selected_items[0]
        hareket_id = int(item_qt.text(0))

        hareket_detayi = self.hareket_detay_map.get(hareket_id)
        if not hareket_detayi:
            QMessageBox.critical(self, "Hata", "İşlem detayları bulunamadı.")
            return
        
        ref_id = hareket_detayi.get('kaynak_id')
        ref_tip = hareket_detayi.get('kaynak')

        if ref_tip in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            if ref_id:
                from pencereler import FaturaGuncellemePenceresi
                FaturaGuncellemePenceresi(self, self.db, ref_id, self._ozet_ve_liste_yenile).exec()
            else:
                QMessageBox.information(self, "Detay", "Fatura referansı bulunamadı.")
        else:
            QMessageBox.information(self, "Bilgi", "Sadece fatura işlemleri güncellenebilir.")

    def on_double_click_hareket_detay(self, item, column):
        if item.text(3) == "DEVİR":
            QMessageBox.warning(self, "Uyarı", "Devir satırı için detay görüntülenemez.")
            return

        hareket_id = int(item.text(0))
        hareket_detay = self.hareket_detay_map.get(hareket_id)

        if not hareket_detay:
            QMessageBox.critical(self, "Hata", "Seçilen işlemin detayları bulunamadı.")
            return

        ref_id = hareket_detay.get('kaynak_id')
        ref_tip_str = hareket_detay.get('kaynak')

        if ref_tip_str in [self.db.KAYNAK_TIP_FATURA, self.db.KAYNAK_TIP_IADE_FATURA, self.db.KAYNAK_TIP_FATURA_SATIS_PESIN, self.db.KAYNAK_TIP_FATURA_ALIS_PESIN]:
            if ref_id:
                from pencereler import FaturaDetayPenceresi
                FaturaDetayPenceresi(self.app, self.db, ref_id).exec()
            else:
                QMessageBox.information(self, "Detay", "Fatura referansı bulunamadı.")
        elif ref_tip_str in [self.db.KAYNAK_TIP_TAHSILAT, self.db.KAYNAK_TIP_ODEME, self.db.KAYNAK_TIP_VERESIYE_BORC_MANUEL]:
            tarih_gosterim = hareket_detay.get('tarih').strftime('%d.%m.%Y') if isinstance(hareket_detay.get('tarih'), date) else str(hareket_detay.get('tarih'))
            tutar_gosterim = self.db._format_currency(hareket_detay.get('tutar'))
            aciklama_gosterim = hareket_detay.get('aciklama') or "Açıklama yok."
            
            QMessageBox.information(self, "İşlem Detayı",
                                 f"Bu bir {ref_tip_str} işlemidir.\n"
                                 f"Tarih: {tarih_gosterim}\n"
                                 f"Tutar: {tutar_gosterim}\n" 
                                 f"Açıklama: {aciklama_gosterim}\n"
                                 f"Referans ID: {hareket_id}")
        else:
            QMessageBox.information(self, "Detay", "Bu işlem tipi için detay görüntüleme mevcut değil.")
            
class FaturaGuncellemePenceresi(QDialog):
    def __init__(self, parent, db_manager, fatura_id_duzenle, yenile_callback_liste=None):
        super().__init__(parent)
        self.app = parent.app if hasattr(parent, 'app') else parent
        self.db = db_manager
        self.yenile_callback_liste = yenile_callback_liste
        self.fatura_id_duzenle = fatura_id_duzenle

        try:
            fatura_ana_bilgileri = self.db.fatura_getir_by_id(self.fatura_id_duzenle)
            if not fatura_ana_bilgileri:
                QMessageBox.critical(self, "Hata", f"ID {self.fatura_id_duzenle} olan fatura bulunamadı.")
                self.reject()
                return

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Fatura bilgileri çekilirken bir hata oluştu: {e}")
            logger.error(f"Fatura bilgileri yüklenirken hata oluştu: {e}", exc_info=True)
            self.reject()
            return

        faturanın_gercek_islem_tipi = fatura_ana_bilgileri.get('fatura_turu')

        self.setWindowTitle(f"Fatura Güncelleme: {fatura_ana_bilgileri.get('fatura_no', 'Bilinmiyor')}")
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)

        dialog_layout = QVBoxLayout(self)

        from arayuz import FaturaOlusturmaSayfasi
        self.fatura_olusturma_form = FaturaOlusturmaSayfasi(
            self,
            self.db,
            self.app,
            faturanın_gercek_islem_tipi,
            duzenleme_id=self.fatura_id_duzenle,
            yenile_callback=self._fatura_guncellendi_callback
        )
        dialog_layout.addWidget(self.fatura_olusturma_form)

        self.fatura_olusturma_form.saved_successfully.connect(self.accept)
        self.fatura_olusturma_form.cancelled_successfully.connect(self.reject)

        self.finished.connect(self.on_dialog_finished)

    def _fatura_guncellendi_callback(self):
        pass

    def on_dialog_finished(self, result):
        if self.yenile_callback_liste:
            self.yenile_callback_liste()

class FaturaPenceresi(QDialog):
    FATURA_TIP_ALIS = "ALIŞ"
    FATURA_TIP_SATIS = "SATIŞ"
    FATURA_TIP_DEVIR_GIRIS = "DEVİR GİRİŞ"
    FATURA_TIP_SATIS_IADE = "SATIŞ İADE"
    FATURA_TIP_ALIS_IADE = "ALIŞ İADE"

    ODEME_TURU_NAKIT = "NAKİT"
    ODEME_TURU_KART = "KART"
    ODEME_TURU_EFT_HAVALE = "EFT/HAVALE"
    ODEME_TURU_CEK = "ÇEK"
    ODEME_TURU_SENET = "SENET"
    ODEME_TURU_ACIK_HESAP = "AÇIK HESAP"
    ODEME_TURU_ETKISIZ_FATURA = "ETKİSİZ FATURA"

    pesin_odeme_turleri = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"]

    CARI_TIP_MUSTERI = "MUSTERI"
    CARI_TIP_TEDARIKCI = "TEDARIKCI"

    def __init__(self, parent=None, db_manager=None, app_ref=None, fatura_tipi=None, duzenleme_id=None, yenile_callback=None, initial_data=None):
        super().__init__(parent)
        
        self.app = app_ref
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.duzenleme_id = duzenleme_id
        self.initial_data = initial_data or {}
        self.islem_tipi = fatura_tipi

        self.iade_modu_aktif = self.initial_data.get('iade_modu', False)
        self.original_fatura_id_for_iade = self.initial_data.get('orijinal_fatura_id')

        if self.iade_modu_aktif:
            if self.islem_tipi == self.FATURA_TIP_SATIS: self.islem_tipi = self.FATURA_TIP_SATIS_IADE
            elif self.islem_tipi == self.FATURA_TIP_ALIS: self.islem_tipi = self.FATURA_TIP_ALIS_IADE

        self.fatura_kalemleri_ui = []
        self.tum_urunler_cache = []
        self.urun_map_filtrelenmis = {}
        self.kasa_banka_map = {}
        self.cari_map_display_to_id = {}
        self.cari_id_to_display_map = {}

        self.secili_cari_id = None
        self.secili_cari_adi = ""
        self.perakende_musteri_id = None

        self.setWindowTitle(self._get_baslik())
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)

        self.main_layout = QVBoxLayout(self)
        
        self._create_ui()
        self._load_initial_data()
        self._connect_signals()

        QTimer.singleShot(0, self._on_iade_modu_changed)

    def _connect_signals(self):
        self.btn_cari_sec.clicked.connect(self._cari_secim_penceresi_ac)
        self.odeme_turu_cb.currentIndexChanged.connect(self._odeme_turu_degisince_event_handler)
        self.genel_iskonto_tipi_cb.currentIndexChanged.connect(self._on_genel_iskonto_tipi_changed)
        self.genel_iskonto_degeri_e.textChanged.connect(self.toplamlari_hesapla_ui)

        self.urun_arama_entry.textChanged.connect(self._delayed_stok_yenile)
        # BURADA DEĞİŞİKLİK YAPILDI: Çift tıklama direkt sepete ekleyecek
        self.urun_arama_sonuclari_tree.itemDoubleClicked.connect(self._double_click_add_to_cart)
        self.urun_arama_sonuclari_tree.itemSelectionChanged.connect(self._secili_urun_bilgilerini_goster_arama_listesinden)

        self.mik_e.textChanged.connect(lambda: format_and_validate_numeric_input(self.mik_e, 2))
        self.birim_fiyat_e.textChanged.connect(lambda: format_and_validate_numeric_input(self.birim_fiyat_e, 2))
        self.iskonto_yuzde_1_e.textChanged.connect(lambda: format_and_validate_numeric_input(self.iskonto_yuzde_1_e, 2))
        self.iskonto_yuzde_2_e.textChanged.connect(lambda: format_and_validate_numeric_input(self.iskonto_yuzde_2_e, 2))
        self.genel_iskonto_degeri_e.textChanged.connect(lambda: format_and_validate_numeric_input(self.genel_iskonto_degeri_e, 2))

        self.btn_sepete_ekle.clicked.connect(self._kalem_ekle_arama_listesinden)
        self.btn_secili_kalemi_sil.clicked.connect(self._secili_kalemi_sil)
        self.btn_sepeti_temizle.clicked.connect(self._sepeti_temizle)
        self.btn_kaydet.clicked.connect(self._kaydet_fatura)

        self.sep_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.sep_tree.customContextMenuRequested.connect(self._open_sepet_context_menu)

    def _load_initial_data(self):
        try:
            self.perakende_musteri_id = self.db.get_perakende_musteri_id()
            self._yukle_carileri()
            self._yukle_kasa_banka_hesaplarini()
            self._urunleri_yukle_ve_cachele_ve_goster()
        except Exception as e:
            self.app.set_status_message(f"Hata: Başlangıç verileri yüklenemedi. Detay: {e}", "red")
            logging.error(f"FaturaPenceresi initial data yükleme hatası: {e}", exc_info=True)
            return

        if self.duzenleme_id:
            self._mevcut_faturayi_yukle()
        elif self.initial_data:
            self._load_data_from_initial_data()
        else:
            self._reset_form_for_new_invoice()
            
    def _get_baslik(self):
        if self.iade_modu_aktif:
            return "İade Faturası Oluştur"
        if self.duzenleme_id:
            return "Fatura Güncelleme"
        if self.islem_tipi == self.FATURA_TIP_SATIS:
            return "Yeni Satış Faturası"
        elif self.islem_tipi == self.FATURA_TIP_ALIS:
            return "Yeni Alış Faturası"
        return "Fatura"

    def _varsayilan_degerleri_yukle(self):
        # Varsayılan Perakende Müşteri'yi otomatik seç
        if self.fatura_tipi == self.db.FATURA_TIP_SATIS:
            perakende_musteri_id = self.db.get_perakende_musteri_id()
            if perakende_musteri_id:
                self._cari_sec_ui(perakende_musteri_id)
        # Varsayılan Genel Tedarikçi'yi otomatik seç (alış faturası için)
        elif self.fatura_tipi == self.db.FATURA_TIP_ALIS:
            genel_tedarikci_id = self.db.get_genel_tedarikci_id()
            if genel_tedarikci_id:
                self._cari_sec_ui(genel_tedarikci_id)

        # Varsayılan kasa/banka hesabını seç (Nakit ödeme türü için)
        # Hata düzeltildi: get_varsayilan_kasa_banka yerine get_kasa_banka_by_odeme_turu kullanıldı
        varsayilan_kb_info = self.db.get_kasa_banka_by_odeme_turu(self.db.ODEME_TURU_NAKIT)
        if varsayilan_kb_info and varsayilan_kb_info[0]: # (id, ad) tuple döndüğü varsayımıyla
            varsayilan_kb_id = varsayilan_kb_info[0]
            # Kasa/Banka combobox'ında bu ID'yi seç
            for i in range(self.kasa_banka_cb.count()):
                if self.kasa_banka_cb.itemData(i) == varsayilan_kb_id:
                    self.kasa_banka_cb.setCurrentIndex(i)
                    break
        else:
            logging.warning("Varsayılan KB çekme hatası: get_kasa_banka_by_odeme_turu metodu boş veya yanlış formatta döndü.")

        # Fatura numarası otomatik doldurma
        if self.fatura_tipi == self.db.FATURA_TIP_SATIS:
            self.fatura_no_input.setText(self.db.son_fatura_no_getir(self.db.FATURA_TIP_SATIS))
        elif self.fatura_tipi == self.db.FATURA_TIP_ALIS:
            self.fatura_no_input.setText(self.db.son_fatura_no_getir(self.db.FATURA_TIP_ALIS))

    def _create_ui(self):
            self.main_layout.setContentsMargins(10, 10, 10, 10)
            self.main_layout.setSpacing(15)

            # Üst ana yatay layout: Fatura Bilgileri ve Ürün Ekleme kısımlarını yan yana tutar
            top_main_h_layout = QHBoxLayout()
            top_main_h_layout.setSpacing(15)
            self.main_layout.addLayout(top_main_h_layout)
            # Ana layout'un üst kısmına daha az dikey esneme vererek alt kısımdaki sepetin büyümesini sağlar
            self.main_layout.setStretchFactor(top_main_h_layout, 1)

            # SOL KISIM: Fatura Bilgileri GroupBox
            fatura_detay_groupbox = QGroupBox("Fatura Bilgileri", self)
            fatura_detay_groupbox.setFont(QFont("Segoe UI", 10, QFont.Bold))
            fatura_detay_groupbox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred) # Expanding yatayda
            fatura_detay_layout = QGridLayout(fatura_detay_groupbox)
            fatura_detay_layout.setContentsMargins(10, 20, 10, 10)
            fatura_detay_layout.setSpacing(8)
            fatura_detay_layout.setHorizontalSpacing(10) # Yatay boşluğu azaltıldı

            top_main_h_layout.addWidget(fatura_detay_groupbox, 2) # Sol kısım, sağdan daha dar (streç faktör 2)

            # Sütun gerilmeleri ve hizalamaları - Sol panelin daha kompakt olması için
            fatura_detay_layout.setColumnStretch(0, 0) # İlk sütun (Label) minimum genişlik
            fatura_detay_layout.setColumnStretch(1, 1) # İkinci sütun (Input) esnesin
            fatura_detay_layout.setColumnStretch(2, 0) # Üçüncü sütun (Label) minimum genişlik
            fatura_detay_layout.setColumnStretch(3, 1) # Dördüncü sütun (Input) esnesin
            fatura_detay_layout.setColumnStretch(4, 0) # Beşinci sütun (Button) minimum genişlik
            fatura_detay_layout.setColumnStretch(5, 1) # En sağdaki boşluğu itmek için yeni bir streç sütun

            # Fatura No ve Tarih
            fatura_detay_layout.addWidget(QLabel("Fatura No:", fatura_detay_groupbox), 0, 0)
            self.f_no_e = QLineEdit(fatura_detay_groupbox)
            self.f_no_e.setFixedWidth(150)
            fatura_detay_layout.addWidget(self.f_no_e, 0, 1)

            fatura_detay_layout.addWidget(QLabel("Tarih:", fatura_detay_groupbox), 0, 2)
            self.fatura_tarihi_entry = QLineEdit(datetime.now().strftime('%Y-%m-%d'), fatura_detay_groupbox)
            self.fatura_tarihi_entry.setReadOnly(True)
            self.fatura_tarihi_entry.setFixedWidth(120)
            fatura_detay_layout.addWidget(self.fatura_tarihi_entry, 0, 3)
            self.btn_fatura_tarihi = QPushButton("🗓️", fatura_detay_groupbox)
            self.btn_fatura_tarihi.setFixedWidth(30)
            self.btn_fatura_tarihi.clicked.connect(lambda: DatePickerDialog(self.app, self.fatura_tarihi_entry))
            fatura_detay_layout.addWidget(self.btn_fatura_tarihi, 0, 4)

            # Cari Seçim
            cari_btn_label_text = "Cari Seç:"
            fatura_detay_layout.addWidget(QLabel(cari_btn_label_text, fatura_detay_groupbox), 1, 0)
            self.btn_cari_sec = QPushButton("Cari Seç...", fatura_detay_groupbox)
            self.btn_cari_sec.clicked.connect(self._cari_secim_penceresi_ac)
            fatura_detay_layout.addWidget(self.btn_cari_sec, 1, 1, 1, 2) # colspan 2

            self.lbl_secili_cari_adi = QLabel("Seçilen Cari: Yok", fatura_detay_groupbox)
            self.lbl_secili_cari_adi.setWordWrap(True)
            self.lbl_secili_cari_adi.setMinimumWidth(150)
            fatura_detay_layout.addWidget(self.lbl_secili_cari_adi, 1, 3, 1, 2, Qt.AlignLeft | Qt.AlignVCenter) # colspan 2

            self.lbl_cari_bakiye = QLabel("Bakiye: ---", fatura_detay_groupbox)
            self.lbl_cari_bakiye.setFont(QFont("Segoe UI", 9, QFont.Bold))
            fatura_detay_layout.addWidget(self.lbl_cari_bakiye, 2, 3, 1, 2, Qt.AlignLeft | Qt.AlignVCenter) # colspan 2

            # Misafir Adı (Gizli Başlangıçta)
            self.misafir_adi_container_frame = QFrame(fatura_detay_groupbox)
            misafir_layout = QHBoxLayout(self.misafir_adi_container_frame)
            misafir_layout.setContentsMargins(0, 0, 0, 0)
            misafir_layout.setSpacing(5)
            misafir_layout.addWidget(QLabel("Misafir Adı:", self.misafir_adi_container_frame))
            self.entry_misafir_adi = QLineEdit(self.misafir_adi_container_frame)
            misafir_layout.addWidget(self.entry_misafir_adi)
            fatura_detay_layout.addWidget(self.misafir_adi_container_frame, 2, 0, 1, 2) # span 2
            self.misafir_adi_container_frame.setVisible(False)

            # Ödeme Türü - Kasa/Banka - Vade Tarihi - Fatura Notları - Genel İskonto
            current_detail_row = 3 # Ödeme Türü'nün başladığı satır

            fatura_detay_layout.addWidget(QLabel("Ödeme Türü:", fatura_detay_groupbox), current_detail_row, 0)
            self.odeme_turu_cb = QComboBox(fatura_detay_groupbox)
            self.odeme_turu_cb.addItems([self.ODEME_TURU_NAKIT, self.ODEME_TURU_KART, self.ODEME_TURU_EFT_HAVALE, self.ODEME_TURU_CEK, self.ODEME_TURU_SENET, self.ODEME_TURU_ACIK_HESAP, self.ODEME_TURU_ETKISIZ_FATURA])
            self.odeme_turu_cb.setFixedWidth(180)
            fatura_detay_layout.addWidget(self.odeme_turu_cb, current_detail_row, 1)

            fatura_detay_layout.addWidget(QLabel("Kasa/Banka:", fatura_detay_groupbox), current_detail_row + 1, 0)
            self.islem_hesap_cb = QComboBox(fatura_detay_groupbox)
            self.islem_hesap_cb.setEnabled(False)
            self.islem_hesap_cb.setFixedWidth(220)
            fatura_detay_layout.addWidget(self.islem_hesap_cb, current_detail_row + 1, 1, 1, 3) # colspan 3

            self.lbl_vade_tarihi = QLabel("Vade Tarihi:", fatura_detay_groupbox)
            fatura_detay_layout.addWidget(self.lbl_vade_tarihi, current_detail_row + 2, 0)
            self.entry_vade_tarihi = QLineEdit(fatura_detay_groupbox)
            self.entry_vade_tarihi.setReadOnly(True)
            self.entry_vade_tarihi.setEnabled(False)
            self.entry_vade_tarihi.setFixedWidth(120)
            fatura_detay_layout.addWidget(self.entry_vade_tarihi, current_detail_row + 2, 1)
            self.btn_vade_tarihi = QPushButton("🗓️", fatura_detay_groupbox)
            self.btn_vade_tarihi.setFixedWidth(30)
            self.btn_vade_tarihi.clicked.connect(lambda: DatePickerDialog(self.app, self.entry_vade_tarihi))
            self.btn_vade_tarihi.setEnabled(False)
            fatura_detay_layout.addWidget(self.btn_vade_tarihi, current_detail_row + 2, 2)
            
            self.lbl_vade_tarihi.hide()
            self.entry_vade_tarihi.hide()
            self.btn_vade_tarihi.hide()

            fatura_detay_layout.addWidget(QLabel("Fatura Notları:", fatura_detay_groupbox), current_detail_row + 3, 0, Qt.AlignTop)
            self.fatura_notlari_text = QTextEdit(fatura_detay_groupbox)
            self.fatura_notlari_text.setFixedHeight(60)
            self.fatura_notlari_text.setMinimumWidth(250)
            fatura_detay_layout.addWidget(self.fatura_notlari_text, current_detail_row + 3, 1, 1, 4) # colspan 4

            # Genel İskonto - Bu kısım fatura_detay_groupbox'ında kalmalı
            fatura_detay_layout.addWidget(QLabel("Genel İskonto Tipi:", fatura_detay_groupbox), current_detail_row + 4, 0)
            self.genel_iskonto_tipi_cb = QComboBox(fatura_detay_groupbox)
            self.genel_iskonto_tipi_cb.addItems(["YOK", "YUZDE", "TUTAR"])
            self.genel_iskonto_tipi_cb.setFixedWidth(120)
            fatura_detay_layout.addWidget(self.genel_iskonto_tipi_cb, current_detail_row + 4, 1, Qt.AlignLeft)

            fatura_detay_layout.addWidget(QLabel("Genel İskonto Değeri:", fatura_detay_groupbox), current_detail_row + 4, 2)
            self.genel_iskonto_degeri_e = QLineEdit("0,00", fatura_detay_groupbox)
            setup_numeric_entry(self.app, self.genel_iskonto_degeri_e, decimal_places=2)
            self.genel_iskonto_degeri_e.setEnabled(False)
            self.genel_iskonto_degeri_e.setFixedWidth(100)
            fatura_detay_layout.addWidget(self.genel_iskonto_degeri_e, current_detail_row + 4, 3)

            # Boş satırları streç ile iterek yukarıya doğru sıkıştır
            fatura_detay_layout.setRowStretch(current_detail_row + 5, 1)


            # SAĞ KISIM: Ürün Ekleme GroupBox
            urun_ekle_groupbox = QGroupBox("Ürün Ekleme", self)
            urun_ekle_groupbox.setFont(QFont("Segoe UI", 10, QFont.Bold))
            urun_ekle_groupbox.setMinimumWidth(350) # Minimum genişlik eklendi
            urun_ekle_groupbox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred) # Genişlemesi için
            urun_ekle_layout = QGridLayout(urun_ekle_groupbox)
            urun_ekle_layout.setContentsMargins(10, 20, 10, 10)
            urun_ekle_layout.setSpacing(8)
            urun_ekle_layout.setHorizontalSpacing(15)

            top_main_h_layout.addWidget(urun_ekle_groupbox, 3) # Streç faktör 3 (sağ panel daha geniş olsun)

            urun_ekle_layout.addWidget(QLabel("Ürün Ara (Kod/Ad):", urun_ekle_groupbox), 0, 0, Qt.AlignRight)
            self.urun_arama_entry = QLineEdit(urun_ekle_groupbox)
            self.urun_arama_entry.setPlaceholderText("Ürün kodu veya adı ile ara...")
            urun_ekle_layout.addWidget(self.urun_arama_entry, 0, 1)
            urun_ekle_layout.setColumnStretch(0, 0) # Etiket için sabit genişlik
            urun_ekle_layout.setColumnStretch(1, 1) # Giriş kutusu için esnek genişlik

            self.urun_arama_sonuclari_tree = QTreeWidget(urun_ekle_groupbox)
            self.urun_arama_sonuclari_tree.setHeaderLabels(["Ürün Adı", "Kod", "Fiyat", "Stok"])
            self.urun_arama_sonuclari_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
            self.urun_arama_sonuclari_tree.setSelectionMode(QAbstractItemView.SingleSelection)
            self.urun_arama_sonuclari_tree.setRootIsDecorated(False)
            self.urun_arama_sonuclari_tree.setAlternatingRowColors(True)
            
            header = self.urun_arama_sonuclari_tree.header()
            header.setSectionResizeMode(0, QHeaderView.Stretch)
            header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(3, QHeaderView.ResizeToContents)

            # Ürün Arama Sonuçları başlık hizalaması
            for i in range(self.urun_arama_sonuclari_tree.columnCount()):
                self.urun_arama_sonuclari_tree.headerItem().setTextAlignment(i, Qt.AlignCenter)


            urun_ekle_layout.addWidget(self.urun_arama_sonuclari_tree, 1, 0, 1, 2)
            urun_ekle_layout.setRowStretch(1, 5) 

            # Miktar, Birim Fiyat, Stok, İskonto 1, İskonto 2 - Tümü Yanyana düzenleme
            urun_input_line_layout = QHBoxLayout()
            urun_input_line_layout.setSpacing(5) # Elemanlar arası küçük boşluk

            urun_input_line_layout.addWidget(QLabel("Mik.:", urun_ekle_groupbox))
            self.mik_e = QLineEdit("1", urun_ekle_groupbox)
            setup_numeric_entry(self.app, self.mik_e, decimal_places=2)
            self.mik_e.setFixedWidth(60) # Daha küçük ve sabit genişlik
            urun_input_line_layout.addWidget(self.mik_e)

            urun_input_line_layout.addWidget(QLabel("B.Fiyat:", urun_ekle_groupbox))
            self.birim_fiyat_e = QLineEdit("0,00", urun_ekle_groupbox)
            setup_numeric_entry(self.app, self.birim_fiyat_e, decimal_places=2)
            self.birim_fiyat_e.setFixedWidth(80) # Daha küçük ve sabit genişlik
            urun_input_line_layout.addWidget(self.birim_fiyat_e)
            
            urun_input_line_layout.addWidget(QLabel("Stok:", urun_ekle_groupbox))
            self.stk_l = QLabel("-", urun_ekle_groupbox)
            self.stk_l.setFont(QFont("Segoe UI", 9, QFont.Bold))
            urun_input_line_layout.addWidget(self.stk_l)

            # İskonto alanları
            urun_input_line_layout.addWidget(QLabel("İsk.1(%):", urun_ekle_groupbox))
            self.iskonto_yuzde_1_e = QLineEdit("0,00", urun_ekle_groupbox)
            setup_numeric_entry(self.app, self.iskonto_yuzde_1_e, decimal_places=2, max_value=100)
            self.iskonto_yuzde_1_e.setFixedWidth(60)
            urun_input_line_layout.addWidget(self.iskonto_yuzde_1_e)

            urun_input_line_layout.addWidget(QLabel("İsk.2(%):", urun_ekle_groupbox))
            self.iskonto_yuzde_2_e = QLineEdit("0,00", urun_ekle_groupbox)
            setup_numeric_entry(self.app, self.iskonto_yuzde_2_e, decimal_places=2, max_value=100)
            self.iskonto_yuzde_2_e.setFixedWidth(60)
            urun_input_line_layout.addWidget(self.iskonto_yuzde_2_e)
            
            urun_input_line_layout.addStretch() # En sağdaki boşluğu itmek için

            urun_ekle_layout.addLayout(urun_input_line_layout, 2, 0, 1, 2) # Yeni row'a ekle, span 2
            urun_ekle_layout.setRowStretch(2, 0)

            self.btn_sepete_ekle = QPushButton("Sepete Ekle", urun_ekle_groupbox)
            self.btn_sepete_ekle.setFont(QFont("Segoe UI", 10, QFont.Bold))
            self.btn_sepete_ekle.setStyleSheet("padding: 8px;")
            urun_ekle_layout.addWidget(self.btn_sepete_ekle, 3, 0, 1, 2)
            urun_ekle_layout.setRowStretch(3, 0)

            urun_ekle_layout.setRowStretch(4, 1) 

            # ALT KISIM: Fatura Kalemleri (Sepet) GroupBox
            sepet_groupbox = QGroupBox("Fatura Kalemleri", self)
            sepet_groupbox.setFont(QFont("Segoe UI", 10, QFont.Bold))
            sepet_layout = QVBoxLayout(sepet_groupbox)
            sepet_layout.setContentsMargins(10, 20, 10, 10)
            sepet_layout.setSpacing(10)
            self.main_layout.addWidget(sepet_groupbox)
            self.main_layout.setStretchFactor(sepet_groupbox, 10) 

            self.sep_tree = QTreeWidget(sepet_groupbox)
            self.sep_tree.setHeaderLabels(["#", "Ürün Adı", "Mik.", "B.Fiyat", "KDV%", "İskonto 1 (%)", "İskonto 2 (%)", "Uyg. İsk. Tutarı", "Tutar(Dah.)", "Fiyat Geçmişi", "Ürün ID"])
            self.sep_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
            self.sep_tree.setSelectionMode(QAbstractItemView.SingleSelection)
            self.sep_tree.setRootIsDecorated(False)
            self.sep_tree.setAlternatingRowColors(True)
            
            header = self.sep_tree.header()
            header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(1, QHeaderView.Stretch)
            header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(7, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(8, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(9, QHeaderView.Fixed)
            header.setSectionResizeMode(10, QHeaderView.Fixed)
            self.sep_tree.setColumnWidth(9, 90)
            self.sep_tree.setColumnHidden(10, True)

            # Fatura Kalemleri başlık hizalaması
            for i in range(self.sep_tree.columnCount()):
                self.sep_tree.headerItem().setTextAlignment(i, Qt.AlignCenter)


            sepet_layout.addWidget(self.sep_tree)

            # Sepet işlemleri butonları
            btn_sepet_islemleri_frame = QFrame(sepet_groupbox)
            btn_sepet_islemleri_layout = QHBoxLayout(btn_sepet_islemleri_frame)
            btn_sepet_islemleri_layout.setContentsMargins(0, 5, 0, 0)
            btn_sepet_islemleri_layout.addStretch()
            self.btn_secili_kalemi_sil = QPushButton("Seçili Kalemi Sil", btn_sepet_islemleri_frame)
            btn_sepet_islemleri_layout.addWidget(self.btn_secili_kalemi_sil)

            self.btn_sepeti_temizle = QPushButton("Tüm Kalemleri Sil", btn_sepet_islemleri_frame)
            btn_sepet_islemleri_layout.addWidget(self.btn_sepeti_temizle)
            sepet_layout.addWidget(btn_sepet_islemleri_frame)


            # EN ALT KISIM: Genel Toplamlar GroupBox ve Kaydet Butonu
            footer_groupbox = QGroupBox("Genel Toplamlar", self)
            footer_groupbox.setFont(QFont("Segoe UI", 10, QFont.Bold))
            
            # FOOTER'IN ANA LAYOUT'U QHBoxLayout olacak
            footer_main_h_layout = QHBoxLayout(footer_groupbox)
            footer_main_h_layout.setContentsMargins(10, 20, 10, 10)
            footer_main_h_layout.setSpacing(15) # Elemanlar arası boşluk
            self.main_layout.addWidget(footer_groupbox) # Dikeyde son element
            self.main_layout.setStretchFactor(footer_groupbox, 0) # Footer'a minimum dikey esneme ver


            # 1. Uygulanan Genel İskonto
            uygulanan_iskonto_frame = QFrame(footer_groupbox)
            uygulanan_iskonto_layout = QVBoxLayout(uygulanan_iskonto_frame)
            uygulanan_iskonto_layout.setContentsMargins(0,0,0,0)
            # NEW: Label for the descriptive text
            uygulanan_iskonto_layout.addWidget(QLabel("Uygulanan Genel İskonto:", uygulanan_iskonto_frame), alignment=Qt.AlignLeft)
            self.lbl_uygulanan_genel_iskonto = QLabel("0,00 TL", uygulanan_iskonto_frame)
            self.lbl_uygulanan_genel_iskonto.setFont(QFont("Segoe UI", 15, italic=True))
            uygulanan_iskonto_layout.addWidget(self.lbl_uygulanan_genel_iskonto, alignment=Qt.AlignLeft)
            footer_main_h_layout.addWidget(uygulanan_iskonto_frame)
            footer_main_h_layout.setStretchFactor(uygulanan_iskonto_frame, 1) # Esnesin


            # 2. KDV Hariç Toplam
            tkh_frame = QFrame(footer_groupbox)
            tkh_layout = QVBoxLayout(tkh_frame)
            tkh_layout.setContentsMargins(0,0,0,0)
            # NEW: Label for the descriptive text
            tkh_layout.addWidget(QLabel("KDV Hariç Toplam:", tkh_frame), alignment=Qt.AlignRight)
            self.tkh_l = QLabel("0,00 TL", tkh_frame)
            self.tkh_l.setFont(QFont("Segoe UI", 15, QFont.Bold))
            tkh_layout.addWidget(self.tkh_l, alignment=Qt.AlignRight)
            footer_main_h_layout.addWidget(tkh_frame)
            footer_main_h_layout.setStretchFactor(tkh_frame, 1) # Esnesin


            # 3. Toplam KDV
            tkdv_frame = QFrame(footer_groupbox)
            tkdv_layout = QVBoxLayout(tkdv_frame)
            tkdv_layout.setContentsMargins(0,0,0,0)
            # NEW: Label for the descriptive text
            tkdv_layout.addWidget(QLabel("Toplam KDV:", tkdv_frame), alignment=Qt.AlignRight)
            self.tkdv_l = QLabel("0,00 TL", tkdv_frame)
            self.tkdv_l.setFont(QFont("Segoe UI", 15, QFont.Bold))
            tkdv_layout.addWidget(self.tkdv_l, alignment=Qt.AlignRight)
            footer_main_h_layout.addWidget(tkdv_frame)
            footer_main_h_layout.setStretchFactor(tkdv_frame, 1) # Esnesin


            # 4. Genel Toplam
            gt_frame = QFrame(footer_groupbox)
            gt_layout = QVBoxLayout(gt_frame)
            gt_layout.setContentsMargins(0,0,0,0)
            # NEW: Label for the descriptive text
            gt_layout.addWidget(QLabel("Genel Toplam:", gt_frame), alignment=Qt.AlignRight)
            self.gt_l = QLabel("0,00 TL", gt_frame)
            self.gt_l.setFont(QFont("Segoe UI", 15, QFont.Bold))
            self.gt_l.setStyleSheet("color: navy;")
            gt_layout.addWidget(self.gt_l, alignment=Qt.AlignRight)
            footer_main_h_layout.addWidget(gt_frame)
            footer_main_h_layout.setStretchFactor(gt_frame, 1) # Esnesin
            
            # Kaydet Butonu çerçevesini ve içeriğini burada oluştur
            btn_kaydet_frame_wrapper = QFrame(footer_groupbox) # Yeni bir wrapper frame
            btn_kaydet_wrapper_layout = QVBoxLayout(btn_kaydet_frame_wrapper)
            btn_kaydet_wrapper_layout.setContentsMargins(0,0,0,0)

            # self.btn_kaydet tanımı burada yapılmalı
            self.btn_kaydet = QPushButton("Kaydet", btn_kaydet_frame_wrapper) 
            self.btn_kaydet.setFont(QFont("Segoe UI", 15, QFont.Bold))
            self.btn_kaydet.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; border-radius: 5px;")
            btn_kaydet_wrapper_layout.addWidget(self.btn_kaydet, alignment=Qt.AlignRight | Qt.AlignVCenter)
            
            footer_main_h_layout.addWidget(btn_kaydet_frame_wrapper) # Wrapper frame'i ekle
            footer_main_h_layout.setStretchFactor(btn_kaydet_frame_wrapper, 0) # Sabit genişlik

            # Diğer alt butonlar için ayrı bir layout
            button_layout = QHBoxLayout()
            self.main_layout.addLayout(button_layout) # Ana layout'a ekle

            self.btn_iptal = QPushButton("İptal")
            self.btn_iptal.clicked.connect(self.reject) # QDialog'u kapatır
            button_layout.addWidget(self.btn_iptal)

    def _mevcut_faturayi_yukle(self):
        """
        Düzenleme modunda mevcut faturanın bilgilerini API'den çeker ve forma yükler.
        """
        try:
            fatura_ana = self.db.fatura_getir_by_id(self.duzenleme_id)
            if not fatura_ana:
                self.app.set_status_message(f"Hata: Fatura ID {self.duzenleme_id} bulunamadı.", "red")
                return

            fatura_kalemleri_api = self.db.fatura_kalemleri_al(self.duzenleme_id)
            
            self.f_no_e.setText(fatura_ana.get('fatura_no', ''))
            self.fatura_tarihi_entry.setText(fatura_ana.get('tarih', ''))

            self.secili_cari_id = fatura_ana.get('cari_id')
            self.lbl_secili_cari_adi.setText(f"Seçilen Cari: {fatura_ana.get('cari_adi', 'Yok')}")
            self._on_cari_selected()

            self.odeme_turu_cb.setCurrentText(fatura_ana.get('odeme_turu', self.ODEME_TURU_NAKIT))
            
            if fatura_ana.get('kasa_banka_id'):
                for i in range(self.islem_hesap_cb.count()):
                    if self.islem_hesap_cb.itemData(i) == fatura_ana.get('kasa_banka_id'):
                        self.islem_hesap_cb.setCurrentIndex(i)
                        break

            self.entry_vade_tarihi.setText(fatura_ana.get('vade_tarihi', ''))
            self.fatura_notlari_text.setPlainText(fatura_ana.get('fatura_notlari', ''))
            self.genel_iskonto_tipi_cb.setCurrentText(fatura_ana.get('genel_iskonto_tipi', "YOK"))
            self.genel_iskonto_degeri_e.setText(f"{fatura_ana.get('genel_iskonto_degeri', 0.0):.2f}".replace('.', ','))
            self._on_genel_iskonto_tipi_changed()

            self.fatura_kalemleri_ui.clear()
            for k_api in fatura_kalemleri_api:
                urun_adi = self._get_urun_adi_by_id(k_api.get('urun_id'))

                self.fatura_kalemleri_ui.append((
                    k_api.get('urun_id'),
                    urun_adi,
                    k_api.get('miktar'),
                    k_api.get('birim_fiyat'),
                    k_api.get('kdv_orani'),
                    k_api.get('kdv_tutari', 0.0),
                    k_api.get('kalem_toplam_kdv_haric', 0.0),
                    k_api.get('kalem_toplam_kdv_dahil', 0.0),
                    k_api.get('alis_fiyati_fatura_aninda', 0.0),
                    k_api.get('kdv_orani_fatura_aninda', k_api.get('kdv_orani')),
                    k_api.get('iskonto_yuzde_1', 0.0),
                    k_api.get('iskonto_yuzde_2', 0.0),
                    k_api.get('iskonto_tipi', "YOK"),
                    k_api.get('iskonto_degeri', 0.0),
                    (k_api.get('kalem_toplam_kdv_dahil') / k_api.get('miktar')) if k_api.get('miktar') else 0.0
                ))
            
            self._sepeti_guncelle_ui()
            self.toplamlari_hesapla_ui()
            self.f_no_e.setEnabled(False)
            self.btn_cari_sec.setEnabled(False)

            if self.iade_modu_aktif:
                self.f_no_e.setEnabled(False)
                self.btn_cari_sec.setEnabled(False)
        
        except Exception as e:
            self.app.set_status_message(f"Hata: Fatura bilgileri yüklenemedi. Detay: {e}", "red")
            logging.error(f"Fatura yükleme hatası: {e}", exc_info=True)

    def _load_data_from_initial_data(self):
        self.f_no_e.setText(self.initial_data.get('fatura_no', self.db.son_fatura_no_getir(self.islem_tipi)))
        self.fatura_tarihi_entry.setText(self.initial_data.get('tarih', datetime.now().strftime('%Y-%m-%d')))
        
        self.odeme_turu_cb.setCurrentText(self.initial_data.get('odeme_turu', self.ODEME_TURU_NAKIT))
        self._odeme_turu_degisince_event_handler()

        self.secili_cari_id = self.initial_data.get('cari_id')
        self.lbl_secili_cari_adi.setText(f"Seçilen Cari: {self.initial_data.get('cari_adi', 'Yok')}")
        self._on_cari_selected()

        self.entry_vade_tarihi.setText(self.initial_data.get('vade_tarihi', ''))
        self.fatura_notlari_text.setPlainText(self.initial_data.get('fatura_notlari', ''))
        self.genel_iskonto_tipi_cb.setCurrentText(self.initial_data.get('genel_iskonto_tipi', "YOK"))
        self.genel_iskonto_degeri_e.setText(f"{self.initial_data.get('genel_iskonto_degeri', 0.0):.2f}".replace('.',','))
        self._on_genel_iskonto_tipi_changed()

        self.fatura_kalemleri_ui.clear()
        for k_init in self.initial_data.get('kalemler', []):
            urun_adi = self._get_urun_adi_by_id(k_init.get('urun_id'))

            kdv_orani_init = k_init.get('kdv_orani', 0.0)
            original_bf_haric_init = k_init.get('birim_fiyat')
            
            iskonto_yuzde_1_init = k_init.get('iskonto_yuzde_1', 0.0)
            iskonto_yuzde_2_init = k_init.get('iskonto_yuzde_2', 0.0)
            
            fiyat_iskonto_1_sonrasi_dahil_calc = original_bf_haric_init * (1 + kdv_orani_init / 100) * (1 - iskonto_yuzde_1_init / 100)
            iskontolu_birim_fiyat_kdv_dahil_calc = fiyat_iskonto_1_sonrasi_dahil_calc * (1 - iskonto_yuzde_2_init / 100)

            alis_fiyati_fatura_aninda_init = k_init.get('alis_fiyati_fatura_aninda', 0.0)

            self.fatura_kalemleri_ui.append((
                k_init.get('urun_id'), urun_adi, k_init.get('miktar'), 
                original_bf_haric_init,
                kdv_orani_init,
                0.0, 0.0, 0.0,
                alis_fiyati_fatura_aninda_init,
                kdv_orani_init,
                iskonto_yuzde_1_init,
                iskonto_yuzde_2_init,
                k_init.get('iskonto_tipi', "YOK"),
                k_init.get('iskonto_degeri', 0.0),
                iskontolu_birim_fiyat_kdv_dahil_calc
            ))
        
        self._sepeti_guncelle_ui()
        self.toplamlari_hesapla_ui()

    def _reset_form_for_new_invoice(self):
        self.duzenleme_id = None
        self.fatura_kalemleri_ui.clear()
        self._sepeti_guncelle_ui()
        self.toplamlari_hesapla_ui()

        # Fatura No'yu ayarla
        self.f_no_e.setText(self.db.son_fatura_no_getir(self.islem_tipi))
        self.fatura_tarihi_entry.setText(datetime.now().strftime('%Y-%m-%d'))
        self.odeme_turu_cb.setCurrentText(self.ODEME_TURU_NAKIT)
        self.entry_vade_tarihi.clear()
        self.fatura_notlari_text.clear()
        self.genel_iskonto_tipi_cb.setCurrentText("YOK")
        self.genel_iskonto_degeri_e.setText("0,00")
        self.btn_cari_sec.setEnabled(True)
        self._on_genel_iskonto_tipi_changed()

        # Cari seçimini temizle
        self._temizle_cari_secimi() 

        # Fatura tipine göre varsayılan cariyi ayarla
        if self.islem_tipi == self.FATURA_TIP_SATIS:
            if self.perakende_musteri_id:
                perakende_display_text = self.cari_id_to_display_map.get(str(self.perakende_musteri_id))
                if perakende_display_text:
                    self.secili_cari_id = self.perakende_musteri_id
                    self.secili_cari_adi = perakende_display_text
                    self.lbl_secili_cari_adi.setText(f"Seçilen Cari: {self.secili_cari_adi}")
                    self._on_cari_selected() # Seçilen cariye göre UI öğelerini güncelle (bakiye vb.)
                else:
                    logging.warning(f"Perakende müşteri ID {self.perakende_musteri_id} için gösterim metni bulunamadı.")
            else:
                logging.warning("Perakende müşteri ID veritabanından alınamadı.")
        elif self.islem_tipi == self.FATURA_TIP_ALIS: # Genel Tedarikçi varsayılan olarak seçilecek
            genel_tedarikci_id = self.db.get_genel_tedarikci_id()
            if genel_tedarikci_id:
                genel_tedarikci_display_text = self.cari_id_to_display_map.get(str(genel_tedarikci_id))
                if genel_tedarikci_display_text:
                    self.secili_cari_id = genel_tedarikci_id
                    self.secili_cari_adi = genel_tedarikci_display_text
                    self.lbl_secili_cari_adi.setText(f"Seçilen Cari: {self.secili_cari_adi}")
                    self._on_cari_selected()
                else:
                    logging.warning(f"Genel Tedarikçi ID {genel_tedarikci_id} için gösterim metni bulunamadı.")
            else:
                logging.warning("Genel Tedarikçi ID veritabanından alınamadı.")


        self.urun_arama_entry.clear()
        self.mik_e.setText("1")
        self.birim_fiyat_e.setText("0,00")
        self.stk_l.setText("-")
        self.iskonto_yuzde_1_e.setText("0,00")
        self.iskonto_yuzde_2_e.setText("0,00")

        QTimer.singleShot(0, self._urunleri_yukle_ve_cachele_ve_goster)
        self.f_no_e.setFocus()

    def _temizle_cari_secimi(self):
        self.secili_cari_id = None
        self.secili_cari_adi = ""
        self.lbl_secili_cari_adi.setText("Seçilen Cari: Yok")
        if hasattr(self, 'misafir_adi_container_frame'):
            self.misafir_adi_container_frame.setVisible(False)
            if hasattr(self, 'entry_misafir_adi'):
                self.entry_misafir_adi.clear()

    def _on_iade_modu_changed(self):
        self.setWindowTitle(self._get_baslik())

        if self.iade_modu_aktif:
            self.f_no_e.setEnabled(False)
            self.btn_cari_sec.setEnabled(False)
            
            self.odeme_turu_cb.setEnabled(True)
            self.islem_hesap_cb.setEnabled(True)
            self.entry_vade_tarihi.setEnabled(True)
            self.btn_vade_tarihi.setEnabled(True)

            self.fatura_notlari_text.setPlainText(f"Orijinal Fatura ID: {self.original_fatura_id_for_iade} için iade faturasıdır.")
            
            if hasattr(self, 'misafir_adi_container_frame'):
                self.misafir_adi_container_frame.setVisible(False)

            self._odeme_turu_degisince_event_handler()
            QMessageBox.information(self, "Bilgi", "İade Faturası modu aktif. Fatura No ve Cari kilitlenmiştir.")
        else:
            self.f_no_e.setEnabled(True)
            self.btn_cari_sec.setEnabled(True)
            self.fatura_notlari_text.clear()
            self._odeme_turu_degisince_event_handler()
    
    def _on_genel_iskonto_tipi_changed(self):
        selected_type = self.genel_iskonto_tipi_cb.currentText()
        if selected_type == "YOK":
            self.genel_iskonto_degeri_e.setEnabled(False)
            self.genel_iskonto_degeri_e.setText("0,00")
        else:
            self.genel_iskonto_degeri_e.setEnabled(True)
        self.toplamlari_hesapla_ui()

    def _odeme_turu_degisince_event_handler(self):
        selected_odeme_turu = self.odeme_turu_cb.currentText()
        
        is_acik_hesap = (selected_odeme_turu == self.ODEME_TURU_ACIK_HESAP)
        self.lbl_vade_tarihi.setVisible(is_acik_hesap)
        self.entry_vade_tarihi.setVisible(is_acik_hesap)
        self.btn_vade_tarihi.setVisible(is_acik_hesap)
        self.entry_vade_tarihi.setEnabled(is_acik_hesap)
        self.btn_vade_tarihi.setEnabled(is_acik_hesap)

        if is_acik_hesap and not self.entry_vade_tarihi.text():
            self.entry_vade_tarihi.setText((datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'))
        elif not is_acik_hesap:
            self.entry_vade_tarihi.clear()

        is_pesin_odeme = (selected_odeme_turu in self.pesin_odeme_turleri)
        self.islem_hesap_cb.setEnabled(is_pesin_odeme)

        if is_pesin_odeme:
            try:
                varsayilan_kb_info = self.db.get_kasa_banka_by_odeme_turu(selected_odeme_turu)
                if varsayilan_kb_info and varsayilan_kb_info[0]:
                    varsayilan_kb_id = varsayilan_kb_info[0]
                    for i in range(self.islem_hesap_cb.count()):
                        if self.islem_hesap_cb.itemData(i) == varsayilan_kb_id:
                            self.islem_hesap_cb.setCurrentIndex(i)
                            break
                elif self.islem_hesap_cb.count() > 0:
                    self.islem_hesap_cb.setCurrentIndex(0)
            except Exception as e:
                QMessageBox.warning(self, "Hata", f"Varsayılan kasa/banka çekilirken hata: {e}")
                logging.warning(f"Varsayılan KB çekme hatası: {e}")
                if self.islem_hesap_cb.count() > 0: self.islem_hesap_cb.setCurrentIndex(0)
                else: self.islem_hesap_cb.clear(); self.islem_hesap_cb.addItem("Hesap Yok", None); self.islem_hesap_cb.setEnabled(False)
        else:
            self.islem_hesap_cb.clear()
            self.islem_hesap_cb.addItem("Hesap Yok", None)
            self.islem_hesap_cb.setEnabled(False)

        is_perakende_satis_current = (self.islem_tipi == self.FATURA_TIP_SATIS and
                                      self.secili_cari_id == self.perakende_musteri_id)
        
        if hasattr(self, 'misafir_adi_container_frame'):
            self.misafir_adi_container_frame.setVisible(is_perakende_satis_current and not self.iade_modu_aktif)
            if hasattr(self, 'entry_misafir_adi'):
                self.entry_misafir_adi.setEnabled(is_perakende_satis_current and not self.iade_modu_aktif)
                if not (is_perakende_satis_current and not self.iade_modu_aktif):
                    self.entry_misafir_adi.clear()

    def _yukle_carileri(self):
        self.tum_cariler_cache_data = []
        self.cari_map_display_to_id = {}
        self.cari_id_to_display_map = {}
        
        api_url = ""
        
        if self.islem_tipi in [self.db.FATURA_TIP_SATIS, self.db.FATURA_TIP_SATIS_IADE]:
            api_url = "/musteriler/"
        elif self.islem_tipi in [self.db.FATURA_TIP_ALIS, self.db.FATURA_TIP_ALIS_IADE, self.db.FATURA_TIP_DEVIR_GIRIS]:
            api_url = "/tedarikciler/"
        
        if api_url:
            try:
                # `cariler_data`'nın `dict` mi yoksa `list` mi olduğunu kontrol ederek robust hale getirildi
                cariler_response = self.db.musteri_listesi_al() if self.islem_tipi in [self.db.FATURA_TIP_SATIS, self.db.FATURA_TIP_SATIS_IADE] else self.db.tedarikci_listesi_al()
                
                cariler = []
                if isinstance(cariler_response, dict) and "items" in cariler_response:
                    cariler = cariler_response["items"]
                elif isinstance(cariler_response, list): # Eğer API doğrudan liste dönüyorsa
                    cariler = cariler_response
                else:
                    logging.warning("Cari listesi API yanıtı beklenen formatta değil.")
                    self.app.set_status_message("Uyarı: Cari listesi API yanıtı beklenen formatta değil.", "orange")
                    return # Veri formatı uyumsuzsa işlemi durdur

                for c in cariler:
                    cari_id = c.get('id')
                    cari_ad = c.get('ad')
                    cari_kodu_gosterim = c.get('kod', "") if self.islem_tipi in [self.db.FATURA_TIP_SATIS, self.db.FATURA_TIP_SATIS_IADE] else c.get('tedarikci_kodu', "")
                    
                    display_text = f"{cari_ad} (Kod: {cari_kodu_gosterim})"
                    self.cari_map_display_to_id[display_text] = str(cari_id)
                    self.cari_id_to_display_map[str(cari_id)] = display_text
                    self.tum_cariler_cache_data.append(c)
                
                self.app.set_status_message(f"{len(cariler)} cari API'den yüklendi.")

            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Cari listesi çekilirken hata: {e}")
                logging.error(f"FaturaPenceresi cari listesi yükleme hatası: {e}", exc_info=True)
        else:
            self.app.set_status_message("Cari listesi yüklenemedi (geçersiz fatura tipi).")

    def _cari_secim_penceresi_ac(self):
        try:
            cari_tip_for_dialog = None
            if self.islem_tipi in [self.db.FATURA_TIP_SATIS, self.db.FATURA_TIP_SATIS_IADE]:
                cari_tip_for_dialog = self.CARI_TIP_MUSTERI
            elif self.islem_tipi in [self.db.FATURA_TIP_ALIS, self.db.FATURA_TIP_ALIS_IADE, self.db.FATURA_TIP_DEVIR_GIRIS]:
                cari_tip_for_dialog = self.CARI_TIP_TEDARIKCI

            from pencereler import CariSecimPenceresi
            dialog = CariSecimPenceresi(self, self.db, cari_tip_for_dialog, self._on_cari_secildi_callback)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Cari seçim penceresi açılırken hata: {e}")
            logging.error(f"Cari seçim penceresi açma hatası: {e}")

    def _on_cari_secildi_callback(self, selected_cari_id, selected_cari_display_text):
        self.secili_cari_id = selected_cari_id
        self.secili_cari_adi = selected_cari_display_text
        self.lbl_secili_cari_adi.setText(f"Seçilen Cari: {self.secili_cari_adi}")
        self._on_cari_selected()

    def _on_cari_selected(self):
        if not self.secili_cari_id:
            self.lbl_cari_bakiye.setText("Bakiye: ---")
            self._odeme_turu_degisince_event_handler()
            return
        
        cari_tip_for_bakiye = None
        if self.islem_tipi in [self.FATURA_TIP_SATIS, self.FATURA_TIP_SATIS_IADE]:
            cari_tip_for_bakiye = self.CARI_TIP_MUSTERI
        elif self.islem_tipi in [self.FATURA_TIP_ALIS, self.FATURA_TIP_ALIS_IADE, self.FATURA_TIP_DEVIR_GIRIS]:
            cari_tip_for_bakiye = self.CARI_TIP_TEDARIKCI
        
        if cari_tip_for_bakiye:
            bakiye_bilgisi = None
            try:
                bakiye_bilgisi = self.db.get_musteri_net_bakiye(self.secili_cari_id) if cari_tip_for_bakiye == self.CARI_TIP_MUSTERI else self.db.get_tedarikci_net_bakiye(self.secili_cari_id)
                
                if bakiye_bilgisi is not None:
                    bakiye_str = self.db._format_currency(bakiye_bilgisi)
                    if bakiye_bilgisi > 0:
                        self.lbl_cari_bakiye.setText(f"Borç: {bakiye_str}")
                        self.lbl_cari_bakiye.setStyleSheet("color: red;")
                    elif bakiye_bilgisi < 0:
                        self.lbl_cari_bakiye.setText(f"Alacak: {bakiye_str}")
                        self.lbl_cari_bakiye.setStyleSheet("color: green;")
                    else:
                        self.lbl_cari_bakiye.setText("Bakiye: 0,00 TL")
                        self.lbl_cari_bakiye.setStyleSheet("color: black;")
                else:
                    self.lbl_cari_bakiye.setText("Bakiye: ---")
                    self.lbl_cari_bakiye.setStyleSheet("color: black;")
            except Exception as e:
                self.lbl_cari_bakiye.setText("Bakiye: Hata")
                self.lbl_cari_bakiye.setStyleSheet("color: red;")
                logging.error(f"Cari bakiye çekilirken hata: {e}")

        self._odeme_turu_degisince_event_handler()

    def _urunleri_yukle_ve_cachele_ve_goster(self):
        try:
            filters = {"limit": 10000, "aktif_durum": True}
            stok_listeleme_sonucu = self.db.stok_listesi_al(**filters) # API çağrısı

            urunler = []
            if isinstance(stok_listeleme_sonucu, dict) and "items" in stok_listeleme_sonucu:
                urunler = stok_listeleme_sonucu["items"]
            elif isinstance(stok_listeleme_sonucu, list):
                urunler = stok_listeleme_sonucu
                self.app.set_status_message("Uyarı: Stok listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                urunler = []
                self.app.set_status_message("Hata: Stok listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Stok listesi API'den beklenen formatta gelmedi: {type(stok_listeleme_sonucu)} - {stok_listeleme_sonucu}")
                return

            self.tum_urunler_cache = urunler
            self.urun_map_filtrelenmis.clear()
            self.urun_arama_sonuclari_tree.clear()

            for urun in urunler: # urunler listesindeki her öğe için döngü
                if urun is not None: # EKLENEN KONTROL: urun'un None olup olmadığını kontrol et
                    birim_obj = urun.get('birim')
                    birim_ad_to_display = birim_obj.get('ad', '') if birim_obj else ''

                    item_text = f"{urun.get('kod', '')} - {urun.get('ad', '')} ({urun.get('miktar', 0):.2f} {birim_ad_to_display})"
                    item = QTreeWidgetItem(self.urun_arama_sonuclari_tree)
                    item.setText(0, urun.get('ad', '')) # Ürün Adı
                    item.setText(1, urun.get('kod', '')) # Kod

                    fiyat_gosterim = 0.0
                    # Fatura tipine göre fiyatı belirle
                    if self.islem_tipi == self.FATURA_TIP_SATIS:
                        fiyat_gosterim = urun.get('satis_fiyati', 0.0)
                    elif self.islem_tipi == self.FATURA_TIP_ALIS or self.islem_tipi == self.FATURA_TIP_DEVIR_GIRIS:
                        fiyat_gosterim = urun.get('alis_fiyati', 0.0)
                    elif self.islem_tipi == self.FATURA_TIP_SATIS_IADE:
                        fiyat_gosterim = urun.get('alis_fiyati', 0.0)
                    elif self.islem_tipi == self.FATURA_TIP_ALIS_IADE:
                        fiyat_gosterim = urun.get('satis_fiyati', 0.0)
                    # else durumu için fiyat_gosterim 0.0 olarak kalır

                    item.setText(2, self.db._format_currency(fiyat_gosterim)) # Fiyat
                    item.setText(3, f"{urun.get('miktar', 0):.2f}".rstrip('0').rstrip('.')) # Stok

                    item.setData(0, Qt.UserRole, urun['id'])
                    item.setData(2, Qt.UserRole, fiyat_gosterim)
                    item.setData(3, Qt.UserRole, urun.get('miktar', 0.0))

                    # Ürün Arama metinlerini hizalama
                    for col_idx in range(item.columnCount()):
                        if col_idx == 0: # Ürün Adı sütunu
                            item.setTextAlignment(col_idx, Qt.AlignLeft)
                        else:
                            item.setTextAlignment(col_idx, Qt.AlignCenter)

                    self.urun_map_filtrelenmis[urun['id']] = {
                        "id": urun['id'],
                        "kod": urun['kod'],
                        "ad": urun['ad'],
                        "alis_fiyati": urun.get('alis_fiyati'),
                        "satis_fiyati": urun.get('satis_fiyati'),
                        "kdv_orani": urun.get('kdv_orani'),
                        "miktar": urun.get('miktar'),
                        "birim": birim_obj # Birim objesini de saklayalım
                    }

            if len(self.urun_map_filtrelenmis) == 1:
                self.urun_arama_sonuclari_tree.setCurrentItem(self.urun_arama_sonuclari_tree.topLevelItem(0))
                self.urun_arama_sonuclari_tree.setFocus()
            
            self.urun_arama_sonuclari_tree.setVisible(bool(self.urun_map_filtrelenmis))

            self.app.set_status_message(f"{len(self.tum_urunler_cache)} ürün API'den önbelleğe alındı.")

        except Exception as e:
            logger.error(f"Ürün listesi yüklenirken hata oluştu: {e}", exc_info=True)
            self.app.set_status_message(f"Hata: Ürünler yüklenemedi. Detay: {e}", "red")     

    def _delayed_stok_yenile(self):
        if hasattr(self, '_delayed_timer') and self._delayed_timer.isActive():
            self._delayed_timer.stop()
        self._delayed_timer = QTimer(self)
        self._delayed_timer.setSingleShot(True)
        self._delayed_timer.timeout.connect(self._urun_listesini_filtrele_anlik)
        self._delayed_timer.start(300)

    def _urun_listesini_filtrele_anlik(self):
        arama_terimi = self.urun_arama_entry.text().lower().strip()
        self.urun_arama_sonuclari_tree.clear()
        self.urun_map_filtrelenmis.clear()

        for urun_item in self.tum_urunler_cache:
            urun_kodu = urun_item['kod']
            urun_adi = urun_item['ad']
            fiyat_to_display = urun_item['satis_fiyati'] if self.islem_tipi == 'SATIŞ' else urun_item['alis_fiyati']
            stok_db = urun_item['miktar']

            if (not arama_terimi or
                (urun_adi and arama_terimi in urun_adi.lower()) or
                (urun_kodu and arama_terimi in urun_kodu.lower())):

                item_iid = f"search_{urun_item['id']}"

                item_qt = QTreeWidgetItem(self.urun_arama_sonuclari_tree)
                item_qt.setText(0, urun_adi)
                item_qt.setText(1, urun_kodu)
                item_qt.setText(2, self.db._format_currency(fiyat_to_display))
                item_qt.setText(3, f"{stok_db:.2f}".rstrip('0').rstrip('.'))

                # Tüm sütunlardaki veriyi ortala ve fontu büyüt
                for col_idx in range(item_qt.columnCount()):
                    item_qt.setTextAlignment(col_idx, Qt.AlignCenter)
                    item_qt.setFont(col_idx, QFont("Segoe UI", 18))

                self.urun_map_filtrelenmis[item_iid] = {"id": urun_item['id'], "kod": urun_kodu, "ad": urun_adi, "fiyat": fiyat_to_display, "kdv": urun_item['kdv_orani'], "stok": stok_db}
        
        if len(self.urun_map_filtrelenmis) == 1:
            self.urun_arama_sonuclari_tree.setCurrentItem(self.urun_arama_sonuclari_tree.topLevelItem(0))
            self.urun_arama_sonuclari_tree.setFocus()
        self.urun_arama_sonuclari_tree.setVisible(bool(self.urun_map_filtrelenmis))
        self.secili_urun_bilgilerini_goster_arama_listesinden()
        
    def _select_product_from_search_list_and_focus_quantity(self, item):
        self._secili_urun_bilgilerini_goster_arama_listesinden(item)
        self.mik_e.setFocus()
        self.mik_e.selectAll()

    def _secili_urun_bilgilerini_goster_arama_listesinden(self): # 'item' parametresi kaldırıldı
        selected_items = self.urun_arama_sonuclari_tree.selectedItems()
        if selected_items:
            urun_id = selected_items[0].data(0, Qt.UserRole)
            if urun_id in self.urun_map_filtrelenmis:
                urun_detaylari = self.urun_map_filtrelenmis[urun_id]
                
                if self.islem_tipi == self.FATURA_TIP_SATIS:
                    birim_fiyat_to_fill = urun_detaylari.get('satis_fiyati', 0.0)
                elif self.islem_tipi == self.FATURA_TIP_ALIS:
                    birim_fiyat_to_fill = urun_detaylari.get('alis_fiyati', 0.0)
                elif self.islem_tipi == self.FATURA_TIP_SATIS_IADE:
                    birim_fiyat_to_fill = urun_detaylari.get('alis_fiyati', 0.0)
                elif self.islem_tipi == self.FATURA_TIP_ALIS_IADE:
                    birim_fiyat_to_fill = urun_detaylari.get('satis_fiyati', 0.0)
                else:
                    birim_fiyat_to_fill = 0.0

                self.birim_fiyat_e.setText(f"{birim_fiyat_to_fill:.2f}".replace('.',','))
                self.stk_l.setText(f"{urun_detaylari['miktar']:.2f}".rstrip('0').rstrip('.'))
                self.stk_l.setStyleSheet("color: black;")
            else:
                self.birim_fiyat_e.setText("0,00")
                self.stk_l.setText("-")
                self.stk_l.setStyleSheet("color: black;")
        else:
            self.birim_fiyat_e.setText("0,00")
            self.stk_l.setText("-")
            self.stk_l.setStyleSheet("color: black;")
            
    def _kalem_ekle_arama_listesinden(self):
        selected_items = self.urun_arama_sonuclari_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Geçersiz Ürün", "Lütfen arama listesinden bir ürün seçin.")
            return
        
        urun_id = selected_items[0].data(0, Qt.UserRole)
        if urun_id not in self.urun_map_filtrelenmis:
            QMessageBox.warning(self.app, "Geçersiz Ürün", "Seçili ürün detayları bulunamadı.")
            return
        
        urun_detaylari = self.urun_map_filtrelenmis[urun_id]
        
        try:
            miktar_str = self.mik_e.text().replace(',', '.')
            eklenecek_miktar = float(miktar_str) if miktar_str else 0.0
            if eklenecek_miktar <= 0:
                QMessageBox.critical(self.app, "Geçersiz Miktar", "Miktar pozitif bir sayı olmalıdır.")
                return

            birim_fiyat_str = self.birim_fiyat_e.text().replace(',', '.')
            birim_fiyat_kdv_dahil_input = float(birim_fiyat_str) if birim_fiyat_str else 0.0

            iskonto_1_str = self.iskonto_yuzde_1_e.text().replace(',', '.')
            iskonto_yuzde_1 = float(iskonto_1_str) if iskonto_1_str else 0.0
            
            iskonto_2_str = self.iskonto_yuzde_2_e.text().replace(',', '.')
            iskonto_yuzde_2 = float(iskonto_2_str) if iskonto_2_str else 0.0

        except ValueError:
            QMessageBox.critical(self.app, "Giriş Hatası", "Miktar veya fiyat alanlarına geçerli sayısal değerler girin.")
            return

        if self.islem_tipi in [self.FATURA_TIP_SATIS, self.FATURA_TIP_ALIS_IADE]:
            mevcut_stok = urun_detaylari.get('miktar', 0.0)
            
            sepetteki_urun_miktari = sum(k[2] for k in self.fatura_kalemleri_ui if k[0] == urun_id)
            
            if self.duzenleme_id:
                original_fatura_kalemleri = self._get_original_invoice_items_from_db(self.duzenleme_id)
                for orig_kalem in original_fatura_kalemleri:
                    if orig_kalem['urun_id'] == urun_id:
                        mevcut_stok += orig_kalem['miktar']
                        break
            
            if (sepetteki_urun_miktari + eklenecek_miktar) > mevcut_stok:
                reply = QMessageBox.question(self.app, "Stok Uyarısı",
                                             f"'{urun_detaylari['ad']}' için stok yetersiz!\n"
                                             f"Mevcut stok: {mevcut_stok:.2f} adet\n"
                                             f"Sepete eklenecek toplam: {sepetteki_urun_miktari + eklenecek_miktar:.2f} adet\n\n"
                                             "Devam etmek negatif stok oluşturacaktır. Emin misiniz?",
                                             QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.No: return

        existing_kalem_index = -1
        for i, kalem in enumerate(self.fatura_kalemleri_ui):
            if kalem[0] == urun_id:
                existing_kalem_index = i
                break

        urun_tam_detay_db = self.db.stok_getir_by_id(urun_id)
        if not urun_tam_detay_db:
            QMessageBox.critical(self.app, "Hata", "Ürün detayları veritabanında bulunamadı. Kalem eklenemiyor.")
            return

        original_birim_fiyat_kdv_haric = urun_tam_detay_db.get('alis_fiyati_kdv_haric') if self.islem_tipi == self.FATURA_TIP_ALIS else urun_tam_detay_db.get('satis_fiyati_kdv_haric')
        kdv_orani = urun_tam_detay_db.get('kdv_orani')
        alis_fiyati_fatura_aninda = urun_tam_detay_db.get('alis_fiyati')

        self.kalem_guncelle(
            kalem_index=existing_kalem_index, 
            yeni_miktar=eklenecek_miktar, 
            yeni_fiyat_kdv_dahil_orijinal=birim_fiyat_kdv_dahil_input,
            yeni_iskonto_yuzde_1=iskonto_yuzde_1, 
            yeni_iskonto_yuzde_2=iskonto_yuzde_2, 
            yeni_alis_fiyati_fatura_aninda=alis_fiyati_fatura_aninda,
            u_id=urun_id, 
            urun_adi=urun_detaylari['ad'],
            kdv_orani=kdv_orani
        )
        self.mik_e.setText("1")
        self.birim_fiyat_e.setText("0,00")
        self.iskonto_yuzde_1_e.setText("0,00")
        self.iskonto_yuzde_2_e.setText("0,00")
        self.urun_arama_entry.clear()
        self.urun_arama_entry.setFocus()
        
    def _double_click_add_to_cart(self, item):
        selected_items = self.urun_arama_sonuclari_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Geçersiz Ürün", "Lütfen sepete eklemek için arama listesinden bir ürün seçin.")
            return

        urun_id = selected_items[0].data(0, Qt.UserRole)
        if urun_id not in self.urun_map_filtrelenmis:
            QMessageBox.warning(self.app, "Geçersiz Ürün", "Seçili ürün detayları bulunamadı.")
            return
        
        urun_detaylari = self.urun_map_filtrelenmis[urun_id]
        
        # Fatura tipine göre varsayılan birim fiyatı belirle
        birim_fiyat_kdv_dahil_input = 0.0
        if self.islem_tipi == self.FATURA_TIP_SATIS or self.islem_tipi == self.FATURA_TIP_DEVIR_GIRIS:
            birim_fiyat_kdv_dahil_input = urun_detaylari.get('satis_fiyati', 0.0)
        elif self.islem_tipi == self.FATURA_TIP_ALIS:
            birim_fiyat_kdv_dahil_input = urun_detaylari.get('alis_fiyati', 0.0)
        elif self.islem_tipi == self.FATURA_TIP_SATIS_IADE:
            birim_fiyat_kdv_dahil_input = urun_detaylari.get('alis_fiyati', 0.0)
        elif self.islem_tipi == self.FATURA_TIP_ALIS_IADE:
            birim_fiyat_kdv_dahil_input = urun_detaylari.get('satis_fiyati', 0.0)

        # Varsayılan miktar 1 ve iskonto 0 olacak
        eklenecek_miktar = 1.0
        iskonto_yuzde_1 = 0.0
        iskonto_yuzde_2 = 0.0

        # Satış ve Satış İade faturalarında stok kontrolü yap
        if self.islem_tipi in [self.FATURA_TIP_SATIS, self.FATURA_TIP_ALIS_IADE]:
            mevcut_stok = urun_detaylari.get('miktar', 0.0)
            sepetteki_urun_miktari = sum(k[2] for k in self.fatura_kalemleri_ui if k[0] == urun_id)
            
            # Eğer mevcut bir fatura düzenleniyorsa, orijinal fatura kalemindeki miktarı mevcut stoka geri ekle
            if self.duzenleme_id:
                original_fatura_kalemleri = self._get_original_invoice_items_from_db(self.duzenleme_id)
                for orig_kalem in original_fatura_kalemleri:
                    if orig_kalem['urun_id'] == urun_id:
                        if self.islem_tipi == self.FATURA_TIP_SATIS:
                            mevcut_stok += orig_kalem['miktar']
                        elif self.islem_tipi == self.FATURA_TIP_ALIS_IADE:
                            mevcut_stok += orig_kalem['miktar']
                        break
            
            if (sepetteki_urun_miktari + eklenecek_miktar) > mevcut_stok:
                reply = QMessageBox.question(self.app, "Stok Uyarısı",
                                            f"'{urun_detaylari['ad']}' için stok yetersiz!\n"
                                            f"Mevcut stok: {mevcut_stok:.2f} adet\n"
                                            f"Sepete eklenecek toplam: {sepetteki_urun_miktari + eklenecek_miktar:.2f} adet\n\n"
                                            "Devam etmek negatif stok oluşturacaktır. Emin misiniz?",
                                            QMessageBox.Yes | QMessageBox.No)
                if reply == QMessageBox.No: return

        # Ürün sepette zaten varsa, sadece miktarını artır
        existing_kalem_index = -1
        for i, kalem in enumerate(self.fatura_kalemleri_ui):
            if kalem[0] == urun_id:
                existing_kalem_index = i
                # Çift tıklamada miktarını 1 artır
                eklenecek_miktar = kalem[2] + 1.0 # Mevcut miktara 1 ekle
                # Birim fiyat ve iskonto oranları aynı kalsın (ilk eklendiği gibi)
                birim_fiyat_kdv_dahil_input = kalem[14] # Güncel iskontolu birim fiyatı al
                iskonto_yuzde_1 = kalem[10]
                iskonto_yuzde_2 = kalem[11]
                break

        # Ürünün orijinal alış fiyatı, eğer satış faturasıysa. Kalem detayına kaydedilecek.
        alis_fiyati_fatura_aninda = urun_detaylari.get('alis_fiyati', 0.0)

        # kalem_guncelle metodunu kullanarak kalemi sepete ekle veya güncelle
        self.kalem_guncelle(
            kalem_index=existing_kalem_index,
            yeni_miktar=eklenecek_miktar,
            yeni_fiyat_kdv_dahil_orijinal=birim_fiyat_kdv_dahil_input,
            yeni_iskonto_yuzde_1=iskonto_yuzde_1,
            yeni_iskonto_yuzde_2=iskonto_yuzde_2,
            yeni_alis_fiyati_fatura_aninda=alis_fiyati_fatura_aninda,
            u_id=urun_id,
            urun_adi=urun_detaylari['ad'],
            kdv_orani=urun_detaylari.get('kdv_orani', 0.0)
        )

        # Alanları temizle ve arama kutusuna odaklan
        self.urun_arama_entry.clear()
        self.mik_e.setText("1")
        self.birim_fiyat_e.setText("0,00")
        self.iskonto_yuzde_1_e.setText("0,00")
        self.iskonto_yuzde_2_e.setText("0,00")
        self.stk_l.setText("-") # Stok etiketini temizle
        self.urun_arama_entry.setFocus()

    def kalem_guncelle(self, kalem_index, yeni_miktar, yeni_fiyat_kdv_dahil_orijinal, yeni_iskonto_yuzde_1, yeni_iskonto_yuzde_2, yeni_alis_fiyati_fatura_aninda, u_id=None, urun_adi=None, kdv_orani=None):
        if kalem_index != -1: # Mevcut kalem, GÜNCELLE
            item_to_update = list(self.fatura_kalemleri_ui[kalem_index])
            urun_id_current = item_to_update[0]
            kdv_orani_current = item_to_update[4]
        else: # Yeni kalem, OLUŞTUR
            if u_id is None or urun_adi is None or kdv_orani is None:
                QMessageBox.critical(self.app, "Hata", "Yeni kalem eklenirken ürün bilgileri eksik.")
                return
            urun_id_current = u_id
            kdv_orani_current = kdv_orani
            
            item_to_update = [
                u_id, urun_adi, 0.0, # miktar daha sonra güncellenecek
                0.0, kdv_orani_current, # birim_fiyat daha sonra güncellenecek
                0.0, 0.0, 0.0, # kdv_tutari, kalem_toplam_kdv_haric, kalem_toplam_kdv_dahil
                0.0, kdv_orani_current, # alis_fiyati_fatura_aninda, kdv_orani_fatura_aninda
                0.0, 0.0, # iskonto_yuzde_1, iskonto_yuzde_2
                "YOK", 0.0, # iskonto_tipi, iskonto_degeri
                0.0 # iskontolu_birim_fiyat_kdv_dahil
            ]

        item_to_update[2] = self.db.safe_float(yeni_miktar)
        item_to_update[10] = self.db.safe_float(yeni_iskonto_yuzde_1)
        item_to_update[11] = self.db.safe_float(yeni_iskonto_yuzde_2)
        item_to_update[8] = self.db.safe_float(yeni_alis_fiyati_fatura_aninda)

        if kdv_orani_current == 0:
            original_birim_fiyat_kdv_haric_calc = self.db.safe_float(yeni_fiyat_kdv_dahil_orijinal)
        else:
            original_birim_fiyat_kdv_haric_calc = self.db.safe_float(yeni_fiyat_kdv_dahil_orijinal) / (1 + self.db.safe_float(kdv_orani_current) / 100)
        item_to_update[3] = original_birim_fiyat_kdv_haric_calc

        fiyat_iskonto_1_sonrasi_dahil = self.db.safe_float(yeni_fiyat_kdv_dahil_orijinal) * (1 - self.db.safe_float(yeni_iskonto_yuzde_1) / 100)
        iskontolu_birim_fiyat_kdv_dahil = fiyat_iskonto_1_sonrasi_dahil * (1 - self.db.safe_float(yeni_iskonto_yuzde_2) / 100)
        
        if iskontolu_birim_fiyat_kdv_dahil < 0: iskontolu_birim_fiyat_kdv_dahil = 0.0
        item_to_update[14] = iskontolu_birim_fiyat_kdv_dahil

        iskontolu_birim_fiyat_kdv_haric = iskontolu_birim_fiyat_kdv_dahil / (1 + self.db.safe_float(kdv_orani_current) / 100) if self.db.safe_float(kdv_orani_current) != 0 else iskontolu_birim_fiyat_kdv_dahil

        item_to_update[5] = (iskontolu_birim_fiyat_kdv_dahil - iskontolu_birim_fiyat_kdv_haric) * self.db.safe_float(yeni_miktar)
        item_to_update[6] = iskontolu_birim_fiyat_kdv_haric * self.db.safe_float(yeni_miktar)
        item_to_update[7] = iskontolu_birim_fiyat_kdv_dahil * self.db.safe_float(yeni_miktar)

        if kalem_index != -1: # Mevcut kalemi güncelle
            self.fatura_kalemleri_ui[kalem_index] = tuple(item_to_update)
        else: # Yeni kalemi ekle
            self.fatura_kalemleri_ui.append(tuple(item_to_update))

        self.sepeti_guncelle_ui()
        self.toplamlari_hesapla_ui()
        
    def sepeti_guncelle_ui(self):
        if not hasattr(self, 'sep_tree'):
            return

        self.sep_tree.clear()

        for i, k in enumerate(self.fatura_kalemleri_ui):
            miktar_f = self.db.safe_float(k[2])
            birim_fiyat_gosterim_f = self.db.safe_float(k[14])
            original_bf_haric_f = self.db.safe_float(k[3])
            kdv_orani_f = self.db.safe_float(k[4])
            iskonto_yuzde_1_f = self.db.safe_float(k[10])
            iskonto_yuzde_2_f = self.db.safe_float(k[11])
            kalem_toplam_dahil_f = self.db.safe_float(k[7])
            
            miktar_gosterim = f"{miktar_f:.2f}".rstrip('0').rstrip('.')
            original_bf_dahil = original_bf_haric_f * (1 + kdv_orani_f / 100)
            uygulanan_iskonto = (original_bf_dahil - birim_fiyat_gosterim_f) * miktar_f

            item_qt = QTreeWidgetItem(self.sep_tree)
            item_qt.setText(0, str(i + 1))
            item_qt.setText(1, k[1])
            item_qt.setText(2, miktar_gosterim)
            item_qt.setText(3, self.db._format_currency(birim_fiyat_gosterim_f))
            item_qt.setText(4, f"%{kdv_orani_f:.0f}")
            item_qt.setText(5, f"{iskonto_yuzde_1_f:.2f}".replace('.',','))
            item_qt.setText(6, f"{iskonto_yuzde_2_f:.2f}".replace('.',','))
            item_qt.setText(7, self.db._format_currency(uygulanan_iskonto))
            item_qt.setText(8, self.db._format_currency(kalem_toplam_dahil_f))
            item_qt.setText(9, "Geçmişi Gör")
            item_qt.setText(10, str(k[0]))

            # Tüm sütunlardaki veriyi ortala ve fontu büyüt
            for col_idx in range(item_qt.columnCount()):
                item_qt.setTextAlignment(col_idx, Qt.AlignCenter)
                item_qt.setFont(col_idx, QFont("Segoe UI", 12))

            item_qt.setData(2, Qt.UserRole, miktar_f)
            item_qt.setData(3, Qt.UserRole, birim_fiyat_gosterim_f)
            item_qt.setData(4, Qt.UserRole, kdv_orani_f)
            item_qt.setData(5, Qt.UserRole, iskonto_yuzde_1_f)
            item_qt.setData(6, Qt.UserRole, iskonto_yuzde_2_f)
            item_qt.setData(7, Qt.UserRole, uygulanan_iskonto)
            item_qt.setData(8, Qt.UserRole, kalem_toplam_dahil_f)
            item_qt.setData(10, Qt.UserRole, k[0])

        self.toplamlari_hesapla_ui()

    def toplamlari_hesapla_ui(self):
        if not hasattr(self, 'tkh_l'): return

        toplam_kdv_haric_kalemler = sum(self.db.safe_float(k[6]) for k in self.fatura_kalemleri_ui)
        toplam_kdv_dahil_kalemler = sum(self.db.safe_float(k[7]) for k in self.fatura_kalemleri_ui)
        toplam_kdv_kalemler = sum(self.db.safe_float(k[5]) for k in self.fatura_kalemleri_ui)

        genel_iskonto_tipi = self.genel_iskonto_tipi_cb.currentText()
        genel_iskonto_degeri_str = self.genel_iskonto_degeri_e.text()
        genel_iskonto_degeri = self.db.safe_float(genel_iskonto_degeri_str) if self.genel_iskonto_degeri_e.isEnabled() else 0.0
        
        uygulanan_genel_iskonto_tutari = 0.0

        if genel_iskonto_tipi == 'YUZDE' and genel_iskonto_degeri > 0:
            uygulanan_genel_iskonto_tutari = toplam_kdv_haric_kalemler * (genel_iskonto_degeri / 100)
        elif genel_iskonto_tipi == 'TUTAR' and genel_iskonto_degeri > 0:
            uygulanan_genel_iskonto_tutari = genel_iskonto_degeri

        nihai_toplam_kdv_dahil = toplam_kdv_dahil_kalemler - uygulanan_genel_iskonto_tutari
        nihai_toplam_kdv_haric = toplam_kdv_haric_kalemler - uygulanan_genel_iskonto_tutari
        nihai_toplam_kdv = nihai_toplam_kdv_dahil - nihai_toplam_kdv_haric

        self.tkh_l.setText(f"KDV Hariç Toplam: {self.db._format_currency(nihai_toplam_kdv_haric)}")
        self.tkdv_l.setText(f"Toplam KDV: {self.db._format_currency(nihai_toplam_kdv)}")
        self.gt_l.setText(f"Genel Toplam: {self.db._format_currency(nihai_toplam_kdv_dahil)}")
        self.lbl_uygulanan_genel_iskonto.setText(f"Uygulanan Genel İskonto: {self.db._format_currency(uygulanan_genel_iskonto_tutari)}")

    def secili_kalemi_sil(self):
        selected_items = self.sep_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir kalem seçin.")
            return
            
        selected_item_qt = selected_items[0]
        kalem_index_str = selected_item_qt.text(0)
        try:
            kalem_index = int(kalem_index_str) - 1
        except ValueError:
            QMessageBox.critical(self.app, "Hata", "Seçili kalemin indeksi okunamadı.")
            return

        del self.fatura_kalemleri_ui[kalem_index]
        
        self.sepeti_guncelle_ui()
        self.toplamlari_hesapla_ui()
        
    def sepeti_temizle(self):
        if self.fatura_kalemleri_ui and QMessageBox.question(self.app, "Onay", "Tüm kalemleri silmek istiyor musunuz?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.fatura_kalemleri_ui.clear()
            self.sepeti_guncelle_ui()
            self.toplamlari_hesapla_ui()

    def _open_sepet_context_menu(self, pos):
        item = self.sep_tree.itemAt(pos)
        if not item: return

        self.sep_tree.setCurrentItem(item)

        context_menu = QMenu(self)

        edit_action = context_menu.addAction("Kalemi Düzenle")
        edit_action.triggered.connect(lambda: self._kalem_duzenle_penceresi_ac(item, None))

        delete_action = context_menu.addAction("Seçili Kalemi Sil")
        delete_action.triggered.connect(self.secili_kalemi_sil)

        history_action = context_menu.addAction("Fiyat Geçmişi")
        history_action.triggered.connect(lambda: self._on_sepet_kalem_click(item, 9))
        
        urun_karti_action = context_menu.addAction("Ürün Kartını Aç")
        urun_karti_action.triggered.connect(lambda: self._open_urun_karti_from_sep_item(item, None))

        context_menu.exec(self.sep_tree.mapToGlobal(pos))

    def _kalem_duzenle_penceresi_ac(self, item, column):
        kalem_index_str = item.text(0)
        try: kalem_index = int(kalem_index_str) - 1
        except ValueError:
            QMessageBox.critical(self.app, "Hata", "Seçili kalemin indeksi okunamadı."); return

        kalem_verisi = self.fatura_kalemleri_ui[kalem_index]
        from pencereler import KalemDuzenlePenceresi
        dialog = KalemDuzenlePenceresi(self, self.db, kalem_index, kalem_verisi, self.islem_tipi, self.duzenleme_id)
        dialog.exec()

    def _on_sepet_kalem_click(self, item, column):
        header_text = self.sep_tree.headerItem().text(column)
        if header_text == "Fiyat Geçmişi":
            urun_id_str = item.text(10)
            kalem_index_str = item.text(0)
            try:
                urun_id = int(urun_id_str)
                kalem_index = int(kalem_index_str) - 1
            except ValueError:
                QMessageBox.critical(self.app, "Hata", "Ürün ID veya kalem indeksi okunamadı."); return

            if not self.secili_cari_id:
                QMessageBox.warning(self.app, "Uyarı", "Fiyat geçmişini görmek için lütfen önce bir cari seçin."); return
            
            from pencereler import FiyatGecmisiPenceresi
            dialog = FiyatGecmisiPenceresi(self.app, self.db, self.secili_cari_id, urun_id, self.islem_tipi, self._update_sepet_kalem_from_history, kalem_index)
            dialog.exec()

    def _update_sepet_kalem_from_history(self, kalem_index, new_price_kdv_dahil, new_iskonto_1, new_iskonto_2):
        if not (0 <= kalem_index < len(self.fatura_kalemleri_ui)): return
        
        current_kalem_data = list(self.fatura_kalemleri_ui[kalem_index])
        
        urun_id = current_kalem_data[0]
        urun_adi = current_kalem_data[1]
        miktar = current_kalem_data[2]
        kdv_orani = current_kalem_data[4]
        alis_fiyati_fatura_aninda = current_kalem_data[8]

        self.kalem_guncelle(kalem_index=kalem_index, yeni_miktar=miktar, yeni_fiyat_kdv_dahil_orijinal=new_price_kdv_dahil, yeni_iskonto_yuzde_1=new_iskonto_1, yeni_iskonto_yuzde_2=new_iskonto_2, yeni_alis_fiyati_fatura_aninda=alis_fiyati_fatura_aninda, u_id=urun_id, urun_adi=urun_adi, kdv_orani=kdv_orani)


    def _get_urun_adi_by_id(self, urun_id):
        for urun in self.tum_urunler_cache:
            if urun.get('id') == urun_id:
                return urun.get('ad')
        return "Bilinmeyen Ürün"

    def _get_urun_full_details_by_id(self, urun_id):
        for urun in self.tum_urunler_cache:
            if urun.get('id') == urun_id:
                return urun
        return None

    def _get_original_invoice_items_from_db(self, fatura_id):
        try:
            return self.db.fatura_kalemleri_al(fatura_id)
        except Exception as e:
            logging.error(f"Orijinal fatura kalemleri çekilirken hata: {e}", exc_info=True)
            return []
        
    def _open_urun_karti_from_sep_item(self, item, column):
        urun_id_str = item.text(10)
        try: urun_id = int(urun_id_str)
        except ValueError: QMessageBox.critical(self.app, "Hata", "Ürün ID okunamadı."); return
        
        try:
            urun_detaylari = self.db.stok_getir_by_id(urun_id)
            if not urun_detaylari:
                QMessageBox.critical(self.app, "Hata", "Ürün detayları bulunamadı.")
                return
            from pencereler import StokKartiPenceresi
            dialog = StokKartiPenceresi(self.app, self.db, urun_duzenle=urun_detaylari, app_ref=self.app)
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün kartı açılamadı: {e}")
            logging.error(f"Ürün kartı açma hatası: {e}", exc_info=True)

    def _secili_kalemi_sil(self):
        selected_items = self.sep_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir kalem seçin.")
            return

        reply = QMessageBox.question(self.app, "Silme Onayı", "Seçili kalemi sepetten silmek istediğinizden emin misiniz?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            item_qt = selected_items[0]
            kalem_sira_no = int(item_qt.text(0))
            kalem_index = kalem_sira_no - 1

            if 0 <= kalem_index < len(self.fatura_kalemleri_ui):
                del self.fatura_kalemleri_ui[kalem_index]
                self._sepeti_guncelle_ui()
                self.toplamlari_hesapla_ui()
                QMessageBox.information(self.app, "Başarılı", "Kalem sepetten silindi.")
            else:
                QMessageBox.critical(self.app, "Hata", "Geçersiz kalem seçimi.")

    def _sepeti_temizle(self):
        if not self.fatura_kalemleri_ui:
            return

        reply = QMessageBox.question(self.app, "Temizleme Onayı", "Tüm kalemleri sepetten silmek istediğinizden emin misiniz?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.fatura_kalemleri_ui.clear()
            self._sepeti_guncelle_ui()
            self.toplamlari_hesapla_ui()
            QMessageBox.information(self.app, "Başarılı", "Sepet temizlendi.")

    def _sepeti_guncelle_ui(self):
        if not hasattr(self, 'sep_tree'):
            return

        self.sep_tree.clear()

        for i, k in enumerate(self.fatura_kalemleri_ui):
            miktar_f = self.db.safe_float(k[2])
            birim_fiyat_gosterim_f = self.db.safe_float(k[14])
            original_bf_haric_f = self.db.safe_float(k[3])
            kdv_orani_f = self.db.safe_float(k[4])
            iskonto_yuzde_1_f = self.db.safe_float(k[10])
            iskonto_yuzde_2_f = self.db.safe_float(k[11])
            kalem_toplam_dahil_f = self.db.safe_float(k[7])
            
            miktar_gosterim = f"{miktar_f:.2f}".rstrip('0').rstrip('.')
            original_bf_dahil = original_bf_haric_f * (1 + kdv_orani_f / 100)
            uygulanan_iskonto = (original_bf_dahil - birim_fiyat_gosterim_f) * miktar_f

            item_qt = QTreeWidgetItem(self.sep_tree)
            item_qt.setText(0, str(i + 1))
            item_qt.setText(1, k[1])
            item_qt.setText(2, miktar_gosterim)
            item_qt.setText(3, self.db._format_currency(birim_fiyat_gosterim_f))
            item_qt.setText(4, f"%{kdv_orani_f:.0f}")
            item_qt.setText(5, f"{iskonto_yuzde_1_f:.2f}".replace('.',','))
            item_qt.setText(6, f"{iskonto_yuzde_2_f:.2f}".replace('.',','))
            item_qt.setText(7, self.db._format_currency(uygulanan_iskonto))
            item_qt.setText(8, self.db._format_currency(kalem_toplam_dahil_f))
            item_qt.setText(9, "Geçmişi Gör")
            item_qt.setText(10, str(k[0]))

            item_qt.setData(2, Qt.UserRole, miktar_f)
            item_qt.setData(3, Qt.UserRole, birim_fiyat_gosterim_f)
            item_qt.setData(4, Qt.UserRole, kdv_orani_f)
            item_qt.setData(5, Qt.UserRole, iskonto_yuzde_1_f)
            item_qt.setData(6, Qt.UserRole, iskonto_yuzde_2_f)
            item_qt.setData(7, Qt.UserRole, uygulanan_iskonto)
            item_qt.setData(8, Qt.UserRole, kalem_toplam_dahil_f)
            item_qt.setData(10, Qt.UserRole, k[0])

        self.toplamlari_hesapla_ui()

    def _format_numeric_line_edit(self, line_edit: QLineEdit, decimals: int):
        text = line_edit.text()
        if not text: return

        # Noktayı virgüle çevir (Türkçe locale için)
        # Bu kısım sadece kullanıcının nokta girmesi durumunda çalışır.
        if '.' in text and ',' not in text:
            cursor_pos = line_edit.cursorPosition()
            line_edit.setText(text.replace('.', ','))
            # İmleç konumunu koru
            if cursor_pos <= len(line_edit.text()):
                line_edit.setCursorPosition(cursor_pos)
            else:
                line_edit.setCursorPosition(len(line_edit.text()))
            text = line_edit.text() # Güncellenmiş metni al

        # Eğer sadece ondalık ayıracı varsa, formatlamadan çık
        if text == ',' or text == '-':
            return
        
        try:
            # Virgülü noktaya çevirerek float'a dönüştür
            value = float(text.replace(',', '.'))
            
            # locale.format_string kullanarak binlik ayraçları ve ondalık basamakları ayarla
            # grouping=True ile binlik ayıracı ekler
            formatted_text = locale.format_string(f"%.{decimals}f", value, grouping=True)

            # Sadece değişiklik varsa güncelle ve imleci koru
            if line_edit.text() != formatted_text:
                current_cursor_pos = line_edit.cursorPosition()
                line_edit.setText(formatted_text)
                # İmleci mümkün olduğunca koru. Eğer formatlama sonucu metin uzadıysa,
                # imleci yeni uzunluğa göre ayarla.
                if current_cursor_pos <= len(formatted_text):
                    line_edit.setCursorPosition(current_cursor_pos)
                else:
                    line_edit.setCursorPosition(len(formatted_text))

        except ValueError:
            # Geçersiz bir değer girildiğinde formatlama yapma, ancak mevcut metni bırak
            pass
        except Exception as e:
            logging.error(f"Sayısal giriş formatlama hatası: {e}", exc_info=True)

    def _kaydet_fatura(self):
        fatura_no = self.f_no_e.text().strip()
        fatura_tarihi = self.fatura_tarihi_entry.text().strip()
        odeme_turu = self.odeme_turu_cb.currentText()
        vade_tarihi = self.entry_vade_tarihi.text().strip() if self.entry_vade_tarihi.isVisible() else None
        fatura_notlari = self.fatura_notlari_text.toPlainText().strip()
        genel_iskonto_tipi = self.genel_iskonto_tipi_cb.currentText()
        genel_iskonto_degeri = float(self.genel_iskonto_degeri_e.text().replace(',', '.')) if self.genel_iskonto_degeri_e.isEnabled() else 0.0
        misafir_adi = self.entry_misafir_adi.text().strip() if self.misafir_adi_container_frame.isVisible() else None

        kasa_banka_id = None
        if self.islem_hesap_cb.isEnabled() and self.islem_hesap_cb.currentData():
            kasa_banka_id = self.islem_hesap_cb.currentData()

        if not fatura_no: QMessageBox.critical(self, "Eksik Bilgi", "Fatura Numarası boş olamaz."); return
        try: datetime.strptime(fatura_tarihi, '%Y-%m-%d')
        except ValueError: QMessageBox.critical(self, "Hata", "Fatura Tarihi formatı (YYYY-AA-GG) olmalıdır."); return

        if not self.secili_cari_id and not misafir_adi: QMessageBox.critical(self, "Eksik Bilgi", "Lütfen bir cari seçin veya Misafir Adı girin."); return
        if odeme_turu == self.ODEME_TURU_ACIK_HESAP and not vade_tarihi: QMessageBox.critical(self, "Eksik Bilgi", "Açık Hesap için Vade Tarihi zorunludur."); return
        if vade_tarihi:
            try: datetime.strptime(vade_tarihi, '%Y-%m-%d')
            except ValueError: QMessageBox.critical(self, "Hata", "Vade Tarihi formatı (YYYY-AA-GG) olmalıdır."); return

        if odeme_turu in self.pesin_odeme_turleri and kasa_banka_id is None: QMessageBox.critical(self, "Eksik Bilgi", "Peşin ödeme türleri için Kasa/Banka seçimi zorunludur."); return
        if not self.fatura_kalemleri_ui: QMessageBox.critical(self, "Eksik Bilgi", "Faturada en az bir kalem olmalıdır."); return

        kalemler_to_send_to_api = []
        for k_ui in self.fatura_kalemleri_ui:
            kalemler_to_send_to_api.append({
                "urun_id": k_ui[0], "miktar": self.db.safe_float(k_ui[2]), "birim_fiyat": self.db.safe_float(k_ui[3]),
                "kdv_orani": self.db.safe_float(k_ui[4]), "alis_fiyati_fatura_aninda": self.db.safe_float(k_ui[8]),
                "iskonto_yuzde_1": self.db.safe_float(k_ui[10]), "iskonto_yuzde_2": self.db.safe_float(k_ui[11]),
                "iskonto_tipi": k_ui[12], "iskonto_degeri": self.db.safe_float(k_ui[13])
            })
    
        fatura_data = {
            "fatura_no": fatura_no, "tarih": fatura_tarihi, "fatura_turu": self.islem_tipi,
            "cari_id": self.secili_cari_id, "odeme_turu": odeme_turu, "kalemler": kalemler_to_send_to_api,
            "kasa_banka_id": kasa_banka_id, "misafir_adi": misafir_adi, "fatura_notlari": fatura_notlari,
            "vade_tarihi": vade_tarihi, "genel_iskonto_tipi": genel_iskonto_tipi,
            "genel_iskonto_degeri": genel_iskonto_degeri,
            "original_fatura_id": self.iade_modu_aktif and self.original_fatura_id_for_iade or None
        }

        try:
            if self.duzenleme_id: response = self.db.fatura_guncelle(self.duzenleme_id, fatura_data)
            else: response = self.db.fatura_ekle(fatura_data)
            
            QMessageBox.information(self, "Başarılı", "Fatura başarıyla kaydedildi!")
            
            if self.yenile_callback: self.yenile_callback()
            
            if not self.duzenleme_id: self.accept()
            else: self._reset_form_for_new_invoice()
            
        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "Hata", f"Fatura kaydedilirken bir hata oluştu:\n{error_detail}")
            logging.error(f"Fatura kaydetme hatası: {e}", exc_info=True)
            
    def _yukle_kasa_banka_hesaplarini(self):
        self.islem_hesap_cb.clear()
        self.kasa_banka_map.clear()
        
        try:
            hesaplar_response = self.db.kasa_banka_listesi_al()
            # API'den gelen yanıtın dict içinde 'items' anahtarı olup olmadığını kontrol et
            if isinstance(hesaplar_response, dict) and "items" in hesaplar_response:
                hesaplar = hesaplar_response["items"]
            elif isinstance(hesaplar_response, list): # Eğer API doğrudan liste dönüyorsa
                hesaplar = hesaplar_response
                self.app.set_status_message("Uyarı: Kasa/Banka listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else: # Beklenmeyen bir format gelirse
                hesaplar = []
                self.app.set_status_message("Hata: Kasa/Banka listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Kasa/Banka listesi API'den beklenen formatta gelmedi: {type(hesaplar_response)} - {hesaplar_response}")
                # Hata durumunda fonksiyonu sonlandır
                self.islem_hesap_cb.addItem("Hesap Yok", None)
                self.islem_hesap_cb.setEnabled(False)
                return


            if hesaplar:
                for h in hesaplar:
                    display_text = f"{h.get('hesap_adi')} ({h.get('tip')})"
                    if h.get('tip') == "BANKA" and h.get('banka_adi'):
                        display_text += f" - {h.get('banka_adi')}"
                    self.kasa_banka_map[display_text] = h.get('id')
                    self.islem_hesap_cb.addItem(display_text, h.get('id'))
                self.islem_hesap_cb.setCurrentIndex(0)
                self.islem_hesap_cb.setEnabled(True)
            else:
                self.islem_hesap_cb.clear()
                self.islem_hesap_cb.addItem("Hesap Yok", None)
                self.islem_hesap_cb.setEnabled(False)

            self.app.set_status_message(f"{len(hesaplar)} kasa/banka hesabı API'den yüklendi.")

        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Kasa/Banka hesapları çekilirken hata: {e}")
            logging.error(f"FaturaPenceresi Kasa/Banka yükleme hatası: {e}", exc_info=True)
            self.islem_hesap_cb.clear()
            self.islem_hesap_cb.addItem("Hesap Yok", None)
            self.islem_hesap_cb.setEnabled(False)

class FaturaDetayPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, fatura_id):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.fatura_id = fatura_id
        
        self.fatura_ana = None
        self.fatura_kalemleri_db = None

        # Fatura tip sabitlerini db_manager'dan al
        self.FATURA_TIP_ALIS = self.db.FATURA_TIP_ALIS
        self.FATURA_TIP_SATIS = self.db.FATURA_TIP_SATIS
        self.FATURA_TIP_DEVIR_GIRIS = self.db.FATURA_TIP_DEVIR_GIRIS
        self.FATURA_TIP_SATIS_IADE = self.db.FATURA_TIP_SATIS_IADE
        self.FATURA_TIP_ALIS_IADE = self.db.FATURA_TIP_ALIS_IADE

        self.ODEME_TURU_NAKIT = self.db.ODEME_TURU_NAKIT
        self.ODEME_TURU_KART = self.db.ODEME_TURU_KART
        self.ODEME_TURU_EFT_HAVALE = self.db.ODEME_TURU_EFT_HAVALE
        self.ODEME_TURU_CEK = self.db.ODEME_TURU_CEK
        self.ODEME_TURU_SENET = self.db.ODEME_TURU_SENET
        self.ODEME_TURU_ACIK_HESAP = self.db.ODEME_TURU_ACIK_HESAP
        self.ODEME_TURU_ETKISIZ_FATURA = self.db.ODEME_TURU_ETKISIZ_FATURA

        self.f_no = "Yükleniyor..."
        self.tip = ""

        self.setWindowTitle(f"Fatura Detayları: {self.f_no}")
        self.setWindowState(Qt.WindowMaximized)
        self.setModal(True)

        self.main_layout = QVBoxLayout(self)
        
        self._create_ui_and_populate_data()

    def _create_ui_and_populate_data(self):
        """Arayüzü oluşturur, API'den verileri çeker ve arayüzü doldurur."""
        if self.main_layout.layout():
            self.clear_layout(self.main_layout)
        try:
            self.fatura_ana = self.db.fatura_getir_by_id(self.fatura_id)
            if not self.fatura_ana:
                raise Exception("Fatura ana bilgileri API'den alınamadı.")
            
            self.fatura_kalemleri_db = self.db.fatura_kalemleri_al(self.fatura_id)
            if not self.fatura_kalemleri_db:
                logging.warning(f"Fatura ID {self.fatura_id} için fatura kalemi bulunamadı.")
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Fatura bilgileri çekilemedi: {e}")
            self.close()
            return
        
        # --- Veri Çekme ve Hazırlama ---
        self.f_no = self.fatura_ana.get('fatura_no', '-')
        self.tip = self.fatura_ana.get('fatura_turu', '-')
        self.setWindowTitle(f"Fatura Detayları: {self.f_no} ({self.tip})")

        tarih_db = self.fatura_ana.get('tarih')
        c_id = self.fatura_ana.get('cari_id')
        toplam_kdv_haric_fatura_ana_db = self.db.safe_float(self.fatura_ana.get('toplam_kdv_haric'))
        toplam_kdv_dahil_fatura_ana_db = self.db.safe_float(self.fatura_ana.get('toplam_kdv_dahil'))
        odeme_turu_db = self.fatura_ana.get('odeme_turu')
        misafir_adi_db = self.fatura_ana.get('misafir_adi')
        kasa_banka_id_db = self.fatura_ana.get('kasa_banka_id')
        olusturma_tarihi_saat = self.fatura_ana.get('olusturma_tarihi_saat')
        olusturan_kullanici_id = self.fatura_ana.get('olusturan_kullanici_id')
        son_guncelleme_tarihi_saat = self.fatura_ana.get('son_guncelleme_tarihi_saat')
        son_guncelleyen_kullanici_id = self.fatura_ana.get('son_guncelleyen_kullanici_id')
        fatura_notlari_db = self.fatura_ana.get('fatura_notlari')
        vade_tarihi_db = self.fatura_ana.get('vade_tarihi')
        genel_iskonto_tipi_db = self.fatura_ana.get('genel_iskonto_tipi')
        genel_iskonto_degeri_db = self.db.safe_float(self.fatura_ana.get('genel_iskonto_degeri'))
        
        try:
            kullanicilar_list = self.db.kullanici_listele()
            kullanicilar_map = {k.get('id'): k.get('kullanici_adi') for k in kullanicilar_list}
        except Exception as e:
            logger.error(f"Kullanıcı listesi API'den alınamadı: {e}")
            kullanicilar_map = {}
        
        olusturan_adi = kullanicilar_map.get(olusturan_kullanici_id, "Bilinmiyor")
        son_guncelleyen_adi = kullanicilar_map.get(son_guncelleyen_kullanici_id, "Bilinmiyor")
        
        cari_adi_text = "Bilinmiyor"
        if str(c_id) == str(self.db.get_perakende_musteri_id()) and self.fatura_ana.get('fatura_turu') == self.db.FATURA_TIP_SATIS:
            cari_adi_text = "Perakende Satış Müşterisi"
            if misafir_adi_db: cari_adi_text += f" (Misafir: {misafir_adi_db})"
        else:
            cari_bilgi_db = None
            if self.fatura_ana.get('fatura_turu') in [self.db.FATURA_TIP_SATIS, self.db.FATURA_TIP_SATIS_IADE]:
                cari_bilgi_db = self.db.musteri_getir_by_id(c_id)
                if cari_bilgi_db and cari_bilgi_db.get('kod'):
                    cari_adi_text = f"{cari_bilgi_db.get('ad')} (Kod: {cari_bilgi_db.get('kod')})"
            elif self.fatura_ana.get('fatura_turu') in [self.db.FATURA_TIP_ALIS, self.db.FATURA_TIP_ALIS_IADE]:
                cari_bilgi_db = self.db.tedarikci_getir_by_id(c_id)
                if cari_bilgi_db and cari_bilgi_db.get('kod'):
                    cari_adi_text = f"{cari_bilgi_db.get('ad')} (Kod: {cari_bilgi_db.get('kod')})"

        # --- Arayüz Oluşturma ---
        font_label = QFont("Segoe UI", 9, QFont.Bold)
        font_value = QFont("Segoe UI", 10)
        font_header = QFont("Segoe UI", 9, QFont.Bold)
        font_groupbox = QFont("Segoe UI", 10, QFont.Bold)
        
        # Ana yatay layout: Üst bilgiler ve toplamlar bir arada
        self.ust_bilgiler_frame = QFrame(self)
        self.ust_bilgiler_layout = QHBoxLayout(self.ust_bilgiler_frame)
        self.ust_bilgiler_layout.setContentsMargins(0, 0, 0, 0)
        self.ust_bilgiler_layout.setSpacing(15)
        self.main_layout.addWidget(self.ust_bilgiler_frame)

        # Sol Panel: Fatura ve Cari Bilgileri
        self.sol_panel_frame = QFrame(self.ust_bilgiler_frame)
        self.sol_panel_layout = QGridLayout(self.sol_panel_frame)
        self.sol_panel_layout.setContentsMargins(0, 0, 0, 0)
        self.sol_panel_layout.setSpacing(5)
        self.ust_bilgiler_layout.addWidget(self.sol_panel_frame)

        try: fatura_tarihi_formatted = datetime.strptime(str(tarih_db), '%Y-%m-%d').strftime('%d.%m.%Y')
        except: fatura_tarihi_formatted = str(tarih_db)
        
        self.sol_panel_layout.addWidget(QLabel("Fatura No:", font=font_label), 0, 0)
        self.sol_panel_layout.addWidget(QLabel(self.f_no, font=font_value), 0, 1)
        self.sol_panel_layout.addWidget(QLabel("Tarih:", font=font_label), 0, 2)
        self.sol_panel_layout.addWidget(QLabel(fatura_tarihi_formatted, font=font_value), 0, 3)

        cari_label_tipi = "Müşteri/Misafir:" if self.fatura_ana.get('fatura_turu') == self.db.FATURA_TIP_SATIS else "Tedarikçi:"
        self.sol_panel_layout.addWidget(QLabel(cari_label_tipi, font=font_label), 1, 0)
        self.sol_panel_layout.addWidget(QLabel(cari_adi_text, font=font_value), 1, 1, 1, 3)

        if kasa_banka_id_db:
            try:
                kb_bilgi = self.db.kasa_banka_getir_by_id(kasa_banka_id_db)
                if kb_bilgi:
                    self.sol_panel_layout.addWidget(QLabel("Kasa/Banka:", font=font_label), 2, 0)
                    self.sol_panel_layout.addWidget(QLabel(kb_bilgi.get('hesap_adi', '-'), font=font_value), 2, 1)
            except Exception as e:
                logging.error(f"Kasa/Banka bilgisi çekilirken hata: {e}")
        
        if odeme_turu_db == self.db.ODEME_TURU_ACIK_HESAP and vade_tarihi_db:
            self.sol_panel_layout.addWidget(QLabel("Vade Tarihi:", font=font_label), 2, 2)
            self.sol_panel_layout.addWidget(QLabel(str(vade_tarihi_db), font=font_value), 2, 3)
        
        genel_iskonto_gosterim_text = "Uygulanmadı"
        if genel_iskonto_tipi_db == 'YUZDE' and genel_iskonto_degeri_db is not None and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = f"Yüzde %{genel_iskonto_degeri_db:.2f}".replace('.', ',').rstrip('0').rstrip(',')
        elif genel_iskonto_tipi_db == 'TUTAR' and genel_iskonto_degeri_db is not None and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = self.db._format_currency(genel_iskonto_degeri_db)
        
        self.sol_panel_layout.addWidget(QLabel("Genel İskonto:", font=font_label), 3, 0)
        self.sol_panel_layout.addWidget(QLabel(genel_iskonto_gosterim_text, font=font_value), 3, 1)

        # Sağ Panel: Toplam Bilgileri
        self.sag_panel_frame = QFrame(self.ust_bilgiler_frame)
        self.sag_panel_layout = QGridLayout(self.sag_panel_frame)
        self.sag_panel_layout.setContentsMargins(0, 0, 0, 0)
        self.sag_panel_layout.setSpacing(5)
        self.ust_bilgiler_layout.addWidget(self.sag_panel_frame)
        self.ust_bilgiler_layout.setStretch(1, 1) # Sağ panelin yatayda daha çok esnemesi için

        toplam_kdv_hesaplanan_detay = toplam_kdv_dahil_fatura_ana_db - toplam_kdv_haric_fatura_ana_db
        toplam_kdv_dahil_kalemler_genel_iskonto_oncesi = sum(self.db.safe_float(k.get('kalem_toplam_kdv_dahil')) for k in self.fatura_kalemleri_db if isinstance(k, dict))
        gercek_uygulanan_genel_iskonto = self.db.safe_float(toplam_kdv_dahil_kalemler_genel_iskonto_oncesi) - self.db.safe_float(toplam_kdv_dahil_fatura_ana_db)
        if gercek_uygulanan_genel_iskonto < 0: gercek_uygulanan_genel_iskonto = 0.0

        self.sag_panel_layout.addWidget(QLabel("Toplam KDV Hariç:", font=font_label), 0, 0, Qt.AlignRight)
        self.tkh_l = QLabel(self.db._format_currency(toplam_kdv_haric_fatura_ana_db), font=font_value)
        self.sag_panel_layout.addWidget(self.tkh_l, 0, 1, Qt.AlignRight)
        
        self.sag_panel_layout.addWidget(QLabel("Toplam KDV:", font=font_label), 1, 0, Qt.AlignRight)
        self.tkdv_l = QLabel(self.db._format_currency(toplam_kdv_hesaplanan_detay), font=font_value)
        self.sag_panel_layout.addWidget(self.tkdv_l, 1, 1, Qt.AlignRight)
        
        self.sag_panel_layout.addWidget(QLabel("Genel Toplam:", font=font_header), 2, 0, Qt.AlignRight)
        self.gt_l = QLabel(self.db._format_currency(toplam_kdv_dahil_fatura_ana_db), font=QFont("Segoe UI", 12, QFont.Bold))
        self.sag_panel_layout.addWidget(self.gt_l, 2, 1, Qt.AlignRight)
        
        self.sag_panel_layout.addWidget(QLabel("Uygulanan Genel İskonto:", font=font_label), 3, 0, Qt.AlignRight)
        self.lbl_uygulanan_genel_iskonto = QLabel(self.db._format_currency(gercek_uygulanan_genel_iskonto), font=font_value)
        self.sag_panel_layout.addWidget(self.lbl_uygulanan_genel_iskonto, 3, 1, Qt.AlignRight)

        # Fatura Kalemleri GroupBox
        kalemler_frame = QGroupBox("Fatura Kalemleri", self)
        kalemler_frame.setFont(font_groupbox)
        kalemler_frame_layout = QVBoxLayout(kalemler_frame)
        self.main_layout.addWidget(kalemler_frame)
        
        cols_kalem = ("Sıra", "Ürün Kodu", "Ürün Adı", "Miktar", "Birim Fiyat", "KDV %", "İskonto 1 (%)", "İskonto 2 (%)", "Uyg. İsk. Tutarı", "Tutar (Dah.)", "Alış Fiyatı (Fatura Anı)")
        self.kalem_tree = QTreeWidget(kalemler_frame)
        self.kalem_tree.setHeaderLabels(cols_kalem)
        self.kalem_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.kalem_tree.setSortingEnabled(True)

        col_defs_kalem = [
            ("Sıra", 40, Qt.AlignCenter), ("Ürün Kodu", 90, Qt.AlignLeft), ("Ürün Adı", 180, Qt.AlignLeft),
            ("Miktar", 60, Qt.AlignRight), ("Birim Fiyat", 90, Qt.AlignRight), ("KDV %", 60, Qt.AlignRight),
            ("İskonto 1 (%)", 75, Qt.AlignRight), ("İskonto 2 (%)", 75, Qt.AlignRight),
            ("Uyg. İsk. Tutarı", 100, Qt.AlignRight), ("Tutar (Dah.)", 110, Qt.AlignRight),
            ("Alış Fiyatı (Fatura Anı)", 120, Qt.AlignRight)
        ]
        for i, (col_name, width, alignment) in enumerate(col_defs_kalem):
            self.kalem_tree.setColumnWidth(i, width)
            self.kalem_tree.headerItem().setTextAlignment(i, alignment)
            self.kalem_tree.headerItem().setFont(i, font_header)
        self.kalem_tree.header().setStretchLastSection(False)
        self.kalem_tree.header().setSectionResizeMode(2, QHeaderView.Stretch)
        kalemler_frame_layout.addWidget(self.kalem_tree)
        self._load_fatura_kalemleri_to_treeview(self.fatura_kalemleri_db)

        # Butonlar
        self._butonlari_olustur()

    def _butonlari_olustur(self):
        button_frame_alt = QFrame(self)
        button_layout_alt = QHBoxLayout(button_frame_alt)
        self.main_layout.addWidget(button_frame_alt)

        btn_guncelle = QPushButton("Güncelle")
        btn_guncelle.clicked.connect(self._open_fatura_guncelleme_penceresi)
        button_layout_alt.addWidget(btn_guncelle)
        
        btn_pdf_yazdir = QPushButton("PDF Yazdır")
        btn_pdf_yazdir.clicked.connect(self._handle_pdf_print)
        button_layout_alt.addWidget(btn_pdf_yazdir)

        button_layout_alt.addStretch()
        
        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        button_layout_alt.addWidget(btn_kapat)

    def _handle_pdf_print(self):
        """Fatura detay penceresinden PDF yazdırma işlemini başlatır."""
        dosya_adi_onek = f"{self.tip.capitalize()}Faturasi"
        file_path, _ = QFileDialog.getSaveFileName(self, f"{self.tip.capitalize()} Faturasını PDF Kaydet", 
                                                 f"{dosya_adi_onek}_{self.f_no.replace('/','_')}.pdf", 
                                                 "PDF Dosyaları (*.pdf);;Tüm Dosyalar (*)")
        if file_path:
            from pencereler import BeklemePenceresi
            bekleme_penceresi = BeklemePenceresi(self, message="Fatura PDF'e aktarılıyor, lütfen bekleyiniz...")
            QTimer.singleShot(0, bekleme_penceresi.exec)

            result_queue = multiprocessing.Queue()
            pdf_process = multiprocessing.Process(target=self.db.fatura_pdf_olustur, args=(self.fatura_id, file_path, result_queue))
            pdf_process.start()

            self.pdf_check_timer = QTimer(self)
            self.pdf_check_timer.timeout.connect(lambda: self._check_pdf_process_completion(result_queue, pdf_process, bekleme_penceresi))
            self.pdf_check_timer.start(100)
        else:
            self.app.set_status_message("PDF kaydetme iptal edildi.")

    def _check_pdf_process_completion(self, result_queue, pdf_process, bekleme_penceresi):
        if not result_queue.empty():
            success, message = result_queue.get()
            bekleme_penceresi.close()
            self.pdf_check_timer.stop()

            if success:
                QMessageBox.information(self, "Başarılı", message)
                self.app.set_status_message(message)
            else:
                QMessageBox.critical(self, "Hata", message)
                self.app.set_status_message(f"PDF kaydetme başarısız: {message}")
            pdf_process.join()
            
        elif not pdf_process.is_alive():
            bekleme_penceresi.close()
            self.pdf_check_timer.stop()
            QMessageBox.critical(self, "Hata", "PDF işlemi beklenmedik şekilde sonlandı.")
            pdf_process.join()

    def _open_fatura_guncelleme_penceresi(self):
        from pencereler import FaturaGuncellemePenceresi
        dialog = FaturaGuncellemePenceresi(
            self.app,
            self.db,
            self.fatura_id,
            yenile_callback_liste=self._fatura_guncellendi_callback_detay
        )
        dialog.exec()
        
    def _fatura_guncellendi_callback_detay(self):
        try:
            self.fatura_ana = self.db.fatura_getir_by_id(self.fatura_id)
            if not self.fatura_ana:
                raise Exception("Fatura ana bilgileri API'den alınamadı.")
            
            self.fatura_kalemleri_db = self.db.fatura_kalemleri_al(self.fatura_id)
            if not self.fatura_kalemleri_db:
                raise Exception("Fatura kalemleri API'den alınamadı.")
            
            self._create_ui_and_populate_data()
            self.app.set_status_message(f"Fatura '{self.f_no}' detayları güncellendi.")

        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Fatura detayları yenilenirken hata: {e}")
            logging.error(f"Fatura detay yenileme hatası: {e}", exc_info=True)
            self.close()
            return
                
        if hasattr(self.app, 'fatura_listesi_sayfasi'):
            if hasattr(self.app.fatura_listesi_sayfasi, 'satis_fatura_frame') and hasattr(self.app.fatura_listesi_sayfasi.satis_fatura_frame, 'fatura_listesini_yukle'):
                self.app.fatura_listesi_sayfasi.satis_fatura_frame.fatura_listesini_yukle()
            if hasattr(self.app.fatura_listesi_sayfasi, 'alis_fatura_frame') and hasattr(self.app.fatura_listesi_sayfasi.alis_fatura_frame, 'fatura_listesini_yukle'):
                self.app.fatura_listesi_sayfasi.alis_fatura_frame.fatura_listesini_yukle()
                
    def _load_fatura_kalemleri_to_treeview(self, kalemler_list):
        """API'den gelen fatura kalemlerini QTreeWidget'a yükler."""
        self.kalem_tree.clear()

        sira_idx = 1
        for kalem_item in kalemler_list:
            miktar_db = self.db.safe_float(kalem_item.get('miktar', 0.0))
            toplam_dahil_db = self.db.safe_float(kalem_item.get('kalem_toplam_kdv_dahil', 0.0))
            original_birim_fiyat_kdv_haric_item = self.db.safe_float(kalem_item.get('birim_fiyat', 0.0))
            original_kdv_orani_item = self.db.safe_float(kalem_item.get('kdv_orani', 0.0))
            
            iskontolu_birim_fiyat_kdv_dahil = 0.0
            uygulanan_toplam_iskonto_tutari_detay = 0.0

            if miktar_db != 0:
                iskontolu_birim_fiyat_kdv_dahil = toplam_dahil_db / miktar_db
                original_birim_fiyat_kdv_dahil_kalem = original_birim_fiyat_kdv_haric_item * (1 + original_kdv_orani_item / 100)
                iskonto_farki_per_birim_detay = original_birim_fiyat_kdv_dahil_kalem - iskontolu_birim_fiyat_kdv_dahil
                uygulanan_toplam_iskonto_tutari_detay = iskonto_farki_per_birim_detay * miktar_db

            item_qt = QTreeWidgetItem(self.kalem_tree)
            item_qt.setText(0, str(sira_idx))
            item_qt.setText(1, kalem_item.get('urun_kodu', ''))
            item_qt.setText(2, kalem_item.get('urun_adi', ''))
            item_qt.setText(3, f"{miktar_db:.2f}".rstrip('0').rstrip('.'))
            item_qt.setText(4, self.db._format_currency(iskontolu_birim_fiyat_kdv_dahil))
            item_qt.setText(5, f"%{kalem_item.get('kdv_orani', 0):.0f}")
            item_qt.setText(6, f"{kalem_item.get('iskonto_yuzde_1', 0):.2f}".replace('.', ',').rstrip('0').rstrip('.'))
            item_qt.setText(7, f"{kalem_item.get('iskonto_yuzde_2', 0):.2f}".replace('.', ',').rstrip('0').rstrip('.'))
            item_qt.setText(8, self.db._format_currency(uygulanan_toplam_iskonto_tutari_detay))
            item_qt.setText(9, self.db._format_currency(toplam_dahil_db))
            item_qt.setText(10, self.db._format_currency(kalem_item.get('alis_fiyati_fatura_aninda', 0.0)))
            
            sira_idx += 1

    def clear_layout(self, layout):
        if layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            if item is None:
                continue
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
            else:
                self.clear_layout(item.layout())

class YeniMusteriEklePenceresi(QDialog):
    def __init__(self, parent, db_manager, yenile_callback, musteri_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager # db_manager, CariService'e aktarılacak
        self.app = app_ref
        self.yenile_callback = yenile_callback
        self.musteri_duzenle_data = musteri_duzenle # API'den gelen düzenleme verisi

        # Eğer düzenleme modundaysak, ID'yi sakla
        self.musteri_duzenle_id = self.musteri_duzenle_data.get('id') if self.musteri_duzenle_data else None

        # CariService örneğini burada oluştur
        from hizmetler import CariService # CariService'i burada import ediyoruz
        self.cari_service = CariService(self.db) # <-- CariService BAŞLATILDI

        title = "Yeni Müşteri Ekle" if not self.musteri_duzenle_id else "Müşteri Düzenle"
        self.setWindowTitle(title)
        self.setMinimumSize(500, 420)
        self.setModal(True) # Bu pencere açıkken ana pencereye tıklamayı engeller

        # Ana layout
        main_layout = QVBoxLayout(self)
        
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        main_layout.addWidget(title_label)

        # Form için grid layout
        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)
        
        # Form elemanları
        self.entries = {}
        labels_entries = {
            "Müşteri Kodu:": "entry_kod",
            "Ad Soyad (*):": "entry_ad",
            "Telefon:": "entry_tel",
            "Adres:": "entry_adres",
            "Vergi Dairesi:": "entry_vd",
            "Vergi No:": "entry_vn"
        }

        for i, (label_text, entry_name) in enumerate(labels_entries.items()):
            form_layout.addWidget(QLabel(label_text), i, 0, alignment=Qt.AlignLeft)
            if entry_name == "entry_adres":
                widget = QTextEdit()
                widget.setFixedHeight(80) # Adres alanı için yükseklik
            else:
                widget = QLineEdit()
            
            self.entries[entry_name] = widget
            form_layout.addWidget(widget, i, 1)

        # Butonlar için yatay layout
        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch() # Butonları sağa yaslamak için boşluk ekle

        self.kaydet_button = QPushButton("Kaydet")
        self.kaydet_button.clicked.connect(self.kaydet)
        button_layout.addWidget(self.kaydet_button)
        
        self.iptal_button = QPushButton("İptal")
        self.iptal_button.clicked.connect(self.reject) # QDialog'u kapatır
        button_layout.addWidget(self.iptal_button)
        
        self._verileri_yukle()

    def _verileri_yukle(self):
        """Mevcut müşteri verilerini düzenleme modunda forma yükler."""
        if self.musteri_duzenle_data:
            self.entries["entry_kod"].setText(self.musteri_duzenle_data.get('kod', ''))
            self.entries["entry_ad"].setText(self.musteri_duzenle_data.get('ad', ''))
            self.entries["entry_tel"].setText(self.musteri_duzenle_data.get('telefon', ''))
            self.entries["entry_adres"].setPlainText(self.musteri_duzenle_data.get('adres', ''))
            self.entries["entry_vd"].setText(self.musteri_duzenle_data.get('vergi_dairesi', ''))
            self.entries["entry_vn"].setText(self.musteri_duzenle_data.get('vergi_no', ''))
            self.entries["entry_kod"].setReadOnly(True)
        else:
            generated_code = self.db.get_next_musteri_kodu()
            self.entries["entry_kod"].setText(generated_code)
            self.entries["entry_kod"].setReadOnly(True)

    def kaydet(self):
        ad = self.entries["entry_ad"].text().strip()
        if not ad:
            QMessageBox.warning(self, "Eksik Bilgi", "Müşteri Adı alanı boş bırakılamaz.")
            return

        data = {
            "ad": ad,
            "kod": self.entries["entry_kod"].text().strip(),
            "telefon": self.entries["entry_tel"].text().strip(),
            "adres": self.entries["entry_adres"].toPlainText().strip(),
            "vergi_dairesi": self.entries["entry_vd"].text().strip(),
            "vergi_no": self.entries["entry_vn"].text().strip()
        }

        try:
            if self.musteri_duzenle_id:
                success, message = self.db.musteri_guncelle(self.musteri_duzenle_id, data)
            else:
                success, message = self.db.musteri_ekle(data)

            if success:
                QMessageBox.information(self, "Başarılı", "Müşteri bilgileri başarıyla kaydedildi.")

                if self.yenile_callback:
                    self.yenile_callback()

                self.accept()
            else:
                QMessageBox.critical(self, "Hata", "Müşteri kaydedilirken bir hata oluştu.")

        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "Hata", f"Müşteri kaydedilirken bir hata oluştu:\n{error_detail}")
            logging.error(f"Müşteri kaydetme hatası: {error_detail}", exc_info=True)

class SiparisDetayPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, siparis_id, yenile_callback=None):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.siparis_id = siparis_id
        self.yenile_callback = yenile_callback

        # Fetch siparis data immediately to check existence
        try:
            # Düzeltildi: API'den sipariş bilgilerini db_manager üzerinden çek
            self.siparis_ana = self.db.siparis_getir_by_id(self.siparis_id)

            # Düzeltildi: API'den sipariş kalemlerini db_manager üzerinden çek
            self.siparis_kalemleri_db = self.db.siparis_kalemleri_al(self.siparis_id)

        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self.app, "API Hatası", f"Sipariş bilgileri çekilemedi: {e}")
            self.close() # Close dialog if data cannot be fetched
            return

        if not self.siparis_ana:
            QMessageBox.critical(self.app, "Sipariş Bulunamadı", "Detayları görüntülenecek sipariş bulunamadı.")
            self.close()
            return

        self.s_no = self.siparis_ana.get('siparis_no')
        durum_db = self.siparis_ana.get('durum')

        self.setWindowTitle(f"Sipariş Detayları: {self.s_no} ({durum_db})")
        self.setWindowState(Qt.WindowMaximized) # Maximize on start
        self.setModal(True)

        self.main_layout = QVBoxLayout(self) # Main layout for the dialog

        self._create_ui_and_populate_data()

        self.finished.connect(self.on_dialog_finished)

    def _create_ui_and_populate_data(self):
        # Bu metod, faturaya ait tüm verileri API'den çeker ve
        # arayüzü sıfırdan oluşturup doldurur.

        # Sipariş Genel Bilgileri
        self.ust_frame = QGroupBox(f"Sipariş Genel Bilgileri: {self.s_no}", self)
        self.ust_frame_layout = QGridLayout(self.ust_frame)
        self.main_layout.addWidget(self.ust_frame)

        # Kullanıcı bilgileri
        # Düzeltildi: Kullanıcı listesi db_manager üzerinden çekildi
        kullanicilar_list = self.db.kullanici_listele() 
        kullanicilar_map = {k.get('id'): k.get('kullanici_adi') for k in kullanicilar_list}

        olusturan_adi = kullanicilar_map.get(self.siparis_ana.get('olusturan_kullanici_id'), "Bilinmiyor")
        son_guncelleyen_adi = kullanicilar_map.get(self.siparis_ana.get('son_guncelleyen_kullanici_id'), "Bilinmiyor")

        # Cari Bilgisi
        cari_adi_text = "Bilinmiyor"
        if self.siparis_ana.get('cari_tip') == 'MUSTERI':
            # Düzeltildi: Müşteri bilgisi db_manager üzerinden çekildi
            cari_bilgi = self.db.musteri_getir_by_id(self.siparis_ana.get('cari_id'))
            cari_adi_text = f"{cari_bilgi.get('ad')} (Kod: {cari_bilgi.get('kod')})" if cari_bilgi else "Bilinmiyor"
        elif self.siparis_ana.get('cari_tip') == 'TEDARIKCI':
            # Düzeltildi: Tedarikçi bilgisi db_manager üzerinden çekildi
            cari_bilgi = self.db.tedarikci_getir_by_id(self.siparis_ana.get('cari_id'))
            cari_adi_text = f"{cari_bilgi.get('ad')} (Kod: {cari_bilgi.get('kod')})" if cari_bilgi else "Bilinmiyor"

        row_idx = 0
        self.ust_frame_layout.addWidget(QLabel("Sipariş No:", font=QFont("Segoe UI", 9, QFont.Bold)), row_idx, 0)
        self.ust_frame_layout.addWidget(QLabel(self.s_no, font=QFont("Segoe UI", 9)), row_idx, 1)
        try: siparis_tarihi_formatted = datetime.strptime(self.siparis_ana.get('tarih'), '%Y-%m-%d').strftime('%d.%m.%Y')
        except: siparis_tarihi_formatted = self.siparis_ana.get('tarih')
        self.ust_frame_layout.addWidget(QLabel("Tarih:", font=QFont("Segoe UI", 9, QFont.Bold)), row_idx, 2)
        self.ust_frame_layout.addWidget(QLabel(siparis_tarihi_formatted, font=QFont("Segoe UI", 9)), row_idx, 3)
        row_idx += 1
        self.ust_frame_layout.addWidget(QLabel("Cari Tipi:", font=QFont("Segoe UI", 9, QFont.Bold)), row_idx, 0)
        self.ust_frame_layout.addWidget(QLabel(self.siparis_ana.get('cari_tip'), font=QFont("Segoe UI", 9)), row_idx, 1)
        self.ust_frame_layout.addWidget(QLabel("Durum:", font=QFont("Segoe UI", 9, QFont.Bold)), row_idx, 2)
        self.ust_frame_layout.addWidget(QLabel(self.siparis_ana.get('durum'), font=QFont("Segoe UI", 9)), row_idx, 3)
        row_idx += 1
        self.ust_frame_layout.addWidget(QLabel("Cari Bilgisi:", font=QFont("Segoe UI", 9, QFont.Bold)), row_idx, 0)
        self.ust_frame_layout.addWidget(QLabel(cari_adi_text, font=QFont("Segoe UI", 9)), row_idx, 1, 1, 3)
        row_idx += 1
        self.ust_frame_layout.addWidget(QLabel("Teslimat Tarihi:", font=QFont("Segoe UI", 9, QFont.Bold)), row_idx, 0)
        try: teslimat_tarihi_formatted = datetime.strptime(self.siparis_ana.get('teslimat_tarihi'), '%Y-%m-%d').strftime('%d.%m.%Y')
        except: teslimat_tarihi_formatted = self.siparis_ana.get('teslimat_tarihi')
        self.ust_frame_layout.addWidget(QLabel(teslimat_tarihi_formatted if teslimat_tarihi_formatted else "-", font=QFont("Segoe UI", 9)), row_idx, 1)
        row_idx += 1
        genel_iskonto_gosterim_text = "Uygulanmadı"
        genel_iskonto_tipi_db = self.siparis_ana.get('genel_iskonto_tipi')
        genel_iskonto_degeri_db = self.siparis_ana.get('genel_iskonto_degeri')
        if genel_iskonto_tipi_db == 'YUZDE' and genel_iskonto_degeri_db is not None and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = f"Yüzde %{genel_iskonto_degeri_db:.2f}".replace('.', ',').rstrip('0').rstrip(',')
        elif genel_iskonto_tipi_db == 'TUTAR' and genel_iskonto_degeri_db is not None and genel_iskonto_degeri_db > 0:
            genel_iskonto_gosterim_text = self.db._format_currency(genel_iskonto_degeri_db)
        self.ust_frame_layout.addWidget(QLabel("Genel İskonto:", font=QFont("Segoe UI", 9, QFont.Bold)), row_idx, 0)
        self.ust_frame_layout.addWidget(QLabel(genel_iskonto_gosterim_text, font=QFont("Segoe UI", 9)), row_idx, 1, 1, 3)
        row_idx += 1
        self.ust_frame_layout.addWidget(QLabel("Oluşturulma:", font=QFont("Segoe UI", 8, QFont.StyleItalic)), row_idx, 0)
        self.ust_frame_layout.addWidget(QLabel(f"{self.siparis_ana.get('olusturma_tarihi_saat', '-') if self.siparis_ana.get('olusturma_tarihi_saat') else '-'} ({olusturan_adi})", font=QFont("Segoe UI", 8, QFont.StyleItalic)), row_idx, 1, 1, 3)
        row_idx += 1
        if self.siparis_ana.get('son_guncelleme_tarihi_saat'):
            self.ust_frame_layout.addWidget(QLabel("Son Güncelleme:", font=QFont("Segoe UI", 8, QFont.StyleItalic)), row_idx, 0)
            self.ust_frame_layout.addWidget(QLabel(f"{self.siparis_ana.get('son_guncelleme_tarihi_saat')} ({son_guncelleyen_adi})", font=QFont("Segoe UI", 8, QFont.StyleItalic)), row_idx, 1, 1, 3)
            row_idx += 1
        self.ust_frame_layout.addWidget(QLabel("Sipariş Notları:", font=QFont("Segoe UI", 9, QFont.Bold)), row_idx, 0, alignment=Qt.AlignTop) 
        siparis_notlari_display = QTextEdit()
        siparis_notlari_display.setPlainText(self.siparis_ana.get('siparis_notlari', '-') if self.siparis_ana.get('siparis_notlari') else "")
        siparis_notlari_display.setReadOnly(True)
        siparis_notlari_display.setFixedHeight(60)
        self.ust_frame_layout.addWidget(siparis_notlari_display, row_idx, 1, 1, 3)

        # Sipariş Kalemleri
        kalemler_frame = QGroupBox("Sipariş Kalemleri", self)
        kalemler_frame_layout = QVBoxLayout(kalemler_frame)
        self.main_layout.addWidget(kalemler_frame)

        cols_kalem = ("Sıra", "Ürün Kodu", "Ürün Adı", "Miktar", "Birim Fiyat", "KDV %", "İskonto 1 (%)", "İskonto 2 (%)", "Uyg. İsk. Tutarı", "Tutar (Dah.)", "Alış Fiyatı (Sipariş Anı)", "Satış Fiyatı (Sipariş Anı)")
        self.kalem_tree = QTreeWidget(kalemler_frame)
        self.kalem_tree.setHeaderLabels(cols_kalem)
        self.kalem_tree.setColumnCount(len(cols_kalem))
        self.kalem_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.kalem_tree.setSortingEnabled(True)

        from PySide6.QtWidgets import QHeaderView
        col_defs_kalem = [
            ("Sıra", 40, Qt.AlignCenter), ("Ürün Kodu", 90, Qt.AlignLeft), ("Ürün Adı", 180, Qt.AlignLeft),
            ("Miktar", 60, Qt.AlignRight), ("Birim Fiyat", 90, Qt.AlignRight), ("KDV %", 60, Qt.AlignRight),
            ("İskonto 1 (%)", 75, Qt.AlignRight), ("İskonto 2 (%)", 75, Qt.AlignRight),
            ("Uyg. İsk. Tutarı", 100, Qt.AlignRight), ("Tutar (Dah.)", 110, Qt.AlignRight),
            ("Alış Fiyatı (Sipariş Anı)", 120, Qt.AlignRight), ("Satış Fiyatı (Sipariş Anı)", 120, Qt.AlignRight)
        ]
        for i, (col_name, width, alignment) in enumerate(col_defs_kalem):
            self.kalem_tree.setColumnWidth(i, width)
            self.kalem_tree.headerItem().setTextAlignment(i, alignment)
            self.kalem_tree.headerItem().setFont(i, QFont("Segoe UI", 9, QFont.Bold))
        self.kalem_tree.header().setStretchLastSection(False)
        self.kalem_tree.header().setSectionResizeMode(2, QHeaderView.Stretch) # Ürün Adı genişlesin

        kalemler_frame_layout.addWidget(self.kalem_tree)
        self._load_siparis_kalemleri_to_treeview(self.siparis_kalemleri_db)


        # Alt Toplamlar ve İskonto Bilgileri
        alt_toplam_iskonto_frame = QFrame(self)
        alt_toplam_iskonto_frame_layout = QGridLayout(alt_toplam_iskonto_frame)
        self.main_layout.addWidget(alt_toplam_iskonto_frame)

        self.lbl_genel_toplam = QLabel(f"Genel Toplam (KDV Dahil): {self.db._format_currency(self.siparis_ana.get('toplam_tutar'))}", font=QFont("Segoe UI", 10, QFont.Bold))
        alt_toplam_iskonto_frame_layout.addWidget(self.lbl_genel_toplam, 0, 1, 1, 2, Qt.AlignRight)
        alt_toplam_iskonto_frame_layout.setColumnStretch(0, 1) # Sol tarafı esnet

        self._butonlari_olustur()

    def _load_siparis_kalemleri_to_treeview(self, kalemler_list):
        self.kalem_tree.clear()
        sira_idx = 1
        for k_db in kalemler_list:
            # Düzeltildi: urun_info db_manager üzerinden çekildi
            urun_info = self.db.stok_getir_by_id(k_db.get('urun_id'))
            if not urun_info:
                urun_kodu_db = "Bilinmiyor"
                urun_adi_db = "Bilinmiyor"
            else:
                urun_kodu_db = urun_info.get('kod', 'Bilinmiyor') # 'urun_kodu' yerine 'kod' kullanıldı
                urun_adi_db = urun_info.get('ad', 'Bilinmiyor') # 'urun_adi' yerine 'ad' kullanıldı

            miktar_gosterim = f"{k_db.get('miktar'):.2f}".rstrip('0').rstrip('.')
            iskontolu_birim_fiyat_kdv_dahil_display = (k_db.get('kalem_toplam_kdv_dahil') / k_db.get('miktar')) if k_db.get('miktar') != 0 else 0.0
            iskonto_yuzde_1_display = f"{k_db.get('iskonto_yuzde_1'):.2f}".replace('.', ',').rstrip('0').rstrip(',')
            iskonto_yuzde_2_display = f"{k_db.get('iskonto_yuzde_2'):.2f}".replace('.', ',').rstrip('0').rstrip(',')

            original_birim_fiyat_kdv_dahil_kalem = k_db.get('birim_fiyat') * (1 + k_db.get('kdv_orani') / 100)
            iskonto_farki_per_birim_detay = original_birim_fiyat_kdv_dahil_kalem - iskontolu_birim_fiyat_kdv_dahil_display
            uygulanan_toplam_iskonto_tutari_detay = iskonto_farki_per_birim_detay * k_db.get('miktar')

            item_qt = QTreeWidgetItem(self.kalem_tree)
            item_qt.setText(0, str(sira_idx))
            item_qt.setText(1, urun_kodu_db)
            item_qt.setText(2, urun_adi_db)
            item_qt.setText(3, miktar_gosterim)
            item_qt.setText(4, self.db._format_currency(iskontolu_birim_fiyat_kdv_dahil_display))
            item_qt.setText(5, f"%{k_db.get('kdv_orani'):.0f}")
            item_qt.setText(6, iskonto_yuzde_1_display)
            item_qt.setText(7, iskonto_yuzde_2_display)
            item_qt.setText(8, self.db._format_currency(uygulanan_toplam_iskonto_tutari_detay))
            item_qt.setText(9, self.db._format_currency(k_db.get('kalem_toplam_kdv_dahil')))
            item_qt.setText(10, self.db._format_currency(k_db.get('alis_fiyati_siparis_aninda')))
            item_qt.setText(11, self.db._format_currency(k_db.get('satis_fiyati_siparis_aninda')))

            sira_idx += 1

    def _butonlari_olustur(self):
        button_frame_alt = QFrame(self)
        button_frame_alt_layout = QHBoxLayout(button_frame_alt)
        self.main_layout.addWidget(button_frame_alt)

        self.faturaya_donustur_button_detail = QPushButton("Faturaya Dönüştür")
        self.faturaya_donustur_button_detail.clicked.connect(self._faturaya_donustur)
        button_frame_alt_layout.addWidget(self.faturaya_donustur_button_detail)

        btn_siparisi_duzenle = QPushButton("Siparişi Düzenle")
        btn_siparisi_duzenle.clicked.connect(self._siparisi_duzenle)
        button_frame_alt_layout.addWidget(btn_siparisi_duzenle)

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close) # QDialog'u kapat
        button_frame_alt_layout.addWidget(btn_kapat)

        if self.siparis_ana.get('fatura_id'):
            self.faturaya_donustur_button_detail.setEnabled(False)
            fatura_no_text = ""
            try:
                # Düzeltildi: db_manager metodu kullanıldı
                fatura_data = self.db.fatura_getir_by_id(self.siparis_ana.get('fatura_id'))
                fatura_no_text = fatura_data.get('fatura_no', '-')
            except Exception: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
                fatura_no_text = "Hata"

            lbl_fatura_iliskisi = QLabel(f"Bu sipariş Fatura No: '{fatura_no_text}' ile ilişkilendirilmiştir.")
            lbl_fatura_iliskisi.setStyleSheet("color: blue; font-style: italic;")
            button_frame_alt_layout.addWidget(lbl_fatura_iliskisi)

    def _faturaya_donustur(self):
        """Bu siparişi satış veya alış faturasına dönüştürür."""

        from pencereler import OdemeTuruSecimDialog

        fatura_tipi_for_dialog = 'SATIŞ' if self.siparis_ana.get('cari_tip') == 'MUSTERI' else 'ALIŞ'

        dialog = OdemeTuruSecimDialog(
            self.app,
            self.db,
            fatura_tipi_for_dialog,
            self.siparis_ana.get('cari_id'),
            self._faturaya_donustur_on_dialog_confirm
        )
        dialog.exec()

    def _faturaya_donustur_on_dialog_confirm(self, selected_odeme_turu, selected_kasa_banka_id, selected_vade_tarihi):
        if selected_odeme_turu is None:
            self.app.set_status_message("Faturaya dönüştürme iptal edildi (ödeme türü seçilmedi).")
            return

        confirm_msg = (f"'{self.s_no}' numaralı siparişi '{selected_odeme_turu}' ödeme türü ile faturaya dönüştürmek istediğinizden emin misiniz?\n"
                    f"Bu işlem sonucunda yeni bir fatura oluşturulacak ve sipariş durumu güncellenecektir.")
        if selected_odeme_turu == "AÇIK HESAP" and selected_vade_tarihi:
            confirm_msg += f"\nVade Tarihi: {selected_vade_tarihi}"
        if selected_kasa_banka_id:
            try:
                kb_bilgi = self.db.kasa_banka_getir_by_id(selected_kasa_banka_id)
                if kb_bilgi:
                    confirm_msg += f"\nİşlem Kasa/Banka: {kb_bilgi.get('hesap_adi')}"
            except Exception as e:
                logging.error(f"Kasa/Banka bilgisi çekilirken hata: {e}")
                confirm_msg += "\nİşlem Kasa/Banka: Bilgi çekilemedi"

        reply = QMessageBox.question(self, "Faturaya Dönüştür Onayı", confirm_msg, QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.No:
            return

        from hizmetler import FaturaService
        fatura_service = FaturaService(self.db)

        success, message = fatura_service.siparis_faturaya_donustur(
            self.siparis_id,
            self.app.current_user[0] if self.app and hasattr(self.app, 'current_user') and self.app.current_user else None,
            selected_odeme_turu,
            selected_kasa_banka_id,
            selected_vade_tarihi
        )

        if success:
            QMessageBox.information(self, "Başarılı", message)
            self.close()
            if hasattr(self.app, 'siparis_listesi_sayfasi'):
                self.app.siparis_listesi_sayfasi.siparis_listesini_yukle()
            if hasattr(self.app, 'fatura_listesi_sayfasi'):
                if hasattr(self.app.fatura_listesi_sayfasi, 'satis_fatura_frame'):
                    self.app.fatura_listesi_sayfasi.satis_fatura_frame.fatura_listesini_yukle()
                if hasattr(self.app.fatura_listesi_sayfasi.alis_fatura_frame, 'fatura_listesini_yukle'):
                    self.app.fatura_listesi_sayfasi.alis_fatura_frame.fatura_listesini_yukle()
        else:
            QMessageBox.critical(self, "Hata", message)

    def _siparisi_duzenle(self):
        """Bu siparişi düzenleme penceresinde açar."""
        from pencereler import SiparisPenceresi # SiparisPenceresi'nin PySide6 versiyonu
        siparis_tipi_db = 'SATIŞ_SIPARIS' if self.siparis_ana.get('cari_tip') == 'MUSTERI' else 'ALIŞ_SIPARIS'
        dialog = SiparisPenceresi(
            parent=self.app, 
            db_manager=self.db, # db_manager'ı geç
            app_ref=self.app,
            siparis_tipi=siparis_tipi_db,
            siparis_id_duzenle=self.siparis_id,
            yenile_callback=self.yenile_callback # Ana listeden gelen yenileme fonksiyonunu aktarıyoruz
        )
        dialog.exec()
        self.close() # Sipariş detay penceresini kapat

    def on_dialog_finished(self, result):
        if self.yenile_callback:
            self.yenile_callback()

class YoneticiAyarlariPenceresi(QDialog):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.setWindowTitle("Yönetici Ayarları ve Veri İşlemleri")
        self.setMinimumSize(600, 500)
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Veri Sıfırlama ve Bakım")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        main_frame = QWidget(self)
        main_frame_layout = QVBoxLayout(main_frame)
        main_layout.addWidget(main_frame)

        buttons_info = [
            ("Geçmiş Hatalı Kayıtları Temizle", "Var olmayan faturalara ait 'hayalet' cari ve gelir/gider hareketlerini siler. (Tek seferlik çalıştırın)", self.db.gecmis_hatali_kayitlari_temizle),
            ("Stok Envanterini Yeniden Hesapla", "Tüm stokları faturalara göre sıfırdan hesaplar. Geçmiş hatalı silme işlemlerini düzeltir.", self.db.stok_envanterini_yeniden_hesapla),
            ("Stok Verilerini Temizle", "Bu işlem tüm ürünleri ve ilişkili kalemleri siler.", self.db.clear_stok_data),
            ("Müşteri Verilerini Temizle", "Bu işlem perakende müşteri hariç tüm müşterileri ve ilişkili hareketlerini siler.", self.db.clear_musteri_data),
            ("Tedarikçi Verilerini Temizle", "Bu işlem tüm tedarikçileri ve ilişkili hareketlerini siler.", self.db.clear_tedarikci_data),
            ("Kasa/Banka Verilerini Temizle", "Bu işlem tüm kasa/banka hesaplarını temizler ve ilişkili referansları kaldırır.", self.db.clear_kasa_banka_data),
            ("Tüm İşlem Verilerini Temizle", "Faturalar, gelir/gider, cari hareketler, siparişler ve teklifler gibi tüm operasyonel verileri siler. Ana kayıtlar korunur.", self.db.clear_all_transaction_data),
            ("Tüm Verileri Temizle (Kullanıcılar Hariç)", "Kullanıcılar ve şirket ayarları hariç tüm veritabanını temizler. Program yeniden başlatılacaktır.", self.db.clear_all_data)
        ]

        for i, (text, desc, func) in enumerate(buttons_info):
            btn_frame = QFrame()
            btn_frame_layout = QHBoxLayout(btn_frame)
            
            btn = QPushButton(text)
            btn.clicked.connect(lambda f=func, t=text: self._confirm_and_run_utility(f, t))
            btn_frame_layout.addWidget(btn)
            
            desc_label = QLabel(desc)
            desc_label.setWordWrap(True)
            desc_label.setStyleSheet("font-size: 8pt;")
            btn_frame_layout.addWidget(desc_label, 1) # Streç faktör 1
            
            main_frame_layout.addWidget(btn_frame)
        
        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignRight)

    def _confirm_and_run_utility(self, utility_function, button_text):
        confirm_message = f"'{button_text}' işlemini gerçekleştirmek istediğinizden emin misiniz?\n\nBU İŞLEM GERİ ALINAMAZ!"
        if "Tüm Verileri Temizle" in button_text:
             confirm_message += "\n\nBu işlemden sonra program yeniden başlatılacaktır."

        reply = QMessageBox.question(self, "Onay Gerekli", confirm_message, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = utility_function()

                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.app.set_status_message(message)
                    
                    if "Tüm Verileri Temizle" in button_text:
                        # self.app.cikis_yap_ve_giris_ekranina_don() # Bu metod app'te yoksa hata verir.
                        QMessageBox.information(self, "Bilgi", "Tüm veriler temizlendi. Uygulama yeniden başlatılıyor.")
                        QApplication.quit() # Uygulamayı kapat

                else:
                    QMessageBox.critical(self, "Hata", message)
                    self.app.set_status_message(f"'{button_text}' işlemi sırasında hata oluştu: {message}")
            except Exception as e:
                QMessageBox.critical(self, "Kritik Hata", f"İşlem sırasında beklenmedik bir hata oluştu: {e}")
                logging.error(f"'{button_text}' yardımcı programı çalıştırılırken hata: {traceback.format_exc()}")
        else:
            self.app.set_status_message(f"'{button_text}' işlemi iptal edildi.")

class SirketBilgileriPenceresi(QDialog):
    def __init__(self, parent, db_manager):
        super().__init__(parent)
        self.db = db_manager
        self.app_parent = parent # Ana App referansı
        self.setWindowTitle("Şirket Bilgileri")
        self.setMinimumSize(550, 400)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Şirket Bilgileri Yönetimi")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        main_frame = QWidget(self)
        main_frame_layout = QGridLayout(main_frame)
        main_layout.addWidget(main_frame)

        self.field_definitions = [
            ("Şirket Adı:", "sirket_adi", QLineEdit),
            ("Adres:", "sirket_adresi", QTextEdit),
            ("Telefon:", "sirket_telefonu", QLineEdit),
            ("E-mail:", "sirket_email", QLineEdit),
            ("Vergi Dairesi:", "sirket_vergi_dairesi", QLineEdit),
            ("Vergi No:", "sirket_vergi_no", QLineEdit),
            ("Logo Yolu:", "sirket_logo_yolu", QLineEdit)
        ]
        self.entries = {}

        for i, (label_text, db_key_name, widget_type) in enumerate(self.field_definitions):
            main_frame_layout.addWidget(QLabel(label_text), i, 0, alignment=Qt.AlignLeft)
            
            if widget_type == QTextEdit:
                widget = QTextEdit()
                widget.setFixedHeight(60) # Yükseklik ayarı
            else: # QLineEdit
                widget = QLineEdit()
            
            self.entries[db_key_name] = widget
            main_frame_layout.addWidget(widget, i, 1)
            
            if db_key_name == "sirket_logo_yolu":
                logo_button = QPushButton("Gözat...")
                logo_button.clicked.connect(self.logo_gozat)
                main_frame_layout.addWidget(logo_button, i, 2)

        main_frame_layout.setColumnStretch(1, 1) # Entry'lerin genişlemesi için

        self.yukle_bilgiler()

        button_layout = QHBoxLayout()
        button_layout.addStretch() # Butonları sağa yasla
        kaydet_button = QPushButton("Kaydet")
        kaydet_button.clicked.connect(self.kaydet_bilgiler)
        button_layout.addWidget(kaydet_button)
        iptal_button = QPushButton("İptal")
        iptal_button.clicked.connect(self.close)
        button_layout.addWidget(iptal_button)
        main_layout.addLayout(button_layout)

    def logo_gozat(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Logo Seçin", "", "Resim Dosyaları (*.png *.jpg *.jpeg);;Tüm Dosyalar (*)")
        if file_path:
            self.entries["sirket_logo_yolu"].setText(file_path)

    def yukle_bilgiler(self):
        mevcut_bilgiler = self.db.sirket_bilgilerini_yukle()
        for db_key_name, entry_widget in self.entries.items():
            if isinstance(entry_widget, QTextEdit):
                entry_widget.setPlainText(mevcut_bilgiler.get(db_key_name, ""))
            else:
                entry_widget.setText(mevcut_bilgiler.get(db_key_name, ""))
    
    def kaydet_bilgiler(self):
        yeni_bilgiler = {}
        for db_key_name, entry_widget in self.entries.items():
            if isinstance(entry_widget, QTextEdit):
                yeni_bilgiler[db_key_name] = entry_widget.toPlainText().strip()
            else:
                yeni_bilgiler[db_key_name] = entry_widget.text().strip()

        print(f"DEBUG: kaydet_bilgiler - yeni_bilgiler sözlüğü: {yeni_bilgiler}")
        success, message = self.db.sirket_bilgilerini_kaydet(yeni_bilgiler)
        if success:
            if hasattr(self.app_parent, 'ana_sayfa') and hasattr(self.app_parent.ana_sayfa, 'guncelle_sirket_adi'):
                self.app_parent.ana_sayfa.guncelle_sirket_adi()
            if hasattr(self.app_parent, 'set_status_message'):
                 self.app_parent.set_status_message(message)
            QMessageBox.information(self, "Başarılı", message)
            self.close()
        else:
            QMessageBox.critical(self, "Hata", message)
            
class StokHareketiPenceresi(QDialog):
    def __init__(self, parent, db_manager, urun_id, urun_adi, mevcut_stok, hareket_yonu, yenile_callback):
        super().__init__(parent)
        self.db = db_manager
        self.urun_id = urun_id
        self.yenile_callback = yenile_callback
        
        # UI elemanları burada tanımlandığından, init metodunun kalanı aynı kalır.

        title = "Stok Girişi" if hareket_yonu == "EKLE" else "Stok Çıkışı"
        self.setWindowTitle(f"{title}: {urun_adi}")
        self.setMinimumWidth(400)
        self.setModal(True)

        self.main_layout = QVBoxLayout(self)
        self.form_layout = QGridLayout()

        self.main_layout.addWidget(QLabel(f"<b>{title}</b><br>Ürün: {urun_adi}<br>Mevcut Stok: {mevcut_stok:.2f}"), alignment=Qt.AlignCenter)
        self.main_layout.addLayout(self.form_layout)

        self.entries = {}
        self.form_layout.addWidget(QLabel("İşlem Tipi:"), 0, 0)
        self.entries['islem_tipi'] = QComboBox()
        if hareket_yonu == "EKLE": self.entries['islem_tipi'].addItems(["Giriş (Manuel)", "Sayım Fazlası", "İade Girişi"])
        else: self.entries['islem_tipi'].addItems(["Çıkış (Manuel)", "Sayım Eksiği", "Zayiat"])
        self.form_layout.addWidget(self.entries['islem_tipi'], 0, 1)

        self.form_layout.addWidget(QLabel("Miktar:"), 1, 0)
        self.entries['miktar'] = QLineEdit("0,00")
        self.entries['miktar'].setValidator(QDoubleValidator(0.01, 999999.0, 2))
        self.form_layout.addWidget(self.entries['miktar'], 1, 1)

        self.form_layout.addWidget(QLabel("Tarih:"), 2, 0)
        self.entries['tarih'] = QLineEdit(datetime.now().strftime('%Y-%m-%d'))
        self.form_layout.addWidget(self.entries['tarih'], 2, 1)

        self.form_layout.addWidget(QLabel("Açıklama:"), 3, 0, alignment=Qt.AlignTop)
        self.entries['aciklama'] = QTextEdit()
        self.form_layout.addWidget(self.entries['aciklama'], 3, 1)

        button_layout = QHBoxLayout()
        button_layout.addStretch()
        kaydet_button = QPushButton("Kaydet")
        kaydet_button.clicked.connect(self.kaydet)
        iptal_button = QPushButton("İptal")
        iptal_button.clicked.connect(self.reject)
        button_layout.addWidget(kaydet_button)
        button_layout.addWidget(iptal_button)
        self.main_layout.addLayout(button_layout)

    def kaydet(self):
        try:
            miktar = float(self.entries['miktar'].text().replace(',', '.'))
            if miktar <= 0: raise ValueError("Miktar pozitif bir değer olmalıdır.")
        except (ValueError, TypeError):
            QMessageBox.warning(self, "Geçersiz Değer", "Lütfen miktar alanına geçerli bir sayı girin.")
            return

        islem_tipi = self.entries['islem_tipi'].currentText()
        if islem_tipi == "Giriş (Manuel)":
            islem_tipi = self.db.STOK_ISLEM_TIP_GIRIS_MANUEL
        elif islem_tipi == "Çıkış (Manuel)":
            islem_tipi = self.db.STOK_ISLEM_TIP_CIKIS_MANUEL
        elif islem_tipi == "Sayım Fazlası":
            islem_tipi = self.db.STOK_ISLEM_TIP_SAYIM_FAZLASI
        elif islem_tipi == "Sayım Eksiği":
            islem_tipi = self.db.STOK_ISLEM_TIP_SAYIM_EKSIGI
        elif islem_tipi == "İade Girişi":
            islem_tipi = self.db.STOK_ISLEM_TIP_IADE_GIRIS
        elif islem_tipi == "Zayiat":
            islem_tipi = self.db.STOK_ISLEM_TIP_ZAYIAT

        data = {
            "islem_tipi": islem_tipi,
            "miktar": miktar, 
            "tarih": self.entries['tarih'].text(),
            "aciklama": self.entries['aciklama'].toPlainText().strip()
        }

        try:
            success, message = self.db.stok_hareket_ekle(self.urun_id, data)
            if success:
                QMessageBox.information(self, "Başarılı", message)
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "API Hatası", f"Stok hareketi kaydedilirken bir hata oluştu:\n{error_detail}")

class IlgiliFaturalarDetayPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, urun_id, urun_adi):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.urun_id = urun_id
        self.urun_adi = urun_adi
        self.setWindowTitle(f"{self.urun_adi} - İlgili Faturalar")
        self.setMinimumSize(1000, 600)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel(f"{self.urun_adi} Ürününün Yer Aldığı Faturalar")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignLeft)
        main_layout.addWidget(title_label)

        filter_frame = QFrame(self)
        filter_layout = QHBoxLayout(filter_frame)
        main_layout.addWidget(filter_frame)

        filter_layout.addWidget(QLabel("Fatura Tipi:"))
        self.fatura_tipi_filter_cb = QComboBox()
        self.fatura_tipi_filter_cb.addItems(["TÜMÜ", "ALIŞ", "SATIŞ"])
        self.fatura_tipi_filter_cb.currentIndexChanged.connect(self._load_ilgili_faturalar)
        filter_layout.addWidget(self.fatura_tipi_filter_cb)
        filter_layout.addStretch() # Sağa yaslama için

        # Filtreleme butonu kaldırıldı, combobox değişince tetikleniyor.
        # btn_filter = QPushButton("Filtrele")
        # btn_filter.clicked.connect(self._load_ilgili_faturalar)
        # filter_layout.addWidget(btn_filter)

        cols_fatura = ("ID", "Fatura No", "Tarih", "Tip", "Cari/Misafir", "KDV Hariç Top.", "KDV Dahil Top.")
        self.ilgili_faturalar_tree = QTreeWidget(self)
        self.ilgili_faturalar_tree.setHeaderLabels(cols_fatura)
        self.ilgili_faturalar_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.ilgili_faturalar_tree.setSortingEnabled(True)

        from PySide6.QtWidgets import QHeaderView
        col_defs_fatura = [
            ("ID", 40, Qt.AlignRight), # Sağa hizala
            ("Fatura No", 120, Qt.AlignLeft),
            ("Tarih", 85, Qt.AlignCenter),
            ("Tip", 70, Qt.AlignCenter),
            ("Cari/Misafir", 200, Qt.AlignLeft),
            ("KDV Hariç Top.", 120, Qt.AlignRight),
            ("KDV Dahil Top.", 120, Qt.AlignRight)
        ]
        for i, (col_name, width, alignment) in enumerate(col_defs_fatura):
            self.ilgili_faturalar_tree.setColumnWidth(i, width)
            self.ilgili_faturalar_tree.headerItem().setTextAlignment(i, alignment)
            self.ilgili_faturalar_tree.headerItem().setFont(i, QFont("Segoe UI", 9, QFont.Bold))
        self.ilgili_faturalar_tree.header().setStretchLastSection(False)
        self.ilgili_faturalar_tree.header().setSectionResizeMode(4, QHeaderView.Stretch) # Cari/Misafir genişlesin

        main_layout.addWidget(self.ilgili_faturalar_tree)

        self.ilgili_faturalar_tree.itemDoubleClicked.connect(self._on_fatura_double_click)

        self._load_ilgili_faturalar() # İlk yükleme

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignRight)

    def _load_ilgili_faturalar(self, index=None): # index parametresi QComboBox'tan gelir, kullanılmıyor
        self.ilgili_faturalar_tree.clear()

        if not self.urun_id:
            item_qt = QTreeWidgetItem(self.ilgili_faturalar_tree)
            item_qt.setText(4, "Ürün seçili değil.")
            return

        fatura_tipi_filtre = self.fatura_tipi_filter_cb.currentText()
        if fatura_tipi_filtre == "TÜMÜ":
            fatura_tipi_filtre = None # API'ye tüm tipleri çekmesi için None gönder

        # API'den veri çek
        try:
            params = {'urun_id': self.urun_id}
            if fatura_tipi_filtre:
                params['fatura_tipi'] = fatura_tipi_filtre

            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            # API endpoint'i: /faturalar/urun_faturalari şeklinde olmalı
            response_data = self.db.get_urun_faturalari(params.get('urun_id'), params.get('fatura_tipi'))

            faturalar = []
            if isinstance(response_data, dict) and "items" in response_data:
                faturalar = response_data["items"]
            elif isinstance(response_data, list): # Eğer API doğrudan liste dönüyorsa
                faturalar = response_data
                self.app.set_status_message("Uyarı: İlgili faturalar API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else: # Beklenmeyen bir format gelirse
                self.app.set_status_message("Hata: İlgili faturalar API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"İlgili faturalar API'den beklenen formatta gelmedi: {type(response_data)} - {response_data}")
                return # Hata durumunda fonksiyonu sonlandır

            if not faturalar:
                item_qt = QTreeWidgetItem(self.ilgili_faturalar_tree)
                item_qt.setText(4, "Bu ürüne ait fatura bulunamadı.")
                return

            for fatura_item in faturalar:
                item_qt = QTreeWidgetItem(self.ilgili_faturalar_tree)

                fatura_id = fatura_item.get('id')
                fatura_no = fatura_item.get('fatura_no')
                tarih_str = fatura_item.get('tarih')
                fatura_tip = fatura_item.get('fatura_turu') # 'tip' yerine 'fatura_turu' kullanıldı
                cari_adi = fatura_item.get('cari_adi') # API'den gelmesi beklenir
                misafir_adi = fatura_item.get('misafir_adi') # API'den gelmesi beklenir
                toplam_kdv_haric = fatura_item.get('toplam_kdv_haric')
                toplam_kdv_dahil = fatura_item.get('toplam_kdv_dahil')

                try:
                    formatted_tarih = datetime.strptime(tarih_str, '%Y-%m-%d').strftime('%d.%m.%Y')
                except ValueError:
                    formatted_tarih = tarih_str

                display_cari_info = cari_adi
                if fatura_tip == self.db.FATURA_TIP_SATIS and misafir_adi: # 'SATIŞ' sabiti kullanıldı
                    display_cari_info = f"Perakende ({misafir_adi})"

                item_qt.setText(0, str(fatura_id))
                item_qt.setText(1, fatura_no)
                item_qt.setText(2, formatted_tarih)
                item_qt.setText(3, fatura_tip)
                item_qt.setText(4, display_cari_info)
                item_qt.setText(5, self.db._format_currency(toplam_kdv_haric))
                item_qt.setText(6, self.db._format_currency(toplam_kdv_dahil))

                self.app.set_status_message(f"Ürün '{self.urun_adi}' için {len(faturalar)} fatura listelendi.")

        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"İlgili faturalar çekilirken hata: {e}")
            logging.error(f"İlgili faturalar yükleme hatası: {e}")

    def _on_fatura_double_click(self, item, column): # item and column from QTreeWidget signal
        fatura_id = item.text(0) # ID ilk sütunda
        if fatura_id:
            from pencereler import FaturaDetayPenceresi
            FaturaDetayPenceresi(self.app, self.db, int(fatura_id)).exec() # fatura_id int olmalı

class KategoriMarkaYonetimiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, refresh_callback=None):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.refresh_callback = refresh_callback
        self.setWindowTitle("Kategori & Marka Yönetimi")
        self.setMinimumSize(800, 500)
        self.setModal(True)
        main_layout = QVBoxLayout(self)
        title_label = QLabel("Kategori & Marka Yönetimi")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignLeft)
        main_layout.addWidget(title_label)
        main_frame = QWidget(self)
        main_frame_layout = QHBoxLayout(main_frame)
        main_layout.addWidget(main_frame)
        main_frame_layout.setStretch(0, 1)
        main_frame_layout.setStretch(1, 1)
        kategori_frame = QGroupBox("Kategori Yönetimi", main_frame)
        kategori_frame_layout = QGridLayout(kategori_frame)
        main_frame_layout.addWidget(kategori_frame)
        kategori_frame_layout.setColumnStretch(1, 1)
        kategori_frame_layout.addWidget(QLabel("Kategori Adı:"), 0, 0)
        self.kategori_entry = QLineEdit()
        kategori_frame_layout.addWidget(self.kategori_entry, 0, 1)
        kategori_frame_layout.addWidget(QPushButton("Ekle", clicked=self._kategori_ekle_ui), 0, 2)
        kategori_frame_layout.addWidget(QPushButton("Güncelle", clicked=self._kategori_guncelle_ui), 0, 3)
        kategori_frame_layout.addWidget(QPushButton("Sil", clicked=self._kategori_sil_ui), 0, 4)
        self.kategori_tree = QTreeWidget()
        self.kategori_tree.setHeaderLabels(["ID", "Kategori Adı"])
        self.kategori_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.kategori_tree.setColumnWidth(0, 50)
        self.kategori_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        self.kategori_tree.itemSelectionChanged.connect(self._on_kategori_select)
        self.kategori_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.kategori_tree.customContextMenuRequested.connect(self._open_kategori_context_menu)
        kategori_frame_layout.addWidget(self.kategori_tree, 1, 0, 1, 5)
        marka_frame = QGroupBox("Marka Yönetimi", main_frame)
        marka_frame_layout = QGridLayout(marka_frame)
        main_frame_layout.addWidget(marka_frame)
        marka_frame_layout.setColumnStretch(1, 1)
        marka_frame_layout.addWidget(QLabel("Marka Adı:"), 0, 0)
        self.marka_entry = QLineEdit()
        marka_frame_layout.addWidget(self.marka_entry, 0, 1)
        marka_frame_layout.addWidget(QPushButton("Ekle", clicked=self._marka_ekle_ui), 0, 2)
        marka_frame_layout.addWidget(QPushButton("Güncelle", clicked=self._marka_guncelle_ui), 0, 3)
        marka_frame_layout.addWidget(QPushButton("Sil", clicked=self._marka_sil_ui), 0, 4)
        self.marka_tree = QTreeWidget()
        self.marka_tree.setHeaderLabels(["ID", "Marka Adı"])
        self.marka_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.marka_tree.setColumnWidth(0, 50)
        self.marka_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        self.marka_tree.itemSelectionChanged.connect(self._on_marka_select)
        self.marka_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.marka_tree.customContextMenuRequested.connect(self._open_marka_context_menu)
        marka_frame_layout.addWidget(self.marka_tree, 1, 0, 1, 5)
        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self._on_close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignRight)
        self._kategori_listesini_yukle()
        self._marka_listesini_yukle()
        
    def _on_close(self):
        if self.refresh_callback:
            self.refresh_callback()
        self.close()

    def _kategori_listesini_yukle(self):
        self.kategori_tree.clear()
        try:
            kategoriler_response = self.db.kategori_listele()
            kategoriler_list = kategoriler_response.get("items", [])
            for kat_item in kategoriler_list:
                item_qt = QTreeWidgetItem(self.kategori_tree)
                item_qt.setText(0, str(kat_item.get('id')))
                item_qt.setText(1, kat_item.get('ad'))
                item_qt.setData(0, Qt.UserRole, kat_item.get('id'))
            self.kategori_tree.sortByColumn(1, Qt.AscendingOrder)
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Kategori listesi çekilirken hata: {e}")
            logging.error(f"Kategori listesi yükleme hatası: {e}", exc_info=True)

    def _on_kategori_select(self):
        selected_items = self.kategori_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.kategori_entry.setText(values)
        else:
            self.kategori_entry.clear()

    def _kategori_ekle_ui(self):
        kategori_adi = self.kategori_entry.text().strip()
        if not kategori_adi:
            QMessageBox.warning(self.app, "Uyarı", "Kategori adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle("kategoriler", {"ad": kategori_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.kategori_entry.clear()
                self._kategori_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                self.app.set_status_message(message)
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Kategori eklenirken bir hata oluştu:\n{e}")
            self.app.set_status_message(f"Kategori ekleme başarısız: {e}")

    def _kategori_guncelle_ui(self):
        selected_items = self.kategori_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir kategori seçin.")
            return
        selected_item = selected_items[0]
        kategori_id = selected_item.data(0, Qt.UserRole)
        yeni_kategori_adi = self.kategori_entry.text().strip()
        if not yeni_kategori_adi:
            QMessageBox.warning(self.app, "Uyarı", "Kategori adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle("kategoriler", kategori_id, {"ad": yeni_kategori_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.kategori_entry.clear()
                self._kategori_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                self.app.set_status_message(message)
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Kategori güncellenirken bir hata oluştu:\n{e}")
            self.app.set_status_message(f"Kategori güncelleme başarısız: {e}")

    def _kategori_sil_ui(self):
        selected_items = self.kategori_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir kategori seçin.")
            return
        selected_item = selected_items[0]
        kategori_id = selected_item.data(0, Qt.UserRole)
        kategori_adi = selected_item.text(1)
        reply = QMessageBox.question(self.app, "Onay", f"'{kategori_adi}' kategorisini silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil("kategoriler", kategori_id)
                if success:
                    QMessageBox.information(self.app, "Başarılı", message)
                    self.kategori_entry.clear()
                    self._kategori_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                    self.app.set_status_message(message)
                else:
                    QMessageBox.critical(self.app, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Kategori silinirken bir hata oluştu:\n{e}")
                self.app.set_status_message(f"Kategori silme başarısız: {e}")

    def _open_kategori_context_menu(self, pos):
        item = self.kategori_tree.itemAt(pos)
        if not item: return
        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._kategori_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._kategori_sil_ui)
        context_menu.exec(self.kategori_tree.mapToGlobal(pos))

    def _marka_listesini_yukle(self):
        self.marka_tree.clear()
        try:
            markalar_response = self.db.marka_listele()
            markalar = markalar_response.get("items", [])
            for mar in markalar:
                item_qt = QTreeWidgetItem(self.marka_tree)
                item_qt.setText(0, str(mar.get('id')))
                item_qt.setText(1, mar.get('ad'))
                item_qt.setData(0, Qt.UserRole, mar.get('id'))
        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Marka listesi çekilirken hata: {e}")
            logging.error(f"Marka listesi yükleme hatası: {e}")

    def _on_marka_select(self):
        selected_items = self.marka_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.marka_entry.setText(values)
        else:
            self.marka_entry.clear()

    def _marka_ekle_ui(self):
        marka_adi = self.marka_entry.text().strip()
        if not marka_adi:
            QMessageBox.warning(self.app, "Uyarı", "Marka adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle("markalar", {"ad": marka_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.marka_entry.clear()
                self._marka_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                self.app.set_status_message(message)
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Marka eklenirken bir hata oluştu:\n{e}")
            self.app.set_status_message(f"Marka ekleme başarısız: {e}")

    def _marka_guncelle_ui(self):
        selected_items = self.marka_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir marka seçin.")
            return
        marka_id = selected_items[0].data(0, Qt.UserRole)
        yeni_marka_adi = self.marka_entry.text().strip()
        if not yeni_marka_adi:
            QMessageBox.warning(self.app, "Uyarı", "Marka adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle("markalar", marka_id, {"ad": yeni_marka_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.marka_entry.clear()
                self._marka_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                self.app.set_status_message(message)
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Marka güncellenirken bir hata oluştu:\n{e}")
            self.app.set_status_message(f"Marka güncelleme başarısız: {e}")

    def _marka_sil_ui(self):
        selected_items = self.marka_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir marka seçin.")
            return
        marka_id = selected_items[0].data(0, Qt.UserRole)
        marka_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self.app, "Onay", f"'{marka_adi}' markasını silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil("markalar", marka_id)
                if success:
                    QMessageBox.information(self.app, "Başarılı", message)
                    self.marka_entry.clear()
                    self._marka_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                    self.app.set_status_message(message)
                else:
                    QMessageBox.critical(self.app, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Marka silinirken bir hata oluştu:\n{e}")
                self.app.set_status_message(f"Marka silme başarısız: {e}")

    def _open_marka_context_menu(self, pos):
        item = self.marka_tree.itemAt(pos)
        if not item: return
        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._marka_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._marka_sil_ui)
        context_menu.exec(self.marka_tree.mapToGlobal(pos))

class UrunNitelikYonetimiPenceresi(QDialog):
    def __init__(self, parent_notebook, db_manager, app_ref, refresh_callback=None):
        super().__init__(parent_notebook)
        self.db = db_manager
        self.app = app_ref
        self.refresh_callback = refresh_callback

        self.setWindowTitle("Ürün Grubu, Birimi ve Menşe Ülke Yönetimi")
        self.setMinimumSize(800, 600)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Ürün Grubu, Birimi ve Menşe Ülke Yönetimi")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignLeft)
        main_layout.addWidget(title_label)

        # Ana içerik çerçevesi (grid gibi düzenleme için)
        main_frame = QWidget(self)
        main_frame_layout = QGridLayout(main_frame)
        main_layout.addWidget(main_frame)
        main_frame_layout.setColumnStretch(0, 1)
        main_frame_layout.setColumnStretch(1, 1)
        main_frame_layout.setRowStretch(0, 1)
        main_frame_layout.setRowStretch(1, 1)


        # --- Ürün Grubu Yönetimi ---
        urun_grubu_frame = QGroupBox("Ürün Grubu Yönetimi", main_frame)
        urun_grubu_frame_layout = QGridLayout(urun_grubu_frame)
        main_frame_layout.addWidget(urun_grubu_frame, 0, 0)
        urun_grubu_frame_layout.setColumnStretch(1, 1)

        urun_grubu_frame_layout.addWidget(QLabel("Grup Adı:"), 0, 0)
        self.urun_grubu_entry = QLineEdit()
        urun_grubu_frame_layout.addWidget(self.urun_grubu_entry, 0, 1)
        urun_grubu_frame_layout.addWidget(QPushButton("Ekle", clicked=self._urun_grubu_ekle_ui), 0, 2)
        urun_grubu_frame_layout.addWidget(QPushButton("Sil", clicked=self._urun_grubu_sil_ui), 0, 3)

        self.urun_grubu_tree = QTreeWidget()
        self.urun_grubu_tree.setHeaderLabels(["ID", "Grup Adı"])
        self.urun_grubu_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.urun_grubu_tree.setColumnWidth(0, 50)
        self.urun_grubu_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        urun_grubu_frame_layout.addWidget(self.urun_grubu_tree, 1, 0, 1, 4)
        self.urun_grubu_tree.itemSelectionChanged.connect(self._on_urun_grubu_select)
        
        self.urun_grubu_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.urun_grubu_tree.customContextMenuRequested.connect(self._open_urun_grubu_context_menu) 
        self._urun_grubu_listesini_yukle()

        # --- Ürün Birimi Yönetimi ---
        urun_birimi_frame = QGroupBox("Ürün Birimi Yönetimi", main_frame)
        urun_birimi_frame_layout = QGridLayout(urun_birimi_frame)
        main_frame_layout.addWidget(urun_birimi_frame, 0, 1)
        urun_birimi_frame_layout.setColumnStretch(1, 1)

        urun_birimi_frame_layout.addWidget(QLabel("Birim Adı:"), 0, 0)
        self.urun_birimi_entry = QLineEdit()
        urun_birimi_frame_layout.addWidget(self.urun_birimi_entry, 0, 1)
        urun_birimi_frame_layout.addWidget(QPushButton("Ekle", clicked=self._urun_birimi_ekle_ui), 0, 2)
        urun_birimi_frame_layout.addWidget(QPushButton("Sil", clicked=self._urun_birimi_sil_ui), 0, 3)

        self.urun_birimi_tree = QTreeWidget()
        self.urun_birimi_tree.setHeaderLabels(["ID", "Birim Adı"])
        self.urun_birimi_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.urun_birimi_tree.setColumnWidth(0, 50)
        self.urun_birimi_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        urun_birimi_frame_layout.addWidget(self.urun_birimi_tree, 1, 0, 1, 4)
        self.urun_birimi_tree.itemSelectionChanged.connect(self._on_urun_birimi_select)
        
        self.urun_birimi_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.urun_birimi_tree.customContextMenuRequested.connect(self._open_birim_context_menu)
        self._urun_birimi_listesini_yukle()

        # --- Ülke (Menşe) Yönetimi ---
        ulke_frame = QGroupBox("Menşe Ülke Yönetimi", main_frame)
        ulke_frame_layout = QGridLayout(ulke_frame)
        main_frame_layout.addWidget(ulke_frame, 1, 0, 1, 2) # İki sütuna yay
        ulke_frame_layout.setColumnStretch(1, 1)

        ulke_frame_layout.addWidget(QLabel("Ülke Adı:"), 0, 0)
        self.ulke_entry = QLineEdit()
        ulke_frame_layout.addWidget(self.ulke_entry, 0, 1)
        ulke_frame_layout.addWidget(QPushButton("Ekle", clicked=self._ulke_ekle_ui), 0, 2)
        ulke_frame_layout.addWidget(QPushButton("Sil", clicked=self._ulke_sil_ui), 0, 3)

        self.ulke_tree = QTreeWidget()
        self.ulke_tree.setHeaderLabels(["ID", "Ülke Adı"])
        self.ulke_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.ulke_tree.setColumnWidth(0, 50)
        self.ulke_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        ulke_frame_layout.addWidget(self.ulke_tree, 1, 0, 1, 4)
        self.ulke_tree.itemSelectionChanged.connect(self._on_ulke_select)
        
        self.ulke_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ulke_tree.customContextMenuRequested.connect(self._open_ulke_context_menu)
        self._ulke_listesini_yukle()

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self._on_close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignRight)

    def _on_close(self):
        if self.refresh_callback:
            self.refresh_callback() # Ürün kartı combobox'larını yenile
        self.close()

    # Ürün Grubu Yönetimi Metotları
    def _urun_grubu_listesini_yukle(self):
        self.urun_grubu_tree.clear()
        try:
            urun_gruplari_response = self.db.urun_grubu_listele() # API'den gelen tam yanıt
            urun_gruplari_list = urun_gruplari_response.get("items", []) # "items" listesini alıyoruz

            for grup_item in urun_gruplari_list: # urun_gruplari_list üzerinde döngü
                item_qt = QTreeWidgetItem(self.urun_grubu_tree)
                item_qt.setText(0, str(grup_item.get('id'))) # .get() ile güvenli erişim
                item_qt.setText(1, grup_item.get('ad')) # .get() ile güvenli erişim
                item_qt.setData(0, Qt.UserRole, grup_item.get('id'))
            self.urun_grubu_tree.sortByColumn(1, Qt.AscendingOrder)
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Ürün grubu listesi çekilirken hata: {e}")
            logging.error(f"Ürün grubu listesi yükleme hatası: {e}", exc_info=True)

    def _on_urun_grubu_select(self):
        selected_items = self.urun_grubu_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.urun_grubu_entry.setText(values)
        else:
            self.urun_grubu_entry.clear()

    def _urun_grubu_ekle_ui(self):
        grup_adi = self.urun_grubu_entry.text().strip()
        if not grup_adi:
            QMessageBox.warning(self.app, "Uyarı", "Ürün grubu adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle("urun_gruplari", {"ad": grup_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.urun_grubu_entry.clear()
                self._urun_grubu_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün grubu eklenirken hata: {e}")
            logging.error(f"Ürün grubu eklenirken hata: {e}", exc_info=True)

    def _urun_grubu_guncelle_ui(self):
        selected_items = self.urun_grubu_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir ürün grubu seçin.")
            return
        grup_id = selected_items[0].data(0, Qt.UserRole)
        yeni_grup_adi = self.urun_grubu_entry.text().strip()
        if not yeni_grup_adi:
            QMessageBox.warning(self.app, "Uyarı", "Grup adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle("urun_gruplari", grup_id, {"ad": yeni_grup_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.urun_grubu_entry.clear()
                self._urun_grubu_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün grubu güncellenirken hata: {e}")
            logging.error(f"Ürün grubu güncellenirken hata: {e}", exc_info=True)

    def _urun_grubu_sil_ui(self):
        selected_items = self.urun_grubu_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir ürün grubu seçin.")
            return
        grup_id = selected_items[0].data(0, Qt.UserRole)
        grup_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self.app, "Onay", f"'{grup_adi}' ürün grubunu silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil("urun_gruplari", grup_id)
                if success:
                    QMessageBox.information(self.app, "Başarılı", message)
                    self.urun_grubu_entry.clear()
                    self._urun_grubu_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                else:
                    QMessageBox.critical(self.app, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Ürün grubu silinirken hata: {e}")
                logging.error(f"Ürün grubu silinirken hata: {e}", exc_info=True)
                
    # Ürün Birimi Yönetimi Metotları
    def _urun_birimi_listesini_yukle(self):
        self.urun_birimi_tree.clear()
        try:
            urun_birimleri_response = self.db.urun_birimi_listele()
            urun_birimleri_list = urun_birimleri_response.get("items", []) # "items" listesini alıyoruz

            for birim_item in urun_birimleri_list: # urun_birimleri_list üzerinde döngü
                item_qt = QTreeWidgetItem(self.urun_birimi_tree)
                item_qt.setText(0, str(birim_item.get('id'))) # .get() ile güvenli erişim
                item_qt.setText(1, birim_item.get('ad')) # .get() ile güvenli erişim
                item_qt.setData(0, Qt.UserRole, birim_item.get('id'))
            self.urun_birimi_tree.sortByColumn(1, Qt.AscendingOrder)
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Ürün birimi listesi çekilirken hata: {e}")
            logging.error(f"Ürün birimi listesi yükleme hatası: {e}", exc_info=True)

    def _on_urun_birimi_select(self):
        selected_items = self.urun_birimi_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.urun_birimi_entry.setText(values)
        else:
            self.urun_birimi_entry.clear()

    def _urun_birimi_ekle_ui(self):
        birim_adi = self.urun_birimi_entry.text().strip()
        if not birim_adi:
            QMessageBox.warning(self.app, "Uyarı", "Ürün birimi adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle("urun_birimleri", {"ad": birim_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.urun_birimi_entry.clear()
                self._urun_birimi_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün birimi eklenirken hata: {e}")
            logging.error(f"Ürün birimi eklenirken hata: {e}", exc_info=True)

    def _urun_birimi_guncelle_ui(self):
        selected_items = self.urun_birimi_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir ürün birimi seçin.")
            return
        birim_id = selected_items[0].data(0, Qt.UserRole)
        yeni_birim_adi = self.urun_birimi_entry.text().strip()
        if not yeni_birim_adi:
            QMessageBox.warning(self.app, "Uyarı", "Birim adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle("urun_birimleri", birim_id, {"ad": yeni_birim_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.urun_birimi_entry.clear()
                self._urun_birimi_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ürün birimi güncellenirken hata: {e}")
            logging.error(f"Ürün birimi güncellenirken hata: {e}", exc_info=True)

    def _urun_birimi_sil_ui(self):
        selected_items = self.urun_birimi_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir ürün birimi seçin.")
            return
        birim_id = selected_items[0].data(0, Qt.UserRole)
        birim_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self.app, "Onay", f"'{birim_adi}' ürün birimini silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil("urun_birimleri", birim_id)
                if success:
                    QMessageBox.information(self.app, "Başarılı", message)
                    self.urun_birimi_entry.clear()
                    self._urun_birimi_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                elif message:
                    QMessageBox.critical(self.app, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Ürün birimi silinirken hata: {e}")
                logging.error(f"Ürün birimi silinirken hata: {e}", exc_info=True)

    def _open_urun_grubu_context_menu(self, pos):
        item = self.urun_grubu_tree.itemAt(pos)
        if not item: return

        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._urun_grubu_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._urun_grubu_sil_ui)
        context_menu.exec(self.urun_grubu_tree.mapToGlobal(pos))

    def _open_birim_context_menu(self, pos):
        item = self.urun_birimi_tree.itemAt(pos)
        if not item: return

        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._urun_birimi_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._urun_birimi_sil_ui)
        context_menu.exec(self.urun_birimi_tree.mapToGlobal(pos))

    def _open_ulke_context_menu(self, pos):
        item = self.ulke_tree.itemAt(pos)
        if not item: return

        context_menu = QMenu(self)
        context_menu.addAction("Güncelle").triggered.connect(self._ulke_guncelle_ui)
        context_menu.addAction("Sil").triggered.connect(self._ulke_sil_ui)
        context_menu.exec(self.ulke_tree.mapToGlobal(pos))

    # Ülke (Menşe) Yönetimi Metotları
    def _ulke_listesini_yukle(self):
        self.ulke_tree.clear()
        try:
            ulkeler_response = self.db.ulke_listele()
            ulkeler_list = ulkeler_response.get("items", []) # "items" listesini alıyoruz

            for ulke_item in ulkeler_list: # ulkeler_list üzerinde döngü
                item_qt = QTreeWidgetItem(self.ulke_tree)
                item_qt.setText(0, str(ulke_item.get('id'))) # .get() ile güvenli erişim
                item_qt.setText(1, ulke_item.get('ad')) # .get() ile güvenli erişim
                item_qt.setData(0, Qt.UserRole, ulke_item.get('id'))
            self.ulke_tree.sortByColumn(1, Qt.AscendingOrder)
        except Exception as e:
            QMessageBox.critical(self.app, "API Hatası", f"Ülke listesi çekilirken hata: {e}")
            logging.error(f"Ülke listesi yükleme hatası: {e}", exc_info=True)
            
    def _on_ulke_select(self):
        selected_items = self.ulke_tree.selectedItems()
        if selected_items:
            values = selected_items[0].text(1)
            self.ulke_entry.setText(values)
        else:
            self.ulke_entry.clear()

    def _ulke_ekle_ui(self):
        ulke_adi = self.ulke_entry.text().strip()
        if not ulke_adi:
            QMessageBox.warning(self.app, "Uyarı", "Ülke adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle("ulkeler", {"ad": ulke_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.ulke_entry.clear()
                self._ulke_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ülke eklenirken hata: {e}")
            logging.error(f"Ülke eklenirken hata: {e}", exc_info=True)
            
    def _ulke_guncelle_ui(self):
        selected_items = self.ulke_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen güncellemek için bir ülke seçin.")
            return
        ulke_id = selected_items[0].data(0, Qt.UserRole)
        yeni_ulke_adi = self.ulke_entry.text().strip()
        if not yeni_ulke_adi:
            QMessageBox.warning(self.app, "Uyarı", "Ülke adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle("ulkeler", ulke_id, {"ad": yeni_ulke_adi})
            if success:
                QMessageBox.information(self.app, "Başarılı", message)
                self.ulke_entry.clear()
                self._ulke_listesini_yukle()
                if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                    self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
            else:
                QMessageBox.critical(self.app, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self.app, "Hata", f"Ülke güncellenirken hata: {e}")
            logging.error(f"Ülke güncellenirken hata: {e}", exc_info=True)

    def _ulke_sil_ui(self):
        selected_items = self.ulke_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self.app, "Uyarı", "Lütfen silmek için bir ülke seçin.")
            return
        ulke_id = selected_items[0].data(0, Qt.UserRole)
        ulke_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self.app, "Onay", f"'{ulke_adi}' ülkesini silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil("ulkeler", ulke_id)
                if success:
                    QMessageBox.information(self.app, "Başarılı", message)
                    self.ulke_entry.clear()
                    self._ulke_listesini_yukle()
                    if hasattr(self.app, 'stok_yonetimi_sayfasi') and hasattr(self.app.stok_yonetimi_sayfasi, '_yukle_filtre_comboboxlari_stok_yonetimi'):
                        self.app.stok_yonetimi_sayfasi._yukle_filtre_comboboxlari_stok_yonetimi()
                else:
                    QMessageBox.critical(self.app, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self.app, "Hata", f"Ülke silinirken hata: {e}")
                logging.error(f"Ülke silinirken hata: {e}", exc_info=True)
                
    def _yukle_kategori_marka_comboboxlari(self):
        # Kategoriler
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            kategoriler_response = self.db.kategori_listele() # <-- BURASI GÜNCELLENDİ
            kategoriler = kategoriler_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.kategoriler_map = {"Seçim Yok": None}
            # kategori_display_values artık kullanılmadığı için kaldırıldı.
            for k in kategoriler:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.kategoriler_map[k.get('ad')] = k.get('id') # <-- 'kategori_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Kategoriler combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Kategoriler yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

        # Markalar
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            markalar_response = self.db.marka_listele() # <-- BURASI GÜNCELLENDİ
            markalar = markalar_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.markalar_map = {"Seçim Yok": None}
            # marka_display_values artık kullanılmadığı için kaldırıldı.
            for m in markalar:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.markalar_map[m.get('ad')] = m.get('id') # <-- 'marka_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Markalar combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Markalar yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

    def _yukle_urun_grubu_birimi_ulke_comboboxlari(self):
        # Ürün Grupları
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            urun_gruplari_response = self.db.urun_grubu_listele() # <-- BURASI GÜNCELLENDİ
            urun_gruplari = urun_gruplari_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.urun_gruplari_map = {"Seçim Yok": None}
            # urun_grubu_display_values artık kullanılmadığı için kaldırıldı.
            for g in urun_gruplari:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.urun_gruplari_map[g.get('ad')] = g.get('id') # <-- 'grup_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Ürün grupları combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Ürün grupları yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

        # Ürün Birimleri
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            urun_birimleri_response = self.db.urun_birimi_listele() # <-- BURASI GÜNCELLENDİ
            urun_birimleri = urun_birimleri_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.urun_birimleri_map = {"Seçim Yok": None}
            # urun_birimi_display_values artık kullanılmadığı için kaldırıldı.
            for b in urun_birimleri:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.urun_birimleri_map[b.get('ad')] = b.get('id') # <-- 'birim_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Ürün birimleri combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Ürün birimleri yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

        # Ülkeler (Menşe)
        try:
            # Doğrudan API çağrısı yerine db_manager metodu kullanıldı
            ulkeler_response = self.db.ulke_listele() # <-- BURASI GÜNCELLENDİ
            ulkeler = ulkeler_response.get("items", []) # <-- Yanıtın 'items' listesini alıyoruz
            
            self.ulkeler_map = {"Seçim Yok": None}
            # ulke_display_values artık kullanılmadığı için kaldırıldı.
            for u in ulkeler:
                # API'den gelen nitelik objelerinde 'ad' alanı varsayılıyor
                self.ulkeler_map[u.get('ad')] = u.get('id') # <-- 'ulke_adi' yerine 'ad' kullanıldı
            
            # combobox'ı dolduran kısım burada yok, muhtemelen başka bir metodda yapılıyor.
            # Sadece map'i güncelliyor.
        except Exception as e: # requests.exceptions.RequestException yerine daha genel hata yakalandı
            logging.error(f"Ülkeler combobox yüklenirken hata: {e}", exc_info=True)
            # QMessageBox.critical(self, "API Hatası", f"Ülkeler yüklenirken hata: {e}") # Eğer hata mesajı göstermek isterseniz açılabilir.

class StokKartiPenceresi(QDialog):
    data_updated = Signal()

    def __init__(self, parent_window, db_manager, refresh_callback=None, urun_duzenle=None, app_ref=None):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.refresh_callback = refresh_callback
        self.app = app_ref
        self.urun_duzenle = urun_duzenle
        self.duzenleme_modu = urun_duzenle is not None
        self.yeni_urun_resmi_yolu = None
        self.mevcut_urun_resmi_yolu = urun_duzenle.get('urun_resmi_yolu') if urun_duzenle and 'urun_resmi_yolu' in urun_duzenle else None
        
        self.stok_id = urun_duzenle.get('id') if urun_duzenle and 'id' in urun_duzenle else None

        self.original_pixmap = None 

        logger.info(f"StokKartiPenceresi başlatılıyor. Düzenleme modu: {self.duzenleme_modu}")

        self.setWindowTitle("Yeni Ürün Ekle" if not self.duzenleme_modu else "Ürün Kartı Düzenle")
        self.setModal(True)
        self.resize(800, 700)

        # UI elemanlarını burada tanımlıyoruz
        self.kod_e = QLineEdit()
        self.ad_e = QLineEdit()
        self.miktar_e = QLineEdit()
        self.alis_fiyat_e = QLineEdit()
        self.satis_fiyat_e = QLineEdit()
        self.kdv_e = QLineEdit()
        self.min_stok_e = QLineEdit()
        self.aktif_cb = QCheckBox()
        self.detay_e = QTextEdit()
        self.resim_label = QLabel("Resim Yok")
        self.kategori_combo = QComboBox()
        self.marka_combo = QComboBox()
        self.urun_grubu_combo = QComboBox()
        self.birim_combo = QComboBox()
        self.mensei_ulke_combo = QComboBox()
        
        # Bu nesneler, alt sekmeleri yönetecek
        self.stok_hareketleri_sekmesi = None
        self.ilgili_faturalar_sekmesi = None

        self._setup_ui() 
        
        if self.duzenleme_modu:
            self._mevcut_urunu_yukle()
        else:
            self._formu_sifirla()

        self._load_combobox_data()
        
    def _setup_ui(self):
        """Pencerenin kullanıcı arayüzü elemanlarını oluşturur ve düzenler."""
        main_layout = QVBoxLayout(self)
        top_frame = QFrame(self)
        top_layout = QHBoxLayout(top_frame)
        main_layout.addWidget(top_frame)
        info_frame = QFrame(top_frame)
        info_layout = QGridLayout(info_frame)
        top_layout.addWidget(info_frame)
        info_layout.addWidget(QLabel("Ürün Kodu:"), 0, 0)
        info_layout.addWidget(self.kod_e, 0, 1)
        info_layout.addWidget(QLabel("Ürün Adı:"), 1, 0)
        info_layout.addWidget(self.ad_e, 1, 1)
        info_layout.addWidget(QLabel("Miktar:"), 2, 0)
        self.miktar_e.setReadOnly(True)
        info_layout.addWidget(self.miktar_e, 2, 1)
        info_layout.addWidget(QLabel("Alış Fiyatı (KDV Dahil):"), 3, 0)
        info_layout.addWidget(self.alis_fiyat_e, 3, 1)
        info_layout.addWidget(QLabel("Satış Fiyatı (KDV Dahil):"), 4, 0)
        info_layout.addWidget(self.satis_fiyat_e, 4, 1)
        info_layout.addWidget(QLabel("KDV Oranı (%):"), 5, 0)
        info_layout.addWidget(self.kdv_e, 5, 1)
        info_layout.addWidget(QLabel("Min. Stok Seviyesi:"), 6, 0)
        info_layout.addWidget(self.min_stok_e, 6, 1)
        info_layout.addWidget(QLabel("Aktif:"), 7, 0)
        info_layout.addWidget(self.aktif_cb, 7, 1)
        info_layout.addWidget(QLabel("Kategori:"), 8, 0)
        info_layout.addWidget(self.kategori_combo, 8, 1)
        info_layout.addWidget(QLabel("Marka:"), 9, 0)
        info_layout.addWidget(self.marka_combo, 9, 1)
        info_layout.addWidget(QLabel("Ürün Grubu:"), 10, 0)
        info_layout.addWidget(self.urun_grubu_combo, 10, 1)
        info_layout.addWidget(QLabel("Birim:"), 11, 0)
        info_layout.addWidget(self.birim_combo, 11, 1)
        info_layout.addWidget(QLabel("Menşei Ülke:"), 12, 0)
        info_layout.addWidget(self.mensei_ulke_combo, 12, 1)
        info_layout.addWidget(QLabel("Detay:"), 13, 0, Qt.AlignTop)
        info_layout.addWidget(self.detay_e, 13, 1)
        image_frame = QFrame(top_frame)
        image_layout = QVBoxLayout(image_frame)
        top_layout.addWidget(image_frame)
        top_layout.setStretch(1, 1)
        self.resim_label.setAlignment(Qt.AlignCenter)
        self.resim_label.setFixedSize(200, 200)
        self.resim_label.setStyleSheet("border: 1px solid gray;")
        image_layout.addWidget(self.resim_label, alignment=Qt.AlignCenter)
        btn_resim_sec = QPushButton("Resim Seç")
        btn_resim_sec.clicked.connect(self._resim_sec)
        image_layout.addWidget(btn_resim_sec)
        btn_resim_sil = QPushButton("Resmi Sil")
        btn_resim_sil.clicked.connect(self._resim_sil)
        image_layout.addWidget(btn_resim_sil)
        self.bottom_tab_widget = QTabWidget(self)
        main_layout.addWidget(self.bottom_tab_widget)
        from arayuz import StokHareketiSekmesi, IlgiliFaturalarSekmesi
        self.stok_hareketleri_sekmesi = StokHareketiSekmesi(
            self.bottom_tab_widget, self.db, self.app, self.stok_id, self.ad_e.text() if self.duzenleme_modu else ""
        )
        self.bottom_tab_widget.addTab(self.stok_hareketleri_sekmesi, "Stok Hareketleri")
        self.ilgili_faturalar_sekmesi = IlgiliFaturalarSekmesi(
            self.bottom_tab_widget, self.db, self.app, self.stok_id, self.ad_e.text() if self.duzenleme_modu else ""
        )
        self.bottom_tab_widget.addTab(self.ilgili_faturalar_sekmesi, "İlgili Faturalar")
        self._add_bottom_buttons()

    def _add_bottom_buttons(self):
        """Pencerenin alt kısmındaki butonları oluşturur ve yerleştirir."""
        button_layout = QHBoxLayout()
        self.layout().addLayout(button_layout)

        self.btn_kaydet = QPushButton("Kaydet")
        self.btn_kaydet.clicked.connect(self.kaydet_urun)
        button_layout.addWidget(self.btn_kaydet)

        self.btn_iptal = QPushButton("İptal")
        self.btn_iptal.clicked.connect(self.reject)
        button_layout.addWidget(self.btn_iptal)

        self.btn_manuel_stok_giris = QPushButton("Manuel Stok Girişi")
        self.btn_manuel_stok_giris.clicked.connect(self._manuel_stok_giris_penceresi_ac)
        button_layout.addWidget(self.btn_manuel_stok_giris)

        self.btn_manuel_stok_cikis = QPushButton("Manuel Stok Çıkışı")
        self.btn_manuel_stok_cikis.clicked.connect(self._manuel_stok_cikis_penceresi_ac)
        button_layout.addWidget(self.btn_manuel_stok_cikis)

        self.btn_sil = QPushButton("Stoku Sil")
        self.btn_sil.clicked.connect(self._stok_sil)
        self.btn_sil.setVisible(bool(self.stok_id))
        button_layout.addWidget(self.btn_sil)

        if not self.duzenleme_modu:
            self.btn_manuel_stok_giris.setEnabled(False)
            self.btn_manuel_stok_cikis.setEnabled(False)
            self.bottom_tab_widget.setEnabled(False)

    def _manuel_stok_giris_penceresi_ac(self):
        """Stok ekleme penceresini açar."""
        if not self.stok_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce stoku kaydedin.")
            return
        
        # Güncel stok miktarını API'den çekerek al
        mevcut_stok = 0.0 # Varsayılan değer
        try:
            current_stok_data = self.db.stok_getir_by_id(self.stok_id)
            if current_stok_data:
                mevcut_stok = current_stok_data.get('miktar', 0.0)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Mevcut stok miktarı alınırken hata oluştu: {e}")
            logging.error(f"Stok miktarı alınırken hata: {e}", exc_info=True)
            return

        from pencereler import StokHareketiPenceresi
        dialog = StokHareketiPenceresi(
            self,
            self.db, # <-- db_manager parametresi eklendi
            self.stok_id,
            self.ad_e.text() if self.ad_e.text() else self.kod_e.text(),
            mevcut_stok,
            "GIRIŞ",
            self.refresh_data_and_ui
        )
        dialog.exec()

    def _manuel_stok_cikis_penceresi_ac(self):
        """Stok eksiltme penceresini açar."""
        if not self.stok_id:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce stoku kaydedin.")
            return

        # Güncel stok miktarını API'den çekerek al
        mevcut_stok = 0.0 # Varsayılan değer
        try:
            current_stok_data = self.db.stok_getir_by_id(self.stok_id)
            if current_stok_data:
                mevcut_stok = current_stok_data.get('miktar', 0.0)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Mevcut stok miktarı alınırken hata oluştu: {e}")
            logging.error(f"Stok miktarı alınırken hata: {e}", exc_info=True)
            return
        
        from pencereler import StokHareketiPenceresi
        dialog = StokHareketiPenceresi(
            self,
            self.db, # <-- db_manager parametresi eklendi
            self.stok_id,
            self.ad_e.text() if self.ad_e.text() else self.kod_e.text(),
            mevcut_stok,
            "CIKIS",
            self.refresh_data_and_ui
        )
        dialog.exec()
        
    def _stok_sil(self):
        reply = QMessageBox.question(self, "Ürün Silme Onayı", "Ürünü silmek istediğinizden emin misiniz? Bu işlem geri alınamaz.", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.stok_sil(self.stok_id)
                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.accept()
                    if self.refresh_callback:
                        self.refresh_callback()
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Ürün silinirken bir hata oluştu: {e}")

    def _resim_sec(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Stok Resmi Seç", "", "Resim Dosyaları (*.png *.jpg *.jpeg)")
        if file_path:
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                data_dir = os.path.join(base_dir, 'data')
                urun_resimleri_klasoru = os.path.join(data_dir, "urun_resimleri")
                
                os.makedirs(urun_resimleri_klasoru, exist_ok=True)
                yeni_path = os.path.join(urun_resimleri_klasoru, os.path.basename(file_path))
                shutil.copy2(file_path, yeni_path)
                self.yeni_urun_resmi_yolu = yeni_path
                self.mevcut_urun_resmi_yolu = None
                logger.info(f"Resim kopyalandı: {yeni_path}")
            except Exception as e:
                QMessageBox.warning(self, "Hata", f"Resim kopyalanamadı: {e}")
                logger.error(f"Resim kopyalama hatası: {e}", exc_info=True)
                self.yeni_urun_resmi_yolu = None
            self._load_urun_resmi()

    def _resim_sil(self):
        self.yeni_urun_resmi_yolu = ""
        self.mevcut_urun_resmi_yolu = None
        self._load_urun_resmi()
        logger.info("Ürün resmi silindi.")
    
    def _load_urun_resmi(self):
        resim_yolu = self.yeni_urun_resmi_yolu or self.mevcut_urun_resmi_yolu
        if resim_yolu and os.path.exists(resim_yolu):
            self.original_pixmap = QPixmap(resim_yolu)
            self._resize_image()
            self.resim_label.setText("")
        else:
            self.original_pixmap = None
            self.resim_label.setText("Resim Yok")
            self.resim_label.setPixmap(QPixmap())

    def _load_combobox_data(self):
        try:
            kategoriler_response = self.db.kategori_listele()
            markalar_response = self.db.marka_listele()
            urun_gruplari_response = self.db.urun_grubu_listele()
            urun_birimleri_response = self.db.urun_birimi_listele()
            ulkeler_response = self.db.ulke_listele()

            kategoriler = kategoriler_response.get("items", [])
            markalar = markalar_response.get("items", [])
            urun_gruplari = urun_gruplari_response.get("items", [])
            urun_birimleri = urun_birimleri_response.get("items", [])
            ulkeler = ulkeler_response.get("items", [])

            self.kategori_combo.clear()
            self.marka_combo.clear()
            self.urun_grubu_combo.clear()
            self.birim_combo.clear()
            self.mensei_ulke_combo.clear()

            self.kategori_combo.addItem("Seçiniz...", userData=None)
            self.marka_combo.addItem("Seçiniz...", userData=None)
            self.urun_grubu_combo.addItem("Seçiniz...", userData=None)
            self.birim_combo.addItem("Seçiniz...", userData=None)
            self.mensei_ulke_combo.addItem("Seçiniz...", userData=None)

            for item in kategoriler: self.kategori_combo.addItem(item.get('ad'), userData=item.get('id'))
            for item in markalar: self.marka_combo.addItem(item.get('ad'), userData=item.get('id'))
            for item in urun_gruplari: self.urun_grubu_combo.addItem(item.get('ad'), userData=item.get('id'))
            for item in urun_birimleri: self.birim_combo.addItem(item.get('ad'), userData=item.get('id'))
            for item in ulkeler: self.mensei_ulke_combo.addItem(item.get('ad'), userData=item.get('id'))
            
            logger.info(f"Nitelik combobox'ları başarıyla yüklendi.")

        except Exception as e:
            logger.error(f"StokKartiPenceresi: Nitelik verileri yüklenirken hata: {e}", exc_info=True)
            QMessageBox.critical(self, "Veri Yükleme Hatası", f"Nitelik verileri yüklenirken bir hata oluştu: {e}")

    def resizeEvent(self, event):
        super().resizeEvent(event)
        QTimer.singleShot(50, self._resize_image)

    def _resize_image(self):
        if self.original_pixmap and not self.original_pixmap.isNull():
            scaled_pixmap = self.original_pixmap.scaled(self.resim_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.resim_label.setPixmap(scaled_pixmap)
            
    def refresh_data_and_ui(self):
        if not self.stok_id: return

        try:
            updated_stok_data = self.db.stok_getir_by_id(self.stok_id)

            self.miktar_e.setText(self.db._format_numeric(updated_stok_data.get('miktar', 0.0), 2))
            self.alis_fiyat_e.setText(self.db._format_numeric(updated_stok_data.get('alis_fiyati', 0.0), 2))
            self.satis_fiyat_e.setText(self.db._format_numeric(updated_stok_data.get('satis_fiyati', 0.0), 2))
            self.kdv_e.setText(self.db._format_numeric(updated_stok_data.get('kdv_orani', 0.0), 0))
            self.min_stok_e.setText(self.db._format_numeric(updated_stok_data.get('min_stok_seviyesi', 0.0), 2))
            self.aktif_cb.setChecked(updated_stok_data.get('aktif', True))

            self.data_updated.emit()
            self.stok_hareketleri_sekmesi._load_stok_hareketleri()
            self.ilgili_faturalar_sekmesi._load_ilgili_faturalar()
            
            logger.info(f"Stok kartı verileri yenilendi: ID {self.stok_id}")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Stok verileri yenilenirken hata oluştu:\n{e}")
            logger.error(f"StokKartiPenceresi refresh_data_and_ui hatası: {e}", exc_info=True)

    def _mevcut_urunu_yukle(self):
        self.kod_e.setText(self.urun_duzenle.get('kod', ''))
        self.ad_e.setText(self.urun_duzenle.get('ad', ''))
        self.miktar_e.setText(self.db._format_numeric(self.urun_duzenle.get('miktar', 0.0), 2))
        self.alis_fiyat_e.setText(self.db._format_numeric(self.urun_duzenle.get('alis_fiyati', 0.0), 2))
        self.satis_fiyat_e.setText(self.db._format_numeric(self.urun_duzenle.get('satis_fiyati', 0.0), 2))
        self.kdv_e.setText(self.db._format_numeric(self.urun_duzenle.get('kdv_orani', 0.0), 0))
        self.min_stok_e.setText(self.db._format_numeric(self.urun_duzenle.get('min_stok_seviyesi', 0.0), 2))
        self.aktif_cb.setChecked(self.urun_duzenle.get('aktif', True))
        self.detay_e.setPlainText(self.urun_duzenle.get('detay', ''))

        self.kategori_combo.setCurrentIndex(self.kategori_combo.findData(self.urun_duzenle.get('kategori_id')))
        self.marka_combo.setCurrentIndex(self.marka_combo.findData(self.urun_duzenle.get('marka_id')))
        self.urun_grubu_combo.setCurrentIndex(self.urun_grubu_combo.findData(self.urun_duzenle.get('urun_grubu_id')))
        self.birim_combo.setCurrentIndex(self.birim_combo.findData(self.urun_duzenle.get('birim_id')))
        self.mensei_ulke_combo.setCurrentIndex(self.mensei_ulke_combo.findData(self.urun_duzenle.get('mense_id')))

        self.mevcut_urun_resmi_yolu = self.urun_duzenle.get('urun_resmi_yolu')
        self._load_urun_resmi()

        urun_adi_for_tabs = self.ad_e.text() if self.ad_e.text() else self.kod_e.text()

        if self.stok_id:
            self.stok_hareketleri_sekmesi.urun_id = self.stok_id
            self.stok_hareketleri_sekmesi.urun_adi = urun_adi_for_tabs
            self.ilgili_faturalar_sekmesi.urun_id = self.stok_id
            self.ilgili_faturalar_sekmesi.urun_adi = urun_adi_for_tabs
            self.stok_hareketleri_sekmesi._load_stok_hareketleri()
            self.ilgili_faturalar_sekmesi._load_ilgili_faturalar()
            self.btn_manuel_stok_giris.setEnabled(True)
            self.btn_manuel_stok_cikis.setEnabled(True)
            self.bottom_tab_widget.setEnabled(True)
        else:
            logger.warning("Mevcut ürünü yüklerken ID bulunamadı, sekmeler devre dışı bırakıldı.")
            self.btn_manuel_stok_giris.setEnabled(False)
            self.btn_manuel_stok_cikis.setEnabled(False)
            self.bottom_tab_widget.setEnabled(False)


        logger.info(f"Ürün ID {self.stok_id} için mevcut ürün verileri yüklendi.")

    def _formu_sifirla(self):
        self.kod_e.clear()
        self.ad_e.clear()
        self.miktar_e.setText("0,00")
        self.alis_fiyat_e.setText("0,00")
        self.satis_fiyat_e.setText("0,00")
        self.kdv_e.setText("20")
        self.min_stok_e.setText("0,00")
        self.aktif_cb.setChecked(True)
        self.detay_e.clear()
        self.resim_label.setText("Resim Yok")
        self.resim_label.setPixmap(QPixmap())
        self.yeni_urun_resmi_yolu = None
        self.mevcut_urun_resmi_yolu = None

        self.kategori_combo.setCurrentIndex(0)
        self.marka_combo.setCurrentIndex(0)
        self.urun_grubu_combo.setCurrentIndex(0)
        self.birim_combo.setCurrentIndex(0)
        self.mensei_ulke_combo.setCurrentIndex(0)

        self.stok_id = None
        self.duzenleme_modu = False
        self.setWindowTitle("Yeni Ürün Ekle")
        self.btn_sil.setVisible(False)
        self.btn_manuel_stok_giris.setEnabled(False)
        self.btn_manuel_stok_cikis.setEnabled(False)
        self.bottom_tab_widget.setEnabled(False)
        
        self.stok_hareketleri_sekmesi.urun_id = None
        self.stok_hareketleri_sekmesi.urun_adi = ""
        self.stok_hareketleri_sekmesi._load_stok_hareketleri()

        self.ilgili_faturalar_sekmesi.urun_id = None
        self.ilgili_faturalar_sekmesi.urun_adi = ""
        self.ilgili_faturalar_sekmesi._load_ilgili_faturalar()

        logger.info("Stok Kartı formu sıfırlandı.")

    def kaydet_urun(self):
        kod = self.kod_e.text().strip()
        ad = self.ad_e.text().strip()
        
        # Miktar readonly olduğu için direkt API'den gelen değeri kullanıyoruz.
        # Bu metot sadece kart bilgilerini günceller, miktar hareketlerle değişir.
        miktar = self.urun_duzenle.get('miktar', 0.0) if self.duzenleme_modu else 0.0
        
        # self.db.safe_float() metodu ile değerleri alıyoruz
        alis_fiyati = self.db.safe_float(self.alis_fiyat_e.text())
        satis_fiyati = self.db.safe_float(self.satis_fiyat_e.text())
        kdv_orani = self.db.safe_float(self.kdv_e.text())
        min_stok = self.db.safe_float(self.min_stok_e.text())
        
        aktif = self.aktif_cb.isChecked()
        detay = self.detay_e.toPlainText().strip()
        
        kategori_id = self.kategori_combo.currentData()
        marka_id = self.marka_combo.currentData()
        urun_grubu_id = self.urun_grubu_combo.currentData()
        birim_id = self.birim_combo.currentData()
        mense_id = self.mensei_ulke_combo.currentData()

        if not ad:
            QMessageBox.critical(self, "Eksik Bilgi", "Ürün Adı boş olamaz.")
            return

        if not self.duzenleme_modu: # Yeni ürün eklenirken kod da zorunlu
            if not kod:
                QMessageBox.critical(self, "Eksik Bilgi", "Yeni ürün için Ürün Kodu boş olamaz.")
                return

        if miktar < 0 or alis_fiyati < 0 or satis_fiyati < 0 or kdv_orani < 0 or min_stok < 0:
            QMessageBox.critical(self, "Geçersiz Değer", "Miktar, fiyatlar, KDV oranı ve minimum stok negatif olamaz.")
            return
        
        urun_data = {
            "kod": kod,
            "ad": ad,
            "miktar": miktar,
            "alis_fiyati": alis_fiyati,
            "satis_fiyati": satis_fiyati,
            "kdv_orani": kdv_orani,
            "min_stok_seviyesi": min_stok,
            "aktif": aktif,
            "detay": detay if detay else None,
            "kategori_id": kategori_id,
            "marka_id": marka_id,
            "urun_grubu_id": urun_grubu_id,
            "birim_id": birim_id,
            "mense_id": mense_id,
            "urun_resmi_yolu": self.yeni_urun_resmi_yolu if self.yeni_urun_resmi_yolu else self.mevcut_urun_resmi_yolu
        }

        try:
            if self.duzenleme_modu and self.stok_id:
                success, message = self.db.stok_guncelle(self.stok_id, urun_data)
            else:
                success, message = self.db.stok_ekle(urun_data)

            if success:
                QMessageBox.information(self, "Başarılı", message)
                self.data_updated.emit()
                self.accept()
                if self.refresh_callback:
                    self.refresh_callback()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            logger.error(f"Ürün kaydedilirken hata oluştu: {e}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Ürün kaydedilirken bir hata oluştu:\n{e}")

class YeniKasaBankaEklePenceresi(QDialog):
    def __init__(self, parent, db_manager, yenile_callback, hesap_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.app = app_ref
        self.yenile_callback = yenile_callback
        self.hesap_duzenle_data = hesap_duzenle

        self.hesap_duzenle_id = self.hesap_duzenle_data.get('id') if self.hesap_duzenle_data else None

        title = "Yeni Kasa/Banka Hesabı" if not self.hesap_duzenle_id else "Hesap Düzenle"
        self.setWindowTitle(title)
        self.setMinimumSize(480, 450)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        main_layout.addWidget(title_label)

        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)
        
        self.entries = {}
        self.odeme_turleri = ["YOK", "NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET", "AÇIK HESAP"]
        
        # Form elemanları
        form_layout.addWidget(QLabel("Hesap Adı (*):"), 0, 0)
        self.entries['hesap_adi'] = QLineEdit()
        form_layout.addWidget(self.entries['hesap_adi'], 0, 1)

        form_layout.addWidget(QLabel("Hesap Tipi (*):"), 1, 0)
        self.entries['tip'] = QComboBox()
        self.entries['tip'].addItems(["KASA", "BANKA"])
        self.entries['tip'].currentTextChanged.connect(self._tip_degisince_banka_alanlarini_ayarla)
        form_layout.addWidget(self.entries['tip'], 1, 1)

        self.banka_labels = {
            'banka_adi': QLabel("Banka Adı:"),
            'sube_adi': QLabel("Şube Adı:"),
            'hesap_no': QLabel("Hesap No/IBAN:")
        }
        form_layout.addWidget(self.banka_labels['banka_adi'], 2, 0)
        self.entries['banka_adi'] = QLineEdit()
        form_layout.addWidget(self.entries['banka_adi'], 2, 1)
        
        form_layout.addWidget(self.banka_labels['sube_adi'], 3, 0)
        self.entries['sube_adi'] = QLineEdit()
        form_layout.addWidget(self.entries['sube_adi'], 3, 1)

        form_layout.addWidget(self.banka_labels['hesap_no'], 4, 0)
        self.entries['hesap_no'] = QLineEdit()
        form_layout.addWidget(self.entries['hesap_no'], 4, 1)
        
        form_layout.addWidget(QLabel("Açılış Bakiyesi:"), 5, 0)
        self.entries['bakiye'] = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.entries['bakiye']) # <-- decimal_places=2 parametresi kaldırıldı
        form_layout.addWidget(self.entries['bakiye'], 5, 1)

        form_layout.addWidget(QLabel("Para Birimi:"), 6, 0)
        self.entries['para_birimi'] = QLineEdit("TL")
        form_layout.addWidget(self.entries['para_birimi'], 6, 1)

        form_layout.addWidget(QLabel("Varsayılan Ödeme Türü:"), 7, 0)
        self.entries['varsayilan_odeme_turu'] = QComboBox()
        self.entries['varsayilan_odeme_turu'].addItems(self.odeme_turleri)
        form_layout.addWidget(self.entries['varsayilan_odeme_turu'], 7, 1)

        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch()
        kaydet_button = QPushButton("Kaydet")
        kaydet_button.clicked.connect(self.kaydet)
        button_layout.addWidget(kaydet_button)
        iptal_button = QPushButton("İptal")
        iptal_button.clicked.connect(self.reject)
        button_layout.addWidget(iptal_button)
        
        self._verileri_yukle()
        self._tip_degisince_banka_alanlarini_ayarla()

    def _tip_degisince_banka_alanlarini_ayarla(self):
        is_banka = self.entries['tip'].currentText() == "BANKA"
        for key, widget in self.banka_labels.items():
            widget.setVisible(is_banka)
        for key in ['banka_adi', 'sube_adi', 'hesap_no']:
            self.entries[key].setVisible(is_banka)
        if not is_banka:
            for key in ['banka_adi', 'sube_adi', 'hesap_no']:
                self.entries[key].clear()

    def _verileri_yukle(self):
        if self.hesap_duzenle_data:
            self.entries['hesap_adi'].setText(self.hesap_duzenle_data.get('hesap_adi', ''))
            self.entries['tip'].setCurrentText(self.hesap_duzenle_data.get('tip', 'KASA'))
            self.entries['banka_adi'].setText(self.hesap_duzenle_data.get('banka_adi', ''))
            self.entries['sube_adi'].setText(self.hesap_duzenle_data.get('sube_adi', ''))
            self.entries['hesap_no'].setText(self.hesap_duzenle_data.get('hesap_no', ''))
            bakiye = self.hesap_duzenle_data.get('bakiye', 0.0)
            self.entries['bakiye'].setText(f"{bakiye:.2f}".replace('.', ','))
            self.entries['para_birimi'].setText(self.hesap_duzenle_data.get('para_birimi', 'TL'))
            varsayilan_odeme_turu = self.hesap_duzenle_data.get('varsayilan_odeme_turu')
            self.entries['varsayilan_odeme_turu'].setCurrentText(varsayilan_odeme_turu if varsayilan_odeme_turu else "YOK")
            self.entries['bakiye'].setReadOnly(True) # Açılış bakiyesi düzenlemede değiştirilemez

    def kaydet(self):
        hesap_adi = self.entries['hesap_adi'].text().strip()
        if not hesap_adi:
            QMessageBox.warning(self, "Eksik Bilgi", "Hesap Adı alanı boş bırakılamaz.")
            return

        bakiye_str = self.entries['bakiye'].text().replace(',', '.')
        
        data = {
            "hesap_adi": hesap_adi,
            "tip": self.entries['tip'].currentText(),
            "bakiye": float(bakiye_str) if bakiye_str else 0.0,
            "banka_adi": self.entries['banka_adi'].text().strip(),
            "sube_adi": self.entries['sube_adi'].text().strip(),
            "hesap_no": self.entries['hesap_no'].text().strip(),
            "para_birimi": self.entries['para_birimi'].text().strip(),
            "varsayilan_odeme_turu": self.entries['varsayilan_odeme_turu'].currentText()
        }
        if data["varsayilan_odeme_turu"] == "YOK":
            data["varsayilan_odeme_turu"] = None

        try:
            if self.hesap_duzenle_id:
                # GÜNCELLEME (PUT isteği)
                success = self.db.kasa_banka_guncelle(self.hesap_duzenle_id, data)
            else:
                # YENİ KAYIT (POST isteği)
                success = self.db.kasa_banka_ekle(data)

            if success:
                QMessageBox.information(self, "Başarılı", "Kasa/Banka hesabı başarıyla kaydedildi.")
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", "Kasa/Banka hesabı kaydedilirken bir hata oluştu.")

        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "Hata", f"Hesap kaydedilirken bir hata oluştu:\n{error_detail}")
            logging.error(f"Kasa/Banka kaydetme hatası: {error_detail}", exc_info=True)

class YeniTedarikciEklePenceresi(QDialog):
    def __init__(self, parent, db_manager, yenile_callback, tedarikci_duzenle=None, app_ref=None):
        super().__init__(parent)
        self.db = db_manager
        self.app = app_ref
        self.yenile_callback = yenile_callback
        self.tedarikci_duzenle_data = tedarikci_duzenle

        self.tedarikci_duzenle_id = self.tedarikci_duzenle_data.get('id') if self.tedarikci_duzenle_data else None

        # CariService örneğini burada oluştur
        from hizmetler import CariService # CariService'i burada import ediyoruz
        self.cari_service = CariService(self.db) # <-- CariService BAŞLATILDI

        title = "Yeni Tedarikçi Ekle" if not self.tedarikci_duzenle_id else "Tedarikçi Düzenle"
        self.setWindowTitle(title)
        self.setMinimumSize(500, 420)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        main_layout.addWidget(title_label)

        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)
        
        self.entries = {}
        labels_entries = {
            "Tedarikçi Kodu:": "entry_kod",
            "Ad Soyad (*):": "entry_ad",
            "Telefon:": "entry_tel",
            "Adres:": "entry_adres",
            "Vergi Dairesi:": "entry_vd",
            "Vergi No:": "entry_vn"
        }

        for i, (label_text, entry_name) in enumerate(labels_entries.items()):
            form_layout.addWidget(QLabel(label_text), i, 0, alignment=Qt.AlignLeft)
            widget = QTextEdit() if entry_name == "entry_adres" else QLineEdit()
            if isinstance(widget, QTextEdit): widget.setFixedHeight(80)
            self.entries[entry_name] = widget
            form_layout.addWidget(widget, i, 1)

        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch()
        self.kaydet_button = QPushButton("Kaydet")
        self.kaydet_button.clicked.connect(self.kaydet)
        button_layout.addWidget(self.kaydet_button)
        self.iptal_button = QPushButton("İptal")
        self.iptal_button.clicked.connect(self.reject)
        button_layout.addWidget(self.iptal_button)
        
        self._verileri_yukle()

    def _verileri_yukle(self):
        """Mevcut tedarikçi verilerini düzenleme modunda forma yükler."""
        if self.tedarikci_duzenle_data:
            self.entries["entry_kod"].setText(self.tedarikci_duzenle_data.get('kod', ''))
            self.entries["entry_ad"].setText(self.tedarikci_duzenle_data.get('ad', ''))
            self.entries["entry_tel"].setText(self.tedarikci_duzenle_data.get('telefon', ''))
            self.entries["entry_adres"].setPlainText(self.tedarikci_duzenle_data.get('adres', ''))
            self.entries["entry_vd"].setText(self.tedarikci_duzenle_data.get('vergi_dairesi', ''))
            self.entries["entry_vn"].setText(self.tedarikci_duzenle_data.get('vergi_no', ''))
            self.entries["entry_kod"].setReadOnly(True)
        else:
            generated_code = self.db.get_next_tedarikci_kodu()
            self.entries["entry_kod"].setText(generated_code)
            self.entries["entry_kod"].setReadOnly(True)

    def kaydet(self):
        ad = self.entries["entry_ad"].text().strip()
        if not ad:
            QMessageBox.warning(self, "Eksik Bilgi", "Tedarikçi Adı alanı boş bırakılamaz.")
            return

        data = {
            "ad": ad,
            "kod": self.entries["entry_kod"].text().strip(),
            "telefon": self.entries["entry_tel"].text().strip(),
            "adres": self.entries["entry_adres"].toPlainText().strip(),
            "vergi_dairesi": self.entries["entry_vd"].text().strip(),
            "vergi_no": self.entries["entry_vn"].text().strip()
        }

        try:
            if self.tedarikci_duzenle_id:
                success, message = self.db.tedarikci_guncelle(self.tedarikci_duzenle_id, data)
            else:
                success, message = self.db.tedarikci_ekle(data)

            if success:
                QMessageBox.information(self, "Başarılı", "Tedarikçi bilgileri başarıyla kaydedildi.")
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept()
            else:
                QMessageBox.critical(self, "Hata", "Tedarikçi kaydedilirken bir hata oluştu.")

        except Exception as e:
            error_detail = str(e)
            QMessageBox.critical(self, "Hata", f"Tedarikçi kaydedilirken bir hata oluştu:\n{error_detail}")
            logging.error(f"Tedarikçi kaydetme hatası: {error_detail}", exc_info=True)

class KalemDuzenlePenceresi(QDialog):
    def __init__(self, parent_page, db_manager, kalem_index, kalem_verisi, islem_tipi, fatura_id_duzenle=None):
        super().__init__(parent_page)
        self.parent_page = parent_page # FaturaPenceresi objesi
        self.db = db_manager # db_manager artık direkt parametre olarak alınıyor
        self.kalem_index = kalem_index
        self.islem_tipi = islem_tipi
        self.fatura_id_duzenle = fatura_id_duzenle

        self.urun_id = kalem_verisi[0]
        self.urun_adi = kalem_verisi[1]
        self.mevcut_miktar = self.db.safe_float(kalem_verisi[2])
        self.orijinal_birim_fiyat_kdv_haric = self.db.safe_float(kalem_verisi[3])
        self.kdv_orani = self.db.safe_float(kalem_verisi[4])
        self.mevcut_alis_fiyati_fatura_aninda = self.db.safe_float(kalem_verisi[8])
        
        self_initial_iskonto_yuzde_1 = self.db.safe_float(kalem_verisi[10])
        self_initial_iskonto_yuzde_2 = self.db.safe_float(kalem_verisi[11])

        self.orijinal_birim_fiyat_kdv_dahil = self.orijinal_birim_fiyat_kdv_haric * (1 + self.kdv_orani / 100)

        self.setWindowTitle(f"Kalem Düzenle: {self.urun_adi}")
        self.setFixedSize(450, 550) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame) # Izgara düzenleyici
        
        main_frame_layout.addWidget(QLabel(f"Ürün: <b>{self.urun_adi}</b>", font=QFont("Segoe UI", 12, QFont.Bold)), 0, 0, 1, 3, Qt.AlignLeft)
        main_frame_layout.setColumnStretch(1, 1) # İkinci sütun genişlesin

        current_row = 1
        main_frame_layout.addWidget(QLabel("Miktar:"), current_row, 0)
        self.miktar_e = QLineEdit()
        setup_numeric_entry(self.parent_page.app, self.miktar_e) # <-- decimal_places=2 kaldırıldı
        self.miktar_e.setText(f"{self.mevcut_miktar:.2f}".replace('.',','))
        self.miktar_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
        main_frame_layout.addWidget(self.miktar_e, current_row, 1)

        current_row += 1
        main_frame_layout.addWidget(QLabel("Birim Fiyat (KDV Dahil):"), current_row, 0)
        self.fiyat_e = QLineEdit()
        setup_numeric_entry(self.parent_page.app, self.fiyat_e) # <-- decimal_places=2 kaldırıldı
        self.fiyat_e.setText(f"{self.orijinal_birim_fiyat_kdv_dahil:.2f}".replace('.',','))
        self.fiyat_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
        main_frame_layout.addWidget(self.fiyat_e, current_row, 1)

        current_row += 1
        self.alis_fiyati_aninda_e = None
        if self.islem_tipi in [self.db.FATURA_TIP_SATIS, self.db.SIPARIS_TIP_SATIS]:
            main_frame_layout.addWidget(QLabel("Fatura Anı Alış Fiyatı (KDV Dahil):"), current_row, 0)
            self.alis_fiyati_aninda_e = QLineEdit()
            setup_numeric_entry(self.parent_page.app, self.alis_fiyati_aninda_e) # <-- decimal_places=2 kaldırıldı
            self.alis_fiyati_aninda_e.setText(f"{self.mevcut_alis_fiyati_fatura_aninda:.2f}".replace('.',','))
            self.alis_fiyati_aninda_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
            main_frame_layout.addWidget(self.alis_fiyati_aninda_e, current_row, 1)
            current_row += 1
        
        main_frame_layout.addWidget(QFrame(), current_row, 0, 1, 3) # Separator yerine boş QFrame
        current_row += 1

        main_frame_layout.addWidget(QLabel("İskonto 1 (%):"), current_row, 0)
        self.iskonto_yuzde_1_e = QLineEdit()
        setup_numeric_entry(self.parent_page.app, self.iskonto_yuzde_1_e) # <-- decimal_places=2 kaldırıldı
        self.iskonto_yuzde_1_e.setText(f"{self_initial_iskonto_yuzde_1:.2f}".replace('.',','))
        self.iskonto_yuzde_1_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
        main_frame_layout.addWidget(self.iskonto_yuzde_1_e, current_row, 1)
        main_frame_layout.addWidget(QLabel("%"), current_row, 2)
        current_row += 1

        main_frame_layout.addWidget(QLabel("İskonto 2 (%):"), current_row, 0)
        self.iskonto_yuzde_2_e = QLineEdit()
        setup_numeric_entry(self.parent_page.app, self.iskonto_yuzde_2_e) # <-- decimal_places=2, max_value=100 kaldırıldı
        self.iskonto_yuzde_2_e.setText(f"{self_initial_iskonto_yuzde_2:.2f}".replace('.',','))
        self.iskonto_yuzde_2_e.textChanged.connect(self._anlik_hesaplama_ve_guncelleme)
        main_frame_layout.addWidget(self.iskonto_yuzde_2_e, current_row, 1)
        main_frame_layout.addWidget(QLabel("%"), current_row, 2)
        current_row += 1

        main_frame_layout.addWidget(QFrame(), current_row, 0, 1, 3) # Separator yerine boş QFrame
        current_row += 1

        main_frame_layout.addWidget(QLabel("Toplam İskonto Yüzdesi:", font=QFont("Segoe UI", 9, QFont.Bold)), current_row, 0)
        self.lbl_toplam_iskonto_yuzdesi = QLabel("0,00 %", font=QFont("Segoe UI", 9))
        main_frame_layout.addWidget(self.lbl_toplam_iskonto_yuzdesi, current_row, 1, 1, 2)
        current_row += 1

        main_frame_layout.addWidget(QLabel("Uygulanan İskonto Tutarı (KDV Dahil):", font=QFont("Segoe UI", 9, QFont.Bold)), current_row, 0)
        self.lbl_uygulanan_iskonto_dahil = QLabel("0,00 TL", font=QFont("Segoe UI", 9))
        main_frame_layout.addWidget(self.lbl_uygulanan_iskonto_dahil, current_row, 1, 1, 2)
        current_row += 1

        main_frame_layout.addWidget(QLabel("İskontolu Birim Fiyat (KDV Dahil):", font=QFont("Segoe UI", 9, QFont.Bold)), current_row, 0)
        self.lbl_iskontolu_bf_dahil = QLabel("0,00 TL", font=QFont("Segoe UI", 9))
        main_frame_layout.addWidget(self.lbl_iskontolu_bf_dahil, current_row, 1, 1, 2)
        current_row += 1

        main_frame_layout.addWidget(QLabel("Kalem Toplam (KDV Dahil):", font=QFont("Segoe UI", 10, QFont.Bold)), current_row, 0)
        self.lbl_kalem_toplam_dahil = QLabel("0,00 TL", font=QFont("Segoe UI", 10, QFont.Bold))
        main_frame_layout.addWidget(self.lbl_kalem_toplam_dahil, current_row, 1, 1, 2)
        current_row += 1

        btn_f = QFrame(self)
        btn_layout = QHBoxLayout(btn_f)
        main_layout.addWidget(btn_f, alignment=Qt.AlignRight)
        
        btn_guncelle = QPushButton("Güncelle")
        btn_guncelle.clicked.connect(self._kalemi_kaydet)
        btn_layout.addWidget(btn_guncelle)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        btn_layout.addWidget(btn_iptal)

        self._anlik_hesaplama_ve_guncelleme()
        self.miktar_e.setFocus()
        self.miktar_e.selectAll()

    def _anlik_hesaplama_ve_guncelleme(self):
        try:
            miktar = self.db.safe_float(self.miktar_e.text())
            birim_fiyat_kdv_dahil_orijinal = self.db.safe_float(self.fiyat_e.text())

            yuzde_iskonto_1 = self.db.safe_float(self.iskonto_yuzde_1_e.text())
            yuzde_iskonto_2 = self.db.safe_float(self.iskonto_yuzde_2_e.text())

            if not (0 <= yuzde_iskonto_1 <= 100):
                self.iskonto_yuzde_1_e.setText("0,00")
                yuzde_iskonto_1 = 0.0

            if not (0 <= yuzde_iskonto_2 <= 100):
                self.iskonto_yuzde_2_e.setText("0,00")
                yuzde_iskonto_2 = 0.0

            fiyat_iskonto_1_sonrasi_dahil = birim_fiyat_kdv_dahil_orijinal * (1 - yuzde_iskonto_1 / 100)
            iskontolu_birim_fiyat_dahil = fiyat_iskonto_1_sonrasi_dahil * (1 - yuzde_iskonto_2 / 100)
            
            if iskontolu_birim_fiyat_dahil < 0:
                iskontolu_birim_fiyat_dahil = 0.0

            toplam_uygulanan_iskonto_dahil = birim_fiyat_kdv_dahil_orijinal - iskontolu_birim_fiyat_dahil
            
            kalem_toplam_dahil = miktar * iskontolu_birim_fiyat_dahil

            if birim_fiyat_kdv_dahil_orijinal > 0:
                toplam_iskonto_yuzdesi = (toplam_uygulanan_iskonto_dahil / birim_fiyat_kdv_dahil_orijinal) * 100
            else:
                toplam_iskonto_yuzdesi = 0.0 

            self.lbl_toplam_iskonto_yuzdesi.setText(f"{toplam_iskonto_yuzdesi:,.2f} %")
            self.lbl_uygulanan_iskonto_dahil.setText(self.db._format_currency(toplam_uygulanan_iskonto_dahil))
            self.lbl_iskontolu_bf_dahil.setText(self.db._format_currency(iskontolu_birim_fiyat_dahil))
            self.lbl_kalem_toplam_dahil.setText(self.db._format_currency(kalem_toplam_dahil))

        except ValueError:
            self.lbl_toplam_iskonto_yuzdesi.setText("0,00 %")
            self.lbl_uygulanan_iskonto_dahil.setText("0,00 TL")
            self.lbl_iskontolu_bf_dahil.setText("0,00 TL")
            self.lbl_kalem_toplam_dahil.setText("0,00 TL")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Hesaplama sırasında beklenmeyen bir hata oluştu: {e}")
            logging.error(f"Anlık hesaplama hatası: {e}", exc_info=True)


    def _kalemi_kaydet(self):
        yeni_miktar = 0.0
        yeni_fiyat_kdv_dahil_orijinal = 0.0
        yuzde_iskonto_1 = 0.0
        yuzde_iskonto_2 = 0.0
        yeni_alis_fiyati_aninda = self.mevcut_alis_fiyati_fatura_aninda

        try:
            yeni_miktar = self.db.safe_float(self.miktar_e.text())
            yeni_fiyat_kdv_dahil_orijinal = self.db.safe_float(self.fiyat_e.text())
            
            yuzde_iskonto_1 = self.db.safe_float(self.iskonto_yuzde_1_e.text())
            yuzde_iskonto_2 = self.db.safe_float(self.iskonto_yuzde_2_e.text())
            
            if (self.islem_tipi == self.db.FATURA_TIP_SATIS or self.islem_tipi == self.db.SIPARIS_TIP_SATIS) and self.alis_fiyati_aninda_e:
                yeni_alis_fiyati_aninda = self.db.safe_float(self.alis_fiyati_aninda_e.text())

            if yeni_miktar <= 0:
                QMessageBox.critical(self, "Geçersiz Miktar", "Miktar pozitif bir sayı olmalıdır.")
                return
            if yeni_fiyat_kdv_dahil_orijinal < 0:
                QMessageBox.critical(self, "Geçersiz Fiyat", "Birim fiyat negatif olamaz.")
                return
            if not (0 <= yuzde_iskonto_1 <= 100):
                QMessageBox.critical(self, "Geçersiz İskonto 1 Yüzdesi", "İskonto 1 yüzdesi 0 ile 100 arasında olmalıdır.")
                return
            if not (0 <= yuzde_iskonto_2 <= 100):
                QMessageBox.critical(self, "Geçersiz İskonto 2 Yüzdesi", "İskonto 2 yüzdesi 0 ile 100 arasında olmalıdır.")
                return
            if (self.islem_tipi == self.db.FATURA_TIP_SATIS or self.islem_tipi == self.db.SIPARIS_TIP_SATIS) and self.alis_fiyati_aninda_e and yeni_alis_fiyati_aninda < 0:
                QMessageBox.critical(self, "Geçersiz Fiyat", "Fatura anı alış fiyatı negatif olamaz.")
                return
            
            self.parent_page.kalem_guncelle( # _kalem_guncelle yerine kalem_guncelle oldu
                self.kalem_index, 
                yeni_miktar, 
                yeni_fiyat_kdv_dahil_orijinal, 
                yuzde_iskonto_1,       
                yuzde_iskonto_2,       
                yeni_alis_fiyati_aninda 
            )
            self.accept() # QDialog'u kapat.

        except ValueError as ve:
            QMessageBox.critical(self, "Giriş Hatası", f"Sayısal alanlarda geçersiz değerler var: {ve}")
            logging.error(f"Kalem Guncelle ValueError: {ve}", exc_info=True)
        except IndexError as ie:
            QMessageBox.critical(self, "Hata", f"Güncellenecek kalem bulunamadı (indeks hatası): {ie}")
            logging.error(f"Kalem Guncelle IndexError: {ie}", exc_info=True)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kalem güncellenirken beklenmeyen bir hata oluştu: {e}")
            logging.error(f"Kalem Guncelle Genel Hata: {e}", exc_info=True)

class FiyatGecmisiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, cari_id, urun_id, fatura_tipi, update_callback, current_kalem_index):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.cari_id = cari_id
        self.urun_id = urun_id
        self.fatura_tipi = fatura_tipi
        self.update_callback = update_callback # FaturaOlusturmaSayfasi'ndaki kalemi güncelleme callback'i
        self.current_kalem_index = current_kalem_index # Sepetteki güncel kalemin indeksi

        self.setWindowTitle("Fiyat Geçmişi Seç")
        self.setFixedSize(600, 400) # Boyut ayarı (resizable=False yerine)
        self.setModal(True) # Diğer pencerelere tıklamayı engeller

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Geçmiş Fiyat Listesi")
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # Fiyat Geçmişi Listesi (Treeview)
        tree_frame = QFrame(self)
        tree_layout = QVBoxLayout(tree_frame)
        main_layout.addWidget(tree_frame)

        # Sütunlar: Fatura No, Tarih, Fiyat (KDV Dahil), İskonto 1 (%), İskonto 2 (%)
        cols = ("Fatura No", "Tarih", "Fiyat (KDV Dahil)", "İskonto 1 (%)", "İskonto 2 (%)")
        self.price_history_tree = QTreeWidget()
        self.price_history_tree.setHeaderLabels(cols)
        self.price_history_tree.setSelectionBehavior(QAbstractItemView.SelectRows) # Tek satır seçimi
        self.price_history_tree.setSortingEnabled(True)

        from PySide6.QtWidgets import QHeaderView # PySide6'ya özel import
        col_defs = [
            ("Fatura No", 120, Qt.AlignLeft),
            ("Tarih", 90, Qt.AlignCenter),
            ("Fiyat (KDV Dahil)", 120, Qt.AlignRight),
            ("İskonto 1 (%)", 90, Qt.AlignRight),
            ("İskonto 2 (%)", 90, Qt.AlignRight)
        ]

        for i, (col_name, width, alignment) in enumerate(col_defs):
            self.price_history_tree.setColumnWidth(i, width)
            self.price_history_tree.headerItem().setTextAlignment(i, alignment)
            self.price_history_tree.headerItem().setFont(i, QFont("Segoe UI", 9, QFont.Bold))
        
        self.price_history_tree.header().setStretchLastSection(True) # Son sütunu esnet

        tree_layout.addWidget(self.price_history_tree)

        # Çift tıklama veya seçip butona basma ile fiyatı seçme
        self.price_history_tree.itemDoubleClicked.connect(self._on_price_selected_double_click)

        self._load_price_history() # Geçmiş fiyatları yükle

        # Alt Butonlar
        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        btn_onayla = QPushButton("Seç ve Uygula")
        btn_onayla.clicked.connect(self._on_price_selected_button)
        button_layout.addWidget(btn_onayla)
        
        button_layout.addStretch() # Sağ tarafa yaslamak için boşluk

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_kapat)

    def _load_price_history(self):
        """Veritabanından geçmiş fiyat bilgilerini çeker ve Treeview'e doldurur."""
        self.price_history_tree.clear()
        # db.get_gecmis_fatura_kalemi_bilgileri metodunu çağır
        history_data = self.db.get_gecmis_fatura_kalemi_bilgileri(self.cari_id, self.urun_id, self.fatura_tipi) 

        if not history_data:
            item_qt = QTreeWidgetItem(self.price_history_tree)
            item_qt.setText(2, "Geçmiş Fiyat Yok")
            return

        for item in history_data:
            # item: (fatura_id, fatura_no, formatted_date, nihai_iskontolu_kdv_dahil_bf, iskonto_yuzde_1, iskonto_yuzde_2)
            fatura_id = item[0]
            fatura_no = item[1]
            tarih = item[2]
            fiyat = self.db._format_currency(item[3])
            iskonto_1 = f"{item[4]:.2f}".replace('.', ',').rstrip('0').rstrip(',')
            iskonto_2 = f"{item[5]:.2f}".replace('.', ',').rstrip('0').rstrip(',')

            item_qt = QTreeWidgetItem(self.price_history_tree)
            item_qt.setText(0, fatura_no)
            item_qt.setText(1, tarih)
            item_qt.setText(2, fiyat)
            item_qt.setText(3, iskonto_1)
            item_qt.setText(4, iskonto_2)
            item_qt.setData(0, Qt.UserRole, fatura_id) # Fatura ID'yi sakla (opsiyonel)


    def _on_price_selected_double_click(self, item, column): # item ve column QTreeWidget sinyalinden gelir
        self._on_price_selected_button()

    def _on_price_selected_button(self):
        """Seçilen fiyatı alır ve FaturaOlusturmaSayfasi'na geri gönderir."""
        selected_items = self.price_history_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen uygulamak için bir geçmiş fiyat seçin.")
            return

        item_values = [selected_items[0].text(i) for i in range(self.price_history_tree.columnCount())]
        
        # item_values formatı: ["Fatura No", "Tarih", "Fiyat (KDV Dahil)", "İskonto 1 (%)", "İskonto 2 (%)"]
        # Fiyatı, İskonto 1 ve İskonto 2'yi al
        selected_price_str = item_values[2] 
        selected_iskonto1_str = item_values[3] 
        selected_iskonto2_str = item_values[4] 

        try:
            cleaned_price_str = selected_price_str.replace(' TL', '').replace('₺', '').strip()
            cleaned_iskonto1_str = selected_iskonto1_str.replace('%', '').strip()
            cleaned_iskonto2_str = selected_iskonto2_str.replace('%', '').strip()

            selected_price = self.db.safe_float(cleaned_price_str)
            selected_iskonto1 = self.db.safe_float(cleaned_iskonto1_str)
            selected_iskonto2 = self.db.safe_float(cleaned_iskonto2_str)

            logging.debug(f"Secilen Fiyat (temizlenmis): '{cleaned_price_str}' -> {selected_price}")
            logging.debug(f"Secilen Iskonto 1 (temizlenmis): '{cleaned_iskonto1_str}' -> {selected_iskonto1}")
            logging.debug(f"Secilen Iskonto 2 (temizlenmis): '{cleaned_iskonto2_str}' -> {selected_iskonto2}")

        except ValueError:
            QMessageBox.critical(self, "Hata", "Seçilen fiyat verisi geçersiz. (Dönüştürme hatası)")
            return
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Fiyat geçmişi verisi işlenirken beklenmeyen bir hata oluştu: {e}")
            logging.error(f"Fiyat geçmişi verisi işleme hatası: {e}", exc_info=True)
            return

        # update_callback metodu, (kalem_index, yeni_birim_fiyat_kdv_dahil, yeni_iskonto_1, yeni_iskonto_2) alacak.
        self.update_callback(self.current_kalem_index, selected_price, selected_iskonto1, selected_iskonto2)
        self.close() # Pencereyi kapat

class KullaniciYonetimiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app # Ana App referansı
        self.setWindowTitle("Kullanıcı Yönetimi")
        self.setMinimumSize(600, 650)
        self.setModal(True)

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Kullanıcı Listesi ve Yönetimi")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # Kullanıcı Listesi
        list_frame = QFrame(self)
        list_layout = QHBoxLayout(list_frame)
        main_layout.addWidget(list_frame)
        
        cols_kul = ("ID", "Kullanıcı Adı", "Yetki")
        self.tree_kul = QTreeWidget()
        self.tree_kul.setHeaderLabels(cols_kul)
        self.tree_kul.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tree_kul.setSortingEnabled(True) # Sıralama özelliği

        self.tree_kul.setColumnWidth(0, 50)
        self.tree_kul.headerItem().setTextAlignment(0, Qt.AlignRight)
        self.tree_kul.headerItem().setTextAlignment(2, Qt.AlignCenter)
        self.tree_kul.header().setSectionResizeMode(1, QHeaderView.Stretch) # Kullanıcı Adı genişlesin

        list_layout.addWidget(self.tree_kul)
        
        self.kullanici_listesini_yenile() # İlk yüklemede listeyi doldur

        # Yeni Kullanıcı Ekleme Formu
        form_frame = QGroupBox("Yeni Kullanıcı Ekle / Güncelle", self)
        form_layout = QGridLayout(form_frame)
        main_layout.addWidget(form_frame)

        form_layout.addWidget(QLabel("Kullanıcı Adı:"), 0, 0, Qt.AlignLeft)
        self.k_adi_yeni_e = QLineEdit()
        form_layout.addWidget(self.k_adi_yeni_e, 0, 1)
        form_layout.setColumnStretch(1, 1) # Genişlesin

        form_layout.addWidget(QLabel("Yeni Şifre:"), 1, 0, Qt.AlignLeft)
        self.sifre_yeni_e = QLineEdit()
        self.sifre_yeni_e.setEchoMode(QLineEdit.Password) # Şifre gizleme
        form_layout.addWidget(self.sifre_yeni_e, 1, 1)

        form_layout.addWidget(QLabel("Yetki:"), 0, 2, Qt.AlignLeft)
        self.yetki_yeni_cb = QComboBox()
        self.yetki_yeni_cb.addItems(["kullanici", "admin"])
        self.yetki_yeni_cb.setCurrentText("kullanici") # Varsayılan
        form_layout.addWidget(self.yetki_yeni_cb, 0, 3)

        # Butonlar
        button_frame_kul = QFrame(self)
        button_layout_kul = QHBoxLayout(button_frame_kul)
        main_layout.addWidget(button_frame_kul)
        
        self.ekle_guncelle_btn = QPushButton("Ekle / Güncelle")
        self.ekle_guncelle_btn.clicked.connect(self.yeni_kullanici_ekle)
        button_layout_kul.addWidget(self.ekle_guncelle_btn)
        
        btn_sil_kul = QPushButton("Seçili Kullanıcıyı Sil")
        btn_sil_kul.clicked.connect(self.secili_kullanici_sil)
        button_layout_kul.addWidget(btn_sil_kul)
        
        button_layout_kul.addStretch() # Sağa yaslama
        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        button_layout_kul.addWidget(btn_kapat)

        self.tree_kul.itemSelectionChanged.connect(self.secili_kullaniciyi_forma_yukle) # Seçim değiştiğinde formu doldur

    def kullanici_listesini_yenile(self):
        self.tree_kul.clear()
        try:
            # API'den kullanıcı listesini çekmek için uygun bir endpoint varsayımı
            # Eğer API'de böyle bir endpoint yoksa, doğrudan db_manager kullanılmalıdır.
            # Şimdilik db_manager'dan çekiliyor.
            kullanicilar = self.db.kullanici_listele()
            
            for kul in kullanicilar:
                item_qt = QTreeWidgetItem(self.tree_kul)
                item_qt.setText(0, str(kul.get('id'))) # 'id' alanı
                item_qt.setText(1, kul.get('kullanici_adi')) # 'kullanici_adi' alanı
                item_qt.setText(2, kul.get('yetki')) # 'yetki' alanı
                item_qt.setData(0, Qt.UserRole, kul.get('id')) # ID'yi UserRole olarak sakla
                
            self.app.set_status_message(f"{len(kullanicilar)} kullanıcı listelendi.")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kullanıcı listesi çekilirken hata: {e}")
            logging.error(f"Kullanıcı listesi yükleme hatası: {e}", exc_info=True)
    
    def secili_kullaniciyi_forma_yukle(self):
        selected_items = self.tree_kul.selectedItems()
        if selected_items:
            item = selected_items[0]
            kullanici_adi = item.text(1)
            yetki = item.text(2)
            self.k_adi_yeni_e.setText(kullanici_adi)
            self.yetki_yeni_cb.setCurrentText(yetki)
            self.sifre_yeni_e.clear() # Şifre alanı temizlensin
            self.ekle_guncelle_btn.setText("Güncelle")
        else: # Seçim yoksa formu temizle
            self.k_adi_yeni_e.clear()
            self.sifre_yeni_e.clear()
            self.yetki_yeni_cb.setCurrentText("kullanici")
            self.ekle_guncelle_btn.setText("Ekle / Güncelle")

    def yeni_kullanici_ekle(self):
        k_adi = self.k_adi_yeni_e.text().strip()
        sifre = self.sifre_yeni_e.text().strip()
        yetki = self.yetki_yeni_cb.currentText()

        if not (k_adi and yetki):
            QMessageBox.critical(self, "Eksik Bilgi", "Kullanıcı adı ve yetki boş bırakılamaz.")
            return

        selected_items = self.tree_kul.selectedItems()
        
        if selected_items: # Güncelleme
            user_id = selected_items[0].data(0, Qt.UserRole)
            mevcut_k_adi = selected_items[0].text(1)

            success_name_update = True
            message_name_update = ""

            if k_adi != mevcut_k_adi:
                try:
                    # API endpoint'i üzerinden kullanıcı adını güncelleme (varsayalım mevcut)
                    # response = requests.put(f"{API_BASE_URL}/kullanicilar/{user_id}/kullanici_adi", json={"kullanici_adi": k_adi})
                    # response.raise_for_status()
                    # success_name_update, message_name_update = True, "Kullanıcı adı güncellendi."
                    success_name_update, message_name_update = self.db.kullanici_adi_guncelle(user_id, k_adi)

                except Exception as e:
                    success_name_update = False
                    message_name_update = f"Kullanıcı adı güncellenirken hata: {e}"
                    logging.error(f"Kullanıcı adı güncelleme hatası: {e}", exc_info=True)
                
                if not success_name_update:
                    QMessageBox.critical(self, "Hata", message_name_update)
                    return

            sifre_to_hash = None
            if sifre:
                sifre_to_hash = self.db._hash_sifre(sifre)
            else: # Şifre boş bırakılırsa mevcut şifreyi koru
                try:
                    # API'den şifre çekme veya doğrudan db_manager'dan çekme
                    # response = requests.get(f"{API_BASE_URL}/kullanicilar/{user_id}/sifre_hash")
                    # response.raise_for_status()
                    # sifre_to_hash = response.json().get('sifre_hash')
                    self.db.c.execute("SELECT sifre FROM kullanicilar WHERE id=?", (user_id,))
                    sifre_to_hash = self.db.c.fetchone()[0]
                except Exception as e:
                    QMessageBox.critical(self, "Hata", f"Mevcut şifre alınırken bir hata oluştu: {e}")
                    logging.error(f"Mevcut şifre alma hatası: {e}", exc_info=True)
                    return

            try:
                # API endpoint'i üzerinden kullanıcıyı güncelleme
                # response = requests.put(f"{API_BASE_URL}/kullanicilar/{user_id}", json={"sifre": sifre_to_hash, "yetki": yetki})
                # response.raise_for_status()
                # success, message = True, "Kullanıcı başarıyla güncellendi."
                success, message = self.db.kullanici_guncelle_sifre_yetki(user_id, sifre_to_hash, yetki)

                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.app.set_status_message(message)
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kullanıcı güncellenirken hata: {e}")
                logging.error(f"Kullanıcı güncelleme hatası: {e}", exc_info=True)

            self.kullanici_listesini_yenile()
            self.k_adi_yeni_e.clear()
            self.sifre_yeni_e.clear()
            self.tree_kul.clearSelection()
            self.secili_kullaniciyi_forma_yukle() # Formu temizle (butonu da "Ekle / Güncelle" yapar)

        else: # Yeni kullanıcı ekleme
            if not sifre:
                QMessageBox.critical(self, "Eksik Bilgi", "Yeni kullanıcı eklerken şifre boş bırakılamaz.")
                return

            try:
                # API endpoint'i üzerinden yeni kullanıcı ekleme
                # response = requests.post(f"{API_BASE_URL}/kullanicilar/", json={"kullanici_adi": k_adi, "sifre": sifre, "yetki": yetki})
                # response.raise_for_status()
                # success, message = True, "Yeni kullanıcı başarıyla eklendi."
                success, message = self.db.kullanici_ekle(k_adi, sifre, yetki)

                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.app.set_status_message(message)
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Yeni kullanıcı eklenirken hata: {e}")
                logging.error(f"Yeni kullanıcı ekleme hatası: {e}", exc_info=True)

            self.kullanici_listesini_yenile()
            self.k_adi_yeni_e.clear()
            self.sifre_yeni_e.clear()
            self.tree_kul.clearSelection()
            self.secili_kullaniciyi_forma_yukle()

    def secili_kullanici_sil(self):
        selected_items = self.tree_kul.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Seçim Yok", "Lütfen silmek istediğiniz kullanıcıyı seçin.")
            return
        
        k_adi_secili = selected_items[0].text(1)
        user_id_to_delete = selected_items[0].data(0, Qt.UserRole)

        if k_adi_secili == self.app.current_user[1]: 
             QMessageBox.warning(self, "Engellendi", "Aktif olarak giriş yapmış olduğunuz kendi kullanıcı hesabınızı silemezsiniz.")
             return

        reply = QMessageBox.question(self, "Onay", f"'{k_adi_secili}' kullanıcısını silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                # API endpoint'i üzerinden kullanıcı silme
                # response = requests.delete(f"{API_BASE_URL}/kullanicilar/{user_id_to_delete}")
                # response.raise_for_status()
                # success, message = True, "Kullanıcı başarıyla silindi."
                success, message = self.db.kullanici_sil(user_id_to_delete)

                if success:
                    QMessageBox.information(self, "Başarılı", message)
                    self.kullanici_listesini_yenile()
                    self.app.set_status_message(message)
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kullanıcı silinirken hata: {e}")
                logging.error(f"Kullanıcı silme hatası: {e}", exc_info=True)

class YeniGelirGiderEklePenceresi(QDialog):
    def __init__(self, parent_app, db_manager, yenile_callback, initial_tip=None):
        super().__init__(parent_app)
        self.db = db_manager
        self.yenile_callback = yenile_callback
        self.app = parent_app # parent_app'i app olarak kaydet

        self.kasa_banka_map = {}
        self.gelir_siniflandirma_map = {}
        self.gider_siniflandirma_map = {}

        self.setWindowTitle("Yeni Manuel Gelir/Gider Kaydı")
        self.setFixedSize(450, 450) # resizable=False yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        entry_frame = QFrame(self)
        main_layout.addWidget(entry_frame)
        entry_frame_layout = QGridLayout(entry_frame)
        
        current_row = 0

        entry_frame_layout.addWidget(QLabel("Tarih (YYYY-AA-GG):"), current_row, 0, Qt.AlignLeft)
        self.tarih_entry = QLineEdit()
        self.tarih_entry.setText(datetime.now().strftime('%Y-%m-%d'))
        # setup_date_entry PySide6 için placeholder, validator ile manuel kontrol daha iyi
        self.tarih_entry.setPlaceholderText("YYYY-AA-GG")
        entry_frame_layout.addWidget(self.tarih_entry, current_row, 1)
        btn_date = QPushButton("🗓️")
        btn_date.setFixedWidth(30)
        btn_date.clicked.connect(lambda: DatePickerDialog(self.app, self.tarih_entry)) # app referansı kullanıldı
        entry_frame_layout.addWidget(btn_date, current_row, 2)
        current_row += 1

        entry_frame_layout.addWidget(QLabel("İşlem Tipi:"), current_row, 0, Qt.AlignLeft)
        self.tip_combo = QComboBox()
        self.tip_combo.addItems(["GELİR", "GİDER"])
        if initial_tip and initial_tip in ["GELİR", "GİDER"]:
            self.tip_combo.setCurrentText(initial_tip)
        else:
            self.tip_combo.setCurrentIndex(0)
        self.tip_combo.currentIndexChanged.connect(self._on_tip_changed)
        entry_frame_layout.addWidget(self.tip_combo, current_row, 1)
        current_row += 1

        entry_frame_layout.addWidget(QLabel("Sınıflandırma:"), current_row, 0, Qt.AlignLeft)
        self.siniflandirma_combo = QComboBox()
        entry_frame_layout.addWidget(self.siniflandirma_combo, current_row, 1)
        current_row += 1

        entry_frame_layout.addWidget(QLabel("Tutar (TL):"), current_row, 0, Qt.AlignLeft)
        self.tutar_entry = QLineEdit("0,00")
        setup_numeric_entry(self.app, self.tutar_entry) # <-- allow_negative=False, decimal_places=2 parametreleri kaldırıldı
        entry_frame_layout.addWidget(self.tutar_entry, current_row, 1)
        current_row += 1

        entry_frame_layout.addWidget(QLabel("İşlem Kasa/Banka (*):"), current_row, 0, Qt.AlignLeft)
        self.kasa_banka_combobox = QComboBox()
        entry_frame_layout.addWidget(self.kasa_banka_combobox, current_row, 1)
        current_row += 1
        
        entry_frame_layout.addWidget(QLabel("Açıklama:"), current_row, 0, Qt.AlignLeft)
        self.aciklama_entry = QLineEdit()
        entry_frame_layout.addWidget(self.aciklama_entry, current_row, 1)
        current_row += 1
        
        entry_frame_layout.setColumnStretch(1, 1) # İkinci sütun genişlesin

        main_layout.addStretch() # Üst kısımdaki elemanları yukarı it

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignCenter) # Butonları ortala

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close)
        button_layout.addWidget(btn_iptal)
        
        self._yukle_siniflandirmalar_comboboxlari_ve_ayarla()
        self.tarih_entry.setFocus()
        self.adjustSize() # Pencere boyutunu içeriğe göre ayarla

    def _yukle_siniflandirmalar_comboboxlari_ve_ayarla(self):
        self._yukle_kasa_banka_hesaplarini() 

        # API'den sınıflandırmaları çek
        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            gelir_siniflandirmalar_api = self.db.gelir_siniflandirma_listele()
            self.gelir_siniflandirma_map = {item.get('ad'): item.get('id') for item in gelir_siniflandirmalar_api} # 'siniflandirma_adi' yerine 'ad' kullanıldı

            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            gider_siniflandirmalar_api = self.db.gider_siniflandirma_listele()
            self.gider_siniflandirma_map = {item.get('ad'): item.get('id') for item in gider_siniflandirmalar_api} # 'siniflandirma_adi' yerine 'ad' kullanıldı

        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Sınıflandırmalar yüklenirken hata: {e}")
            logging.error(f"Sınıflandırma yükleme hatası: {e}", exc_info=True)

        self._on_tip_changed()

    def _on_tip_changed(self):
        selected_tip = self.tip_combo.currentText()
        display_values = ["Seçim Yok"]
        selected_map = {}

        if selected_tip == "GELİR":
            selected_map = self.gelir_siniflandirma_map
        elif selected_tip == "GİDER":
            selected_map = self.gider_siniflandirma_map

        display_values.extend(sorted(selected_map.keys()))
        self.siniflandirma_combo.clear()
        self.siniflandirma_combo.addItems(display_values)
        self.siniflandirma_combo.setCurrentText("Seçim Yok")
        # combobox'ın state'i QComboBox'ta otomatik olarak readonly'dir.

    def _yukle_kasa_banka_hesaplarini(self):
        self.kasa_banka_combobox.clear()
        self.kasa_banka_map.clear()

        try:
            hesaplar_response = self.db.kasa_banka_listesi_al(limit=10000)

            hesaplar = []
            if isinstance(hesaplar_response, dict) and "items" in hesaplar_response:
                hesaplar = hesaplar_response["items"]
            elif isinstance(hesaplar_response, list):
                hesaplar = hesaplar_response
                self.app.set_status_message("Uyarı: Kasa/Banka listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                hesaplar = []
                self.app.set_status_message("Hata: Kasa/Banka listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Kasa/Banka listesi API'den beklenen formatta gelmedi: {type(hesaplar_response)} - {hesaplar_response}")
                self.kasa_banka_combobox.addItem("Hesap Yok", None)
                self.kasa_banka_combobox.setEnabled(False)
                return

            if hesaplar:
                for h in hesaplar:
                    display_text = f"{h.get('hesap_adi')} ({h.get('tip')})"
                    if h.get('tip') == "BANKA" and h.get('banka_adi'):
                        display_text += f" - {h.get('banka_adi')}"
                    if h.get('bakiye') is not None:
                        display_text += f" (Bakiye: {self.db._format_currency(h.get('bakiye'))})"

                    self.kasa_banka_map[display_text] = h.get('id')
                    self.kasa_banka_combobox.addItem(display_text, h.get('id'))

                default_hesap_text = None
                for text in self.kasa_banka_map.keys():
                    if text.strip().startswith("NAKİT KASA"):
                        default_hesap_text = text
                        break

                if default_hesap_text:
                    self.kasa_banka_combobox.setCurrentText(default_hesap_text)
                elif self.kasa_banka_combobox.count() > 0:
                    self.kasa_banka_combobox.setCurrentIndex(0)
                else:
                    self.kasa_banka_combobox.clear()
                    self.kasa_banka_combobox.addItem("Hesap Yok", None)
                    self.kasa_banka_combobox.setEnabled(False)

            self.app.set_status_message(f"{len(hesaplar)} kasa/banka hesabı API'den yüklendi.")

        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Kasa/Banka hesapları yüklenirken hata: {e}")
            logging.error(f"Kasa/Banka yükleme hatası: {e}", exc_info=True)
            self.kasa_banka_combobox.addItem("Hesap Yok", None)
            self.kasa_banka_combobox.setEnabled(False)

    def _kaydet(self):
        tarih_str = self.tarih_entry.text().strip()
        tip_str = self.tip_combo.currentText()
        tutar_str = self.tutar_entry.text().strip()
        aciklama_str = self.aciklama_entry.text().strip()

        secili_hesap_id = self.kasa_banka_combobox.currentData()

        secili_siniflandirma_adi = self.siniflandirma_combo.currentText()
        gelir_siniflandirma_id_val = None
        gider_siniflandirma_id_val = None

        if secili_siniflandirma_adi == "Seçim Yok" or not secili_siniflandirma_adi:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir sınıflandırma seçin.")
            return

        if tip_str == "GELİR":
            gelir_siniflandirma_id_val = self.gelir_siniflandirma_map.get(secili_siniflandirma_adi)
        elif tip_str == "GİDER":
            gider_siniflandirma_id_val = self.gider_siniflandirma_map.get(secili_siniflandirma_adi)

        if secili_hesap_id is None:
            QMessageBox.critical(self, "Eksik Bilgi", "Lütfen bir İşlem Kasa/Banka hesabı seçin.")
            return

        # Tarih formatı kontrolü
        try:
            datetime.strptime(tarih_str, '%Y-%m-%d')
        except ValueError:
            QMessageBox.critical(self, "Hata", "Tarih formatı 'YYYY-AA-GG' şeklinde olmalıdır.")
            return

        if not all([tarih_str, tutar_str, aciklama_str]):
            QMessageBox.critical(self, "Eksik Bilgi", "Lütfen tüm zorunlu (*) alanları doldurun.")
            return

        try:
            tutar_f = float(tutar_str.replace(',', '.'))
            if tutar_f <= 0:
                QMessageBox.critical(self, "Geçersiz Tutar", "Tutar pozitif bir sayı olmalıdır.")
                return
        except ValueError:
            QMessageBox.critical(self, "Giriş Hatası", "Tutar sayısal bir değer olmalıdır.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            data = {
                "tarih": tarih_str,
                "tip": tip_str,
                "tutar": tutar_f,
                "aciklama": aciklama_str,
                "kaynak": "MANUEL", # Manuel işlemler için kaynak her zaman MANUEL'dir.
                "kasa_banka_id": secili_hesap_id,
                "gelir_siniflandirma_id": gelir_siniflandirma_id_val,
                "gider_siniflandirma_id": gider_siniflandirma_id_val
            }
            success = self.db.gelir_gider_ekle(data)

            if success:
                QMessageBox.information(self, "Başarılı", "Gelir/Gider kaydı başarıyla eklendi.")
                if self.yenile_callback:
                    self.yenile_callback()
                self.accept() # QDialog'u kapat
            else:
                QMessageBox.critical(self, "Hata", "Gelir/Gider kaydı eklenirken bir hata oluştu.")

        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "Hata", f"Kaydedilirken bir hata oluştu:\n{e}")
            logging.error(f"Gelir/Gider kaydetme hatası: {e}", exc_info=True)

class TarihAraligiDialog(QDialog): # simpledialog.Dialog yerine QDialog kullanıldı
    def __init__(self, parent_app, title=None, baslangic_gun_sayisi=30):
        super().__init__(parent_app)
        self.app = parent_app # Ana uygulama referansını tut
        self.bas_tarih_str = (datetime.now() - timedelta(days=baslangic_gun_sayisi)).strftime('%Y-%m-%d')
        self.bitis_tarih_str = datetime.now().strftime('%Y-%m-%d')
        self.sonuc = None # Kullanıcının seçtiği tarih aralığını tutacak

        self.setWindowTitle(title if title else "Tarih Aralığı Seçin")
        self.setFixedSize(350, 180) # Sabit boyut
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        form_layout = QGridLayout()
        main_layout.addLayout(form_layout)

        form_layout.addWidget(QLabel("Başlangıç Tarihi (YYYY-AA-GG):"), 0, 0, Qt.AlignLeft)
        self.bas_tarih_entry_dialog = QLineEdit()
        self.bas_tarih_entry_dialog.setText(self.bas_tarih_str)
        form_layout.addWidget(self.bas_tarih_entry_dialog, 0, 1)
        btn_bas_tarih = QPushButton("🗓️")
        btn_bas_tarih.setFixedWidth(30)
        btn_bas_tarih.clicked.connect(lambda: DatePickerDialog(self.app, self.bas_tarih_entry_dialog)) # app referansı kullanıldı
        form_layout.addWidget(btn_bas_tarih, 0, 2)

        form_layout.addWidget(QLabel("Bitiş Tarihi (YYYY-AA-GG):"), 1, 0, Qt.AlignLeft)
        self.bit_tarih_entry_dialog = QLineEdit()
        self.bit_tarih_entry_dialog.setText(self.bitis_tarih_str)
        form_layout.addWidget(self.bit_tarih_entry_dialog, 1, 1)
        btn_bit_tarih = QPushButton("🗓️")
        btn_bit_tarih.setFixedWidth(30)
        btn_bit_tarih.clicked.connect(lambda: DatePickerDialog(self.app, self.bit_tarih_entry_dialog)) # app referansı kullanıldı
        form_layout.addWidget(btn_bit_tarih, 1, 2)

        button_layout = QHBoxLayout()
        main_layout.addLayout(button_layout)
        button_layout.addStretch()

        btn_ok = QPushButton("Onayla")
        btn_ok.clicked.connect(self._apply)
        button_layout.addWidget(btn_ok)

        btn_cancel = QPushButton("İptal")
        btn_cancel.clicked.connect(self.reject) # QDialog'u reject ile kapat
        button_layout.addWidget(btn_cancel)

        self.bas_tarih_entry_dialog.setFocus() # İlk odaklanılacak widget

    def _apply(self): 
        bas_t_str_dialog = self.bas_tarih_entry_dialog.text()
        bit_t_str_dialog = self.bit_tarih_entry_dialog.text()
        try:
            bas_dt_dialog = datetime.strptime(bas_t_str_dialog, '%Y-%m-%d')
            bit_dt_dialog = datetime.strptime(bit_t_str_dialog, '%Y-%m-%d')
            if bas_dt_dialog > bit_dt_dialog:
                QMessageBox.critical(self, "Tarih Hatası", "Başlangıç tarihi, bitiş tarihinden sonra olamaz.")
                self.sonuc = None 
                return
            self.sonuc = (bas_t_str_dialog, bit_t_str_dialog) 
            self.accept() # QDialog'u accept ile kapat
        except ValueError:
            QMessageBox.critical(self, "Format Hatası", "Tarih formatı YYYY-AA-GG olmalıdır (örn: 2023-12-31).")
            self.sonuc = None
            return

class OdemeTuruSecimDialog(QDialog):
    def __init__(self, parent_app, db_manager, fatura_tipi, initial_cari_id, callback_func):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.fatura_tipi = fatura_tipi
        self.initial_cari_id = initial_cari_id
        self.callback_func = callback_func
        self.setWindowTitle("Ödeme Türü Seçimi")
        self.setFixedSize(400, 300)
        self.setModal(True)
        self.kasa_banka_map = {}
        main_layout = QVBoxLayout(self)
        title_label = QLabel("Fatura Ödeme Türünü Seçin")
        title_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        form_frame = QFrame(self)
        form_layout = QGridLayout(form_frame)
        main_layout.addWidget(form_frame)
        form_layout.addWidget(QLabel("Ödeme Türü (*):"), 0, 0, Qt.AlignLeft)
        self.odeme_turu_cb = QComboBox()
        self._set_odeme_turu_values()
        form_layout.addWidget(self.odeme_turu_cb, 0, 1)
        self.odeme_turu_cb.currentIndexChanged.connect(self._odeme_turu_degisince_hesap_combobox_ayarla)
        self.odeme_turu_cb.setCurrentIndex(0)
        form_layout.addWidget(QLabel("İşlem Kasa/Banka (*):"), 1, 0, Qt.AlignLeft)
        self.islem_hesap_cb = QComboBox()
        self.islem_hesap_cb.setEnabled(False)
        form_layout.addWidget(self.islem_hesap_cb, 1, 1)
        self.lbl_vade_tarihi = QLabel("Vade Tarihi:")
        self.entry_vade_tarihi = QLineEdit()
        self.entry_vade_tarihi.setEnabled(False) 
        self.btn_vade_tarihi = QPushButton("🗓️")
        self.btn_vade_tarihi.setFixedWidth(30)
        self.btn_vade_tarihi.clicked.connect(lambda: DatePickerDialog(self.app, self.entry_vade_tarihi))
        self.btn_vade_tarihi.setEnabled(False)
        form_layout.addWidget(self.lbl_vade_tarihi, 2, 0, Qt.AlignLeft)
        form_layout.addWidget(self.entry_vade_tarihi, 2, 1)
        form_layout.addWidget(self.btn_vade_tarihi, 2, 2)
        self.lbl_vade_tarihi.hide()
        self.entry_vade_tarihi.hide()
        self.btn_vade_tarihi.hide()
        form_layout.setColumnStretch(1, 1)
        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)
        btn_onayla = QPushButton("Onayla")
        btn_onayla.clicked.connect(self._onayla)
        button_layout.addWidget(btn_onayla)
        button_layout.addStretch()
        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close)
        button_layout.addWidget(btn_iptal)
        self._yukle_kasa_banka_hesaplarini()
        self._odeme_turu_degisince_hesap_combobox_ayarla()

    def _set_odeme_turu_values(self):
        all_payment_values = [self.db.ODEME_TURU_NAKIT, self.db.ODEME_TURU_KART, 
                              self.db.ODEME_TURU_EFT_HAVALE, self.db.ODEME_TURU_CEK, 
                              self.db.ODEME_TURU_SENET, self.db.ODEME_TURU_ACIK_HESAP, 
                              self.db.ODEME_TURU_ETKISIZ_FATURA]
        is_perakende_musteri = (self.fatura_tipi == 'SATIŞ' and self.initial_cari_id is not None and 
                                str(self.initial_cari_id) == str(self.db.get_perakende_musteri_id()))
        if is_perakende_musteri:
            self.odeme_turu_cb.addItems([p for p in all_payment_values if p != self.db.ODEME_TURU_ACIK_HESAP and p != self.db.ODEME_TURU_ETKISIZ_FATURA])
        else:
            self.odeme_turu_cb.addItems([p for p in all_payment_values if p != self.db.ODEME_TURU_ETKISIZ_FATURA])

    def _yukle_kasa_banka_hesaplarini(self):
        self.islem_hesap_cb.clear()
        self.kasa_banka_map.clear()
        try:
            hesaplar_response = self.db.kasa_banka_listesi_al(limit=10000)
            hesaplar = []
            if isinstance(hesaplar_response, dict) and "items" in hesaplar_response:
                hesaplar = hesaplar_response["items"]
            elif isinstance(hesaplar_response, list):
                hesaplar = hesaplar_response
                self.app.set_status_message("Uyarı: Kasa/Banka listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                hesaplar = []
                self.app.set_status_message("Hata: Kasa/Banka listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Kasa/Banka listesi API'den beklenen formatta gelmedi: {type(hesaplar_response)} - {hesaplar_response}")
                self.islem_hesap_cb.addItem("Hesap Yok", None)
                self.islem_hesap_cb.setEnabled(False)
                return
            if hesaplar:
                for h in hesaplar:
                    display_text = f"{h.get('hesap_adi')} ({h.get('tip')})"
                    if h.get('tip') == "BANKA" and h.get('banka_adi'):
                        display_text += f" - {h.get('banka_adi')}"
                    if h.get('bakiye') is not None:
                        display_text += f" (Bakiye: {self.db._format_currency(h.get('bakiye'))})"
                    self.kasa_banka_map[display_text] = h.get('id')
                    self.islem_hesap_cb.addItem(display_text, h.get('id'))
                self.islem_hesap_cb.setEnabled(True)
                self.islem_hesap_cb.setCurrentIndex(0)
            else:
                self.islem_hesap_cb.addItem("Hesap Yok", None)
                self.islem_hesap_cb.setEnabled(False)
            self.app.set_status_message(f"{len(hesaplar)} kasa/banka hesabı API'den yüklendi.")
        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Kasa/Banka hesapları yüklenirken hata: {e}")
            logging.error(f"Kasa/Banka yükleme hatası: {e}", exc_info=True)
            self.islem_hesap_cb.addItem("Hesap Yok", None)
            self.islem_hesap_cb.setEnabled(False)

    def _odeme_turu_degisince_hesap_combobox_ayarla(self):
        secili_odeme_turu = self.odeme_turu_cb.currentText()
        pesin_odeme_turleri = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"]
        if secili_odeme_turu == "AÇIK HESAP":
            self.lbl_vade_tarihi.show()
            self.entry_vade_tarihi.show()
            self.btn_vade_tarihi.show()
            self.entry_vade_tarihi.setEnabled(True)
            self.btn_vade_tarihi.setEnabled(True)
            if not self.entry_vade_tarihi.text():
                self.entry_vade_tarihi.setText(datetime.now().strftime('%Y-%m-%d'))
        else:
            self.lbl_vade_tarihi.hide()
            self.entry_vade_tarihi.hide()
            self.btn_vade_tarihi.hide()
            self.entry_vade_tarihi.setEnabled(False)
            self.entry_vade_tarihi.clear()
        if secili_odeme_turu in pesin_odeme_turleri:
            self.islem_hesap_cb.setEnabled(True)
            try:
                varsayilan_kb_info = self.db.get_kasa_banka_by_odeme_turu(secili_odeme_turu)
                if varsayilan_kb_info and varsayilan_kb_info[0]:
                    varsayilan_kb_id = varsayilan_kb_info[0]
                    index = self.islem_hesap_cb.findData(varsayilan_kb_id)
                    if index != -1:
                        self.islem_hesap_cb.setCurrentIndex(index)
                    elif self.islem_hesap_cb.count() > 0:
                        self.islem_hesap_cb.setCurrentIndex(0)
                elif self.islem_hesap_cb.count() > 0:
                    self.islem_hesap_cb.setCurrentIndex(0)
                else:
                    self.islem_hesap_cb.clear()
                    self.islem_hesap_cb.addItem("Hesap Yok", None)
                    self.islem_hesap_cb.setEnabled(False)
            except Exception as e:
                logging.warning(f"Varsayılan kasa/banka çekilirken hata: {e}")
                if self.islem_hesap_cb.count() > 0:
                    self.islem_hesap_cb.setCurrentIndex(0)
                else:
                    self.islem_hesap_cb.clear()
                    self.islem_hesap_cb.addItem("Hesap Yok", None)
                    self.islem_hesap_cb.setEnabled(False)
        else:
            self.islem_hesap_cb.clear()
            self.islem_hesap_cb.addItem("Hesap Yok", None)
            self.islem_hesap_cb.setEnabled(False)

    def _onayla(self):
        secili_odeme_turu = self.odeme_turu_cb.currentText()
        secili_hesap_display = self.islem_hesap_cb.currentText()
        vade_tarihi_val = self.entry_vade_tarihi.text().strip()
        kasa_banka_id_val = None
        if secili_hesap_display and secili_hesap_display != "Hesap Yok":
            kasa_banka_id_val = self.kasa_banka_map.get(secili_hesap_display)
        if not secili_odeme_turu:
            QMessageBox.critical(self, "Eksik Bilgi", "Lütfen bir Ödeme Türü seçin.")
            return
        pesin_odeme_turleri = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"]
        if secili_odeme_turu in pesin_odeme_turleri and kasa_banka_id_val is None:
            QMessageBox.critical(self, "Eksik Bilgi", "Peşin ödeme türleri için bir İşlem Kasa/Banka hesabı seçmelisiniz.")
            return
        if secili_odeme_turu == "AÇIK HESAP":
            if not vade_tarihi_val:
                QMessageBox.critical(self, "Eksik Bilgi", "Açık Hesap ödeme türü için Vade Tarihi boş olamaz.")
                return
            try:
                datetime.strptime(vade_tarihi_val, '%Y-%m-%d')
            except ValueError:
                QMessageBox.critical(self, "Tarih Formatı Hatası", "Vade Tarihi formatı (YYYY-AA-GG) olmalıdır.")
                return
        self.callback_func(secili_odeme_turu, kasa_banka_id_val, vade_tarihi_val)
        self.accept()

class TopluVeriEklePenceresi(QDialog):
    def __init__(self, parent_app, db_manager):
        super().__init__(parent_app)
        self.app = parent_app
        self.db = db_manager
        self.setWindowTitle("Toplu Veri Ekleme (Excel)")
        self.setFixedSize(600, 650) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Toplu Veri Ekleme (Excel)")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        main_frame = QFrame(self)
        main_frame_layout = QGridLayout(main_frame)
        main_layout.addWidget(main_frame)

        main_frame_layout.addWidget(QLabel("Veri Tipi:"), 0, 0, Qt.AlignLeft)
        self.veri_tipi_combo = QComboBox()
        self.veri_tipi_combo.addItems(["Müşteri", "Tedarikçi", "Stok/Ürün Ekle/Güncelle"])
        self.veri_tipi_combo.setCurrentText("Müşteri")
        self.veri_tipi_combo.currentIndexChanged.connect(self._show_template_info_and_options)
        main_frame_layout.addWidget(self.veri_tipi_combo, 0, 1)

        main_frame_layout.addWidget(QLabel("Excel Dosyası:"), 1, 0, Qt.AlignLeft)
        self.dosya_yolu_entry = QLineEdit()
        main_frame_layout.addWidget(self.dosya_yolu_entry, 1, 1)
        btn_gozat = QPushButton("Gözat...")
        btn_gozat.clicked.connect(self._gozat_excel_dosyasi)
        main_frame_layout.addWidget(btn_gozat, 1, 2)

        self.stok_guncelleme_options_frame = QGroupBox("Stok/Ürün Güncelleme Seçenekleri", main_frame)
        self.stok_guncelleme_options_layout = QVBoxLayout(self.stok_guncelleme_options_frame)
        main_frame_layout.addWidget(self.stok_guncelleme_options_frame, 2, 0, 1, 3) # Tüm sütunlara yay
        self.stok_guncelleme_options_frame.hide() # Başlangıçta gizli

        self.cb_vars = {} # Boolean değişkenleri için sözlük gibi kullanılacak
        self.cb_vars['fiyat_bilgileri'] = QCheckBox("Fiyat Bilgileri (Alış/Satış/KDV)")
        self.stok_guncelleme_options_layout.addWidget(self.cb_vars['fiyat_bilgileri'])
        self.cb_vars['urun_nitelikleri'] = QCheckBox("Ürün Nitelikleri (Kategori/Marka/Grup/Birim/Menşe/Detay)")
        self.stok_guncelleme_options_layout.addWidget(self.cb_vars['urun_nitelikleri'])
        self.cb_vars['stok_miktari'] = QCheckBox("Stok Miktarı (Mevcut/Minimum)")
        self.stok_guncelleme_options_layout.addWidget(self.cb_vars['stok_miktari'])
        
        self.cb_tumu = QCheckBox("Tümü (Yukarıdakilerin hepsi)")
        self.cb_tumu.stateChanged.connect(self._toggle_all_checkboxes)
        self.stok_guncelleme_options_layout.addWidget(self.cb_tumu)
        
        self.template_info_label = QLabel()
        self.template_info_label.setWordWrap(True)
        self.template_info_label.setAlignment(Qt.AlignLeft)
        main_frame_layout.addWidget(self.template_info_label, 3, 0, 1, 2) # İki sütuna yay

        self.detayli_aciklama_button = QPushButton("Detaylı Bilgi / Şablon Açıklaması")
        self.detayli_aciklama_button.clicked.connect(self._show_detayli_aciklama_penceresi)
        main_frame_layout.addWidget(self.detayli_aciklama_button, 3, 2, Qt.AlignRight | Qt.AlignTop)
        self.detayli_aciklama_button.hide() # Başlangıçta gizli

        main_frame_layout.setColumnStretch(1, 1) # Excel Dosyası entry'sinin genişlemesi için

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        btn_yukle = QPushButton("Verileri Yükle")
        btn_yukle.clicked.connect(self._verileri_yukle)
        button_layout.addWidget(btn_yukle)
        
        btn_sablon_indir = QPushButton("Örnek Şablon İndir")
        btn_sablon_indir.clicked.connect(self._excel_sablonu_indir)
        button_layout.addWidget(btn_sablon_indir)
        
        button_layout.addStretch() # Sağa yaslama
        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close)
        button_layout.addWidget(btn_iptal)

        self.analysis_results = None
        self._show_template_info_and_options() # Başlangıç durumunu ayarla
        self.adjustSize() # Pencere boyutunu içeriğe göre ayarla


    def _show_template_info_and_options(self):
        selected_type = self.veri_tipi_combo.currentText()
        short_info_text = ""
        if selected_type == "Stok/Ürün Ekle/Güncelle":
            self.stok_guncelleme_options_frame.show()
            self.detayli_aciklama_button.show()
        else:
            self.stok_guncelleme_options_frame.hide()
            self.detayli_aciklama_button.hide()
            self.cb_tumu.setChecked(False) # "Tümü" checkbox'ını kaldır
            self._toggle_all_checkboxes(Qt.Unchecked, force_off=True) # Tüm diğer checkbox'ları kapat
            
        if selected_type == "Müşteri":
            short_info_text = "Müşteri Excel dosyası:\n`Müşteri Kodu`, `Ad Soyad` (ZORUNLU) ve diğer detaylar."
        elif selected_type == "Tedarikçi":
            short_info_text = "Tedarikçi Excel dosyası:\n`Tedarikçi Kodu`, `Ad Soyad` (ZORUNLU) ve diğer detaylar."
        elif selected_type == "Stok/Ürün Ekle/Güncelle":
            short_info_text = "Stok/Ürün Excel dosyası:\n`Ürün Kodu`, `Ürün Adı` (ZORUNLU) ve diğer detaylar.\nGüncellemek istediğiniz alanları yukarıdan seçin. Detaylı şablon bilgisi için butona tıklayın."
        self.template_info_label.setText(short_info_text)

    def _excel_sablonu_indir(self):
        veri_tipi = self.veri_tipi_combo.currentText()
        if not veri_tipi: 
            QMessageBox.warning(self, "Uyarı", "Lütfen şablon indirmek için bir veri tipi seçin.")
            return
        
        file_name_prefix, headers = "", []
        if veri_tipi == "Müşteri": file_name_prefix, headers = "Musteri_Sablonu", ["Müşteri Kodu", "Ad Soyad", "Telefon", "Adres", "Vergi Dairesi", "Vergi No"]
        elif veri_tipi == "Tedarikçi": file_name_prefix, headers = "Tedarikci_Sablonu", ["Tedarikçi Kodu", "Ad Soyad", "Telefon", "Adres", "Vergi Dairesi", "Vergi No"]
        elif veri_tipi == "Stok/Ürün Ekle/Güncelle": file_name_prefix, headers = "Stok_Urun_Sablonu", ["Ürün Kodu", "Ürün Adı", "Miktar", "Alış Fiyatı (KDV Dahil)", "Satış Fiyatı (KDV Dahil)", "KDV Oranı (%)", "Minimum Stok Seviyesi", "Kategori Adı", "Marka Adı", "Ürün Grubu Adı", "Ürün Birimi Adı", "Menşe Ülke Adı", "Ürün Detayı", "Ürün Resmi Yolu"]
        else: 
            QMessageBox.critical(self, "Hata", "Geçersiz veri tipi seçimi.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Excel Şablonunu Kaydet", 
                                                    f"{file_name_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx", 
                                                    "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)")
        if file_path:
            try:
                workbook = openpyxl.Workbook(); sheet = workbook.active; sheet.title = "Veri Şablonu"; sheet.append(headers)
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_idx); cell.font = openpyxl.styles.Font(bold=True)
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) + 2, 15)
                workbook.save(file_path)
                QMessageBox.information(self, "Başarılı", f"'{veri_tipi}' şablonu başarıyla oluşturuldu:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Şablon oluşturulurken bir hata oluştu: {e}")

    def _show_detayli_aciklama_penceresi(self):
        selected_type = self.veri_tipi_combo.currentText()
        title = f"{selected_type} Şablon Açıklaması"
        message = ""
        if selected_type == "Müşteri": message = "Müşteri Veri Şablonu Detayları:\n\nExcel dosyasının ilk satırı başlık (header) olmalıdır. Veriler ikinci satırdan başlamalıdır.\n\nSütun Sırası ve Açıklamaları:\n1.  **Müşteri Kodu (ZORUNLU):** Müşterinin benzersiz kodu.\n2.  **Ad Soyad (ZORUNLU):** Müşterinin tam adı veya şirket adı.\n3.  **Telefon (İsteğe Bağlı)**\n4.  **Adres (İsteğe Bağlı)**\n5.  **Vergi Dairesi (İsteğe Bağlı)**\n6.  **Vergi No (İsteğe Bağlı)**"
        elif selected_type == "Tedarikçi": message = "Tedarikçi Veri Şablonu Detayları:\n\nExcel dosyasının ilk satırı başlık (header) olmalıdır. Veriler ikinci satırdan başlamalıdır.\n\nSütun Sırası ve Açıklamaları:\n1.  **Tedarikçi Kodu (ZORUNLU):** Tedarikçinin benzersiz kodu.\n2.  **Ad Soyad (ZORUNLU):** Tedarikçinin tam adı veya şirket adı.\n3.  **Telefon (İsteğe Bağlı)**\n4.  **Adres (İsteğe Bağlı)**\n5.  **Vergi Dairesi (İsteğe Bağlı)**\n6.  **Vergi No (İsteğe Bağlı)**"
        elif selected_type == "Stok/Ürün Ekle/Güncelle": message = "Stok/Ürün Veri Şablonu Detayları:\n\n'Ürün Kodu' eşleşirse güncelleme, eşleşmezse yeni kayıt yapılır.\n\nSütunlar:\n1.  **Ürün Kodu (ZORUNLU)**\n2.  **Ürün Adı (Yeni ürün için ZORUNLU)**\n3.  **Miktar (İsteğe Bağlı):** Pozitif girilirse, mevcut stoğa eklemek için bir 'ALIŞ' faturası oluşturulur.\n4.  **Alış Fiyatı (KDV Dahil) (İsteğe Bağlı)**\n5.  **Satış Fiyatı (KDV Dahil) (İsteğe Bağlı)**\n6.  **KDV Oranı (%) (İsteğe Bağlı)**\n7.  **Minimum Stok Seviyesi (İsteğe Bağlı)**\n8.  **Kategori Adı (İsteğe Bağlı)**\n9.  **Marka Adı (İsteğe Bağlı)**\n10. **Ürün Grubu Adı (İsteğe Bağlı)**\n11. **Ürün Birimi Adı (İsteğe Bağlı)**\n12. **Menşe Ülke Adı (İsteğe Bağlı)**\n13. **Ürün Detayı (İsteğe Bağlı)**\n14. **Ürün Resmi Yolu (İsteğe Bağlı):** Resim dosyasının tam yolu (ör: C:/resimler/urun1.png)."
        from pencereler import AciklamaDetayPenceresi # PySide6 dialog
        AciklamaDetayPenceresi(self, title, message).exec()

    def _gozat_excel_dosyasi(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Excel Dosyası Seç", "", "Excel Dosyaları (*.xlsx;*.xls);;Tüm Dosyalar (*)")
        if file_path:
            self.dosya_yolu_entry.setText(file_path)

    def _toggle_all_checkboxes(self, state, force_off=False):
        # state QCheckBox.Checked (2) veya QCheckBox.Unchecked (0) olabilir
        is_checked = (state == Qt.Checked) if not force_off else False
        for key, checkbox in self.cb_vars.items():
            checkbox.setChecked(is_checked)

    def _verileri_yukle(self):
        dosya_yolu = self.dosya_yolu_entry.text().strip()
        veri_tipi = self.veri_tipi_combo.currentText()
        if not dosya_yolu or not os.path.exists(dosya_yolu):
            QMessageBox.critical(self, "Dosya Hatası", "Lütfen geçerli bir Excel dosyası seçin.")
            return
        
        # Seçili güncelleme alanlarını al
        selected_update_fields = []
        if self.cb_tumu.isChecked():
            # "Tümü" seçiliyse, tüm alt seçenekleri ekle
            selected_update_fields = [key for key in self.cb_vars.keys()]
        else:
            # "Tümü" seçili değilse, tek tek seçili olanları ekle
            selected_update_fields = [key for key, checkbox in self.cb_vars.items() if checkbox.isChecked()]
            
        from pencereler import BeklemePenceresi # PySide6 dialog
        bekleme_penceresi = BeklemePenceresi(self, message="Excel okunuyor ve veriler analiz ediliyor...")
        # PySide6'da QTimer.singleShot ile UI güncellemeleri main thread'de yapılmalı.
        # Threading başlatmadan önce bekleme penceresini göster
        QTimer.singleShot(0, bekleme_penceresi.exec) # Modalı olarak göster

        # Analizi ayrı bir thread'de çalıştır
        threading.Thread(target=self._analiz_et_ve_onizle_threaded, 
                         args=(dosya_yolu, veri_tipi, selected_update_fields, bekleme_penceresi)).start()

    def _analiz_et_ve_onizle_threaded(self, dosya_yolu, veri_tipi, selected_update_fields, bekleme_penceresi):
        analysis_results = {}
        try:
            workbook = openpyxl.load_workbook(dosya_yolu, data_only=True)
            sheet = workbook.active
            
            raw_data_from_excel_list = []
            for row_obj in sheet.iter_rows(min_row=2):
                if any(cell.value is not None and str(cell.value).strip() != '' for cell in row_obj):
                    row_values = [cell.value for cell in row_obj]
                    raw_data_from_excel_list.append(row_values)

            if not raw_data_from_excel_list:
                raise ValueError("Excel dosyasında okunacak geçerli veri bulunamadı.")
            
            from hizmetler import TopluIslemService # TopluIslemService'i import et
            local_db_manager = self.db.__class__(data_dir=self.db.data_dir) # Aynı db örneğini yeniden yarat
            from hizmetler import FaturaService 
            local_fatura_service = FaturaService(local_db_manager)
            local_toplu_islem_service = TopluIslemService(local_db_manager, local_fatura_service)

            if veri_tipi == "Müşteri":
                analysis_results = local_toplu_islem_service.toplu_musteri_analiz_et(raw_data_from_excel_list)
            elif veri_tipi == "Tedarikçi":
                analysis_results = local_toplu_islem_service.toplu_tedarikci_analiz_et(raw_data_from_excel_list)
            elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                analysis_results = local_toplu_islem_service.toplu_stok_analiz_et(raw_data_from_excel_list, selected_update_fields)
            
            # UI güncellemeleri ana thread'e gönderilmeli
            QTimer.singleShot(0, bekleme_penceresi.close) # Bekleme penceresini kapat
            QTimer.singleShot(0, lambda: self._onizleme_penceresini_ac(veri_tipi, analysis_results))

        except Exception as e:
            QTimer.singleShot(0, bekleme_penceresi.close)
            QTimer.singleShot(0, lambda: QMessageBox.critical(self, "Hata", f"Veri analizi başarısız oldu:\n{e}"))
            logging.error(f"Toplu veri analizi thread'inde hata: {e}", exc_info=True)
        finally:
            if 'local_db_manager' in locals() and local_db_manager.conn:
                local_db_manager.conn.close() # Thread'e özgü DB bağlantısını kapat

    def _onizleme_penceresini_ac(self, veri_tipi, analysis_results):
        from pencereler import TopluVeriOnizlemePenceresi
        dialog = TopluVeriOnizlemePenceresi(self.app, self.db, veri_tipi, analysis_results, 
                                            callback_on_confirm=self._gercek_yazma_islemini_yap_threaded_from_onizleme)
        dialog.exec() # Modalı olarak göster

    def _gercek_yazma_islemini_yap_threaded_from_onizleme(self, veri_tipi, analysis_results):
        from pencereler import BeklemePenceresi # PySide6 dialog
        bekleme_penceresi_gercek_islem = BeklemePenceresi(
            self.app, 
            message=f"Toplu {veri_tipi} veritabanına yazılıyor, lütfen bekleyiniz..."
        )
        QTimer.singleShot(0, bekleme_penceresi_gercek_islem.exec)

        threading.Thread(target=self._yazma_islemi_threaded, args=(
            veri_tipi, 
            analysis_results, 
            bekleme_penceresi_gercek_islem
        )).start()

    def _yazma_islemi_threaded(self, veri_tipi, analysis_results, bekleme_penceresi):
        local_db_manager = None
        try:
            from veritabani import OnMuhasebe 
            from hizmetler import FaturaService, TopluIslemService

            local_db_manager = OnMuhasebe(data_dir=self.db.data_dir) # db_name parametresi kaldırıldı, data_dir yeterli
            local_db_manager.app = self.app 

            local_fatura_service = FaturaService(local_db_manager)
            local_toplu_islem_service = TopluIslemService(local_db_manager, local_fatura_service)

            # Transaction'ı burada, bu thread içinde başlat
            local_db_manager.conn.execute("BEGIN TRANSACTION")

            data_to_process = analysis_results.get('all_processed_data', [])
            success, message = False, f"Bilinmeyen veri tipi: {veri_tipi}"
            
            if veri_tipi == "Müşteri":
                success, message = local_toplu_islem_service.toplu_musteri_ekle_guncelle(data_to_process)
            elif veri_tipi == "Tedarikçi":
                success, message = local_toplu_islem_service.toplu_tedarikci_ekle_guncelle(data_to_process)
            elif veri_tipi == "Stok/Ürün Ekle/Güncelle":
                success, message = local_toplu_islem_service.toplu_stok_ekle_guncelle(data_to_process, analysis_results.get('selected_update_fields_from_ui', []))
            
            if success:
                local_db_manager.conn.commit() 
            else:
                local_db_manager.conn.rollback() 

            QTimer.singleShot(0, bekleme_penceresi.close)
            if success:
                QTimer.singleShot(0, lambda: QMessageBox.information(self, "Başarılı", f"Toplu {veri_tipi} işlemi tamamlandı:\n{message}"))
                QTimer.singleShot(0, lambda: self._refresh_related_lists(veri_tipi))
                QTimer.singleShot(0, self.accept) # Pencereyi kapat
            else:
                QTimer.singleShot(0, lambda: QMessageBox.critical(self, "Hata", f"Toplu {veri_tipi} işlemi başarısız oldu:\n{message}"))
        
        except Exception as e:
            if local_db_manager and local_db_manager.conn: 
                local_db_manager.conn.rollback()
            QTimer.singleShot(0, bekleme_penceresi.close)
            QTimer.singleShot(0, lambda: QMessageBox.critical(self, "Kritik Hata", f"Yazma işlemi sırasında beklenmedik bir hata oluştu: {e}"))
            logging.error(f"Toplu yazma işlemi thread'inde hata: {e}", exc_info=True)
        
        finally:
            if local_db_manager and local_db_manager.conn:
                local_db_manager.conn.close()

    def _refresh_related_lists(self, veri_tipi):
        # UI'daki sekme sayfalarını yenileme
        if veri_tipi == "Müşteri" and hasattr(self.app, 'musteri_yonetimi_sayfasi'): 
            self.app.musteri_yonetimi_sayfasi.musteri_listesini_yenile()
        elif veri_tipi == "Tedarikçi" and hasattr(self.app, 'tedarikci_yonetimi_sayfasi'):
            self.app.tedarikci_yonetimi_sayfasi.tedarikci_listesini_yenile()
        elif veri_tipi == "Stok/Ürün Ekle/Güncelle" and hasattr(self.app, 'stok_yonetimi_sayfasi'):
            self.app.stok_yonetimi_sayfasi.stok_listesini_yenile()
        if hasattr(self.app, 'ana_sayfa'):
            self.app.ana_sayfa.guncelle_ozet_bilgiler()

class AciklamaDetayPenceresi(QDialog):
    def __init__(self, parent_app, title="Detaylı Bilgi", message_text=""):
        super().__init__(parent_app)
        self.setWindowTitle(title)
        self.setFixedSize(600, 400) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        # Pencereyi ortalamak için
        self.move(parent_app.pos() + parent_app.rect().center() - self.rect().center())

        main_layout = QVBoxLayout(self)
        self.text_widget = QTextEdit() # tk.Text yerine QTextEdit kullanıldı
        self.text_widget.setPlainText(message_text)
        self.text_widget.setReadOnly(True) # config(state=tk.DISABLED) yerine setReadOnly
        
        main_layout.addWidget(self.text_widget)

        # QScrollArea içinde QTextEdit otomatik kaydırma çubuklarını yönetir, ek scrollbar gerekmez
        # tk.Text'teki vsb kısmı kaldırıldı

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close) # QDialog'u kapat
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignCenter) # Ortala

class CariSecimPenceresi(QDialog):
    def __init__(self, parent_window, db_manager, cari_selection_type, callback_func): # Parametre adı değiştirildi
        super().__init__(parent_window) 
        self.app = parent_window.app # parent_window'un içindeki app referansını al
        self.db = db_manager
        
        # Doğrudan cari tipini sakla (MUSTERI veya TEDARIKCI)
        self.cari_to_select_type = cari_selection_type 
        self.callback_func = callback_func

        self.setWindowTitle("Cari Seçimi")
        self.setFixedSize(600, 450) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Diğer pencerelere tıklamayı engeller

        self.tum_cariler_cache_data = [] 
        self.cari_map_display_to_id = {} 

        # Pencere başlığını doğru cari tipine göre ayarla
        if self.cari_to_select_type == self.db.CARI_TIP_MUSTERI:
            baslik_text = "Müşteri Seçimi"
        elif self.cari_to_select_type == self.db.CARI_TIP_TEDARIKCI:
            baslik_text = "Tedarikçi Seçimi"
        else: 
            baslik_text = "Cari Seçimi (Bilinmeyen Tip)" # Bilinmeyen bir tip gelirse
            logging.warning(f"CariSecimPenceresi bilinmeyen tip ile başlatıldı: {cari_selection_type}")

        title_label = QLabel(baslik_text)
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(title_label)

        # Arama Çerçevesi
        search_frame = QFrame(self)
        search_layout = QHBoxLayout(search_frame)
        main_layout.addWidget(search_frame)

        search_layout.addWidget(QLabel("Ara (Ad/Kod):"), Qt.AlignLeft)
        self.search_entry = QLineEdit()
        self.search_entry.textChanged.connect(self._filtre_liste)
        search_layout.addWidget(self.search_entry)
        search_layout.setStretchFactor(self.search_entry, 1) # Genişlemesi için

        # Cari Listesi Treeview
        tree_frame = QFrame(self)
        tree_layout = QVBoxLayout(tree_frame)
        main_layout.addWidget(tree_frame)

        self.cari_tree = QTreeWidget()
        self.cari_tree.setHeaderLabels(["Cari Adı", "Kodu"])
        self.cari_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.cari_tree.setSortingEnabled(True) # Sıralama özelliği
        
        self.cari_tree.setColumnWidth(0, 300) # Cari Adı sütun genişliği
        self.cari_tree.header().setSectionResizeMode(0, QHeaderView.Stretch) # Cari Adı genişlesin
        self.cari_tree.setColumnWidth(1, 100) # Kodu sütun genişliği
        self.cari_tree.headerItem().setTextAlignment(1, Qt.AlignCenter) # Kodu sütununu ortala

        tree_layout.addWidget(self.cari_tree)
        
        self.cari_tree.itemDoubleClicked.connect(self._sec) # Çift tıklama ile seçim

        # Butonlar
        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        btn_sec = QPushButton("Seç")
        btn_sec.clicked.connect(self._sec)
        button_layout.addWidget(btn_sec)
        
        button_layout.addStretch() # Sağ tarafa yaslamak için boşluk

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

        # Başlangıç yüklemesi
        self._yukle_carileri()
        self.search_entry.setFocus()
    
    def _yukle_carileri(self):
        """Tüm carileri (müşteri veya tedarikçi) API'den çeker ve listeler."""
        self.tum_cariler_cache_data = [] 
        self.cari_map_display_to_id = {} 

        try:
            kod_anahtari_db = ''

            # self.cari_to_select_type'ı doğrudan kullanarak doğru API çağrısını yap
            if self.cari_to_select_type == self.db.CARI_TIP_MUSTERI: 
                cariler_response = self.db.musteri_listesi_al(limit=10000)
                kod_anahtari_db = 'kod' 
            elif self.cari_to_select_type == self.db.CARI_TIP_TEDARIKCI: 
                cariler_response = self.db.tedarikci_listesi_al(limit=10000)
                kod_anahtari_db = 'tedarikci_kodu' # Tedarikçi modelinde 'tedarikci_kodu' alanı varsa
            else: # Beklenmeyen veya geçersiz tip
                self.app.set_status_message("Hata: CariSecimPenceresi için geçersiz tip belirtildi.", "red")
                logging.error(f"CariSecimPenceresi._yukle_carileri: Geçersiz cari_to_select_type: {self.cari_to_select_type}")
                return # Fonksiyonu sonlandır

            cariler = []
            if isinstance(cariler_response, dict) and "items" in cariler_response:
                cariler = cariler_response["items"]
            elif isinstance(cariler_response, list): # Eğer API doğrudan liste dönüyorsa
                cariler = cariler_response
                self.app.set_status_message("Uyarı: Cari listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                self.app.set_status_message("Hata: Cari listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Cari listesi API'den beklenen formatta gelmedi: {type(cariler_response)} - {cariler_response}")
                return # Hata durumunda fonksiyonu sonlandır

            for c in cariler: # c: dict objesi
                cari_id = c.get('id')
                cari_ad = c.get('ad')

                # Cari kodunu almak için dinamik anahtar kullan
                cari_kodu = c.get(kod_anahtari_db, "") 

                display_text = f"{cari_ad} (Kod: {cari_kodu})" 
                self.cari_map_display_to_id[display_text] = str(cari_id) 
                self.tum_cariler_cache_data.append(c) 

            self._filtre_liste() 

            # Varsayılan seçimi yap
            default_id_str = None
            if self.cari_to_select_type == self.db.CARI_TIP_MUSTERI and self.db.get_perakende_musteri_id() is not None:
                default_id_str = str(self.db.get_perakende_musteri_id())
            elif self.cari_to_select_type == self.db.CARI_TIP_TEDARIKCI and self.db.get_genel_tedarikci_id() is not None:
                default_id_str = str(self.db.get_genel_tedarikci_id())

            if default_id_str:
                # Treeview'deki item'ı ID'sine göre bul ve seç
                for i in range(self.cari_tree.topLevelItemCount()):
                    item = self.cari_tree.topLevelItem(i)
                    if item.data(0, Qt.UserRole) == int(default_id_str): # UserRole'a kaydettiğimiz ID ile karşılaştır
                        item.setSelected(True)
                        self.cari_tree.scrollToItem(item)
                        break

        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Cari listesi çekilirken hata: {e}")
            logging.error(f"Cari listesi yükleme hatası: {e}", exc_info=True)
            
    def _filtre_liste(self): # event parametresi kaldırıldı
        """Arama terimine göre cari listesini filtreler."""
        arama_terimi = self.search_entry.text().lower().strip()
        normalized_arama_terimi = normalize_turkish_chars(arama_terimi) 

        self.cari_tree.clear()

        for cari_row in self.tum_cariler_cache_data: # cari_row: dict objesi
            cari_id = cari_row.get('id')
            cari_ad = cari_row.get('ad')
            
            cari_kodu = ""
            # self.cari_to_select_type'a göre doğru kod alanını kullan
            if self.cari_to_select_type == self.db.CARI_TIP_MUSTERI: 
                cari_kodu = cari_row.get('kod', '')
            elif self.cari_to_select_type == self.db.CARI_TIP_TEDARIKCI: 
                cari_kodu = cari_row.get('tedarikci_kodu', '')
            
            normalized_cari_ad = normalize_turkish_chars(cari_ad) if cari_ad else ''
            normalized_cari_kodu = normalize_turkish_chars(cari_kodu) if cari_kodu else ''

            if (not normalized_arama_terimi or
                (normalized_cari_ad and normalized_arama_terimi in normalized_cari_ad) or
                (normalized_cari_kodu and normalized_arama_terimi in normalized_cari_kodu)
               ):
                item_qt = QTreeWidgetItem(self.cari_tree)
                item_qt.setText(0, cari_ad)
                item_qt.setText(1, cari_kodu)
                item_qt.setData(0, Qt.UserRole, cari_id) # ID'yi UserRole olarak sakla

    def _sec(self, item=None, column=None): # item ve column QTreeWidget sinyalinden gelir
        selected_items = self.cari_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Seçim Yok", "Lütfen bir cari seçin.")
            return

        selected_cari_id = selected_items[0].data(0, Qt.UserRole) # UserRole'dan ID'yi al
        selected_cari_display_text = selected_items[0].text(0) # Cari Adı sütunu
        
        # YENİ EKLENEN SATIR: Sinyali kopararak tekrar açılmayı engelle
        self.cari_tree.itemDoubleClicked.disconnect(self._sec) 

        self.callback_func(selected_cari_id, selected_cari_display_text) # Callback'i çağır
        self.accept() # QDialog'u kapat

class TedarikciSecimDialog(QDialog):
    def __init__(self, parent_window, db_manager, callback_func):
        super().__init__(parent_window) 
        self.app = parent_window.app # parent_window'un içindeki app referansını al
        self.db = db_manager
        self.callback_func = callback_func

        self.setWindowTitle("Tedarikçi Seçimi")
        self.setFixedSize(600, 400) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        self.tum_tedarikciler_cache = [] # Data dict'lerini saklar

        main_layout = QVBoxLayout(self)
        title_label = QLabel("Tedarikçi Seçimi")
        title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # Arama Çerçevesi
        search_frame = QFrame(self)
        search_layout = QHBoxLayout(search_frame)
        main_layout.addWidget(search_frame)

        search_layout.addWidget(QLabel("Ara (Ad/Kod):"), Qt.AlignLeft)
        self.search_entry = QLineEdit()
        self.search_entry.textChanged.connect(self._filtre_liste)
        search_layout.addWidget(self.search_entry)
        search_layout.setStretchFactor(self.search_entry, 1) # Genişlemesi için

        # Tedarikçi Listesi Treeview
        tree_frame = QFrame(self)
        tree_layout = QVBoxLayout(tree_frame)
        main_layout.addWidget(tree_frame)

        self.tedarikci_tree = QTreeWidget()
        self.tedarikci_tree.setHeaderLabels(["Tedarikçi Adı", "Kodu"])
        self.tedarikci_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tedarikci_tree.setSortingEnabled(True)

        self.tedarikci_tree.setColumnWidth(0, 300) # Tedarikçi Adı sütun genişliği
        self.tedarikci_tree.header().setSectionResizeMode(0, QHeaderView.Stretch) # Tedarikçi Adı genişlesin
        self.tedarikci_tree.setColumnWidth(1, 100) # Kodu sütun genişliği
        self.tedarikci_tree.headerItem().setTextAlignment(1, Qt.AlignCenter) # Kodu sütununu ortala

        tree_layout.addWidget(self.tedarikci_tree)
        
        self.tedarikci_tree.itemDoubleClicked.connect(self._sec) # Çift tıklama ile seçim

        # Butonlar
        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame)

        btn_sec = QPushButton("Seç")
        btn_sec.clicked.connect(self._sec)
        button_layout.addWidget(btn_sec)
        
        button_layout.addStretch() # Sağ tarafa yaslamak için boşluk

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

        # Başlangıç yüklemesi
        self._yukle_tedarikcileri()
        self.search_entry.setFocus()
    
    def _yukle_tedarikcileri(self):
        """Tüm tedarikçileri API'den çeker ve listeler."""
        self.tum_tedarikciler_cache = [] 

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            tedarikciler_response = self.db.tedarikci_listesi_al(limit=10000)

            tedarikciler = []
            if isinstance(tedarikciler_response, dict) and "items" in tedarikciler_response:
                tedarikciler = tedarikciler_response["items"]
            elif isinstance(tedarikciler_response, list): # Eğer API doğrudan liste dönüyorsa
                tedarikciler = tedarikciler_response
                self.app.set_status_message("Uyarı: Tedarikçi listesi API yanıtı beklenen formatta değil. Doğrudan liste olarak işleniyor.", "orange")
            else:
                self.app.set_status_message("Hata: Tedarikçi listesi API'den alınamadı veya formatı geçersiz.", "red")
                logging.error(f"Tedarikçi listesi API'den beklenen formatta gelmedi: {type(tedarikciler_response)} - {tedarikciler_response}")
                return # Hata durumunda fonksiyonu sonlandır

            self.tum_tedarikciler_cache = tedarikciler
            self._filtre_liste() 

        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Tedarikçi listesi çekilirken hata: {e}")
            logging.error(f"Tedarikçi listesi yükleme hatası: {e}", exc_info=True)

    def _filtre_liste(self): # event parametresi kaldırıldı
        """Arama kutusuna yazıldıkça tedarikçi listesini filtreler."""
        # Arama terimini al ve normalleştir
        arama_terimi = self.search_entry.text().lower().strip()
        normalized_arama_terimi = normalize_turkish_chars(arama_terimi) 
        
        # Treeview'i temizle
        self.tedarikci_tree.clear()
        
        # Önbelleğe alınmış tedarikçi verileri üzerinde döngü.
        for tedarikci_row in self.tum_tedarikciler_cache: # tedarikci_row: dict objesi
            tedarikci_id = tedarikci_row.get('id')
            tedarikci_kodu = tedarikci_row.get('tedarikci_kodu', '')
            tedarikci_ad = tedarikci_row.get('ad')
            
            # Tedarikçi adını ve kodunu normalleştirerek karşılaştırma yapalım.
            normalized_tedarikci_ad = normalize_turkish_chars(tedarikci_ad) if tedarikci_ad else ''
            normalized_tedarikci_kodu = normalize_turkish_chars(tedarikci_kodu) if tedarikci_kodu else ''
            
            # Filtreleme koşulu
            if (not normalized_arama_terimi or
                (normalized_tedarikci_ad and normalized_arama_terimi in normalized_tedarikci_ad) or
                (normalized_tedarikci_kodu and normalized_arama_terimi in normalized_tedarikci_kodu)
               ):
                item_qt = QTreeWidgetItem(self.tedarikci_tree)
                item_qt.setText(0, tedarikci_ad)
                item_qt.setText(1, tedarikci_kodu)
                item_qt.setData(0, Qt.UserRole, tedarikci_id) # ID'yi UserRole olarak sakla

    def _sec(self, item=None, column=None): # item ve column QTreeWidget sinyalinden gelir
        """Seçili tedarikçiyi onaylar ve callback fonksiyonunu çağırır."""
        selected_items = self.tedarikci_tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Seçim Yok", "Lütfen bir tedarikçi seçin.")
            return

        selected_tedarikci_id = selected_items[0].data(0, Qt.UserRole) # UserRole'dan ID'yi al
        selected_tedarikci_ad = selected_items[0].text(0) # Tedarikçi Adı sütunu
        
        self.callback_func(selected_tedarikci_id, selected_tedarikci_ad) # Callback'i çağır
        self.accept() # Pencereyi kapat  

class BeklemePenceresi(QDialog):
    def __init__(self, parent_app, title="İşlem Devam Ediyor...", message="Lütfen bekleyiniz..."):
        super().__init__(parent_app)
        self.setWindowTitle(title)
        self.setFixedSize(300, 120) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla ve diğer etkileşimleri engelle

        # Pencereyi ana pencerenin ortasına konumlandır
        if parent_app:
            parent_rect = parent_app.geometry()
            x = parent_rect.x() + (parent_rect.width() - self.width()) // 2
            y = parent_rect.y() + (parent_rect.height() - self.height()) // 2
            self.move(x, y)

        main_layout = QVBoxLayout(self)
        message_label = QLabel(message)
        message_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
        message_label.setWordWrap(True)
        message_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(message_label)
        
        self.progressbar = QProgressBar() # ttk.Progressbar yerine QProgressBar kullanıldı
        self.progressbar.setRange(0, 0) # Belirsiz (indeterminate) mod için
        main_layout.addWidget(self.progressbar, alignment=Qt.AlignCenter)
        
        self.setWindowFlags(Qt.FramelessWindowHint) # Çerçevesiz pencere
        self.setAttribute(Qt.WA_DeleteOnClose) # Kapatıldığında otomatik sil
        
        # Pencere kapatma olayını engelle (kullanıcının kapatmasını önle)
        self.setWindowModality(Qt.ApplicationModal)
        self.closeEvent = self._do_nothing_close_event

    def _do_nothing_close_event(self, event):
        # Kullanıcının pencereyi kapatmasını engelle
        event.ignore()

    def kapat(self):
        self.close() # QDialog'u kapat
        
class GelirGiderSiniflandirmaYonetimiPenceresi(QDialog):
    def __init__(self, parent_app, db_manager, yenile_callback):
        super().__init__(parent_app)
        self.db = db_manager
        self.app = parent_app
        self.yenile_callback = yenile_callback # Ana pencereyi yenilemek için

        self.setWindowTitle("Gelir/Gider Sınıflandırma Yönetimi")
        self.setMinimumSize(600, 450)
        self.setModal(True)

        main_layout = QVBoxLayout(self)

        # Notebook (Sekmeler) oluştur
        self.notebook = QTabWidget(self)
        main_layout.addWidget(self.notebook)

        # Gelir Sınıflandırmaları Sekmesi
        self.gelir_frame = QWidget()
        self.notebook.addTab(self.gelir_frame, "Gelir Sınıflandırmaları")
        self._setup_siniflandirma_sekmesi(self.gelir_frame, "GELİR")

        # Gider Sınıflandırmaları Sekmesi
        self.gider_frame = QWidget()
        self.notebook.addTab(self.gider_frame, "Gider Sınıflandırmaları")
        self._setup_siniflandirma_sekmesi(self.gider_frame, "GİDER")

        btn_kapat = QPushButton("Kapat")
        btn_kapat.clicked.connect(self.close)
        main_layout.addWidget(btn_kapat, alignment=Qt.AlignRight)

        # Sağ tık menüsü (Ortak olabilir)
        self.context_menu = QMenu(self)
        self.context_menu.addAction("Güncelle").triggered.connect(self._siniflandirma_guncelle)
        self.context_menu.addAction("Sil").triggered.connect(self._siniflandirma_sil)

    def _setup_siniflandirma_sekmesi(self, parent_frame, tip):
        frame_layout = QVBoxLayout(parent_frame) # Çerçeveye bir layout ata

        # Arama ve Ekleme alanı
        top_frame = QFrame(parent_frame)
        top_layout = QHBoxLayout(top_frame)
        frame_layout.addWidget(top_frame)

        top_layout.addWidget(QLabel("Sınıflandırma Adı:")) # "Yeni Sınıflandırma Adı:" yerine "Sınıflandırma Adı:"
        entry = QLineEdit()
        top_layout.addWidget(entry)

        add_button = QPushButton("Ekle")
        add_button.clicked.connect(lambda: self._siniflandirma_ekle(tip, entry.text().strip(), entry))
        top_layout.addWidget(add_button)

        # Yeni Eklendi: Güncelle butonu ve fonksiyonu
        update_button = QPushButton("Güncelle")
        update_button.clicked.connect(lambda: self._siniflandirma_guncelle_dogrudan(tip, entry.text().strip(), entry)) # Yeni doğrudan güncelleme metodu
        top_layout.addWidget(update_button)

        # Yeni Eklendi: Sil butonu ve fonksiyonu
        delete_button = QPushButton("Sil")
        delete_button.clicked.connect(lambda: self._siniflandirma_sil(tip)) # Sil metodu doğrudan çağrıldı
        top_layout.addWidget(delete_button)

        # Treeview alanı
        tree_frame = QFrame(parent_frame)
        tree_layout = QVBoxLayout(tree_frame)
        frame_layout.addWidget(tree_frame)

        tree = QTreeWidget()
        tree.setHeaderLabels(["ID", "Sınıflandırma Adı"])
        tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        tree.setColumnWidth(0, 50)
        tree.header().setSectionResizeMode(1, QHeaderView.Stretch) # Sınıflandırma Adı genişlesin
        tree_layout.addWidget(tree)

        # Treeview'i kaydet
        if tip == "GELİR":
            self.gelir_tree = tree
            self.gelir_entry = entry # Gelir giriş alanını sınıf özelliği olarak kaydet
        else:
            self.gider_tree = tree
            self.gider_entry = entry # Gider giriş alanını sınıf özelliği olarak kaydet

        # Seçim değiştiğinde giriş alanını doldurmak için bağlantı
        tree.itemSelectionChanged.connect(lambda: self._on_siniflandirma_select(tree, entry))

        # Sağ tık menüsünü treeview'e bağla
        tree.setContextMenuPolicy(Qt.CustomContextMenu)
        tree.customContextMenuRequested.connect(self._on_treeview_right_click)

        self._load_siniflandirmalar(tip)

    def _load_siniflandirmalar(self, tip):
        tree = self.gelir_tree if tip == "GELİR" else self.gider_tree

        tree.clear() # Mevcut öğeleri temizle

        siniflandirmalar = []
        try:
            if tip == "GELİR":
                # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
                siniflandirmalar = self.db.gelir_siniflandirma_listele()
            else:
                # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
                siniflandirmalar = self.db.gider_siniflandirma_listele()
        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"{tip} sınıflandırmaları çekilirken hata: {e}")
            logging.error(f"{tip} sınıflandırma yükleme hatası: {e}", exc_info=True)
            return

        for s_item in siniflandirmalar:
            item_qt = QTreeWidgetItem(tree)
            item_qt.setText(0, str(s_item.get('id')))
            item_qt.setText(1, s_item.get('ad')) # 'siniflandirma_adi' yerine 'ad' kullanıldı
            item_qt.setData(0, Qt.UserRole, s_item.get('id')) # ID'yi UserRole olarak sakla

    def _siniflandirma_ekle(self, tip, siniflandirma_adi, entry_widget):
        if not siniflandirma_adi:
            QMessageBox.warning(self, "Uyarı", "Sınıflandırma adı boş olamaz.")
            return
        try:
            success, message = self.db.nitelik_ekle(f"{tip.lower()}_siniflandirmalari", {"ad": siniflandirma_adi})
            if success:
                QMessageBox.information(self, "Başarılı", "Sınıflandırma başarıyla eklendi.")
                entry_widget.clear()
                self._load_siniflandirmalar(tip)
                if self.yenile_callback:
                    self.yenile_callback()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Sınıflandırma eklenirken hata: {e}")
            logging.error(f"Sınıflandırma ekleme hatası: {e}", exc_info=True)

    def _on_siniflandirma_select(self, tree, entry_widget):
        selected_items = tree.selectedItems()
        if selected_items:
            # Sınıflandırma adını al ve giriş alanına yerleştir
            siniflandirma_adi = selected_items[0].text(1) 
            entry_widget.setText(siniflandirma_adi)
        else:
            entry_widget.clear()

    def _on_treeview_right_click(self, pos):
        current_tab_text = self.notebook.tabText(self.notebook.currentIndex())

        tree = None
        entry_widget = None # Güncelleme için giriş alanını alacağız
        if "Gelir Sınıflandırmaları" in current_tab_text:
            tree = self.gelir_tree
            entry_widget = self.gelir_entry
        else:
            tree = self.gider_tree
            entry_widget = self.gider_entry

        item = tree.itemAt(pos) # Position'dan öğeyi al

        if item:
            tree.setCurrentItem(item) # Öğeyi seçili hale getir (sağ tıklama ile seçilmemiş olabilir)

            context_menu = QMenu(self)

            # Düzeltildi: _siniflandirma_guncelle metodunu doğrudan çağırıldı
            update_action = context_menu.addAction("Güncelle")
            update_action.triggered.connect(self._siniflandirma_guncelle)

            # Silme işlemi
            delete_action = context_menu.addAction("Sil")
            delete_action.triggered.connect(lambda: self._siniflandirma_sil(
                "GELİR" if "Gelir" in current_tab_text else "GİDER"
            ))

            context_menu.exec(tree.mapToGlobal(pos))
        else:
            # Boş alana tıklandığında menüyü gizle/kapat (eğer açıksa)
            if hasattr(self, 'context_menu') and self.context_menu.isVisible():
                self.context_menu.hide()

    def _siniflandirma_guncelle(self):
        current_tab_text = self.notebook.tabText(self.notebook.currentIndex())
        tree = self.gelir_tree if "Gelir Sınıflandırmaları" in current_tab_text else self.gider_tree
        entry_widget = self.gelir_entry if "Gelir Sınıflandırmaları" in current_tab_text else self.gider_entry
        tip = "GELİR" if "Gelir" in current_tab_text else "GİDER"

        selected_items = tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen güncellemek istediğiniz sınıflandırmayı seçin.")
            return

        siniflandirma_id = selected_items[0].data(0, Qt.UserRole)
        mevcut_siniflandirma_adi = selected_items[0].text(1)
        yeni_siniflandirma_adi = entry_widget.text().strip()

        if not yeni_siniflandirma_adi or yeni_siniflandirma_adi == mevcut_siniflandirma_adi:
            QMessageBox.warning(self, "Uyarı", "Yeni sınıflandırma adı boş olamaz veya mevcut adla aynı olamaz.")
            return
        try:
            success, message = self.db.nitelik_guncelle(f"{tip.lower()}_siniflandirmalari", siniflandirma_id, {"ad": yeni_siniflandirma_adi})
            if success:
                QMessageBox.information(self, "Başarılı", "Sınıflandırma başarıyla güncellendi.")
                entry_widget.clear()
                self._load_siniflandirmalar(tip)
                if self.yenile_callback:
                    self.yenile_callback()
            else:
                QMessageBox.critical(self, "Hata", message)
        except Exception as e:
            QMessageBox.critical(self, "API Hatası", f"Sınıflandırma güncellenirken hata: {e}")
            logging.error(f"Sınıflandırma güncellenirken hata: {e}", exc_info=True)

    def _siniflandirma_sil(self):
        current_tab_text = self.notebook.tabText(self.notebook.currentIndex())
        
        tree = None
        tip = ""
        if "Gelir Sınıflandırmaları" in current_tab_text:
            tree = self.gelir_tree
            tip = "GELİR"
        else:
            tree = self.gider_tree
            tip = "GİDER"

        selected_items = tree.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek istediğiniz sınıflandırmayı seçin.")
            return
        siniflandirma_id = selected_items[0].data(0, Qt.UserRole)
        siniflandirma_adi = selected_items[0].text(1)
        reply = QMessageBox.question(self, "Onay", f"'{siniflandirma_adi}' sınıflandırmasını silmek istediğinizden emin misiniz?", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                success, message = self.db.nitelik_sil(f"{tip.lower()}_siniflandirmalari", siniflandirma_id)
                if success:
                    QMessageBox.information(self, "Başarılı", "Sınıflandırma başarıyla silindi.")
                    self._load_siniflandirmalar(tip)
                    if self.yenile_callback:
                        self.yenile_callback()
                else:
                    QMessageBox.critical(self, "Hata", message)
            except Exception as e:
                QMessageBox.critical(self, "API Hatası", f"Sınıflandırma silinirken hata: {e}")
                logging.error(f"Sınıflandırma silme hatası: {e}", exc_info=True)

class BirimDuzenlePenceresi(QDialog):
    def __init__(self, parent_window, db_manager, birim_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.birim_id = birim_info['id']
        self.mevcut_birim_adi = birim_info['birim_adi']
        self.yenile_callback = yenile_callback

        self.setWindowTitle(f"Birim Düzenle: {self.mevcut_birim_adi}")
        self.setFixedSize(350, 200) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame)

        main_frame_layout.addWidget(QLabel("Birim Adı:"), 0, 0, Qt.AlignLeft)
        self.birim_adi_entry = QLineEdit()
        self.birim_adi_entry.setText(self.mevcut_birim_adi)
        main_frame_layout.addWidget(self.birim_adi_entry, 0, 1)
        main_frame_layout.setColumnStretch(1, 1) # Genişlesin

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignRight) # Butonları sağa yasla

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

    def _kaydet(self):
        yeni_birim_adi = self.birim_adi_entry.text().strip()
        if not yeni_birim_adi:
            QMessageBox.warning(self, "Uyarı", "Birim adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            response = self.db.nitelik_guncelle("urun_birimleri", self.birim_id, {"ad": yeni_birim_adi})

            QMessageBox.information(self, "Başarılı", "Birim başarıyla güncellendi.")
            self.yenile_callback() # Ana listedeki birimleri yenile
            self.accept() # Pencereyi kapat
        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Birim güncellenirken hata: {e}")
            logging.error(f"Birim güncelleme hatası: {e}", exc_info=True)

class GrupDuzenlePenceresi(QDialog):
    def __init__(self, parent_window, db_manager, grup_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.grup_id = grup_info['id']
        self.mevcut_grup_adi = grup_info['grup_adi']
        self.yenile_callback = yenile_callback

        self.setWindowTitle(f"Grup Düzenle: {self.mevcut_grup_adi}")
        self.setFixedSize(350, 200) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame)

        main_frame_layout.addWidget(QLabel("Grup Adı:"), 0, 0, Qt.AlignLeft)
        self.grup_adi_entry = QLineEdit()
        self.grup_adi_entry.setText(self.mevcut_grup_adi)
        main_frame_layout.addWidget(self.grup_adi_entry, 0, 1)
        main_frame_layout.setColumnStretch(1, 1) # Genişlesin

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignRight) # Butonları sağa yasla

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

    def _kaydet(self):
        yeni_grup_adi = self.grup_adi_entry.text().strip()
        if not yeni_grup_adi:
            QMessageBox.warning(self, "Uyarı", "Grup adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            response = self.db.nitelik_guncelle("urun_gruplari", self.grup_id, {"ad": yeni_grup_adi})

            QMessageBox.information(self, "Başarılı", "Grup başarıyla güncellendi.")
            self.yenile_callback()
            self.accept()
        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Grup güncellenirken hata: {e}")
            logging.error(f"Grup güncelleme hatası: {e}", exc_info=True)

# UlkeDuzenlePenceresi sınıfı
class UlkeDuzenlePenceresi(QDialog):
    def __init__(self, parent_window, db_manager, ulke_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.ulke_id = ulke_info['id']
        self.mevcut_ulke_adi = ulke_info['ulke_adi']
        self.yenile_callback = yenile_callback

        self.setWindowTitle(f"Ülke Düzenle: {self.mevcut_ulke_adi}")
        self.setFixedSize(350, 200) # geometry yerine setFixedSize kullanıldı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame)

        main_frame_layout.addWidget(QLabel("Ülke Adı:"), 0, 0, Qt.AlignLeft)
        self.ulke_adi_entry = QLineEdit()
        self.ulke_adi_entry.setText(self.mevcut_ulke_adi)
        main_frame_layout.addWidget(self.ulke_adi_entry, 0, 1)
        main_frame_layout.setColumnStretch(1, 1) # Genişlesin

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignRight) # Butonları sağa yasla

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

    def _kaydet(self):
        yeni_ulke_adi = self.ulke_adi_entry.text().strip()
        if not yeni_ulke_adi:
            QMessageBox.warning(self, "Uyarı", "Ülke adı boş olamaz.")
            return

        try:
            # Düzeltildi: Doğrudan requests yerine db_manager metodu kullanıldı
            response = self.db.nitelik_guncelle("ulkeler", self.ulke_id, {"ad": yeni_ulke_adi})

            QMessageBox.information(self, "Başarılı", "Ülke başarıyla güncellendi.")
            self.yenile_callback()
            self.accept()
        except Exception as e: # Düzeltildi: requests.exceptions.RequestException yerine daha genel hata yakalandı
            QMessageBox.critical(self, "API Hatası", f"Ülke güncellenirken hata: {e}")
            logging.error(f"Ülke güncelleme hatası: {e}", exc_info=True)

class SiniflandirmaDuzenlePenceresi(QDialog):
    def __init__(self, parent_window, db_manager, tip, siniflandirma_info, yenile_callback):
        super().__init__(parent_window)
        self.db = db_manager
        self.parent_window = parent_window
        self.tip = tip # "GELİR" veya "GİDER"
        self.siniflandirma_id = siniflandirma_info['id']
        self.mevcut_siniflandirma_adi = siniflandirma_info['siniflandirma_adi']
        self.yenile_callback = yenile_callback

        self.setWindowTitle(f"{tip.capitalize()} Sınıflandırma Düzenle: {self.mevcut_siniflandirma_adi}")
        self.setFixedSize(400, 220) # Boyut ayarı
        self.setModal(True) # Modalı olarak ayarla

        main_layout = QVBoxLayout(self)
        main_frame = QFrame(self)
        main_layout.addWidget(main_frame)
        main_frame_layout = QGridLayout(main_frame)

        main_frame_layout.addWidget(QLabel("Sınıflandırma Adı:"), 0, 0, Qt.AlignLeft)
        self.siniflandirma_adi_entry = QLineEdit()
        self.siniflandirma_adi_entry.setText(self.mevcut_siniflandirma_adi)
        main_frame_layout.addWidget(self.siniflandirma_adi_entry, 0, 1)
        main_frame_layout.setColumnStretch(1, 1) # Genişlesin

        button_frame = QFrame(self)
        button_layout = QHBoxLayout(button_frame)
        main_layout.addWidget(button_frame, alignment=Qt.AlignRight) # Butonları sağa yasla

        btn_kaydet = QPushButton("Kaydet")
        btn_kaydet.clicked.connect(self._kaydet)
        button_layout.addWidget(btn_kaydet)

        btn_iptal = QPushButton("İptal")
        btn_iptal.clicked.connect(self.close) # QDialog'u kapat
        button_layout.addWidget(btn_iptal)

    def _kaydet(self):
        yeni_siniflandirma_adi = self.siniflandirma_adi_entry.text().strip()
        if not yeni_siniflandirma_adi:
            QMessageBox.warning(self, "Uyarı", "Sınıflandırma adı boş olamaz.")
            return

        try:
            # API endpointleri kullanıldı
            if self.tip == "GELİR":
                response = requests.put(f"{API_BASE_URL}/nitelikler/gelir_siniflandirmalari/{self.siniflandirma_id}", json={"siniflandirma_adi": yeni_siniflandirma_adi})
            else: # GİDER
                response = requests.put(f"{API_BASE_URL}/nitelikler/gider_siniflandirmalari/{self.siniflandirma_id}", json={"siniflandirma_adi": yeni_siniflandirma_adi})
            response.raise_for_status()

            QMessageBox.information(self, "Başarılı", "Sınıflandırma başarıyla güncellendi.")
            self.yenile_callback() # Ana listedeki sınıflandırmaları yenile
            self.accept() # Pencereyi kapat
        except requests.exceptions.RequestException as e:
            error_detail = str(e)
            if e.response is not None:
                try: error_detail = e.response.json().get('detail', str(e.response.content))
                except ValueError: pass
            QMessageBox.critical(self, "API Hatası", f"Sınıflandırma güncellenirken hata: {error_detail}")
            logging.error(f"Sınıflandırma güncelleme hatası: {error_detail}", exc_info=True)