# veritabani.py dosyasının içeriği
from sqlalchemy import create_engine, Column, Integer, String, Numeric, TIMESTAMP, Date, Boolean, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship
from sqlalchemy import func, case, and_, or_ 
from datetime import datetime, date, timedelta
import os
import shutil
import hashlib
import locale
import logging
import traceback
import openpyxl
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate, Spacer
from reportlab.lib import colors
from reportlab.pdfgen import canvas as rp_canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics, ttfonts
import sqlite3
# --- GLOBAL VERİTABANI AYARLARI (UI tarafından doğrudan kullanılacaksa) ---
# Bu ayarlar, sadece UI'dan OnMuhasebe objesi direkt başlatıldığında kullanılır.
# API tarafında veritabani.py'nin ayrı bir engine ve SessionLocal'ı vardır.

# `main.py` dosyasının bulunduğu dizini bulalım ve 'data' klasörünü varsayalım.
# Eğer 'data' klasörü `main.py` ile aynı dizinde değilse bu yolu düzeltmelisiniz.
_base_dir = os.path.dirname(os.path.abspath(__file__)) # veritabani.py'nin bulunduğu dizin
_project_root_dir = os.path.abspath(os.path.join(_base_dir, os.pardir)) # Projenin ana kök dizini (onmuhasebe)
_data_dir = os.path.join(_project_root_dir, 'data') # 'onmuhasebe/data'
if not os.path.exists(_data_dir):
    os.makedirs(_data_dir)

# Geçici olarak SQLite Engine'i. Eğer UI tarafı da PostgreSQL kullanacaksa bu kısım değiştirilmeli.
# Şu anki hata, OnMuhasebe'nin doğru Engine'i bulamamasından kaynaklandığı için burayı sağlıyoruz.
_SQL_URL = f"sqlite:///{os.path.join(_data_dir, 'on_muhasebe.db')}" 
_engine = create_engine(_SQL_URL, connect_args={"check_same_thread": False})
_SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=_engine)

# --- PDF FONT AYARLARI ---
# Font dosyalarının proje kök dizinindeki 'fonts' klasöründe olduğunu varsayalım.
# Eğer bu klasör yoksa veya fontlar farklı bir yerde ise, yolu güncelleyin.
_font_dir = os.path.join(_project_root_dir, 'fonts')
if not os.path.exists(_font_dir):
    os.makedirs(_font_dir) # fonts klasörünü oluştur

TURKISH_FONT_NORMAL = "DejaVuSans" 
TURKISH_FONT_BOLD = "DejaVuSans-Bold"

try:
    _dejavu_sans_ttf = os.path.join(_font_dir, 'DejaVuSans.ttf')
    _dejavu_sans_bold_ttf = os.path.join(_font_dir, 'DejaVuSans-Bold.ttf')

    if not os.path.exists(_dejavu_sans_ttf):
        logging.warning(f"Font dosyası bulunamadı: {_dejavu_sans_ttf}. PDF'lerde Türkçe karakter sorunları yaşanabilir.")
    else:
        pdfmetrics.registerFont(ttfonts.TTFont(TURKISH_FONT_NORMAL, _dejavu_sans_ttf))

    if not os.path.exists(_dejavu_sans_bold_ttf):
        logging.warning(f"Font dosyası bulunamadı: {_dejavu_sans_bold_ttf}. PDF'lerde Türkçe karakter sorunları yaşanabilir.")
    else:
        pdfmetrics.registerFont(ttfonts.TTFont(TURKISH_FONT_BOLD, _dejavu_sans_bold_ttf))

except Exception as e:
    logging.warning(f"PDF fontları yüklenirken hata oluştu: {e}. PDF'lerde Türkçe karakter sorunları yaşanabilir.")


# --- SQLAlchemy ORM Temeli ---
Base = declarative_base()

# --- SQLAlchemy ORM Modellerinin Import Edilmesi ---
# semalar.py'deki TÜM ORM MODELLERİ BURADAN IMPORT EDİLECEK.
# Bu, modelleri tekrar kopyalamak yerine daha temiz bir yaklaşımdır.
# Ancak semalar.py'nin de doğru relationship tanımlamalarına sahip olduğundan emin olun.
from api.semalar import (
    Kullanici, SirketBilgileri, Musteri, Tedarikci, UrunKategorileri, 
    UrunMarkalari, UrunGruplari, UrunBirimleri, UrunUlkeleri, Stok, StokHareketleri, 
    KasaBanka, Fatura, FaturaKalemleri, Siparis, SiparisKalemleri, 
    GelirGider, GelirSiniflandirma, GiderSiniflandirma, CariHareketler
)
class OnMuhasebe:
    PERAKENDE_MUSTERI_KODU = "M-000"
    GENEL_TEDARIKCI_KODU = "T-000"
    FATURA_TIP_ALIS = "ALIŞ"
    FATURA_TIP_SATIS = "SATIŞ"
    FATURA_TIP_DEVIR_GIRIS = "DEVİR_GİRİŞ"
    FATURA_TIP_SATIS_IADE = "SATIŞ İADE"
    FATURA_TIP_ALIS_IADE = "ALIŞ İADE"

    # Ödeme Türleri
    ODEME_TURU_NAKIT = "NAKİT"
    ODEME_TURU_KART = "KART"
    ODEME_TURU_EFT_HAVALE = "EFT/HAVALE"
    ODEME_TURU_CEK = "ÇEK"
    ODEME_TURU_SENET = "SENET"
    ODEME_TURU_ACIK_HESAP = "AÇIK HESAP"
    ODEME_TURU_ETKISIZ_FATURA = "ETKİSİZ FATURA"
    # Peşin Ödeme Türleri (Liste olarak tanımlanmıştır)
    # self.pesin_odeme_turleri = ["NAKİT", "KART", "EFT/HAVALE", "ÇEK", "SENET"] # Bu, __init__ içinde tanımlanmalı

    # Cari Tipleri
    CARI_TIP_MUSTERI = "MUSTERI"
    CARI_TIP_TEDARIKCI = "TEDARIKCI"

    # İşlem Tipleri (Cari Hareketler ve Gelir/Gider)
    ISLEM_TIP_ALACAK = "ALACAK"
    ISLEM_TIP_BORC = "BORC"
    ISLEM_TIP_TAHSILAT = "TAHSILAT"
    ISLEM_TIP_ODEME = "ODEME"
    ISLEM_TIP_GELIR = "GELİR"
    ISLEM_TIP_GIDER = "GİDER"

    # Kaynak Tipleri (Stok Hareketleri, Gelir/Gider, Cari Hareketler)
    KAYNAK_TIP_MANUEL = "MANUEL"
    KAYNAK_TIP_FATURA = "FATURA"
    KAYNAK_TIP_SIPARIS = "SIPARIS"
    KAYNAK_TIP_TAHSILAT = "TAHSILAT"
    KAYNAK_TIP_ODEME = "ODEME"
    KAYNAK_TIP_VERESIYE_BORC_MANUEL = "VERESIYE_BORC_MANUEL"
    KAYNAK_TIP_FATURA_SATIS_PESIN = "FATURA_SATIS_PESIN"
    KAYNAK_TIP_FATURA_ALIS_PESIN = "FATURA_ALIS_PESIN"
    KAYNAK_TIP_IADE_FATURA = "İADE_FATURA"
    KAYNAK_TIP_IADE_FATURA_SATIS_PESIN = "İADE_FATURA_SATIS_PESIN"
    KAYNAK_TIP_IADE_FATURA_ALIS_PESIN = "İADE_FATURA_ALIS_PESIN"

    # Stok Hareketleri İşlem Tipleri (Daha spesifik, açıklamada kullanılabilir)
    STOK_ISLEM_TIP_GIRIS_MANUEL = "Giriş (Manuel)"
    STOK_ISLEM_TIP_CIKIS_MANUEL = "Çıkış (Manuel)"
    STOK_ISLEM_TIP_SAYIM_FAZLASI = "Sayım Fazlası"
    STOK_ISLEM_TIP_SAYIM_EKSIGI = "Sayım Eksiği"
    STOK_ISLEM_TIP_ZAYIAT = "Zayiat"
    STOK_ISLEM_TIP_IADE_GIRIS = "İade Girişi"
    STOK_ISLEM_TIP_FATURA_ALIS = "Fatura Alış"
    STOK_ISLEM_TIP_FATURA_SATIS = "Fatura Satış"
    STOK_ISLEM_TIP_FATURA_SATIS_IADE = "Fatura Satış İade"
    STOK_ISLEM_TIP_FATURA_ALIS_IADE = "Fatura Alış İade"
    STOK_ISLEM_TIP_DEVIR_GIRIS = "Devir Giriş" # Genellikle 'DEVİR_GİRİŞ' fatura tipi için kullanılır.
    STOK_ISLEM_TIP_GIRIS_MANUEL_DUZELTME = "Giriş (Manuel Düzeltme)" # Stok kartından manuel düzeltme
    STOK_ISLEM_TIP_CIKIS_MANUEL_DUZELTME = "Çıkış (Manuel Düzeltme)" # Stok kartından manuel düzeltme

    # Sipariş Durumları
    SIPARIS_DURUM_BEKLEMEDE = "BEKLEMEDE"
    SIPARIS_DURUM_TAMAMLANDI = "TAMAMLANDI"
    SIPARIS_DURUM_KISMİ_TESLIMAT = "KISMİ_TESLİMAT"
    SIPARIS_DURUM_IPTAL_EDILDI = "İPTAL_EDİLDİ"

    # Sipariş Tipleri (arayüzdeki combobox için)
    SIPARIS_TIP_SATIS = "SATIŞ_SIPARIS"
    SIPARIS_TIP_ALIS = "ALIŞ_SIPARIS"

    # Genel İskonto Tipleri
    ISKONTO_TIP_YOK = "YOK"
    ISKONTO_TIP_YUZDE = "YUZDE"
    ISKONTO_TIP_TUTAR = "TUTAR"
    
    def __init__(self, data_dir='data'):
        """
        Hem eski (sqlite3) hem de yeni (SQLAlchemy) veritabanı bağlantılarını başlatır.
        """
        # --- YENİ EKLENEN KISIM: SQLAlchemy Oturumu ---
        from api.veritabani import SessionLocal
        self._db_session = SessionLocal()
        # ---------------------------------------------

        # --- ESKİ YAPI (GEÇİŞ SÜRECİNDE GEREKLİ) ---
        self.data_dir = data_dir
        self.db_name = os.path.join(self.data_dir, 'on_muhasebe.db')
        self.conn = sqlite3.connect(self.db_name, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
        self.conn.row_factory = sqlite3.Row
        self.c = self.conn.cursor()
        
        # Diğer başlangıç ayarları
        self.sirket_bilgileri = self.sirket_bilgilerini_yukle()
        self.perakende_musteri_id = self.get_perakende_musteri_id()
        self.genel_tedarikci_id = None # Bu özellik kaldırılabilir veya ayarlanabilir.
        
        logging.info(f"OnMuhasebe sınıfı başlatıldı. Veritabanı yolu: {self.db_name}")

    # --- Müşteri Yönetimi ---
    def musteri_ekle(self, kod, ad, telefon=None, adres=None, vergi_dairesi=None, vergi_no=None):
        try:
            db_musteri = self._db_session.query(Musteri).filter(Musteri.kod == kod).first()
            if db_musteri:
                return False, f"Müşteri kodu '{kod}' zaten mevcut."
            
            yeni_musteri = Musteri(
                kod=kod, ad=ad, telefon=telefon, adres=adres,
                vergi_dairesi=vergi_dairesi, vergi_no=vergi_no,
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_musteri)
            self._db_session.commit()
            self._db_session.refresh(yeni_musteri)
            return True, yeni_musteri.id # Başarılı olursa ID döndür
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Müşteri eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Müşteri eklenirken bir hata oluştu: {e}"

    def musteri_guncelle(self, musteri_id, kod, ad, telefon=None, adres=None, vergi_dairesi=None, vergi_no=None):
        try:
            db_musteri = self._db_session.query(Musteri).filter(Musteri.id == musteri_id).first()
            if db_musteri is None:
                return False, "Müşteri bulunamadı."
            
            # Kod değişmişse, yeni kodun benzersizliğini kontrol et
            if db_musteri.kod != kod:
                existing_musteri_with_new_kod = self._db_session.query(Musteri).filter(Musteri.kod == kod).first()
                if existing_musteri_with_new_kod:
                    return False, f"Müşteri kodu '{kod}' zaten başka bir müşteriye ait."

            db_musteri.kod = kod
            db_musteri.ad = ad
            db_musteri.telefon = telefon
            db_musteri.adres = adres
            db_musteri.vergi_dairesi = vergi_dairesi
            db_musteri.vergi_no = vergi_no
            db_musteri.son_guncelleme_tarihi_saat = datetime.now()
            db_musteri.son_guncelleyen_kullanici_id = self._get_current_user_id()
            
            self._db_session.commit()
            self._db_session.refresh(db_musteri)
            return True, f"Müşteri '{ad}' başarıyla güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Müşteri güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Müşteri güncellenirken bir hata oluştu: {e}"

    def musteri_sil(self, musteri_id):
        try:
            if musteri_id == self.perakende_musteri_id:
                return False, "Perakende satış müşterisi silinemez."

            db_musteri = self._db_session.query(Musteri).filter(Musteri.id == musteri_id).first()
            if db_musteri is None:
                return False, "Müşteri bulunamadı."
            
            # Müşteriye ait faturalar veya cari hareketler varsa silme engellenebilir (iş mantığına göre)
            # Şu an için doğrudan silme işlemi yapılıyor, ancak ilişkili kayıtlar hata verebilir.
            # CASCADE DELETE ayarlanmadıysa foreign key hatası alınabilir.
            # Ancak biz ORM kullandığımız için, ilişkileri silmek için önce bağlı objeleri silmemiz gerekir.
            # VEYA semalar.py'de cascade='all, delete-orphan' gibi ayarlamalar yapmalıyız.
            
            # İlişkili cari hareketlerin kontrolü (daha önce manuel SQL ile yapılmıştı, şimdi ORM)
            if self._db_session.query(CariHareketler).filter(CariHareketler.cari_id == musteri_id, CariHareketler.cari_tip == self.CARI_TIP_MUSTERI).first():
                return False, "Bu müşteriye ait cari hareketler (fatura, tahsilat vb.) bulunmaktadır.\nBir müşteriyi silebilmek için öncelikle tüm ilişkili kayıtların (faturalar, tahsilatlar vb.) silinmesi gerekir."

            self._db_session.delete(db_musteri)
            self._db_session.commit()
            return True, f"Müşteri '{db_musteri.ad}' başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Müşteri silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Müşteri silinirken bir hata oluştu: {e}"

    def musteri_listesi_al(self, arama_terimi=None, perakende_haric=False, limit=None, offset=None):
        query = self._db_session.query(Musteri)
        if perakende_haric:
            query = query.filter(Musteri.kod != self.PERAKENDE_MUSTERI_KODU)
        if arama_terimi:
            # SQLAlchemy'de Türkçe karakter normalizasyonu doğrudan SQL ile zor.
            # Python tarafında normalizasyon yapılıp, like sorgusu gönderilebilir.
            # Veya veritabanında normalize edilmiş sütunlar tutulabilir.
            normalized_term = normalize_turkish_chars(arama_terimi) # yardimcilar.py'den gelen fonksiyonu kullanırız
            query = query.filter(
                or_(
                    Musteri.ad.ilike(f"%{normalized_term}%"),
                    Musteri.kod.ilike(f"%{normalized_term}%"),
                    Musteri.telefon.ilike(f"%{normalized_term}%"),
                    Musteri.adres.ilike(f"%{normalized_term}%")
                )
            )
        
        query = query.order_by(Musteri.ad) # Varsayılan sıralama
        
        if limit is not None:
            query = query.limit(limit)
        if offset is not None:
            query = query.offset(offset)

        return query.all() # ORM objelerini döndür

    def musteri_getir_by_id(self, musteri_id):
        try:
            return self._db_session.query(Musteri).filter(Musteri.id == musteri_id).first()
        except Exception as e:
            logging.error(f"Müşteri ID ile getirilirken hata: {e}\n{traceback.format_exc()}")
            return None
            
    def get_next_musteri_kodu(self, length=4): # Uzunluk 4 olarak ayarlandı (örn: M0001)
        try:
            # Sadece 'M' ile başlayan ve sonu rakam olan kodları bul
            last_musteri = self._db_session.query(Musteri).filter(
                Musteri.kod.like('M%')
            ).order_by(
                Musteri.kod.desc()
            ).first()

            if last_musteri and last_musteri.kod and len(last_musteri.kod) > 1 and last_musteri.kod[0].upper() == 'M' and last_musteri.kod[1:].isdigit():
                last_num = int(last_musteri.kod[1:])
                return f"M{last_num + 1:0{length}d}" # Dinamik uzunluk için format
            return f"M{1:0{length}d}"
        except Exception as e:
            logging.error(f"Sonraki müşteri kodu oluşturulurken hata: {e}\n{traceback.format_exc()}")
            return f"M{1:0{length}d}" # Hata durumunda varsayılan dön

    def _ensure_perakende_musteri(self):
        """Varsayılan perakende satış müşterisini oluşturur ve ID'sini döndürür."""
        try:
            perakende = self._db_session.query(Musteri).filter(Musteri.kod == self.PERAKENDE_MUSTERI_KODU).first()
            if not perakende:
                new_perakende = Musteri(
                    kod=self.PERAKENDE_MUSTERI_KODU,
                    ad="Perakende Satış Müşterisi",
                    adres="Genel Müşteri",
                    olusturma_tarihi_saat=datetime.now(),
                    olusturan_kullanici_id=self._get_current_user_id()
                )
                self._db_session.add(new_perakende)
                self._db_session.commit()
                self._db_session.refresh(new_perakende)
                logging.info("Varsayılan 'Perakende Satış Müşterisi' oluşturuldu.")
                return new_perakende.id
            return perakende.id
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Perakende müşteri oluşturulurken hata: {e}\n{traceback.format_exc()}")
            return None # Hata durumunda None dön

    def get_perakende_musteri_id(self):
        perakende = self._db_session.query(Musteri).filter(Musteri.kod == self.PERAKENDE_MUSTERI_KODU).first()
        return perakende.id if perakende else None
    
    def get_musteri_net_bakiye(self, musteri_id):
        # Müşteri: Alacak (+) Borç (-)
        try:
            # SQLAlchemy ORM ile sorgu
            # func.sum ve case kullanımları SQLAlchemy'de de benzerdir.
            # Ancak, referans tipleri ve odeme_turu için JOIN gerekebilir.
            # Bu, API'deki cari_hareketler endpoint'inden gelmelidir, OnMuhasebe'nin kendisi hesaplamamalıdır.
            # Şimdilik, bu metot API'ye taşınana kadar doğrudan ORM ile simüle edelim.

            # Bu hesaplama çok kritik ve detaylı, direkt API'den alınması daha doğru.
            # Ama geçici olarak ORM ile yapalım.

            toplam_alacak = self._db_session.query(func.sum(CariHareketler.tutar)).filter(
                CariHareketler.cari_id == musteri_id,
                CariHareketler.cari_tip == self.CARI_TIP_MUSTERI,
                or_(
                    CariHareketler.islem_tipi == self.ISLEM_TIP_ALACAK,
                    CariHareketler.referans_tip == self.KAYNAK_TIP_FATURA,
                    CariHareketler.referans_tip == self.KAYNAK_TIP_VERESIYE_BORC_MANUEL
                )
            ).scalar() or 0.0

            toplam_borc = self._db_session.query(func.sum(CariHareketler.tutar)).filter(
                CariHareketler.cari_id == musteri_id,
                CariHareketler.cari_tip == self.CARI_TIP_MUSTERI,
                or_(
                    CariHareketler.islem_tipi == self.ISLEM_TIP_TAHSILAT,
                    CariHareketler.referans_tip == self.KAYNAK_TIP_FATURA_SATIS_PESIN,
                    CariHareketler.referans_tip == self.KAYNAK_TIP_IADE_FATURA
                )
            ).scalar() or 0.0

            # Müşterinin net borcu: Alacaklar (bizim ona borcumuz) - Tahsilatlar (onun bize olan borcu)
            # Bu projede cari net bakiye (Musteri: Bize Borçlu (+), Biz Ona Borçlu (-)) şeklinde tanımlanmış.
            # Yani müşteri bize borçluysa pozitif, biz ona borçluysak negatif.
            # ALACAK (müşteri bize borçlu) - BORÇ (biz müşteriye borçluyuz)
            return float(toplam_alacak - toplam_borc)
        except Exception as e:
            logging.error(f"Müşteri net bakiye hesaplanırken hata: {e}\n{traceback.format_exc()}")
            return 0.0

    # --- Tedarikçi Yönetimi ---
    def tedarikci_ekle(self, tedarikci_kodu, ad, telefon=None, adres=None, vergi_dairesi=None, vergi_no=None):
        try:
            db_tedarikci = self._db_session.query(Tedarikci).filter(Tedarikci.tedarikci_kodu == tedarikci_kodu).first()
            if db_tedarikci:
                return False, f"Tedarikçi kodu '{tedarikci_kodu}' zaten mevcut."
            
            yeni_tedarikci = Tedarikci(
                tedarikci_kodu=tedarikci_kodu, ad=ad, telefon=telefon, adres=adres,
                vergi_dairesi=vergi_dairesi, vergi_no=vergi_no,
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_tedarikci)
            self._db_session.commit()
            self._db_session.refresh(yeni_tedarikci)
            return True, yeni_tedarikci.id # Başarılı olursa ID döndür
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Tedarikçi eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Tedarikçi eklenirken bir hata oluştu: {e}"

    def tedarikci_guncelle(self, tedarikci_id, tedarikci_kodu, ad, telefon=None, adres=None, vergi_dairesi=None, vergi_no=None):
        try:
            db_tedarikci = self._db_session.query(Tedarikci).filter(Tedarikci.id == tedarikci_id).first()
            if db_tedarikci is None:
                return False, "Tedarikçi bulunamadı."
            
            # Kod değişmişse, yeni kodun benzersizliğini kontrol et
            if db_tedarikci.tedarikci_kodu != tedarikci_kodu:
                existing_tedarikci_with_new_kod = self._db_session.query(Tedarikci).filter(Tedarikci.tedarikci_kodu == tedarikci_kodu).first()
                if existing_tedarikci_with_new_kod:
                    return False, f"Tedarikçi kodu '{tedarikci_kodu}' zaten başka bir tedarikçiye ait."

            db_tedarikci.tedarikci_kodu = tedarikci_kodu
            db_tedarikci.ad = ad
            db_tedarikci.telefon = telefon
            db_tedarikci.adres = adres
            db_tedarikci.vergi_dairesi = vergi_dairesi
            db_tedarikci.vergi_no = vergi_no
            db_tedarikci.son_guncelleme_tarihi_saat = datetime.now()
            db_tedarikci.son_guncelleyen_kullanici_id = self._get_current_user_id()
            
            self._db_session.commit()
            self._db_session.refresh(db_tedarikci)
            return True, f"Tedarikçi '{ad}' başarıyla güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Tedarikçi güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Tedarikçi güncellenirken bir hata oluştu: {e}"

    def tedarikci_sil(self, tedarikci_id):
        try:
            if tedarikci_id == self.genel_tedarikci_id:
                return False, "Genel tedarikçi silinemez."

            db_tedarikci = self._db_session.query(Tedarikci).filter(Tedarikci.id == tedarikci_id).first()
            if db_tedarikci is None:
                return False, "Tedarikçi bulunamadı."
            
            # Tedarikçiye ait faturalar veya cari hareketler varsa silme engellenebilir (iş mantığına göre)
            if self._db_session.query(CariHareketler).filter(CariHareketler.cari_id == tedarikci_id, CariHareketler.cari_tip == self.CARI_TIP_TEDARIKCI).first():
                 return False, "Bu tedarikçiye ait cari hareketler (fatura, ödeme vb.) bulunmaktadır.\nBir tedarikçiyi silebilmek için öncelikle tüm ilişkili kayıtların (faturalar, ödemeler vb.) silinmesi gerekir."

            self._db_session.delete(db_tedarikci)
            self._db_session.commit()
            return True, f"Tedarikçi '{db_tedarikci.ad}' başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Tedarikçi silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Tedarikçi silinirken bir hata oluştu: {e}"

    def tedarikci_listesi_al(self, arama_terimi=None, limit=None, offset=None):
        query = self._db_session.query(Tedarikci)
        if arama_terimi:
            normalized_term = normalize_turkish_chars(arama_terimi)
            query = query.filter(
                or_(
                    Tedarikci.ad.ilike(f"%{normalized_term}%"),
                    Tedarikci.tedarikci_kodu.ilike(f"%{normalized_term}%"),
                    Tedarikci.telefon.ilike(f"%{normalized_term}%"),
                    Tedarikci.adres.ilike(f"%{normalized_term}%")
                )
            )
        
        query = query.order_by(Tedarikci.ad) # Varsayılan sıralama

        if limit is not None:
            query = query.limit(limit)
        if offset is not None:
            query = query.offset(offset)

        return query.all() # ORM objelerini döndür

    def tedarikci_getir_by_id(self, tedarikci_id):
        try:
            return self._db_session.query(Tedarikci).filter(Tedarikci.id == tedarikci_id).first()
        except Exception as e:
            logging.error(f"Tedarikçi ID ile getirilirken hata: {e}\n{traceback.format_exc()}")
            return None

    def get_next_tedarikci_kodu(self, length=4): # Uzunluk 4 olarak ayarlandı (örn: T0001)
        try:
            last_tedarikci = self._db_session.query(Tedarikci).filter(
                Tedarikci.tedarikci_kodu.like('T%')
            ).order_by(
                Tedarikci.tedarikci_kodu.desc()
            ).first()

            if last_tedarikci and last_tedarikci.tedarikci_kodu and len(last_tedarikci.tedarikci_kodu) > 1 and last_tedarikci.tedarikci_kodu[0].upper() == 'T' and last_tedarikci.tedarikci_kodu[1:].isdigit():
                last_num = int(last_tedarikci.tedarikci_kodu[1:])
                return f"T{last_num + 1:0{length}d}"
            return f"T{1:0{length}d}"
        except Exception as e:
            logging.error(f"Sonraki tedarikçi kodu oluşturulurken hata: {e}\n{traceback.format_exc()}")
            return f"T{1:0{length}d}"

    def _ensure_genel_tedarikci(self):
        """Varsayılan genel tedarikçiyi oluşturur ve ID'sini döndürür."""
        try:
            genel = self._db_session.query(Tedarikci).filter(Tedarikci.tedarikci_kodu == self.GENEL_TEDARIKCI_KODU).first()
            if not genel:
                new_genel = Tedarikci(
                    tedarikci_kodu=self.GENEL_TEDARIKCI_KODU,
                    ad="Genel Tedarikçi",
                    adres="Genel",
                    olusturma_tarihi_saat=datetime.now(),
                    olusturan_kullanici_id=self._get_current_user_id()
                )
                self._db_session.add(new_genel)
                self._db_session.commit()
                self._db_session.refresh(new_genel)
                logging.info("Varsayılan 'Genel Tedarikçi' oluşturuldu.")
                return new_genel.id
            return genel.id
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Genel tedarikçi oluşturulurken hata: {e}\n{traceback.format_exc()}")
            return None

    def get_genel_tedarikci_id(self):
        genel = self._db_session.query(Tedarikci).filter(Tedarikci.tedarikci_kodu == self.GENEL_TEDARIKCI_KODU).first()
        return genel.id if genel else None

    def get_tedarikci_net_bakiye(self, tedarikci_id):
        # Tedarikçi: Borç (+) Alacak (-)
        try:
            # Buradaki hesaplama, API'den gelmesi gereken bir iş mantığıdır.
            # Geçici olarak ORM ile yapıyoruz.
            toplam_borc = self._db_session.query(func.sum(CariHareketler.tutar)).filter(
                CariHareketler.cari_id == tedarikci_id,
                CariHareketler.cari_tip == self.CARI_TIP_TEDARIKCI,
                or_(
                    CariHareketler.islem_tipi == self.ISLEM_TIP_BORC,
                    CariHareketler.referans_tip == self.KAYNAK_TIP_FATURA,
                    CariHareketler.referans_tip == self.KAYNAK_TIP_VERESIYE_BORC_MANUEL
                )
            ).scalar() or 0.0

            toplam_alacak = self._db_session.query(func.sum(CariHareketler.tutar)).filter(
                CariHareketler.cari_id == tedarikci_id,
                CariHareketler.cari_tip == self.CARI_TIP_TEDARIKCI,
                or_(
                    CariHareketler.islem_tipi == self.ISLEM_TIP_ODEME,
                    CariHareketler.referans_tip == self.KAYNAK_TIP_FATURA_ALIS_PESIN,
                    CariHareketler.referans_tip == self.KAYNAK_TIP_IADE_FATURA # Alış iadesi borcu azaltır
                )
            ).scalar() or 0.0

            # Tedarikçinin net borcu: Borçlar (Bizim tedarikçiye borcumuz) - Alacaklar (Tedarikçiden aldıklarımız)
            # Eğer sonuç pozitifse biz borçluyuz, negatifse tedarikçi bize borçlu (yani biz alacaklıyız).
            return float(toplam_borc - toplam_alacak)
        except Exception as e:
            logging.error(f"Tedarikçi net bakiye hesaplanırken hata: {e}\n{traceback.format_exc()}")
            return 0.0

    # --- Ürün Kategori / Marka / Grup / Birim / Ülke Yönetimi ---
    def kategori_ekle(self, kategori_adi):
        try:
            db_kategori = self._db_session.query(UrunKategorileri).filter(UrunKategorileri.kategori_adi == kategori_adi).first()
            if db_kategori:
                return False, "Bu kategori adı zaten mevcut."
            yeni_kategori = UrunKategorileri(
                kategori_adi=kategori_adi,
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_kategori)
            self._db_session.commit()
            self._db_session.refresh(yeni_kategori)
            return True, f"'{kategori_adi}' kategorisi başarıyla eklendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Kategori eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Kategori eklenirken bir hata oluştu: {e}"

    def kategori_guncelle(self, kategori_id, yeni_kategori_adi):
        try:
            db_kategori = self._db_session.query(UrunKategorileri).filter(UrunKategorileri.id == kategori_id).first()
            if db_kategori is None:
                return False, "Kategori bulunamadı."
            
            if db_kategori.kategori_adi != yeni_kategori_adi:
                existing_kategori = self._db_session.query(UrunKategorileri).filter(UrunKategorileri.kategori_adi == yeni_kategori_adi).first()
                if existing_kategori:
                    return False, "Bu kategori adı zaten mevcut."

            db_kategori.kategori_adi = yeni_kategori_adi
            db_kategori.son_guncelleme_tarihi_saat = datetime.now()
            db_kategori.son_guncelleyen_kullanici_id = self._get_current_user_id()
            self._db_session.commit()
            self._db_session.refresh(db_kategori)
            return True, f"Kategori başarıyla '{yeni_kategori_adi}' olarak güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Kategori güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Kategori güncellenirken bir hata oluştu: {e}"

    def kategori_sil(self, kategori_id):
        try:
            db_kategori = self._db_session.query(UrunKategorileri).filter(UrunKategorileri.id == kategori_id).first()
            if db_kategori is None:
                return False, "Kategori bulunamadı."
            
            # Kategoriye bağlı ürün olup olmadığını kontrol et
            if self._db_session.query(Stok).filter(Stok.kategori_id == kategori_id).first():
                return False, "Bu kategoriye bağlı ürünler olduğu için silinemez."

            self._db_session.delete(db_kategori)
            self._db_session.commit()
            return True, f"'{db_kategori.kategori_adi}' kategorisi başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Kategori silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Kategori silinirken bir hata oluştu: {e}"

    def kategori_listele(self):
        try:
            return self._db_session.query(UrunKategorileri).order_by(UrunKategorileri.kategori_adi).all()
        except Exception as e:
            logging.error(f"Kategoriler listelenirken hata: {e}\n{traceback.format_exc()}")
            return []

    def kategori_getir_by_id(self, kategori_id):
        try:
            return self._db_session.query(UrunKategorileri).filter(UrunKategorileri.id == kategori_id).first()
        except Exception as e:
            logging.error(f"Kategori ID ile getirilirken hata: {e}\n{traceback.format_exc()}")
            return None

    def marka_ekle(self, marka_adi):
        try:
            db_marka = self._db_session.query(UrunMarkalari).filter(UrunMarkalari.marka_adi == marka_adi).first()
            if db_marka:
                return False, "Bu marka adı zaten mevcut."
            yeni_marka = UrunMarkalari(
                marka_adi=marka_adi,
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_marka)
            self._db_session.commit()
            self._db_session.refresh(yeni_marka)
            return True, f"'{marka_adi}' markası başarıyla eklendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Marka eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Marka eklenirken bir hata oluştu: {e}"

    def marka_guncelle(self, marka_id, yeni_marka_adi):
        try:
            db_marka = self._db_session.query(UrunMarkalari).filter(UrunMarkalari.id == marka_id).first()
            if db_marka is None:
                return False, "Marka bulunamadı."
            
            if db_marka.marka_adi != yeni_marka_adi:
                existing_marka = self._db_session.query(UrunMarkalari).filter(UrunMarkalari.marka_adi == yeni_marka_adi).first()
                if existing_marka:
                    return False, "Bu marka adı zaten mevcut."

            db_marka.marka_adi = yeni_marka_adi
            db_marka.son_guncelleme_tarihi_saat = datetime.now()
            db_marka.son_guncelleyen_kullanici_id = self._get_current_user_id()
            self._db_session.commit()
            self._db_session.refresh(db_marka)
            return True, f"Marka başarıyla '{yeni_marka_adi}' olarak güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Marka güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Marka güncellenirken bir hata oluştu: {e}"

    def marka_sil(self, marka_id):
        try:
            db_marka = self._db_session.query(UrunMarkalari).filter(UrunMarkalari.id == marka_id).first()
            if db_marka is None:
                return False, "Marka bulunamadı."
            
            # Markaya bağlı ürün olup olmadığını kontrol et
            if self._db_session.query(Stok).filter(Stok.marka_id == marka_id).first():
                return False, "Bu markaya bağlı ürünler olduğu için silinemez."

            self._db_session.delete(db_marka)
            self._db_session.commit()
            return True, f"'{db_marka.marka_adi}' markası başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Marka silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Marka silinirken bir hata oluştu: {e}"

    def marka_getir_by_id(self, marka_id):
        try:
            return self._db_session.query(UrunMarkalari).filter(UrunMarkalari.id == marka_id).first()
        except Exception as e:
            logging.error(f"Marka ID ile getirilirken hata: {e}\n{traceback.format_exc()}")
            return None

    def marka_listele(self):
        try:
            return self._db_session.query(UrunMarkalari).order_by(UrunMarkalari.marka_adi).all()
        except Exception as e:
            logging.error(f"Markalar listelenirken hata: {e}\n{traceback.format_exc()}")
            return []

    def urun_grubu_ekle(self, grup_adi):
        try:
            db_grup = self._db_session.query(UrunGruplari).filter(UrunGruplari.grup_adi == grup_adi).first()
            if db_grup:
                return False, "Bu ürün grubu adı zaten mevcut."
            yeni_grup = UrunGruplari(
                grup_adi=grup_adi,
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_grup)
            self._db_session.commit()
            self._db_session.refresh(yeni_grup)
            return True, f"'{grup_adi}' ürün grubu başarıyla eklendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ürün grubu eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ürün grubu eklenirken bir hata oluştu: {e}"

    def urun_grubu_guncelle(self, grup_id, yeni_grup_adi):
        try:
            db_grup = self._db_session.query(UrunGruplari).filter(UrunGruplari.id == grup_id).first()
            if db_grup is None:
                return False, "Ürün grubu bulunamadı."
            
            if db_grup.grup_adi != yeni_grup_adi:
                existing_grup = self._db_session.query(UrunGruplari).filter(UrunGruplari.grup_adi == yeni_grup_adi).first()
                if existing_grup:
                    return False, "Bu ürün grubu adı zaten mevcut."

            db_grup.grup_adi = yeni_grup_adi
            db_grup.son_guncelleme_tarihi_saat = datetime.now()
            db_grup.son_guncelleyen_kullanici_id = self._get_current_user_id()
            self._db_session.commit()
            self._db_session.refresh(db_grup)
            return True, f"Ürün grubu başarıyla '{yeni_grup_adi}' olarak güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ürün grubu güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ürün grubu güncellenirken bir hata oluştu: {e}"

    def urun_grubu_sil(self, grup_id):
        try:
            db_grup = self._db_session.query(UrunGruplari).filter(UrunGruplari.id == grup_id).first()
            if db_grup is None:
                return False, "Ürün grubu bulunamadı."
            
            if self._db_session.query(Stok).filter(Stok.urun_grubu_id == grup_id).first():
                return False, "Bu ürün grubuna bağlı ürünler olduğu için silinemez."

            self._db_session.delete(db_grup)
            self._db_session.commit()
            return True, f"'{db_grup.grup_adi}' ürün grubu başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ürün grubu silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ürün grubu silinirken bir hata oluştu: {e}"

    def urun_grubu_listele(self):
        try:
            return self._db_session.query(UrunGruplari).order_by(UrunGruplari.grup_adi).all()
        except Exception as e:
            logging.error(f"Ürün grupları listelenirken hata: {e}\n{traceback.format_exc()}")
            return []

    def urun_grubu_getir_by_id(self, grup_id):
        try:
            return self._db_session.query(UrunGruplari).filter(UrunGruplari.id == grup_id).first()
        except Exception as e:
            logging.error(f"Ürün grubu ID ile getirilirken hata: {e}\n{traceback.format_exc()}")
            return None

    def urun_birimi_ekle(self, birim_adi):
        try:
            db_birim = self._db_session.query(UrunBirimleri).filter(UrunBirimleri.birim_adi == birim_adi).first()
            if db_birim:
                return False, "Bu ürün birimi adı zaten mevcut."
            yeni_birim = UrunBirimleri(
                birim_adi=birim_adi,
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_birim)
            self._db_session.commit()
            self._db_session.refresh(yeni_birim)
            return True, f"'{birim_adi}' ürün birimi başarıyla eklendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ürün birimi eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ürün birimi eklenirken bir hata oluştu: {e}"

    def urun_birimi_guncelle(self, birim_id, yeni_birim_adi):
        try:
            db_birim = self._db_session.query(UrunBirimleri).filter(UrunBirimleri.id == birim_id).first()
            if db_birim is None:
                return False, "Ürün birimi bulunamadı."
            
            if db_birim.birim_adi != yeni_birim_adi:
                existing_birim = self._db_session.query(UrunBirimleri).filter(UrunBirimleri.birim_adi == yeni_birim_adi).first()
                if existing_birim:
                    return False, "Bu ürün birimi adı zaten mevcut."

            db_birim.birim_adi = yeni_birim_adi
            db_birim.son_guncelleme_tarihi_saat = datetime.now()
            db_birim.son_guncelleyen_kullanici_id = self._get_current_user_id()
            self._db_session.commit()
            self._db_session.refresh(db_birim)
            return True, f"Ürün birimi başarıyla '{yeni_birim_adi}' olarak güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ürün birimi güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ürün birimi güncellenirken bir hata oluştu: {e}"

    def urun_birimi_sil(self, birim_id):
        try:
            db_birim = self._db_session.query(UrunBirimleri).filter(UrunBirimleri.id == birim_id).first()
            if db_birim is None:
                return False, "Ürün birimi bulunamadı."
            
            if self._db_session.query(Stok).filter(Stok.urun_birimi_id == birim_id).first():
                return False, "Bu ürün birimine bağlı ürünler olduğu için silinemez."

            self._db_session.delete(db_birim)
            self._db_session.commit()
            return True, f"'{db_birim.birim_adi}' ürün birimi başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ürün birimi silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ürün birimi silinirken bir hata oluştu: {e}"

    def urun_birimi_listele(self):
        try:
            return self._db_session.query(UrunBirimleri).order_by(UrunBirimleri.birim_adi).all()
        except Exception as e:
            logging.error(f"Ürün birimleri listelenirken hata: {e}\n{traceback.format_exc()}")
            return []

    def urun_birimi_getir_by_id(self, birim_id):
        try:
            return self._db_session.query(UrunBirimleri).filter(UrunBirimleri.id == birim_id).first()
        except Exception as e:
            logging.error(f"Ürün birimi ID ile getirilirken hata: {e}\n{traceback.format_exc()}")
            return None

    def ulke_ekle(self, ulke_adi):
        try:
            db_ulke = self._db_session.query(UrunUlkeleri).filter(UrunUlkeleri.ulke_adi == ulke_adi).first()
            if db_ulke:
                return False, "Bu ülke adı zaten mevcut."
            yeni_ulke = UrunUlkeleri(
                ulke_adi=ulke_adi,
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_ulke)
            self._db_session.commit()
            self._db_session.refresh(yeni_ulke)
            return True, f"'{ulke_adi}' ülkesi başarıyla eklendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ülke eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ülke eklenirken bir hata oluştu: {e}"

    def ulke_guncelle(self, ulke_id, yeni_ulke_adi):
        try:
            db_ulke = self._db_session.query(UrunUlkeleri).filter(UrunUlkeleri.id == ulke_id).first()
            if db_ulke is None:
                return False, "Ülke bulunamadı."
            
            if db_ulke.ulke_adi != yeni_ulke_adi:
                existing_ulke = self._db_session.query(UrunUlkeleri).filter(UrunUlkeleri.ulke_adi == yeni_ulke_adi).first()
                if existing_ulke:
                    return False, "Bu ülke adı zaten mevcut."

            db_ulke.ulke_adi = yeni_ulke_adi
            db_ulke.son_guncelleme_tarihi_saat = datetime.now()
            db_ulke.son_guncelleyen_kullanici_id = self._get_current_user_id()
            self._db_session.commit()
            self._db_session.refresh(db_ulke)
            return True, f"Ülke başarıyla '{yeni_ulke_adi}' olarak güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ülke güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ülke güncellenirken bir hata oluştu: {e}"

    def ulke_sil(self, ulke_id):
        try:
            db_ulke = self._db_session.query(UrunUlkeleri).filter(UrunUlkeleri.id == ulke_id).first()
            if db_ulke is None:
                return False, "Ülke bulunamadı."
            
            if self._db_session.query(Stok).filter(Stok.ulke_id == ulke_id).first():
                return False, "Bu ülkeye bağlı ürünler olduğu için silinemez."

            self._db_session.delete(db_ulke)
            self._db_session.commit()
            return True, f"'{db_ulke.ulke_adi}' ülkesi başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Ülke silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Ülke silinirken bir hata oluştu: {e}"

    def ulke_listele(self):
        try:
            return self._db_session.query(UrunUlkeleri).order_by(UrunUlkeleri.ulke_adi).all()
        except Exception as e:
            logging.error(f"Ülkeler listelenirken hata: {e}\n{traceback.format_exc()}")
            return []

    def ulke_getir_by_id(self, ulke_id):
        try:
            return self._db_session.query(UrunUlkeleri).filter(UrunUlkeleri.id == ulke_id).first()
        except Exception as e:
            logging.error(f"Ülke ID ile getirilirken hata: {e}\n{traceback.format_exc()}")
            return None
#------------------------------------------
# --- Ürün Kategori / Marka / Grup / Birim / Ülke Yönetimi --- SONU
#------------------------------------------

    def stok_ekle(self, urun_kodu, urun_adi, stok_miktari=0.0, alis_fiyati_kdv_haric=0.0, satis_fiyati_kdv_haric=0.0, kdv_orani=0.0, min_stok_seviyesi=0.0, alis_fiyati_kdv_dahil=0.0, satis_fiyati_kdv_dahil=0.0, kategori_id=None, marka_id=None, urun_detayi=None, urun_resmi_yolu=None, fiyat_degisiklik_tarihi=None, urun_grubu_id=None, urun_birimi_id=None, ulke_id=None):
        try:
            db_stok = self._db_session.query(Stok).filter(Stok.urun_kodu == urun_kodu).first()
            if db_stok:
                return False, "Ürün kodu zaten mevcut."
            
            if fiyat_degisiklik_tarihi is None:
                fiyat_degisiklik_tarihi = date.today()

            yeni_urun = Stok(
                urun_kodu=urun_kodu,
                urun_adi=urun_adi,
                stok_miktari=self.safe_float(stok_miktari),
                alis_fiyati_kdv_haric=self.safe_float(alis_fiyati_kdv_haric),
                satis_fiyati_kdv_haric=self.safe_float(satis_fiyati_kdv_haric),
                kdv_orani=self.safe_float(kdv_orani),
                min_stok_seviyesi=self.safe_float(min_stok_seviyesi),
                alis_fiyati_kdv_dahil=self.safe_float(alis_fiyati_kdv_dahil),
                satis_fiyati_kdv_dahil=self.safe_float(satis_fiyati_kdv_dahil),
                kategori_id=kategori_id,
                marka_id=marka_id,
                urun_detayi=urun_detayi,
                urun_resmi_yolu=urun_resmi_yolu,
                fiyat_degisiklik_tarihi=fiyat_degisiklik_tarihi,
                urun_grubu_id=urun_grubu_id,
                urun_birimi_id=urun_birimi_id,
                ulke_id=ulke_id,
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_urun)
            self._db_session.commit()
            self._db_session.refresh(yeni_urun)
            return True, yeni_urun.id
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Stok eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Stok eklenirken bir hata oluştu: {e}"

    def stok_guncelle(self, urun_id, urun_kodu, urun_adi, stok_miktari, alis_fiyati_kdv_haric, satis_fiyati_kdv_haric, kdv_orani, min_stok_seviyesi, alis_fiyati_kdv_dahil, satis_fiyati_kdv_dahil, kategori_id=None, marka_id=None, urun_detayi=None, urun_resmi_yolu=None, fiyat_degisiklik_tarihi=None, urun_grubu_id=None, urun_birimi_id=None, ulke_id=None):
        try:
            db_stok = self._db_session.query(Stok).filter(Stok.id == urun_id).first()
            if db_stok is None:
                return False, "Ürün bulunamadı."
            
            # Kod değişmişse, yeni kodun benzersizliğini kontrol et
            if db_stok.urun_kodu != urun_kodu:
                existing_stok_with_new_kod = self._db_session.query(Stok).filter(Stok.urun_kodu == urun_kodu).first()
                if existing_stok_with_new_kod:
                    return False, f"Ürün kodu '{urun_kodu}' zaten başka bir ürüne ait."

            # Eski stok miktarını al (değişim hesabı için)
            eski_stok_miktari_f = float(db_stok.stok_miktari)

            db_stok.urun_kodu = urun_kodu
            db_stok.urun_adi = urun_adi
            db_stok.alis_fiyati_kdv_haric = self.safe_float(alis_fiyati_kdv_haric)
            db_stok.satis_fiyati_kdv_haric = self.safe_float(satis_fiyati_kdv_haric)
            db_stok.kdv_orani = self.safe_float(kdv_orani)
            db_stok.min_stok_seviyesi = self.safe_float(min_stok_seviyesi)
            db_stok.alis_fiyati_kdv_dahil = self.safe_float(alis_fiyati_kdv_dahil)
            db_stok.satis_fiyati_kdv_dahil = self.safe_float(satis_fiyati_kdv_dahil)
            db_stok.kategori_id = kategori_id
            db_stok.marka_id = marka_id
            db_stok.urun_detayi = urun_detayi
            db_stok.urun_resmi_yolu = urun_resmi_yolu
            db_stok.fiyat_degisiklik_tarihi = fiyat_degisiklik_tarihi
            db_stok.urun_grubu_id = urun_grubu_id
            db_stok.urun_birimi_id = urun_birimi_id
            db_stok.ulke_id = ulke_id
            db_stok.son_guncelleme_tarihi_saat = datetime.now()
            db_stok.son_guncelleyen_kullanici_id = self._get_current_user_id()

            # Stok miktarını güncelle (bu, eğer API'den geliyorsa doğrudan güncellenecek alan)
            db_stok.stok_miktari = self.safe_float(stok_miktari) # API'den gelen son stok miktarı

            # Yeni stok miktarı ile eski stok miktarını karşılaştır ve fark varsa stok hareketi oluştur.
            stok_farki = self.safe_float(stok_miktari) - eski_stok_miktari_f

            if stok_farki != 0:
                islem_tipi = ""
                if stok_farki > 0:
                    islem_tipi = self.STOK_ISLEM_TIP_GIRIS_MANUEL_DUZELTME
                else:
                    islem_tipi = self.STOK_ISLEM_TIP_CIKIS_MANUEL_DUZELTME

                # _stok_guncelle_ve_hareket_kaydet metodunu çağır (bu metot stoğu güncellediği için burada db_stok.stok_miktari'nı tekrar güncellemiyoruz)
                # Ancak _stok_guncelle_ve_hareket_kaydet transaction yönettiği için, burada yeni bir yaklaşım gerekli.
                # Ya _stok_guncelle_ve_hareket_kaydet'in transaction'ı kaldırılır, ya da burada manuel olarak yapılır.
                # Şimdilik, _stok_guncelle_ve_hareket_kaydet'i sadece hareketi kaydetmek için kullanalım.

                # Sadece hareket kaydı oluştur (stok zaten db_stok.stok_miktari ile güncellendi)
                yeni_hareket = StokHareketleri(
                    urun_id=urun_id,
                    tarih=date.today(),
                    islem_tipi=islem_tipi,
                    miktar=abs(stok_farki),
                    onceki_stok=eski_stok_miktari_f,
                    sonraki_stok=self.safe_float(stok_miktari),
                    kaynak=self.KAYNAK_TIP_MANUEL,
                    kaynak_id=None,
                    kaynak_no="Ürün Kartı Düzeltme",
                    olusturma_tarihi_saat=datetime.now(),
                    olusturan_kullanici_id=self._get_current_user_id()
                )
                self._db_session.add(yeni_hareket)

            self._db_session.commit()
            self._db_session.refresh(db_stok)
            return True, f"Ürün '{urun_adi}' başarıyla güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Stok güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Stok güncellenirken bir hata oluştu: {e}"

    def stok_sil(self, urun_id):
        try:
            db_stok = self._db_session.query(Stok).filter(Stok.id == urun_id).first()
            if db_stok is None:
                return False, "Ürün bulunamadı."
            
            # Ürüne bağlı diğer tablolarda kayıt olup olmadığını kontrol et
            if self._db_session.query(StokHareketleri).filter(StokHareketleri.urun_id == urun_id).first():
                return False, "Bu ürüne bağlı stok hareketleri olduğu için silinemez."
            if self._db_session.query(FaturaKalemleri).filter(FaturaKalemleri.urun_id == urun_id).first():
                return False, "Bu ürüne bağlı fatura kalemleri olduğu için silinemez."
            if self._db_session.query(SiparisKalemleri).filter(SiparisKalemleri.urun_id == urun_id).first():
                return False, "Bu ürüne bağlı sipariş kalemleri olduğu için silinemez."

            self._db_session.delete(db_stok)
            self._db_session.commit()
            return True, f"Ürün '{db_stok.urun_adi}' başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Stok silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Stok silinirken bir hata oluştu: {e}"

    def stok_listele(self, arama_terimi=None, limit=None, offset=None, kategori_id_filter=None, marka_id_filter=None, urun_grubu_id_filter=None, urun_birimi_id_filter=None, ulke_id_filter=None):
        # SQLAlchemy ORM ile JOIN'li sorgu
        query = self._db_session.query(
            Stok,
            UrunKategorileri.kategori_adi,
            UrunMarkalari.marka_adi,
            UrunGruplari.grup_adi,
            UrunBirimleri.birim_adi,
            UrunUlkeleri.ulke_adi
        ).outerjoin(UrunKategorileri, Stok.kategori_id == UrunKategorileri.id)\
         .outerjoin(UrunMarkalari, Stok.marka_id == UrunMarkalari.id)\
         .outerjoin(UrunGruplari, Stok.urun_grubu_id == UrunGruplari.id)\
         .outerjoin(UrunBirimleri, Stok.urun_birimi_id == UrunBirimleri.id)\
         .outerjoin(UrunUlkeleri, Stok.ulke_id == UrunUlkeleri.id)
        
        if arama_terimi:
            normalized_term = normalize_turkish_chars(arama_terimi)
            query = query.filter(
                or_(
                    Stok.urun_adi.ilike(f"%{normalized_term}%"),
                    Stok.urun_kodu.ilike(f"%{normalized_term}%")
                )
            )
        if kategori_id_filter is not None:
            query = query.filter(Stok.kategori_id == kategori_id_filter)
        if marka_id_filter is not None:
            query = query.filter(Stok.marka_id == marka_id_filter)
        if urun_grubu_id_filter is not None:
            query = query.filter(Stok.urun_grubu_id == urun_grubu_id_filter)
        if urun_birimi_id_filter is not None:
            query = query.filter(Stok.urun_birimi_id == urun_birimi_id_filter)
        if ulke_id_filter is not None:
            query = query.filter(Stok.ulke_id == ulke_id_filter)
        
        query = query.order_by(Stok.urun_adi) # Varsayılan sıralama
        
        if limit is not None:
            query = query.limit(limit)
        if offset is not None:
            query = query.offset(offset)
        
        # Sorgu sonuçlarını al
        stok_result = query.all()
        
        results = []
        for stok_obj, kategori_adi, marka_adi, grup_adi, birim_adi, ulke_adi in stok_result:
            stok_dict = {
                'id': stok_obj.id,
                'urun_kodu': stok_obj.urun_kodu,
                'urun_adi': stok_obj.urun_adi,
                'stok_miktari': float(stok_obj.stok_miktari),
                'alis_fiyati_kdv_haric': float(stok_obj.alis_fiyati_kdv_haric),
                'alis_fiyati_kdv_dahil': float(stok_obj.alis_fiyati_kdv_dahil),
                'satis_fiyati_kdv_haric': float(stok_obj.satis_fiyati_kdv_haric),
                'satis_fiyati_kdv_dahil': float(stok_obj.satis_fiyati_kdv_dahil),
                'kdv_orani': float(stok_obj.kdv_orani),
                'min_stok_seviyesi': float(stok_obj.min_stok_seviyesi),
                'urun_detayi': stok_obj.urun_detayi,
                'urun_resmi_yolu': stok_obj.urun_resmi_yolu,
                'fiyat_degisiklik_tarihi': stok_obj.fiyat_degisiklik_tarihi.strftime('%Y-%m-%d') if stok_obj.fiyat_degisiklik_tarihi else None,
                'kategori_id': stok_obj.kategori_id,
                'marka_id': stok_obj.marka_id,
                'urun_grubu_id': stok_obj.urun_grubu_id,
                'urun_birimi_id': stok_obj.urun_birimi_id,
                'ulke_id': stok_obj.ulke_id,
                'kategori_adi': kategori_adi,
                'marka_adi': marka_adi,
                'urun_grubu_adi': grup_adi,
                'urun_birimi_adi': birim_adi,
                'ulke_adi': ulke_adi,
            }
            results.append(stok_dict)
        return results

    def stok_getir_by_id(self, urun_id):
        try:
            stok_obj = self._db_session.query(Stok).filter(Stok.id == urun_id).first()
            if not stok_obj: return None
            
            # ORM ilişkilerini kullanarak ilgili adları al
            kategori_adi = stok_obj.kategori.kategori_adi if stok_obj.kategori else None
            marka_adi = stok_obj.marka.marka_adi if stok_obj.marka else None
            urun_grubu_adi = stok_obj.urun_grubu.grup_adi if stok_obj.urun_grubu else None
            urun_birimi_adi = stok_obj.urun_birimi.birim_adi if stok_obj.urun_birimi else None
            ulke_adi = stok_obj.ulke.ulke_adi if stok_obj.ulke else None

            stok_dict = {
                'id': stok_obj.id, 'urun_kodu': stok_obj.urun_kodu, 'urun_adi': stok_obj.urun_adi,
                'stok_miktari': float(stok_obj.stok_miktari),
                'alis_fiyati_kdv_haric': float(stok_obj.alis_fiyati_kdv_haric),
                'alis_fiyati_kdv_dahil': float(stok_obj.alis_fiyati_kdv_dahil),
                'satis_fiyati_kdv_haric': float(stok_obj.satis_fiyati_kdv_haric),
                'satis_fiyati_kdv_dahil': float(stok_obj.satis_fiyati_kdv_dahil),
                'kdv_orani': float(stok_obj.kdv_orani), 'min_stok_seviyesi': float(stok_obj.min_stok_seviyesi),
                'urun_detayi': stok_obj.urun_detayi, 'urun_resmi_yolu': stok_obj.urun_resmi_yolu,
                'fiyat_degisiklik_tarihi': stok_obj.fiyat_degisiklik_tarihi.strftime('%Y-%m-%d') if stok_obj.fiyat_degisiklik_tarihi else None,
                'kategori_id': stok_obj.kategori_id, 'marka_id': stok_obj.marka_id, 'urun_grubu_id': stok_obj.urun_grubu_id,
                'urun_birimi_id': stok_obj.urun_birimi_id, 'ulke_id': stok_obj.ulke_id,
                'kategori_adi': kategori_adi,
                'marka_adi': marka_adi,
                'urun_grubu_adi': urun_grubu_adi,
                'urun_birimi_adi': urun_birimi_adi,
                'ulke_adi': ulke_adi,
                'olusturma_tarihi_saat': stok_obj.olusturma_tarihi_saat.strftime('%Y-%m-%d %H:%M:%S') if stok_obj.olusturma_tarihi_saat else None,
                'olusturan_kullanici_id': stok_obj.olusturan_kullanici_id,
                'son_guncelleme_tarihi_saat': stok_obj.son_guncelleme_tarihi_saat.strftime('%Y-%m-%d %H:%M:%S') if stok_obj.son_guncelleme_tarihi_saat else None,
                'son_guncelleyen_kullanici_id': stok_obj.son_guncelleyen_kullanici_id,
            }
            return stok_dict
        except Exception as e:
            logging.error(f"Stok ID ile getirilirken hata: {e}\n{traceback.format_exc()}")
            return None

    def get_next_stok_kodu(self, length=4): # Uzunluk 4 olarak ayarlandı (örn: UR0001)
        try:
            last_stok = self._db_session.query(Stok).filter(
                Stok.urun_kodu.like('UR%')
            ).order_by(
                Stok.urun_kodu.desc()
            ).first()

            if last_stok and last_stok.urun_kodu and len(last_stok.urun_kodu) > 1 and last_stok.urun_kodu[0:2].upper() == 'UR' and last_stok.urun_kodu[2:].isdigit():
                last_num = int(last_stok.urun_kodu[2:])
                return f"UR{last_num + 1:0{length}d}" # Dinamik uzunluk için format
            return f"UR{1:0{length}d}"
        except Exception as e:
            logging.error(f"Sonraki stok kodu oluşturulurken hata: {e}\n{traceback.format_exc()}")
            return f"UR{1:0{length}d}" # Hata durumunda varsayılan dön

    def _stok_guncelle_ve_hareket_kaydet(self, urun_id, miktar_degisimi_net, islem_tipi_aciklamasi, kaynak_tipi, kaynak_id=None, referans_no=None):
        """
        Belirli bir ürünün stok miktarını günceller ve ilgili stok hareketi kaydını oluşturur.
        Bu metodun çağrıldığı ana transaction içinde çalışır.
        
        Args:
            urun_id (int): Stok miktarı güncellenecek ürünün ID'si.
            miktar_degisimi_net (float): Stoğa eklenecek (+) veya çıkarılacak (-) net miktar.
            islem_tipi_aciklamasi (str): Stok hareketi tipi (örn: 'Fatura Satış', 'Manuel Giriş').
            kaynak_tipi (str): Stok hareketinin kaynağı (örn: 'FATURA', 'MANUEL', 'İADE_FATURA').
            kaynak_id (int, optional): Kaynak faturanın/siparişin/işlemin ID'si. None olabilir.
            referans_no (str, optional): Kaynak faturanın/siparişin/işlemin numarası/açıklaması.
        
        Returns:
            tuple: (bool success, str message)
        """
        if miktar_degisimi_net == 0:
            return True, "Stok değişimi sıfır, işlem yapılmadı."

        try:
            # Ürünün mevcut stok miktarını kilitli olarak al (optimistik kilit yerine pesimistik)
            # with_for_update(), SELECT FOR UPDATE gibi çalışır ve bu satırı kilitler.
            db_stok = self._db_session.query(Stok).filter(Stok.id == urun_id).with_for_update().first()
            if db_stok is None:
                return False, "Stok bulunamadı."

            stok_oncesi = float(db_stok.stok_miktari)
            urun_adi = db_stok.urun_adi
            stok_sonrasi = stok_oncesi + miktar_degisimi_net

            # db_stok objesinin stok_miktari alanını güncelle
            db_stok.stok_miktari = stok_sonrasi
            
            # Stok hareketi kaydını oluştur
            yeni_hareket = StokHareketleri(
                urun_id=urun_id,
                tarih=date.today(),
                islem_tipi=islem_tipi_aciklamasi,
                miktar=abs(miktar_degisimi_net), # Miktar her zaman pozitif kaydedilir
                onceki_stok=stok_oncesi,
                sonraki_stok=stok_sonrasi,
                aciklama=f"{islem_tipi_aciklamasi} - Ürün: {urun_adi}. Ref No: {referans_no if referans_no else ''}",
                kaynak=kaynak_tipi,
                kaynak_id=kaynak_id,
                kaynak_no=referans_no, # referans_no'yu da kaydedelim
                olusturma_tarihi_saat=datetime.now(),
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_hareket)
            
            # Commit işlemi bu metodun dışında, ana transaction tarafından yönetilir.
            return True, "Stok hareketi başarıyla kaydedildi."
        except Exception as e:
            logging.error(f"Stok ve hareket kaydedilirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Stok ve hareket kaydedilirken bir hata oluştu: {e}"

    def stok_hareketi_ekle(self, urun_id, islem_tipi, miktar, tarih_str, aciklama=None):
        try:
            tarih_obj = datetime.strptime(tarih_str, '%Y-%m-%d').date()
        except ValueError:
            return False, "Geçersiz tarih formatı. YYYY-AA-GG olmalıdır."
        
        if self.safe_float(miktar) <= 0:
            return False, "Miktar pozitif bir sayı olmalıdır."

        try:
            # Transaction'ı burada yönetiyoruz çünkü bu tek başına bir işlemdir.
            # _stok_guncelle_ve_hareket_kaydet kendi transaction'ını yönetmediği için güvenli.
            self._db_session.begin_nested() # İç içe transaction başlat
            
            success, message = self._stok_guncelle_ve_hareket_kaydet(
                urun_id=urun_id,
                miktar_degisimi_net=self.safe_float(miktar) if islem_tipi in [self.STOK_ISLEM_TIP_GIRIS_MANUEL, self.STOK_ISLEM_TIP_SAYIM_FAZLASI, self.STOK_ISLEM_TIP_IADE_GIRIS] else -self.safe_float(miktar),
                islem_tipi_aciklamasi=islem_tipi,
                kaynak_tipi=self.KAYNAK_TIP_MANUEL,
                kaynak_id=None,
                referans_no=aciklama if aciklama else f"Manuel {islem_tipi}"
            )
            
            if success:
                self._db_session.commit() # İç içe transaction'ı commit et
                return True, f"Stok hareketi başarıyla kaydedildi."
            else:
                self._db_session.rollback() # İç içe transaction'ı rollback et
                return False, message
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Stok hareketi eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Stok hareketi eklenirken bir hata oluştu: {e}"

    def stok_hareketleri_listele(self, urun_id=None, islem_tipi=None, baslangic_tarih=None, bitis_tarih=None):
        query = self._db_session.query(StokHareketleri)
        if urun_id:
            query = query.filter(StokHareketleri.urun_id == urun_id)
        if islem_tipi and islem_tipi != "TÜMÜ":
            query = query.filter(StokHareketleri.islem_tipi == islem_tipi)
        if baslangic_tarih:
            query = query.filter(StokHareketleri.tarih >= baslangic_tarih)
        if bitis_tarih:
            query = query.filter(StokHareketleri.tarih <= bitis_tarih)
        
        hareketler = query.order_by(StokHareketleri.tarih.desc(), StokHareketleri.olusturma_tarihi_saat.desc()).all()
        
        results = []
        for h in hareketler:
            results.append({
                'id': h.id,
                'urun_id': h.urun_id,
                'tarih': h.tarih,
                'islem_tipi': h.islem_tipi,
                'miktar': float(h.miktar),
                'onceki_stok': float(h.onceki_stok),
                'sonraki_stok': float(h.sonraki_stok),
                'aciklama': h.aciklama,
                'kaynak': h.kaynak,
                'kaynak_id': h.kaynak_id,
                'kaynak_no': h.kaynak_no,
                'olusturma_tarihi_saat': h.olusturma_tarihi_saat
            })
        return results

    def get_stok_miktari_for_kontrol(self, urun_id, fatura_id_hariç=None):
        """
        Bir ürünün güncel stok miktarını döndürür.
        Eğer fatura_id_hariç belirtilirse (yani bir fatura düzenleniyorsa),
        o faturadaki ürün miktarını mevcut stoktan düşülmüş gibi hesaplar
        (çünkü faturadaki miktar zaten stoktan düşülmüştür, düzenleme anında geri eklemiş gibi oluruz).
        """
        try:
            current_stok = self._db_session.query(Stok.stok_miktari).filter(Stok.id == urun_id).scalar()
            if current_stok is None: return 0.0
            
            if fatura_id_hariç:
                # Düzenlenen faturadaki bu ürünün miktarını ve fatura tipini al.
                # Sadece SATIŞ ve ALIŞ İADE faturaları stoku azaltır (bu yüzden miktarı geri ekleriz)
                # Sadece ALIŞ ve SATIŞ İADE faturaları stoku artırır (bu yüzden miktarı geri çıkarırız)
                fatura_kalem_data = self._db_session.query(
                    func.sum(FaturaKalemleri.miktar), Fatura.tip
                ).join(Fatura, FaturaKalemleri.fatura_id == Fatura.id).filter(
                    FaturaKalemleri.fatura_id == fatura_id_hariç,
                    FaturaKalemleri.urun_id == urun_id
                ).group_by(Fatura.tip).first() # Fatura tipi tek olduğu için group_by uygun değil, direkt tip alınmalıydı.

                if fatura_kalem_data:
                    miktar_bu_faturada = float(fatura_kalem_data[0])
                    fatura_tipi_bu_faturada = fatura_kalem_data[1]

                    if fatura_tipi_bu_faturada == self.FATURA_TIP_SATIS or \
                       fatura_tipi_bu_faturada == self.FATURA_TIP_ALIS_IADE:
                        # Bu fatura stoku azaltmıştır, kontrol için geri ekle
                        current_stok += miktar_bu_faturada
                    elif fatura_tipi_bu_faturada == self.FATURA_TIP_ALIS or \
                         fatura_tipi_bu_faturada == self.FATURA_TIP_SATIS_IADE:
                        # Bu fatura stoku artırmıştır, kontrol için geri çıkar
                        current_stok -= miktar_bu_faturada
            return float(current_stok)
        except Exception as e:
            logging.error(f"Stok miktarı kontrolü sırasında hata: {e}\n{traceback.format_exc()}")
            return 0.0

    def stok_getir_for_fatura(self, fatura_tipi, arama_terimi=None):
        query = self._db_session.query(Stok)
        
        if arama_terimi:
            normalized_term = normalize_turkish_chars(arama_terimi)
            query = query.filter(
                or_(
                    Stok.urun_adi.ilike(f"%{normalized_term}%"),
                    Stok.urun_kodu.ilike(f"%{normalized_term}%")
                )
            )
        
        results = []
        stok_list = query.order_by(Stok.urun_adi).all()

        for s in stok_list:
            fiyat = 0.0
            if fatura_tipi == self.FATURA_TIP_SATIS or fatura_tipi == self.SIPARIS_TIP_SATIS:
                fiyat = float(s.satis_fiyati_kdv_dahil)
            elif fatura_tipi == self.FATURA_TIP_ALIS or fatura_tipi == self.SIPARIS_TIP_ALIS:
                fiyat = float(s.alis_fiyati_kdv_dahil)
            elif fatura_tipi == self.FATURA_TIP_SATIS_IADE: 
                fiyat = float(s.alis_fiyati_kdv_dahil) 
            elif fatura_tipi == self.FATURA_TIP_ALIS_IADE: 
                fiyat = float(s.alis_fiyati_kdv_dahil) 
            
            results.append({
                'id': s.id,
                'urun_kodu': s.urun_kodu,
                'urun_adi': s.urun_adi,
                'fiyat': fiyat, 
                'kdv_orani': float(s.kdv_orani),
                'stok': float(s.stok_miktari),
                'alis_fiyati_kdv_haric': float(s.alis_fiyati_kdv_haric),
                'alis_fiyati_kdv_dahil': float(s.alis_fiyati_kdv_dahil),
                'satis_fiyati_kdv_haric': float(s.satis_fiyati_kdv_haric),
                'satis_fiyati_kdv_dahil': float(s.satis_fiyati_kdv_dahil)
            })
        return results

    def _ensure_default_urun_birimi(self):
        try:
            db_birim = self._db_session.query(UrunBirimleri).filter(UrunBirimleri.birim_adi == "Adet").first()
            if not db_birim:
                new_birim = UrunBirimleri(
                    birim_adi="Adet",
                    olusturma_tarihi_saat=datetime.now(),
                    olusturan_kullanici_id=self._get_current_user_id()
                )
                self._db_session.add(new_birim)
                self._db_session.commit()
                self._db_session.refresh(new_birim)
                logging.info("Varsayılan 'Adet' ürün birimi oluşturuldu.")
                return new_birim.id
            return db_birim.id
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Varsayılan ürün birimi oluşturulurken hata: {e}\n{traceback.format_exc()}")
            return None

    def _ensure_default_ulke(self):
        try:
            db_ulke = self._db_session.query(UrunUlkeleri).filter(UrunUlkeleri.ulke_adi == "Türkiye").first()
            if not db_ulke:
                new_ulke = UrunUlkeleri(
                    ulke_adi="Türkiye",
                    olusturma_tarihi_saat=datetime.now(),
                    olusturan_kullanici_id=self._get_current_user_id()
                )
                self._db_session.add(new_ulke)
                self._db_session.commit()
                self._db_session.refresh(new_ulke)
                logging.info("Varsayılan 'Türkiye' ülkesi oluşturuldu.")
                return new_ulke.id
            return db_ulke.id
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Varsayılan ülke oluşturulurken hata: {e}\n{traceback.format_exc()}")
            return None

    def geriye_donuk_stok_hareketlerini_olustur(self):
        """
        Tüm mevcut faturaları tarar ve eksik olan stok hareketlerini oluşturur.
        Bu işlem, kendi geçici veritabanı oturumunu oluşturarak thread-safe hale getirilmiştir.
        """
        # OnMuhasebe sınıfı artık _db_session alıyor. Bu metodun içindeki
        # sqlite3.connect ve cursor kullanımlarını _db_session'a çevirmeliyiz.
        
        # Bu metod API tarafından çağrılacaksa (bir endpoint üzerinden),
        # o zaman orada zaten Session objesi sağlanır.
        # Eğer UI'dan direkt çağrılacaksa (main.py'de olduğu gibi),
        # o zaman _db_session kullanılacaktır.

        db_session_local_for_thread = None # Eğer UI'dan çağrılırsa ve ana session kullanılmazsa
        try:
            # Eğer self._db_session mevcut değilse (API dışı, bağımsız çağrı), kendi session'ımızı oluşturalım.
            # Normalde UI'dan _db_session zaten dolu gelir.
            if self._db_session is None or not self._db_session.is_active:
                db_session_local_for_thread = _SessionLocal() # Global _SessionLocal'ı kullan
                db_to_use = db_session_local_for_thread
            else:
                db_to_use = self._db_session

            # İç içe transaction başlat (Savepoint gibi çalışır)
            with db_to_use.begin_nested():
                # ÖNLEM: Mükerrer kayıt olmaması için mevcut fatura kaynaklı stok hareketlerini temizle.
                db_to_use.query(StokHareketleri).filter(
                    or_(
                        StokHareketleri.kaynak == self.KAYNAK_TIP_FATURA,
                        StokHareketleri.kaynak == self.KAYNAK_TIP_IADE_FATURA
                    )
                ).delete(synchronize_session=False) # Toplu silme için False
                logging.info("Mevcut fatura kaynaklı stok hareketleri temizlendi.")

                # Stok miktarlarını sıfırla (önceki hatalı hesaplamaları sıfırdan başlatmak için)
                # Toplu güncelleme için synchronize_session=False kullanılabilir.
                db_to_use.query(Stok).update({Stok.stok_miktari: 0.0}, synchronize_session=False)
                logging.info("Tüm ürünlerin stok miktarları sıfırlandı.")

                # Tüm faturaları ve kalemlerini çek (oluşturulma tarihine göre sıralı)
                fatura_kalemleri = db_to_use.query(
                    Fatura.id, Fatura.fatura_no, Fatura.tarih, Fatura.tip,
                    FaturaKalemleri.urun_id, FaturaKalemleri.miktar
                ).join(FaturaKalemleri, Fatura.id == FaturaKalemleri.fatura_id)\
                 .order_by(Fatura.tarih.asc(), Fatura.olusturma_tarihi_saat.asc(), Fatura.id.asc()).all()

                if not fatura_kalemleri:
                    db_to_use.commit() # Nested transaction'ı commit et (hiçbir şey yapılmadı)
                    return True, "İşlenecek fatura bulunamadı. Stok hareketi oluşturulmadı."

                hareket_sayisi = 0
                for fatura_id, fatura_no, tarih_obj, tip, urun_id, miktar in fatura_kalemleri:
                    tarih_str = tarih_obj.strftime('%Y-%m-%d') # Date objesini string'e çevir

                    # Stok hareketinin tipini ve miktar değişimini belirle
                    islem_tipi_hareket = ""
                    miktar_degisimi_net = 0.0
                    kaynak_tipi_hareket = self.KAYNAK_TIP_FATURA

                    if tip == self.FATURA_TIP_SATIS:
                        islem_tipi_hareket = self.STOK_ISLEM_TIP_FATURA_SATIS
                        miktar_degisimi_net = -float(miktar) # Satışta stok azalır
                    elif tip == self.FATURA_TIP_ALIS:
                        islem_tipi_hareket = self.STOK_ISLEM_TIP_FATURA_ALIS
                        miktar_degisimi_net = float(miktar) # Alışta stok artar
                    elif tip == self.FATURA_TIP_SATIS_IADE:
                        islem_tipi_hareket = self.STOK_ISLEM_TIP_FATURA_SATIS_IADE
                        miktar_degisimi_net = float(miktar) # Satış iadesinde stok artar
                        kaynak_tipi_hareket = self.KAYNAK_TIP_IADE_FATURA
                    elif tip == self.FATURA_TIP_ALIS_IADE:
                        islem_tipi_hareket = self.STOK_ISLEM_TIP_FATURA_ALIS_IADE
                        miktar_degisimi_net = -float(miktar) # Alış iadesinde stok azalır
                        kaynak_tipi_hareket = self.KAYNAK_TIP_IADE_FATURA
                    elif tip == self.FATURA_TIP_DEVIR_GIRIS:
                        islem_tipi_hareket = self.STOK_ISLEM_TIP_DEVIR_GIRIS
                        miktar_degisimi_net = float(miktar)
                        kaynak_tipi_hareket = self.KAYNAK_TIP_FATURA

                    # Ürünün mevcut stok miktarını al (sıfırlandıktan sonraki anlık miktar)
                    urun_obj_current = db_to_use.query(Stok).filter(Stok.id == urun_id).first()
                    if not urun_obj_current:
                        logging.warning(f"Ürün ID {urun_id} bulunamadı, stok hareketi oluşturulamadı.")
                        continue # Bu kalemi atla

                    onceki_stok = float(urun_obj_current.stok_miktari)
                    sonraki_stok = onceki_stok + miktar_degisimi_net

                    # Stok miktarını güncelle
                    urun_obj_current.stok_miktari = sonraki_stok

                    # Stok hareketini kaydet
                    yeni_hareket = StokHareketleri(
                        urun_id=urun_id,
                        tarih=tarih_obj,
                        islem_tipi=islem_tipi_hareket,
                        miktar=abs(miktar_degisimi_net),
                        onceki_stok=onceki_stok,
                        sonraki_stok=sonraki_stok,
                        aciklama=f"Geçmiş Fatura No: {fatura_no}",
                        kaynak=kaynak_tipi_hareket,
                        kaynak_id=fatura_id,
                        kaynak_no=fatura_no,
                        olusturma_tarihi_saat=datetime.now(),
                        olusturan_kullanici_id=self._get_current_user_id()
                    )
                    db_to_use.add(yeni_hareket)
                    hareket_sayisi += 1

                db_to_use.commit() # Nested transaction'ı commit et
                return True, f"Geçmişe dönük {hareket_sayisi} adet stok hareketi başarıyla oluşturuldu ve stoklar yeniden hesaplandı."

        except Exception as e:
            if db_session_local_for_thread: # Kendi oluşturduğumuz session'ı rollback et
                db_session_local_for_thread.rollback()
            elif self._db_session: # Dışarıdan gelen session'da nested transaction'ı rollback et
                self._db_session.rollback()
            error_details = traceback.format_exc()
            logging.error(f"Geçmiş stok hareketleri oluşturulurken hata: {e}\nDetaylar: {error_details}")
            return False, f"Geçmiş stok hareketleri oluşturulurken hata: {e}\n{error_details}"
        finally:
            if db_session_local_for_thread: # Kendi oluşturduğumuz session'ı kapat
                db_session_local_for_thread.close()

    def manuel_stok_hareketi_sil(self, hareket_id):
        try:
            db_hareket = self._db_session.query(StokHareketleri).filter(StokHareketleri.id == hareket_id).first()
            if db_hareket is None:
                return False, "Stok hareketi bulunamadı."
            
            if db_hareket.kaynak != self.KAYNAK_TIP_MANUEL:
                return False, "Sadece manuel stok hareketleri silinebilir."

            self._db_session.begin_nested() 

            # Stok miktarını tersine çevir
            urun_id = db_hareket.urun_id
            # Miktar değişimini tersine çeviriyoruz (örn: +5 olanı -5 yapar)
            # islem_tipi içinde 'Giriş' veya 'Fazlası' varsa miktar pozitif demektir, tersi negatif olmalı.
            miktar_degisimi_net_tersi = -float(db_hareket.miktar) if "Giriş" in db_hareket.islem_tipi or "Fazlası" in db_hareket.islem_tipi or "Alış" in db_hareket.islem_tipi else float(db_hareket.miktar)

            success, message = self._stok_guncelle_ve_hareket_kaydet(
                urun_id=urun_id,
                miktar_degisimi_net=miktar_degisimi_net_tersi,
                islem_tipi_aciklamasi=f"{db_hareket.islem_tipi} (Silme Geri Alma)",
                kaynak_tipi="MANUEL_SILME_GERI_ALMA",
                kaynak_id=db_hareket.id,
                referans_no=f"Orijinal Hareket ID: {db_hareket.id}"
            )
            if not success:
                self._db_session.rollback()
                return False, f"Stok hareketi geri alınamadı: {message}"

            self._db_session.delete(db_hareket)
            self._db_session.commit()
            return True, "Stok hareketi başarıyla silindi ve stok güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Manuel stok hareketi silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Manuel stok hareketi silinirken bir hata oluştu: {e}"

    def get_fatura_servisi(self):
        """
        API endpoint'lerinin FaturaService'e kolayca erişmesini sağlayan yardımcı metot.
        """
        if not hasattr(self, '_fatura_servisi_instance'):
            # FaturaService'i hizmetler.py dosyasından import ettiğinizden emin olun
            from hizmetler import FaturaService
            self._fatura_servisi_instance = FaturaService(self)
        return self._fatura_servisi_instance

    def get_cari_bakiye_snapshot(self, cari_id, cari_tip, tarih_str):
        """
        Belirli bir tarihteki cari bakiyesini hesaplar.
        Bu metod, bir faturanın kesildiği tarihteki anlık bakiye durumunu almak için kullanılır.
        tarih_str: 'YYYY-MM-DD' formatında tarih.
        Dönüş: {'onceki_bakiye': float, 'bugun_odenen': float, 'kalan_borc': float}
        """
        onceki_bakiye = 0.0 # Belirtilen tarihten önceki bakiyeler için
        bugun_odenen_tahsil_edilen = 0.0 # Sadece fatura tarihiyle aynı günkü tahsilat/ödeme hareketleri

        try:
            # 1. Adım: Önceki Bakiye - Belirtilen tarihten önceki tüm hareketlerin net toplamı
            # Bu sorgu, 'tarih_str' ile belirtilen günden önceki tüm işlemleri kapsar.
            query_onceki = """
                SELECT islem_tipi, tutar, referans_tip FROM cari_hareketler
                WHERE cari_id = ? AND cari_tip = ? AND tarih < ?
            """
            self.c.execute(query_onceki, (cari_id, cari_tip, tarih_str))
            onceki_hareketler = self.c.fetchall()

            for h in onceki_hareketler:
                tutar_h = h['tutar'] or 0.0
                islem_tipi_h = h['islem_tipi']
                referans_tip_h = h['referans_tip']

                if cari_tip == self.CARI_TIP_MUSTERI:
                    # Müşteri bize borçlandı (bizim alacağımız)
                    if islem_tipi_h == self.ISLEM_TIP_ALACAK or referans_tip_h == self.KAYNAK_TIP_FATURA or referans_tip_h == self.KAYNAK_TIP_VERESIYE_BORC_MANUEL:
                        onceki_bakiye += tutar_h
                    # Müşteriden tahsilat veya satış iadesi
                    elif islem_tipi_h == self.ISLEM_TIP_TAHSILAT or referans_tip_h == self.KAYNAK_TIP_FATURA_SATIS_PESIN or referans_tip_h == self.KAYNAK_TIP_IADE_FATURA:
                        onceki_bakiye -= tutar_h
                elif cari_tip == self.CARI_TIP_TEDARIKCI:
                    # Biz tedarikçiye borçlandık (bizim borcumuz)
                    if islem_tipi_h == self.ISLEM_TIP_BORC or referans_tip_h == self.KAYNAK_TIP_FATURA or referans_tip_h == self.KAYNAK_TIP_VERESIYE_BORC_MANUEL:
                        onceki_bakiye += tutar_h
                    # Tedarikçiye ödeme veya alış iadesi
                    elif islem_tipi_h == self.ISLEM_TIP_ODEME or referans_tip_h == self.KAYNAK_TIP_FATURA_ALIS_PESIN or referans_tip_h == self.KAYNAK_TIP_IADE_FATURA:
                        onceki_bakiye -= tutar_h

            # 2. Adım: Bugün Yapılan Ödeme/Tahsilat - Sadece fatura tarihiyle AYNI GÜNE ait olan tahsilat/ödeme hareketleri
            # Bu sorgu, 'tarih_str' ile belirtilen GÜNDEKİ tüm işlemleri kapsar.
            query_bugun = """
                SELECT islem_tipi, tutar, referans_tip FROM cari_hareketler
                WHERE cari_id = ? AND cari_tip = ? AND tarih = ?
            """
            self.c.execute(query_bugun, (cari_id, cari_tip, tarih_str))
            bugun_hareketler = self.c.fetchall()

            # Kalan borç/alacak: Önceki bakiye + (bugünkü tüm ALACAK/BORÇ hareketleri) - (bugünkü tüm TAHSİLAT/ÖDEME hareketleri)
            kalan_borc_anlik = onceki_bakiye
            for h in bugun_hareketler:
                tutar_h = h['tutar'] or 0.0
                islem_tipi_h = h['islem_tipi']
                referans_tip_h = h['referans_tip']

                if cari_tip == self.CARI_TIP_MUSTERI:
                    if islem_tipi_h == self.ISLEM_TIP_ALACAK or referans_tip_h == self.KAYNAK_TIP_FATURA or referans_tip_h == self.KAYNAK_TIP_VERESIYE_BORC_MANUEL:
                        kalan_borc_anlik += tutar_h
                    elif islem_tipi_h == self.ISLEM_TIP_TAHSILAT or referans_tip_h == self.KAYNAK_TIP_FATURA_SATIS_PESIN or referans_tip_h == self.KAYNAK_TIP_IADE_FATURA:
                        kalan_borc_anlik -= tutar_h
                        # Bugüne ait tahsilatları toplama
                        if islem_tipi_h == self.ISLEM_TIP_TAHSILAT or referans_tip_h == self.KAYNAK_TIP_FATURA_SATIS_PESIN:
                            bugun_odenen_tahsil_edilen += tutar_h

                elif cari_tip == self.CARI_TIP_TEDARIKCI:
                    if islem_tipi_h == self.ISLEM_TIP_BORC or referans_tip_h == self.KAYNAK_TIP_FATURA or referans_tip_h == self.KAYNAK_TIP_VERESIYE_BORC_MANUEL:
                        kalan_borc_anlik += tutar_h
                    elif islem_tipi_h == self.ISLEM_TIP_ODEME or referans_tip_h == self.KAYNAK_TIP_FATURA_ALIS_PESIN or referans_tip_h == self.KAYNAK_TIP_IADE_FATURA:
                        kalan_borc_anlik -= tutar_h
                        # Bugüne ait ödemeleri toplama
                        if islem_tipi_h == self.ISLEM_TIP_ODEME or referans_tip_h == self.KAYNAK_TIP_FATURA_ALIS_PESIN:
                            bugun_odenen_tahsil_edilen += tutar_h

            return {
                'onceki_bakiye': onceki_bakiye,
                'bugun_odenen': bugun_odenen_tahsil_edilen,
                'kalan_borc': kalan_borc_anlik
            }
        except Exception as e:
            import logging
            import traceback
            logging.error(f"Cari bakiye anlık görüntü hatası: {e}\n{traceback.format_exc()}")
            return {
                'onceki_bakiye': 0.0,
                'bugun_odenen': 0.0,
                'kalan_borc': 0.0
            }

    def kasa_banka_hareket_ekle(self, kasa_banka_id, tutar, aciklama, referans_id):
        """
        Kasa/Banka hareketini gelir_gider tablosuna kaydeder ve kasa/banka bakiyesini günceller.
        Bu metot, dış bir transaction (örneğin fatura_olustur) içinde çağrılmalıdır.
        Kendi içinde BEGIN/COMMIT yapmaz.
        """
        if kasa_banka_id is None:
            logging.warning(f"Kasa/Banka ID boş olduğu için hareket kaydedilemedi: {aciklama}")
            return False, "Kasa/Banka hesabı belirtilmediği için hareket kaydedilemedi."

        current_time = self.get_current_datetime_str()
        olusturan_id = self._get_current_user_id() 

        try:
            # Gelir/Gider tablosuna kaydı ekle
            gg_tip = self.ISLEM_TIP_GELIR if tutar >= 0 else self.ISLEM_TIP_GIDER 
            
            self.c.execute("""
                INSERT INTO gelir_gider (
                    tarih, tip, tutar, aciklama, kaynak, kaynak_id, kasa_banka_id,
                    olusturma_tarihi_saat, olusturan_kullanici_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().strftime('%Y-%m-%d'),
                gg_tip,
                abs(tutar), 
                aciklama,
                'KASA_BANKA_HAREKET', # Bu kaydın kaynağının Kasa/Banka hareketi olduğunu belirt
                referans_id, 
                kasa_banka_id,
                current_time,
                olusturan_id
            ))
            
            # Kasa/Banka bakiyesini güncelle
            is_bakiye_artir = (tutar >= 0)
            # kasa_banka_bakiye_guncelle metodu artık kendi transaction'ını yönetmiyor.
            self.kasa_banka_bakiye_guncelle(kasa_banka_id, abs(tutar), artir=is_bakiye_artir)
            
            return True, "Kasa/Banka hareketi ve bakiye başarıyla güncellendi."
        except Exception as e:
            # Bu hata durumunda dış transaction tarafından rollback yapılacağı için burada ayrıca rollback yapmıyoruz.
            logging.error(f"Kasa/Banka hareketi eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Kasa/Banka hareketi eklenirken bir hata oluştu: {e}"

    def _get_current_user_id(self):
        # Bu metod App sınıfından mevcut kullanıcı ID'sini almalı.
        # Geçici olarak 1 (admin) dönelim, veya App objesi üzerinden alalım.
        if self.app and hasattr(self.app, 'current_user') and self.app.current_user:
            return self.app.current_user[0]
        return 1 # Default Admin ID (Üretimde daha güvenli bir mekanizma olmalı)
       
    def clear_log_file(self):
        try:
            log_file_path = os.path.join(self.data_dir, 'application.log')
            if os.path.exists(log_file_path):
                with open(log_file_path, 'w') as f:
                    f.truncate(0) # Dosyanın içeriğini sıfırlar
                return True, "Log dosyası başarıyla sıfırlandı."
            else:
                return True, "Log dosyası bulunamadı, sıfırlama gerekmiyor."
        except Exception as e:
            logging.error(f"Log dosyası sıfırlanırken hata oluştu: {e}")
            return False, f"Log dosyası sıfırlanırken hata oluştu: {e}"

    def get_monthly_sales_summary(self, baslangic_tarih, bitis_tarih):
        """Belirtilen tarih aralığındaki aylık satış toplamlarını döndürür."""
        query = """
            SELECT
                STRFTIME('%Y-%m', tarih) AS ay,
                SUM(toplam_kdv_dahil) AS toplam_satis
            FROM faturalar
            WHERE tip = 'SATIŞ' AND tarih BETWEEN ? AND ?
            GROUP BY ay
            ORDER BY ay;
        """
        self.c.execute(query, (baslangic_tarih, bitis_tarih))
        return self.c.fetchall()

    def get_monthly_income_expense_summary(self, baslangic_tarih, bitis_tarih):
        """Belirtilen tarih aralığındaki aylık toplam gelir ve giderleri döndürür."""
        query = """
            SELECT
                STRFTIME('%Y-%m', tarih) AS ay,
                SUM(CASE WHEN tip = 'GELİR' THEN tutar ELSE 0 END) AS toplam_gelir,
                SUM(CASE WHEN tip = 'GİDER' THEN tutar ELSE 0 END) AS toplam_gider
            FROM gelir_gider
            WHERE tarih BETWEEN ? AND ?
            GROUP BY ay
            ORDER BY ay;
        """
        self.c.execute(query, (baslangic_tarih, bitis_tarih))
        return self.c.fetchall()

    def get_monthly_gross_profit_summary(self, baslangic_tarih, bitis_tarih):
        """
        Belirtilen tarih aralığındaki aylık brüt kâr ve satılan malın maliyetini döndürür.
        """
        query = """
            SELECT
                STRFTIME('%Y-%m', f.tarih) AS ay,
                SUM(f.toplam_kdv_dahil) AS toplam_satis_geliri,
                SUM(fk.miktar * fk.alis_fiyati_fatura_aninda) AS toplam_alis_maliyeti
            FROM faturalar f
            JOIN fatura_kalemleri fk ON f.id = fk.fatura_id
            WHERE f.tip = 'SATIŞ' AND f.tarih BETWEEN ? AND ?
            GROUP BY ay
            ORDER BY ay;
        """
        self.c.execute(query, (baslangic_tarih, bitis_tarih))
        return self.c.fetchall()

    def get_monthly_cash_flow_summary(self, baslangic_tarih, bitis_tarih):
        """
        Belirtilen tarih aralığındaki aylık toplam nakit girişi ve çıkışını (kasa/banka ile ilişkili) döndürür.
        """
        query = """
            SELECT
                STRFTIME('%Y-%m', gg.tarih) AS ay,
                SUM(CASE WHEN gg.tip = 'GELİR' THEN gg.tutar ELSE 0 END) AS toplam_nakit_giris,
                SUM(CASE WHEN gg.tip = 'GİDER' THEN gg.tutar ELSE 0 END) AS toplam_nakit_cikis
            FROM gelir_gider gg
            WHERE gg.kasa_banka_id IS NOT NULL AND gg.tarih BETWEEN ? AND ?
            GROUP BY ay
            ORDER BY ay;
        """
        self.c.execute(query, (baslangic_tarih, bitis_tarih))
        return self.c.fetchall()

    def get_top_selling_product_of_month(self):
        """
        Mevcut ayda en çok satılan ürünü (miktara göre) döndürür.
        Dönüş: (urun_adi, toplam_miktar) veya None
        """
        current_month_start = datetime.now().strftime('%Y-%m-01')
        current_month_end_day = calendar.monthrange(datetime.now().year, datetime.now().month)[1]
        current_month_end = datetime.now().strftime(f'%Y-%m-{current_month_end_day}')

        query = """
            SELECT 
                s.urun_adi, 
                SUM(fk.miktar) AS toplam_miktar
            FROM fatura_kalemleri fk
            JOIN faturalar f ON fk.fatura_id = f.id
            JOIN tbl_stoklar s ON fk.urun_id = s.id
            WHERE f.tip = 'SATIŞ' AND f.tarih BETWEEN ? AND ?
            GROUP BY s.urun_adi
            ORDER BY toplam_miktar DESC
            LIMIT 1
        """
        self.c.execute(query, (current_month_start, current_month_end))
        result = self.c.fetchone()
        return result if result else None

    def get_today_transaction_summary(self):
        """
        Bugün yapılan tüm önemli hareketlerin (satış/alış faturaları, tahsilatlar, ödemeler) özetini döndürür.
        Dönüş: {'toplam_satis_tutari': float, 'toplam_alis_tutari': float, 'toplam_tahsilat_tutari': float, 'toplam_odeme_tutari': float, 'toplam_fatura_sayisi': int, 'toplam_cari_hareket_sayisi': int}
        """
        today_str = datetime.now().strftime('%Y-%m-%d')
        
        # Bugün kesilen satış ve alış faturalarının toplam tutarı ve sayısı
        self.c.execute("SELECT SUM(toplam_kdv_dahil) FROM faturalar WHERE tarih = ? AND tip = ?", (today_str, self.FATURA_TIP_SATIS))
        toplam_satis_tutari = self.c.fetchone()[0] or 0.0

        self.c.execute("SELECT SUM(toplam_kdv_dahil) FROM faturalar WHERE tarih = ? AND tip = ?", (today_str, self.FATURA_TIP_ALIS))
        toplam_alis_tutari = self.c.fetchone()[0] or 0.0

        self.c.execute("SELECT COUNT(id) FROM faturalar WHERE tarih = ? AND tip IN (?, ?)", (today_str, self.FATURA_TIP_SATIS, self.FATURA_TIP_ALIS))
        toplam_fatura_sayisi = self.c.fetchone()[0] or 0

        # Bugün yapılan tüm tahsilatların toplam tutarı (manuel veya peşin fatura kaynaklı)
        self.c.execute("SELECT SUM(tutar) FROM cari_hareketler WHERE tarih = ? AND islem_tipi = ?", (today_str, self.ISLEM_TIP_TAHSILAT))
        toplam_tahsilat_tutari = self.c.fetchone()[0] or 0.0

        # Bugün yapılan tüm ödemelerin toplam tutarı (manuel veya peşin fatura kaynaklı)
        self.c.execute("SELECT SUM(tutar) FROM cari_hareketler WHERE tarih = ? AND islem_tipi = ?", (today_str, self.ISLEM_TIP_ODEME))
        toplam_odeme_tutari = self.c.fetchone()[0] or 0.0

        # Bugün yapılan tüm cari hareketlerin (faturalar, tahsilatlar, ödemeler, veresiye borçlar) sayısı
        self.c.execute("SELECT COUNT(id) FROM cari_hareketler WHERE tarih = ?", (today_str,))
        toplam_cari_hareket_sayisi = self.c.fetchone()[0] or 0


        return {
            'toplam_satis_tutari': toplam_satis_tutari,
            'toplam_alis_tutari': toplam_alis_tutari, # Alışları da ekledim
            'toplam_tahsilat_tutari': toplam_tahsilat_tutari,
            'toplam_odeme_tutari': toplam_odeme_tutari,
            'toplam_fatura_sayisi': toplam_fatura_sayisi,
            'toplam_cari_hareket_sayisi': toplam_cari_hareket_sayisi
        }

    def get_top_selling_products(self, baslangic_tarih, bitis_tarih, limit=5):
        """
        Belirtilen tarih aralığında en çok satılan ürünleri (miktara göre) döndürür.
        """
        query = """
            SELECT
                s.urun_adi,
                SUM(fk.miktar) AS toplam_miktar
            FROM faturalar f
            JOIN fatura_kalemleri fk ON f.id = fk.fatura_id
            JOIN tbl_stoklar s ON fk.urun_id = s.id
            WHERE f.tip = 'SATIŞ' AND f.tarih BETWEEN ? AND ?
            GROUP BY s.urun_adi
            ORDER BY toplam_miktar DESC
            LIMIT ?;
        """
        self.c.execute(query, (baslangic_tarih, bitis_tarih, limit))
        return self.c.fetchall()

    def get_stock_value_by_category(self):
        """
        Mevcut stoktaki ürünlerin kategori bazında toplam KDV dahil alış değerini döndürür.
        """
        query = """
            SELECT
                uk.kategori_adi,
                SUM(s.stok_miktari * s.alis_fiyati_kdv_dahil) AS toplam_deger
            FROM tbl_stoklar s
            LEFT JOIN urun_kategorileri uk ON s.kategori_id = uk.id
            GROUP BY uk.kategori_adi
            HAVING SUM(s.stok_miktari * s.alis_fiyati_kdv_dahil) > 0 -- Değeri olan kategorileri al
            ORDER BY toplam_deger DESC;
        """
        self.c.execute(query)
        return self.c.fetchall()


    def son_fatura_no_getir(self, fatura_tipi):
        """
        Belirtilen fatura tipi için benzersiz bir fatura numarası oluşturur.
        Format: TIP-YYYYAA_HHMMSSms (örn: SAT-20250703150649123456)
        """
        prefix = ""
        if fatura_tipi == 'SATIŞ':
            prefix = 'SAT'
        elif fatura_tipi == 'ALIŞ':
            prefix = 'AL'
        else: # Diğer tipler için varsayılan
            prefix = 'FAT'
        
        # YYYYMMDDHHMMSSms formatında benzersiz bir zaman damgası kullanarak fatura numarası oluştur
        return f"{prefix}-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"

    def get_son_fatura_kalemi_bilgisi(self, cari_id, urun_id, fatura_tipi):
        """
        Belirli bir cari (müşteri/tedarikçi) ve ürün için son fatura kalemi bilgilerini döndürür.
        Bu, son fiyatlandırma ve iskonto bilgilerini hatırlamak için kullanılır.
        Dönüş: (birim_fiyat, kdv_orani, iskonto_yuzde_1, iskonto_yuzde_2) veya None
        Not: birim_fiyat burada iskonto uygulanmış KDV Dahil Birim Fiyatıdır.
        """
        query = """
            SELECT
                fk.birim_fiyat,           -- Orijinal KDV Hariç Birim Fiyatı (3)
                fk.kdv_orani,             -- KDV Oranı (4)
                fk.iskonto_yuzde_1,       -- İskonto Yüzde 1 (10)
                fk.iskonto_yuzde_2,       -- İskonto Yüzde 2 (11)
                f.tarih                   -- Fatura Tarihi (Sıralama için)
            FROM fatura_kalemleri fk
            JOIN faturalar f ON fk.fatura_id = f.id
            WHERE f.cari_id = ? AND fk.urun_id = ? AND f.tip = ?
            ORDER BY f.tarih DESC, f.id DESC -- En yeni faturadan başla
            LIMIT 1
        """
        params = (cari_id, urun_id, fatura_tipi)
        self.c.execute(query, params)
        result = self.c.fetchone()

        if result:
            # birim_fiyat (original_kdv_haric_bf) , kdv_orani, iskonto_yuzde_1, iskonto_yuzde_2
            original_kdv_haric_bf = result[0]
            kdv_orani = result[1]
            iskonto_yuzde_1 = result[2]
            iskonto_yuzde_2 = result[3]

            # Fatura oluşturma ekranında Birim Fiyat (KDV Dahil) gösterildiği için
            # ve bu fiyat aynı zamanda iskontoları da içereceği için,
            # burada iskonto uygulanmış KDV Dahil birim fiyatını hesaplayıp dönelim.
            
            # 1. Orijinal (iskontosuz) KDV Dahil Birim Fiyatı
            original_kdv_dahil_bf = original_kdv_haric_bf * (1 + kdv_orani / 100)

            # 2. İskonto 1'i uygula
            fiyat_iskonto_1_sonrasi_dahil = original_kdv_dahil_bf * (1 - iskonto_yuzde_1 / 100)
            
            # 3. İskonto 2'yi uygula
            nihai_iskontolu_kdv_dahil_bf = fiyat_iskonto_1_sonrasi_dahil * (1 - iskonto_yuzde_2 / 100)

            return (nihai_iskontolu_kdv_dahil_bf, kdv_orani, iskonto_yuzde_1, iskonto_yuzde_2)
        else:
            return None # Kayıt bulunamazsa None döndür

    def get_gecmis_fatura_kalemi_bilgileri(self, cari_id, urun_id, fatura_tipi, limit=5):
        """
        Belirli bir cari (müşteri/tedarikçi) ve ürün için geçmiş fatura kalemi bilgilerini döndürür.
        Bu, "Fiyat Geçmişi" butonu pop-up'ı için kullanılır.
        Dönüş: [(fatura_id, fatura_no, fatura_tarihi, iskontolu_kdv_dahil_bf, iskonto_yuzde_1, iskonto_yuzde_2), ...]
        Not: birim_fiyat burada iskonto uygulanmış KDV Dahil Birim Fiyatıdır.
        """
        query = """
            SELECT
                    f.id,                      -- Fatura ID
                f.fatura_no,               -- Fatura Numarası
                f.tarih,                   -- Fatura Tarihi
                fk.birim_fiyat,            -- Orijinal KDV Hariç Birim Fiyatı
                fk.kdv_orani,              -- KDV Oranı
                fk.iskonto_yuzde_1,        -- İskonto Yüzde 1
                fk.iskonto_yuzde_2         -- İskonto Yüzde 2
            FROM fatura_kalemleri fk
            JOIN faturalar f ON fk.fatura_id = f.id
            WHERE f.cari_id = ? AND fk.urun_id = ? AND f.tip = ?
            ORDER BY f.tarih DESC, f.id DESC -- En yeni faturalardan başla
            LIMIT ?
        """
        params = (cari_id, urun_id, fatura_tipi, limit)
        self.c.execute(query, params)
        raw_results = self.c.fetchall()

        formatted_results = []
        for result in raw_results:
            fatura_id = result['id']
            fatura_no = result['fatura_no']
            fatura_tarihi_obj = result['tarih'] # Veritabanından gelen tarih nesnesi
            original_kdv_haric_bf = result['birim_fiyat']
            kdv_orani = result['kdv_orani']
            iskonto_yuzde_1 = result['iskonto_yuzde_1']
            iskonto_yuzde_2 = result['iskonto_yuzde_2']

            # Fatura oluşturma ekranında Birim Fiyat (KDV Dahil) gösterildiği için
            # ve bu fiyat aynı zamanda iskontoları da içereceği için,
            # burada iskonto uygulanmış KDV Dahil birim fiyatını hesaplayıp dönelim.
        
            # 1. Orijinal (iskontosuz) KDV Dahil Birim Fiyatı
            original_kdv_dahil_bf = original_kdv_haric_bf * (1 + kdv_orani / 100)

            # 2. İskonto 1'i uygula
            fiyat_iskonto_1_sonrasi_dahil = original_kdv_dahil_bf * (1 - iskonto_yuzde_1 / 100)

            # 3. İskonto 2'yi uygula
            nihai_iskontolu_kdv_dahil_bf = fiyat_iskonto_1_sonrasi_dahil * (1 - iskonto_yuzde_2 / 100)

            # ### HATA DÜZELTMESİ BURADA ###
            # Gelen veri zaten bir tarih nesnesi olduğu için strptime kullanmıyoruz.
            # Doğrudan strftime ile formatlıyoruz.
            if isinstance(fatura_tarihi_obj, (datetime, date)):
                formatted_date = fatura_tarihi_obj.strftime('%d.%m.%Y')
            else:
                # Beklenmedik bir durum (veri string ise veya None ise) için yedek kontrol
                formatted_date = str(fatura_tarihi_obj)

            formatted_results.append((
                fatura_id,
                fatura_no,
                formatted_date,
                nihai_iskontolu_kdv_dahil_bf,
                iskonto_yuzde_1,
                iskonto_yuzde_2
            ))
        return formatted_results

    def get_recent_cari_hareketleri(self, cari_tip, cari_id, limit=10):
        """
        Belirli bir cari hesabın son işlemlerini döndürür.
        """
        query = """
            SELECT 
                ch.tarih, 
                ch.islem_tipi, 
                ch.tutar, 
                ch.aciklama, 
                kb.hesap_adi 
            FROM cari_hareketler ch
            LEFT JOIN kasalar_bankalar kb ON ch.kasa_banka_id = kb.id
            WHERE ch.cari_tip = ? AND ch.cari_id = ?
            ORDER BY ch.tarih DESC, ch.olusturma_tarihi_saat DESC
            LIMIT ?
        """
        self.c.execute(query, (cari_tip, cari_id, limit))
        return self.c.fetchall()

    def get_current_datetime_str(self):
        """Geçerli tarih ve saati 'YYYY-AA-GG HH:MM:SS' formatında döndürür."""
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def get_current_date_str(self):
        """Mevcut tarihi YYYY-AA-GG formatında döndürür."""
        return date.today().strftime('%Y-%m-%d')

    def _get_config_path(self):
        return os.path.join(data_dir, 'config.json')

    def _log_audit_action(self, action_type, table_name, record_id, details):
        """
        Veritabanı üzerinde yapılan önemli eylemleri denetim kaydına (audit log) kaydeder.
        Şimdilik sadece konsola yazdırıyoruz. Gerçek bir sistemde ayrı bir tabloya kaydedilmelidir.
        """
        # self.app'in var olup olmadığını kontrol edin
        user_info = self.app.current_user if self.app and hasattr(self.app, 'current_user') and self.app.current_user else (None, "Sistem", "Bilinmeyen")
        user_id = user_info[0]
        username = user_info[1]

        log_message = f"AUDIT: [{self.get_current_datetime_str()}] User: {username} (ID: {user_id}) | " \
                      f"Action: {action_type} | Table: {table_name} | Record ID: {record_id} | Details: {details}"
        print(log_message)
        # TODO: İleride denetim kaydını ayrı bir veritabanı tablosuna kaydetmek için burayı geliştirin.

    def get_cari_genel_bakiyeler(self, cari_id, cari_tip):
        """
        Bir cari hesabın tüm zamanlara ait toplam alacak, borç, tahsilat ve ödeme tutarlarını döndürür.
        `toplam_alacak` = Müşterinin bize olan borcu (bizim alacağımız).
        `toplam_borc` = Bizim cariye olan borcumuz (bizim borcumuz).
        `toplam_tahsilat` = Müşteriden alınan tahsilatların toplamı.
        `toplam_odeme` = Tedarikçiye yapılan ödemelerin toplamı.
        """
        # Müşterinin bize olan borcu (Bizim alacağımız)
        # MUSTERI: islem_tipi = 'ALACAK' veya referans_tip = 'FATURA' (SATIŞ) veya 'VERESIYE_BORC_MANUEL' (Müşteriden)
        query_musteri_alacaklari_bize = """
            SELECT SUM(tutar) FROM cari_hareketler
            WHERE cari_id = ? AND cari_tip = 'MUSTERI' 
            AND (islem_tipi = 'ALACAK' OR referans_tip = 'FATURA' OR referans_tip = 'VERESIYE_BORC_MANUEL')
        """
        self.c.execute(query_musteri_alacaklari_bize, (cari_id,))
        musteri_borcu_bize = self.c.fetchone()[0] or 0.0

        # Müşteriden gelen tahsilatlar (bizim tahsilatımız)
        # MUSTERI: islem_tipi = 'TAHSILAT' veya referans_tip = 'FATURA_SATIS_PESIN'
        query_musteri_tahsilatlari = """
            SELECT SUM(tutar) FROM cari_hareketler
            WHERE cari_id = ? AND cari_tip = 'MUSTERI' 
            AND (islem_tipi = 'TAHSILAT' OR referans_tip = 'FATURA_SATIS_PESIN')
        """
        self.c.execute(query_musteri_tahsilatlari, (cari_id,))
        musteri_tahsilati = self.c.fetchone()[0] or 0.0

        # Tedarikçiye olan bizim borcumuz
        # TEDARIKCI: islem_tipi = 'BORC' veya referans_tip = 'FATURA' (ALIŞ) veya 'VERESIYE_BORC_MANUEL' (Tedarikçiden)
        query_tedarikci_borclari_bize = """
            SELECT SUM(tutar) FROM cari_hareketler
            WHERE cari_id = ? AND cari_tip = 'TEDARIKCI' 
            AND (islem_tipi = 'BORC' OR referans_tip = 'FATURA' OR referans_tip = 'VERESIYE_BORC_MANUEL')
        """
        self.c.execute(query_tedarikci_borclari_bize, (cari_id,))
        tedarikci_borcu_bize = self.c.fetchone()[0] or 0.0

        # Tedarikçiye yapılan ödemeler
        # TEDARIKCI: islem_tipi = 'ODEME' veya referans_tip = 'FATURA_ALIS_PESIN'
        query_tedarikci_odemeleri = """
            SELECT SUM(tutar) FROM cari_hareketler
            WHERE cari_id = ? AND cari_tip = 'TEDARIKCI' 
            AND (islem_tipi = 'ODEME' OR referans_tip = 'FATURA_ALIS_PESIN')
        """
        self.c.execute(query_tedarikci_odemeleri, (cari_id,))
        tedarikci_odemesi = self.c.fetchone()[0] or 0.0

        # Nihai dönüş değerleri:
        # `toplam_alacak` = Müşterinin bize olan NET borcu (bizim genel alacağımız)
        # `toplam_borc` = Bizim tedarikçiye olan NET borcumuz (bizim genel borcumuz)
        # BU HESAPLAMA DOĞRU, MUSTERI VE TEDARIKCI İÇİN AYRI HESAPLANIYOR
        # RETURN DEĞERLERİNDE: musteri_borcu_bize - musteri_tahsilati KULLANILMIŞ, bu zaten NET alacak.
        # tedarikci_borcu_bize - tedarikci_odemesi KULLANILMIŞ, bu zaten NET borç.
        return musteri_borcu_bize, tedarikci_borcu_bize, musteri_tahsilati, tedarikci_odemesi # musterinin_borcu_bize yerine toplam_alacak kullanıldı.

    def load_config(self):
        config_path = self._get_config_path()
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except json.JSONDecodeError as e:
                print(f"Hata: Config dosyası bozuk veya boş. Yeni dosya oluşturulacak. Detay: {e}")
                return {}
            except Exception as e:
                print(f"Hata: Config dosyası okunurken beklenmeyen bir hata oluştu: {e}")
                return {}
        return {}

    def save_config(self, config_data):
        config_path = self._get_config_path()
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, indent=4)
        except Exception as e:
            print(f"Hata: Config dosyası yazılırken hata oluştu: {e}")


    def get_critical_stock_items(self):
        """
        Minimum stok seviyesinin altında olan ürünleri döndürür.
        (id, urun_kodu, urun_adi, stok_miktari, alis_fiyati, satis_fiyati, kdv_orani, min_stok_seviyesi, alis_fiyati_kdv_dahil)
        """
        # alis_fiyati_kdv_dahil sütununu da sorguya ekleyin (9. eleman, indeks 8 olacaktır)
        self.c.execute("SELECT id, urun_kodu, urun_adi, stok_miktari, alis_fiyati_kdv_haric, satis_fiyati_kdv_haric, kdv_orani, min_stok_seviyesi, alis_fiyati_kdv_dahil FROM tbl_stoklar WHERE stok_miktari < min_stok_seviyesi ORDER BY urun_adi ASC")
        return self.c.fetchall()
    
    def kullanici_adi_guncelle(self, user_id, new_username):
        try:
            db_user = self._db_session.query(Kullanici).filter(Kullanici.id == user_id).first()
            if not db_user:
                return False, "Kullanıcı bulunamadı."
            
            existing_user_with_new_name = self._db_session.query(Kullanici).filter(Kullanici.kullanici_adi == new_username).first()
            if existing_user_with_new_name and existing_user_with_new_name.id != user_id:
                return False, "Bu kullanıcı adı zaten başka bir kullanıcı tarafından kullanılıyor."

            db_user.kullanici_adi = new_username
            db_user.son_guncelleme_tarihi_saat = datetime.now() # TIMESTAMP
            db_user.son_guncelleyen_kullanici_id = self._get_current_user_id()
            self._db_session.commit()
            self._db_session.refresh(db_user)
            return True, f"Kullanıcı adı başarıyla '{new_username}' olarak güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Kullanıcı adı güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Kullanıcı adı güncellenirken bir hata oluştu: {e}"
                
    def get_faturalar_by_urun_id(self, urun_id, fatura_tipi=None): 
        """
        Belirli bir ürün ID'sinin yer aldığı faturaları (hem alış hem satış) döndürür.
        İsteğe bağlı olarak fatura tipine göre (ALIŞ/SATIŞ) filtreler.
        Dönüş formatı: (fatura_id, fatura_no, tarih, tip, cari_adi, toplam_kdv_haric, toplam_kdv_dahil)
        """
        query = """
            SELECT
                f.id,
                f.fatura_no,
                f.tarih,
                f.tip,
                CASE
                    WHEN f.cari_id = ? AND f.tip = 'SATIŞ' THEN IFNULL(f.misafir_adi, 'Perakende Satış')
                    WHEN f.tip = 'ALIŞ' THEN ted.ad
                    WHEN f.tip = 'SATIŞ' THEN mus.ad
                    ELSE 'Bilinmeyen Cari'
                END AS cari_adi,
                f.toplam_kdv_haric,
                f.toplam_kdv_dahil
            FROM faturalar f
            JOIN fatura_kalemleri fk ON f.id = fk.fatura_id
            LEFT JOIN musteriler mus ON f.cari_id = mus.id AND f.tip = 'SATIŞ'
            LEFT JOIN tedarikciler ted ON f.cari_id = ted.id AND f.tip = 'ALIŞ'
            WHERE fk.urun_id = ?
        """
        # Perakende müşteri ID'si her zaman sorgudaki ilk parametre olacak
        perakende_id_param = self.perakende_musteri_id if self.perakende_musteri_id is not None else -999
        
        params = [perakende_id_param, urun_id] # Parametre listesi

        if fatura_tipi and fatura_tipi != "TÜMÜ":
            query += " AND f.tip = ?"
            params.append(fatura_tipi)

        query += " ORDER BY f.tarih DESC, f.fatura_no DESC"
        
        self.c.execute(query, params)
        return self.c.fetchall()


    def get_kasa_banka_by_odeme_turu(self, odeme_turu):
        """
        Belirtilen ödeme türüne varsayılan olarak atanmış kasa/banka hesabını döndürür.
        Dönüş: (id, hesap_adi, hesap_no, bakiye, para_birimi, tip, acilis_tarihi, banka_adi, sube_adi, varsayilan_odeme_turu)
        Veya bulunamazsa None.
        """
        self.c.execute("SELECT id, hesap_adi, hesap_no, bakiye, para_birimi, tip, acilis_tarihi, banka_adi, sube_adi, varsayilan_odeme_turu FROM kasalar_bankalar WHERE varsayilan_odeme_turu = ?", (odeme_turu,))
        return self.c.fetchone()
    
    def get_total_sales(self, baslangic_tarih, bitis_tarih):
        """Belirtilen tarih aralığındaki toplam satış tutarını (KDV Dahil) döndürür."""
        query = "SELECT SUM(toplam_kdv_dahil) FROM faturalar WHERE tip = 'SATIŞ' AND tarih BETWEEN ? AND ?"
        self.c.execute(query, (baslangic_tarih, bitis_tarih))
        result = self.c.fetchone()[0]
        return result if result is not None else 0.0

    def get_sales_by_payment_type(self, baslangic_tarih, bitis_tarih):
        """Belirtilen tarih aralığındaki satışları ödeme türüne göre gruplayarak toplam tutarları döndürür."""
        query = """
            SELECT odeme_turu, SUM(toplam_kdv_dahil)
            FROM faturalar
            WHERE tip = 'SATIŞ' AND tarih BETWEEN ? AND ?
            GROUP BY odeme_turu
            ORDER BY odeme_turu
        """
        self.c.execute(query, (baslangic_tarih, bitis_tarih))
        return self.c.fetchall()

    def get_cari_yaslandirma_verileri(self, rapor_tarihi_str=None):
        """
        Belirtilen rapor tarihine göre müşteri alacaklarını ve tedarikçi borçlarını yaşlandırma yapar.
        Dönüş: {
            'musteri_alacaklari': {
                '0-30': [], '31-60': [], '61-90': [], '90+': []
            },
            'tedarikci_borclari': {
                '0-30': [], '31-60': [], '61-90': [], '90+': []
            }
        }
        Her bir liste: (cari_id, cari_adi, tutar, vadesi_gecen_gun_sayisi)
        """
        if rapor_tarihi_str:
            rapor_tarihi = datetime.strptime(rapor_tarihi_str, '%Y-%m-%d')
        else:
            rapor_tarihi = datetime.now()
    
        yaslandirma_sonuclari = {
            'musteri_alacaklari': {'0-30': [], '31-60': [], '61-90': [], '90+': []},
            'tedarikci_borclari': {'0-30': [], '31-60': [], '61-90': [], '90+': []}
        }
    
        # SQL sorgusu güncellendi: islem_tipi = 'ALACAK' ve 'TAHSILAT' olarak netleştirildi.
        # Ayrıca FATURA referans tipleri de dahil edildi.
        self.c.execute("""
            SELECT
                c.cari_id,
                CASE
                    WHEN c.cari_tip = 'MUSTERI' THEN m.ad
                    WHEN c.cari_tip = 'TEDARIKCI' THEN t.ad
                END AS cari_adi,
                -- Müşteri Alacakları (Bize olan borçları: Fatura, Manuel Alacak, Manuel Veresiye Borç)
                SUM(CASE WHEN c.cari_tip = 'MUSTERI' AND (c.islem_tipi = 'ALACAK' OR c.referans_tip = 'FATURA' OR c.referans_tip = 'VERESIYE_BORC_MANUEL') THEN c.tutar ELSE 0 END) AS musteri_toplam_alacak,
                -- Müşteri Tahsilatları (Müşteriden aldıklarımız: Tahsilat, Peşin Satış Faturası)
                SUM(CASE WHEN c.cari_tip = 'MUSTERI' AND (c.islem_tipi = 'TAHSILAT' OR c.referans_tip = 'FATURA_SATIS_PESIN') THEN c.tutar ELSE 0 END) AS musteri_toplam_tahsilat,
                -- Tedarikçi Borçları (Bizim onlara olan borcumuz: Fatura, Manuel Borç, Manuel Veresiye Borç)
                SUM(CASE WHEN c.cari_tip = 'TEDARIKCI' AND (c.islem_tipi = 'BORC' OR c.referans_tip = 'FATURA' OR c.referans_tip = 'VERESIYE_BORC_MANUEL') THEN c.tutar ELSE 0 END) AS tedarikci_toplam_borc,
                -- Tedarikçi Ödemeleri (Tedarikçiye ödediklerimiz: Ödeme, Peşin Alış Faturası)
                SUM(CASE WHEN c.cari_tip = 'TEDARIKCI' AND (c.islem_tipi = 'ODEME' OR c.referans_tip = 'FATURA_ALIS_PESIN') THEN c.tutar ELSE 0 END) AS tedarikci_toplam_odeme,
                c.cari_tip,
                MAX(c.tarih) AS son_islem_tarihi
            FROM cari_hareketler c
            LEFT JOIN musteriler m ON c.cari_id = m.id AND c.cari_tip = 'MUSTERI'
            LEFT JOIN tedarikciler t ON c.cari_id = t.id AND c.cari_tip = 'TEDARIKCI'
            WHERE c.tarih <= ?
            GROUP BY c.cari_id, c.cari_tip
            HAVING (musteri_toplam_alacak - musteri_toplam_tahsilat) != 0 OR (tedarikci_toplam_borc - tedarikci_toplam_odeme) != 0
        """, (rapor_tarihi_str,))
    
        raw_results = self.c.fetchall()
    
        # DÜZELTME: Artık satırlara indeks ile değil, sütun adlarıyla erişiyoruz.
        for row in raw_results:
            cari_id = row['cari_id']
            cari_adi = row['cari_adi']
            musteri_toplam_alacak = row['musteri_toplam_alacak'] or 0.0
            musteri_toplam_tahsilat = row['musteri_toplam_tahsilat'] or 0.0
            tedarikci_toplam_borc = row['tedarikci_toplam_borc'] or 0.0
            tedarikci_toplam_odeme = row['tedarikci_toplam_odeme'] or 0.0
            cari_tip_from_db = row['cari_tip']
            son_islem_tarihi_str = row['son_islem_tarihi']

            if not cari_adi: continue
    
            net_bakiye_musteri = musteri_toplam_alacak - musteri_toplam_tahsilat
            net_bakiye_tedarikci = tedarikci_toplam_borc - tedarikci_toplam_odeme
    
            if net_bakiye_musteri == 0 and net_bakiye_tedarikci == 0: continue
    
            try:
                son_islem_tarihi = datetime.strptime(son_islem_tarihi_str, '%Y-%m-%d')
            except (ValueError, TypeError):
                son_islem_tarihi = rapor_tarihi
    
            vadesi_gecen_gun_sayisi = (rapor_tarihi - son_islem_tarihi).days
    
            if cari_tip_from_db == 'MUSTERI' and net_bakiye_musteri > 0: # Müşteri bize borçlu
                if vadesi_gecen_gun_sayisi <= 30:
                    yaslandirma_sonuclari['musteri_alacaklari']['0-30'].append((cari_id, cari_adi, net_bakiye_musteri, vadesi_gecen_gun_sayisi))
                elif vadesi_gecen_gun_sayisi <= 60:
                    yaslandirma_sonuclari['musteri_alacaklari']['31-60'].append((cari_id, cari_adi, net_bakiye_musteri, vadesi_gecen_gun_sayisi))
                elif vadesi_gecen_gun_sayisi <= 90:
                    yaslandirma_sonuclari['musteri_alacaklari']['61-90'].append((cari_id, cari_adi, net_bakiye_musteri, vadesi_gecen_gun_sayisi))
                else:
                    yaslandirma_sonuclari['musteri_alacaklari']['90+'].append((cari_id, cari_adi, net_bakiye_musteri, vadesi_gecen_gun_sayisi))
    
            elif cari_tip_from_db == 'TEDARIKCI' and net_bakiye_tedarikci > 0: # Biz tedarikçiye borçluyuz
                if vadesi_gecen_gun_sayisi <= 30:
                    yaslandirma_sonuclari['tedarikci_borclari']['0-30'].append((cari_id, cari_adi, net_bakiye_tedarikci, vadesi_gecen_gun_sayisi))
                elif vadesi_gecen_gun_sayisi <= 60:
                    yaslandirma_sonuclari['tedarikci_borclari']['31-60'].append((cari_id, cari_adi, net_bakiye_tedarikci, vadesi_gecen_gun_sayisi))
                elif vadesi_gecen_gun_sayisi <= 90:
                    yaslandirma_sonuclari['tedarikci_borclari']['61-90'].append((cari_id, cari_adi, net_bakiye_tedarikci, vadesi_gecen_gun_sayisi))
                else:
                    yaslandirma_sonuclari['tedarikci_borclari']['90+'].append((cari_id, cari_adi, net_bakiye_tedarikci, vadesi_gecen_gun_sayisi))
    
        return yaslandirma_sonuclari

    def get_total_collections(self, baslangic_tarih, bitis_tarih):
        """Belirtilen tarih aralığındaki toplam tahsilat tutarını döndürür."""
        # Tahsilatlar hem manuel tahsilatlardan hem de peşin satış faturalarından gelir.
        query = """
            SELECT SUM(tutar) FROM gelir_gider
            WHERE tip = 'GELİR' AND (kaynak = 'TAHSILAT' OR kaynak = 'FATURA') AND tarih BETWEEN ? AND ?
        """
        self.c.execute(query, (baslangic_tarih, bitis_tarih))
        result = self.c.fetchone()[0]
        return result if result is not None else 0.0

    def get_total_payments(self, baslangic_tarih, bitis_tarih):
        """Belirtilen tarih aralığındaki toplam ödeme tutarını döndürür."""
        # Ödemeler hem manuel ödemelerden hem de peşin alış faturalarından gelir.
        query = """
            SELECT SUM(tutar) FROM gelir_gider
            WHERE tip = 'GİDER' AND (kaynak = 'ODEME' OR kaynak = 'FATURA') AND tarih BETWEEN ? AND ?
        """
        self.c.execute(query, (baslangic_tarih, bitis_tarih))
        result = self.c.fetchone()[0]
        return result if result is not None else 0.0

    def get_manual_income_expenses(self, baslangic_tarih, bitis_tarih):
        """Belirtilen tarih aralığındaki manuel gelir ve gider toplamlarını döndürür."""
        query_gelir = "SELECT SUM(tutar) FROM gelir_gider WHERE tip = 'GELİR' AND kaynak = 'MANUEL' AND tarih BETWEEN ? AND ?"
        self.c.execute(query_gelir, (baslangic_tarih, bitis_tarih))
        total_manual_income = self.c.fetchone()[0]
        total_manual_income = total_manual_income if total_manual_income is not None else 0.0

        query_gider = "SELECT SUM(tutar) FROM gelir_gider WHERE tip = 'GİDER' AND kaynak = 'MANUEL' AND tarih BETWEEN ? AND ?"
        self.c.execute(query_gider, (baslangic_tarih, bitis_tarih))
        total_manual_expense = self.c.fetchone()[0]
        total_manual_expense = total_manual_expense if total_manual_expense is not None else 0.0

        return total_manual_income, total_manual_expense
    
    def get_gross_profit_and_cost(self, baslangic_tarih, bitis_tarih):
        """
        Belirtilen tarih aralığındaki satışlardan elde edilen brüt kârı (KDV Dahil Fiyatlar Üzerinden),
        satılan malın maliyetini (KDV Dahil Alış Fiyatları Üzerinden) ve brüt kâr oranını hesaplar.
        """
        # Toplam Satış Geliri (KDV Dahil)
        query_sales_revenue = """
            SELECT SUM(toplam_kdv_dahil) FROM faturalar
            WHERE tip = 'SATIŞ' AND tarih BETWEEN ? AND ?
        """
        self.c.execute(query_sales_revenue, (baslangic_tarih, bitis_tarih))
        total_sales_revenue = self.c.fetchone()[0]
        total_sales_revenue = total_sales_revenue if total_sales_revenue is not None else 0.0

        # Satılan Malın Maliyeti (COGS - Cost of Goods Sold)
        # Her satış kalemindeki ürünün ALIŞ FİYATI (KDV DAHİL) üzerinden hesaplanır.
        # Stok tablosundaki alis_fiyati KDV hariçtir, bu yüzden KDV oranını kullanarak KDV dahil alış fiyatını hesaplamalıyız.
        query_cogs = """
            SELECT SUM(fk.miktar * fk.alis_fiyati_fatura_aninda)
            FROM fatura_kalemleri fk
            JOIN faturalar f ON fk.fatura_id = f.id
            WHERE f.tip = 'SATIŞ' AND f.tarih BETWEEN ? AND ?
        """
        self.c.execute(query_cogs, (baslangic_tarih, bitis_tarih))
        total_cogs = self.c.fetchone()[0]
        total_cogs = total_cogs if total_cogs is not None else 0.0

        gross_profit = total_sales_revenue - total_cogs
        gross_profit_rate = (gross_profit / total_sales_revenue * 100) if total_sales_revenue > 0 else 0.0

        return gross_profit, total_cogs, gross_profit_rate    
            
    def _add_column_if_not_exists(self, table_name, column_name, column_type, is_unique=False):
        try:
            self.c.execute(f"PRAGMA table_info({table_name})")
            columns = [col[1] for col in self.c.fetchall()]
            if column_name not in columns:
                self.c.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")
                self.conn.commit()
                # print(f"DEBUG: Sütun eklendi: Tablo={table_name}, Sütun={column_name}") # Bu satır kaldırıldı
                
            if is_unique:
                try:
                    index_name = f"idx_{table_name}_{column_name}_unique"
                    self.c.execute(f"CREATE UNIQUE INDEX IF NOT EXISTS {index_name} ON {table_name} ({column_name})")
                    self.conn.commit()
                except sqlite3.OperationalError as e:
                    logging.warning(f"UYARI: UNIQUE INDEX oluşturulurken hata oluştu: {e}. Bu genellikle '{table_name}.{column_name}' sütununda zaten yinelenen veri olduğu anlamına gelir.") # print yerine logging.warning
                except Exception as e:
                    logging.error(f"UYARI: UNIQUE INDEX oluşturulurken beklenmeyen bir hata oluştu: {e}", exc_info=True) # print yerine logging.error
        except sqlite3.OperationalError as e:
            logging.error(f"HATA: Sütun eklenirken veritabanı hatası: {e}", exc_info=True) # print yerine logging.error
        except Exception as e:
            logging.error(f"HATA: Sütun ekleme kontrolü sırasında beklenmeyen hata: {e}", exc_info=True) # print yerine logging.error

    def create_tables(self, cursor):
        """
        Veritabanı tablolarını PostgreSQL formatında oluşturur.
        Bu metot artık dışarıdan bir cursor objesi alır.
        """
        # --- Şirket Ayarları ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS sirket_ayarlari(
                            anahtar VARCHAR(255) PRIMARY KEY,
                            deger TEXT
                           )''')

        # --- Kullanıcılar ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS kullanicilar (
                            id SERIAL PRIMARY KEY, 
                            kullanici_adi VARCHAR(255) UNIQUE NOT NULL, 
                            sifre VARCHAR(255) NOT NULL, 
                            yetki VARCHAR(50) CHECK(yetki IN ('admin', 'kullanici')) NOT NULL,
                            olusturma_tarihi_saat TIMESTAMP,
                            olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP,
                            son_guncelleyen_kullanici_id INTEGER
                           )''')

        # --- Müşteriler ---
        cursor.execute("""CREATE TABLE IF NOT EXISTS musteriler (
                            id SERIAL PRIMARY KEY,
                            ad VARCHAR(255) NOT NULL,
                            soyad VARCHAR(255),
                            kod VARCHAR(50) UNIQUE,
                            vergi_dairesi VARCHAR(255),
                            vergi_no VARCHAR(50),
                            adres TEXT,
                            telefon VARCHAR(50),
                            email VARCHAR(255),
                            notlar TEXT,
                            bakiye NUMERIC(15, 2) DEFAULT 0.0,
                            olusturma_tarihi_saat TIMESTAMP,
                            olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP,
                            son_guncelleyen_kullanici_id INTEGER
                        )""")

        # --- Tedarikçiler ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS tedarikciler (
                            id SERIAL PRIMARY KEY, 
                            tedarikci_kodu VARCHAR(50) UNIQUE,
                            ad VARCHAR(255) NOT NULL, 
                            telefon VARCHAR(50), 
                            adres TEXT, 
                            vergi_dairesi VARCHAR(255), 
                            vergi_no VARCHAR(50),
                            bakiye NUMERIC(15, 2) DEFAULT 0.0,
                            olusturma_tarihi_saat TIMESTAMP,
                            olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP,
                            son_guncelleyen_kullanici_id INTEGER
                           )''')

        # --- Ürün Nitelik Tabloları ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS urun_kategorileri(
                                id SERIAL PRIMARY KEY,
                                kategori_adi VARCHAR(255) UNIQUE NOT NULL,
                                olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                                son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS urun_markalari(
                                id SERIAL PRIMARY KEY,
                                marka_adi VARCHAR(255) UNIQUE NOT NULL,
                                olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                                son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS urun_gruplari(
                                id SERIAL PRIMARY KEY,
                                grup_adi VARCHAR(255) UNIQUE NOT NULL,
                                olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                                son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS urun_birimleri(
                                id SERIAL PRIMARY KEY,
                                birim_adi VARCHAR(255) UNIQUE NOT NULL,
                                olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                                son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS urun_ulkeleri(
                                id SERIAL PRIMARY KEY,
                                ulke_adi VARCHAR(255) UNIQUE NOT NULL,
                                olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                                son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
                                 
        # --- Stoklar ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS tbl_stoklar (
                            id SERIAL PRIMARY KEY,
                            urun_kodu VARCHAR(50) UNIQUE NOT NULL,
                            urun_adi VARCHAR(255) NOT NULL,
                            stok_miktari NUMERIC(15, 2) DEFAULT 0.0,
                            alis_fiyati_kdv_haric NUMERIC(15, 2) DEFAULT 0.0,
                            satis_fiyati_kdv_haric NUMERIC(15, 2) DEFAULT 0.0,
                            kdv_orani NUMERIC(5, 2) DEFAULT 20.0,
                            min_stok_seviyesi NUMERIC(15, 2) DEFAULT 0.0,
                            alis_fiyati_kdv_dahil NUMERIC(15, 2) DEFAULT 0.0,
                            satis_fiyati_kdv_dahil NUMERIC(15, 2) DEFAULT 0.0,
                            kategori_id INTEGER REFERENCES urun_kategorileri(id) ON DELETE SET NULL,
                            marka_id INTEGER REFERENCES urun_markalari(id) ON DELETE SET NULL,
                            urun_detayi TEXT,
                            urun_resmi_yolu TEXT,
                            fiyat_degisiklik_tarihi DATE,
                            urun_grubu_id INTEGER REFERENCES urun_gruplari(id) ON DELETE SET NULL,
                            urun_birimi_id INTEGER REFERENCES urun_birimleri(id) ON DELETE SET NULL,
                            ulke_id INTEGER REFERENCES urun_ulkeleri(id) ON DELETE SET NULL,
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')

        # --- Kasa ve Bankalar ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS kasalar_bankalar (
                            id SERIAL PRIMARY KEY,
                            hesap_adi VARCHAR(255) NOT NULL UNIQUE,
                            hesap_no VARCHAR(255), 
                            bakiye NUMERIC(15, 2) DEFAULT 0.0,
                            para_birimi VARCHAR(10) DEFAULT 'TL',
                            tip VARCHAR(50) CHECK(tip IN ('KASA', 'BANKA')) NOT NULL,
                            acilis_tarihi DATE, 
                            banka_adi VARCHAR(255), 
                            sube_adi VARCHAR(255),
                            varsayilan_odeme_turu VARCHAR(50),
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        
        # --- Faturalar ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS faturalar (
                            id SERIAL PRIMARY KEY, 
                            fatura_no VARCHAR(255) UNIQUE NOT NULL, 
                            tarih DATE NOT NULL, 
                            tip VARCHAR(50) CHECK(tip IN ('ALIŞ', 'SATIŞ', 'DEVİR_GİRİŞ', 'SATIŞ İADE', 'ALIŞ İADE')) NOT NULL,
                            cari_id INTEGER NOT NULL,
                            toplam_kdv_haric NUMERIC(15, 2) NOT NULL, 
                            toplam_kdv_dahil NUMERIC(15, 2) NOT NULL, 
                            odeme_turu VARCHAR(50),
                            misafir_adi VARCHAR(255),
                            kasa_banka_id INTEGER REFERENCES kasalar_bankalar(id),
                            fatura_notlari TEXT,
                            vade_tarihi DATE,
                            genel_iskonto_tipi VARCHAR(50) DEFAULT 'YOK',
                            genel_iskonto_degeri NUMERIC(15, 2) DEFAULT 0.0,
                            original_fatura_id INTEGER,
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        
        # --- Fatura Kalemleri ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS fatura_kalemleri (
                            id SERIAL PRIMARY KEY, 
                            fatura_id INTEGER NOT NULL REFERENCES faturalar(id) ON DELETE CASCADE, 
                            urun_id INTEGER NOT NULL REFERENCES tbl_stoklar(id), 
                            miktar NUMERIC(15, 2) NOT NULL,
                            birim_fiyat NUMERIC(15, 2) NOT NULL, 
                            kdv_orani NUMERIC(5, 2) NOT NULL, 
                            kdv_tutari NUMERIC(15, 2) NOT NULL, 
                            kalem_toplam_kdv_haric NUMERIC(15, 2) NOT NULL, 
                            kalem_toplam_kdv_dahil NUMERIC(15, 2) NOT NULL,
                            alis_fiyati_fatura_aninda NUMERIC(15, 2) DEFAULT 0.0,
                            kdv_orani_fatura_aninda NUMERIC(5, 2) DEFAULT 0.0,
                            iskonto_yuzde_1 NUMERIC(5, 2) DEFAULT 0.0,
                            iskonto_yuzde_2 NUMERIC(5, 2) DEFAULT 0.0,
                            iskonto_tipi VARCHAR(50) DEFAULT 'YOK',
                            iskonto_degeri NUMERIC(15, 2) DEFAULT 0.0,
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')

        # --- Gelir/Gider Sınıflandırmaları ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS gelir_siniflandirmalari(
                                id SERIAL PRIMARY KEY,
                                siniflandirma_adi VARCHAR(255) UNIQUE NOT NULL,
                                olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                                son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS gider_siniflandirmalari(
                                id SERIAL PRIMARY KEY,
                                siniflandirma_adi VARCHAR(255) UNIQUE NOT NULL,
                                olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                                son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
                                 
        # --- Gelir & Gider ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS gelir_gider (
                            id SERIAL PRIMARY KEY, 
                            tarih DATE NOT NULL, 
                            tip VARCHAR(50) CHECK(tip IN ('GELİR', 'GİDER')) NOT NULL, 
                            tutar NUMERIC(15, 2) NOT NULL, 
                            aciklama TEXT,
                            kaynak VARCHAR(255) DEFAULT 'MANUEL',
                            kaynak_id INTEGER,
                            kasa_banka_id INTEGER REFERENCES kasalar_bankalar(id),
                            gelir_siniflandirma_id INTEGER REFERENCES gelir_siniflandirmalari(id) ON DELETE SET NULL,
                            gider_siniflandirma_id INTEGER REFERENCES gider_siniflandirmalari(id) ON DELETE SET NULL,
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        
        # --- Cari Hareketler ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS cari_hareketler (
                            id SERIAL PRIMARY KEY, 
                            tarih DATE NOT NULL, 
                            cari_tip VARCHAR(50) CHECK(cari_tip IN ('MUSTERI', 'TEDARIKCI')) NOT NULL, 
                            cari_id INTEGER NOT NULL, 
                            islem_tipi VARCHAR(50) CHECK(islem_tipi IN ('ALACAK', 'BORC', 'TAHSILAT', 'ODEME')) NOT NULL, 
                            tutar NUMERIC(15, 2) NOT NULL, 
                            aciklama TEXT,
                            referans_id INTEGER,
                            referans_tip VARCHAR(255),
                            kasa_banka_id INTEGER REFERENCES kasalar_bankalar(id),
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')

        # --- Siparişler ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS siparisler (
                            id SERIAL PRIMARY KEY, 
                            siparis_no VARCHAR(255) UNIQUE NOT NULL, 
                            tarih DATE NOT NULL, 
                            cari_tip VARCHAR(50) CHECK(cari_tip IN ('MUSTERI', 'TEDARIKCI')) NOT NULL, 
                            cari_id INTEGER NOT NULL, 
                            toplam_tutar NUMERIC(15, 2) NOT NULL, 
                            durum VARCHAR(50) CHECK(durum IN ('BEKLEMEDE', 'TAMAMLANDI', 'İPTAL_EDİLDİ', 'KISMİ_TESLİMAT')) NOT NULL,
                            fatura_id INTEGER,
                            siparis_notlari TEXT,
                            onay_durumu VARCHAR(50) DEFAULT 'ONAY_BEKLIYOR',
                            teslimat_tarihi DATE,
                            genel_iskonto_tipi VARCHAR(50) DEFAULT 'YOK',
                            genel_iskonto_degeri NUMERIC(15, 2) DEFAULT 0.0,
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
        
        # --- Sipariş Kalemleri ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS siparis_kalemleri (
                            id SERIAL PRIMARY KEY, 
                            siparis_id INTEGER NOT NULL REFERENCES siparisler(id) ON DELETE CASCADE, 
                            urun_id INTEGER NOT NULL REFERENCES tbl_stoklar(id), 
                            miktar NUMERIC(15, 2) NOT NULL, 
                            birim_fiyat NUMERIC(15, 2) NOT NULL, 
                            kdv_orani NUMERIC(5, 2) NOT NULL, 
                            kdv_tutari NUMERIC(15, 2) NOT NULL, 
                            kalem_toplam_kdv_haric NUMERIC(15, 2) NOT NULL, 
                            kalem_toplam_kdv_dahil NUMERIC(15, 2) NOT NULL,
                            alis_fiyati_siparis_aninda NUMERIC(15, 2) DEFAULT 0.0,
                            satis_fiyati_siparis_aninda NUMERIC(15, 2) DEFAULT 0.0,
                            iskonto_yuzde_1 NUMERIC(5, 2) DEFAULT 0.0,
                            iskonto_yuzde_2 NUMERIC(5, 2) DEFAULT 0.0,
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')

        # --- Stok Hareketleri ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS stok_hareketleri (
                            id SERIAL PRIMARY KEY, 
                            urun_id INTEGER NOT NULL REFERENCES tbl_stoklar(id) ON DELETE CASCADE, 
                            tarih DATE NOT NULL, 
                            islem_tipi VARCHAR(255) NOT NULL, 
                            miktar NUMERIC(15, 2) NOT NULL, 
                            onceki_stok NUMERIC(15, 2) NOT NULL,
                            sonraki_stok NUMERIC(15, 2) NOT NULL,
                            aciklama TEXT,
                            kaynak VARCHAR(255), 
                            kaynak_id INTEGER, 
                            olusturma_tarihi_saat TIMESTAMP,
                            olusturan_kullanici_id INTEGER
                           )''')
                           
        # --- Teklifler ve Teklif Kalemleri ---
        cursor.execute('''CREATE TABLE IF NOT EXISTS teklifler (
                            id SERIAL PRIMARY KEY, 
                            teklif_no VARCHAR(255) UNIQUE NOT NULL, 
                            tarih DATE NOT NULL, 
                            musteri_id INTEGER NOT NULL, 
                            toplam_tutar NUMERIC(15, 2) NOT NULL, 
                            durum VARCHAR(50) CHECK(durum IN ('BEKLEMEDE', 'KABUL EDİLDİ', 'REDDEDİLDİ')) NOT NULL,
                            siparis_id INTEGER,
                            teklif_notlari TEXT,
                            genel_iskonto_tipi VARCHAR(50) DEFAULT 'YOK',
                            genel_iskonto_degeri NUMERIC(15, 2) DEFAULT 0.0,
                            gecerlilik_tarihi DATE,
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS teklif_kalemleri (
                            id SERIAL PRIMARY KEY, 
                            teklif_id INTEGER NOT NULL REFERENCES teklifler(id) ON DELETE CASCADE, 
                            urun_id INTEGER NOT NULL REFERENCES tbl_stoklar(id), 
                            miktar NUMERIC(15, 2) NOT NULL,
                            birim_fiyat NUMERIC(15, 2) NOT NULL, 
                            kdv_orani NUMERIC(5, 2) NOT NULL, 
                            kdv_tutari NUMERIC(15, 2) NOT NULL, 
                            kalem_toplam_kdv_haric NUMERIC(15, 2) NOT NULL, 
                            kalem_toplam_kdv_dahil NUMERIC(15, 2) NOT NULL,
                            alis_fiyati_teklif_aninda NUMERIC(15, 2) DEFAULT 0.0,
                            satis_fiyati_teklif_aninda NUMERIC(15, 2) DEFAULT 0.0,
                            iskonto_yuzde_1 NUMERIC(5, 2) DEFAULT 0.0,
                            iskonto_yuzde_2 NUMERIC(5, 2) DEFAULT 0.0,
                            olusturma_tarihi_saat TIMESTAMP, olusturan_kullanici_id INTEGER,
                            son_guncelleme_tarihi_saat TIMESTAMP, son_guncelleyen_kullanici_id INTEGER
                           )''')
                
    def sirket_bilgilerini_yukle(self):
        defaults = {
            "sirket_adi": "ŞİRKET ADINIZ",
            "sirket_adresi": "Şirket Adresiniz, Şehir, Posta Kodu",
            "sirket_telefonu": "Şirket Telefon Numaranız",
            "sirket_email": "sirket@emailadresiniz.com",
            "sirket_vergi_dairesi": "Bağlı Olduğunuz Vergi Dairesi",
            "sirket_vergi_no": "Vergi Numaranız",
            "sirket_logo_yolu": "" # Logo dosya yolu
        }
        ayarlar = {}
        for anahtar, varsayilan_deger in defaults.items():
            self.c.execute("SELECT deger FROM sirket_ayarlari WHERE anahtar=?", (anahtar,))
            sonuc = self.c.fetchone()
            ayarlar[anahtar] = sonuc[0] if sonuc else varsayilan_deger
            if not sonuc: # Eğer ayar yoksa, varsayılan değerle ekle
                self.c.execute("INSERT INTO sirket_ayarlari (anahtar, deger) VALUES (?,?)", (anahtar, varsayilan_deger))
        return ayarlar

    def ensure_admin_user(self):
        try:
            self.c.execute("SELECT COUNT(*) FROM kullanicilar WHERE yetki='admin'")
            if self.c.fetchone()[0] == 0:
                # Sadece admin kullanıcısı yoksa şifreyi hash'le ve ekle
                admin_password_hashed = self._hash_sifre("admin123")
                self.c.execute("INSERT INTO kullanicilar (kullanici_adi, sifre, yetki, olusturma_tarihi_saat) VALUES (?,?,?,?)",
                               ("admin", admin_password_hashed, "admin", self.get_current_datetime_str()))
                self.conn.commit()
                return True, "Varsayılan 'admin' kullanıcısı başarıyla oluşturuldu."
            return True, "Varsayılan 'admin' kullanıcısı zaten mevcut."
        except Exception as e:
            return False, f"Admin kullanıcısı oluşturulurken/kontrol edilirken hata: {e}"

    def sirket_bilgilerini_yukle(self):
        try:
            sirket = self._db_session.query(SirketBilgileri).first()
            if not sirket:
                return {} # Boş dictionary dön
            return {
                "sirket_adi": sirket.sirket_adi,
                "sirket_adresi": sirket.sirket_adresi,
                "sirket_telefonu": sirket.sirket_telefonu,
                "sirket_email": sirket.sirket_email,
                "sirket_vergi_dairesi": sirket.sirket_vergi_dairesi,
                "sirket_vergi_no": sirket.sirket_vergi_no,
                "sirket_logo_yolu": sirket.sirket_logo_yolu
            }
        except Exception as e:
            logging.error(f"Şirket bilgileri yüklenirken hata: {e}\n{traceback.format_exc()}")
            return {}

    def sirket_bilgilerini_kaydet(self, bilgiler):
        try:
            sirket = self._db_session.query(SirketBilgileri).first()
            if not sirket:
                sirket = SirketBilgileri()
                self._db_session.add(sirket)
            
            for key, value in bilgiler.items():
                setattr(sirket, key, value)
            
            sirket.son_guncelleme_tarihi_saat = datetime.now() # TIMESTAMP
            self._db_session.commit()
            self._db_session.refresh(sirket)
            return True, "Şirket bilgileri başarıyla kaydedildi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Şirket bilgileri kaydedilirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Şirket bilgileri kaydedilirken bir hata oluştu: {e}"

    def _ensure_sirket_bilgileri(self):
        """Varsayılan şirket bilgilerini oluşturur."""
        try:
            if self._db_session.query(SirketBilgileri).count() == 0:
                default_sirket = SirketBilgileri(
                    sirket_adi="DEMO ŞİRKETİ",
                    olusturma_tarihi_saat=datetime.now()
                )
                self._db_session.add(default_sirket)
                self._db_session.commit()
                self._db_session.refresh(default_sirket)
                logging.info("Varsayılan 'DEMO ŞİRKETİ' oluşturuldu.")
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Varsayılan şirket bilgileri oluşturulurken hata: {e}\n{traceback.format_exc()}")

    def _hash_sifre(self, sifre):
        return hashlib.sha256(sifre.encode()).hexdigest()

    def kullanici_dogrula(self, kullanici_adi, sifre):
        try:
            db_kullanici = self._db_session.query(Kullanici).filter(Kullanici.kullanici_adi == kullanici_adi).first()
            if db_kullanici and db_kullanici.sifre == self._hash_sifre(sifre):
                db_kullanici.son_giris_tarihi_saat = datetime.now() # TIMESTAMP
                self._db_session.commit()
                self._db_session.refresh(db_kullanici)
                return (db_kullanici.id, db_kullanici.kullanici_adi, db_kullanici.yetki)
            return None
        except Exception as e:
            logging.error(f"Kullanıcı doğrulama hatası: {e}\n{traceback.format_exc()}")
            return None

    def kullanici_ekle(self, kullanici_adi, sifre, yetki="kullanici"):
        try:
            hashed_sifre = self._hash_sifre(sifre)
            # SQLAlchemy ORM ile sorgu
            db_kullanici = self._db_session.query(Kullanici).filter(Kullanici.kullanici_adi == kullanici_adi).first()
            if db_kullanici:
                return False, "Bu kullanıcı adı zaten mevcut."
            
            yeni_kullanici = Kullanici(
                kullanici_adi=kullanici_adi,
                sifre=hashed_sifre,
                yetki=yetki,
                olusturma_tarihi_saat=datetime.now(), # TIMESTAMP
                olusturan_kullanici_id=self._get_current_user_id()
            )
            self._db_session.add(yeni_kullanici)
            self._db_session.commit()
            self._db_session.refresh(yeni_kullanici) # Yeni eklenen objeyi güncelleyerek ID'sini alırız
            return True, f"'{kullanici_adi}' kullanıcısı başarıyla eklendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Kullanıcı eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Kullanıcı eklenirken bir hata oluştu: {e}"
        
    def kullanici_guncelle_sifre_yetki(self, user_id, hashed_sifre=None, yetki=None):
        try:
            db_user = self._db_session.query(Kullanici).filter(Kullanici.id == user_id).first()
            if not db_user:
                return False, "Kullanıcı bulunamadı."
            
            if hashed_sifre:
                db_user.sifre = hashed_sifre
            if yetki:
                db_user.yetki = yetki
            
            db_user.son_guncelleme_tarihi_saat = datetime.now() # TIMESTAMP
            db_user.son_guncelleyen_kullanici_id = self._get_current_user_id()
            self._db_session.commit()
            self._db_session.refresh(db_user)
            return True, f"Kullanıcı bilgileri başarıyla güncellendi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Kullanıcı bilgileri güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Kullanıcı bilgileri güncellenirken bir hata oluştu: {e}"

    def kullanici_listele(self):
        try:
            return self._db_session.query(Kullanici).all()
        except Exception as e:
            logging.error(f"Kullanıcı listelenirken hata: {e}\n{traceback.format_exc()}")
            return []
        
    def kullanici_sil(self, user_id):
        try:
            if self.app and hasattr(self.app, 'current_user') and self.app.current_user and user_id == self.app.current_user[0]:
                return False, "Kendi kullanıcı hesabınızı silemezsiniz."

            db_user = self._db_session.query(Kullanici).filter(Kullanici.id == user_id).first()
            if db_user is None:
                return False, "Kullanıcı bulunamadı."
            
            # Admin kullanıcı sayısını kontrol et (en az 1 admin olmalı)
            if db_user.yetki == 'admin':
                admin_count = self._db_session.query(Kullanici).filter(Kullanici.yetki == 'admin').count()
                if admin_count <= 1:
                    return False, "Sistemde en az bir admin kullanıcısı bulunmalıdır. Bu admin kullanıcısını silemezsiniz."

            self._db_session.delete(db_user)
            self._db_session.commit()
            return True, f"'{db_user.kullanici_adi}' kullanıcısı başarıyla silindi."
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Kullanıcı silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Kullanıcı silinirken bir hata oluştu: {e}"

    def _ensure_admin_user(self):
        """Uygulama başladığında varsayılan bir admin kullanıcısı oluşturur."""
        try:
            if self._db_session.query(Kullanici).count() == 0:
                admin_user = Kullanici(
                    kullanici_adi="admin",
                    sifre=self._hash_sifre("admin"), # Varsayılan şifre "admin"
                    yetki="admin",
                    olusturma_tarihi_saat=datetime.now()
                )
                self._db_session.add(admin_user)
                self._db_session.commit()
                self._db_session.refresh(admin_user)
                logging.info("Varsayılan 'admin' kullanıcısı oluşturuldu.")
        except Exception as e:
            self._db_session.rollback()
            logging.error(f"Varsayılan admin kullanıcısı oluşturulurken hata: {e}\n{traceback.format_exc()}")

    def get_cari_list_summary_data(self, cari_tip, arama_terimi=None, limit=None, offset=None, perakende_haric=False):
        """
        Belirtilen cari tipi için özet listeleme verilerini (fatura sayısı, açık hesap, tahsilat/ödeme, kalan borç, vadesi geçmiş, son işlem tarihi) döndürür.
        """
        data_list = []
        
        # Ana carileri filtreleyerek alalım (arama terimi ve sayfalama ile)
        if cari_tip == self.CARI_TIP_MUSTERI:
            cariler = self.musteri_listesi_al(arama_terimi=arama_terimi, perakende_haric=perakende_haric, limit=limit, offset=offset)
        elif cari_tip == self.CARI_TIP_TEDARIKCI:
            cariler = self.tedarikci_listesi_al(arama_terimi=arama_terimi, limit=limit, offset=offset)
        else:
            return []

        for cari in cariler:
            cari_id = cari['id']
            cari_adi = cari['ad']

            # Fatura Sayısı (Bu kısım doğru, tüm faturaları sayar)
            query_fatura_sayisi = "SELECT COUNT(id) FROM faturalar WHERE cari_id = ? AND (tip = ? OR tip = ?)"
            if cari_tip == self.CARI_TIP_MUSTERI:
                self.c.execute(query_fatura_sayisi, (cari_id, self.FATURA_TIP_SATIS, self.FATURA_TIP_SATIS_IADE))
            else: # TEDARIKCI
                self.c.execute(query_fatura_sayisi, (cari_id, self.FATURA_TIP_ALIS, self.FATURA_TIP_ALIS_IADE))
            fatura_sayisi = self.c.fetchone()[0] or 0

            # AÇIK HESAP FATURA TOPLAMI (TÜM ZAMANLAR) (Sadece 'AÇIK HESAP' olanları toplar)
            query_acik_hesap_toplam = """
                SELECT SUM(toplam_kdv_dahil) FROM faturalar
                WHERE cari_id = ? AND odeme_turu = ? AND (tip = ? OR tip = ?)
            """
            if cari_tip == self.CARI_TIP_MUSTERI:
                self.c.execute(query_acik_hesap_toplam, (cari_id, self.ODEME_TURU_ACIK_HESAP, self.FATURA_TIP_SATIS, self.FATURA_TIP_SATIS_IADE))
            else: # TEDARIKCI
                self.c.execute(query_acik_hesap_toplam, (cari_id, self.ODEME_TURU_ACIK_HESAP, self.FATURA_TIP_ALIS, self.FATURA_TIP_ALIS_IADE))
            acik_hesap_fatura_toplam = self.c.fetchone()[0] or 0.0

            # ÖDEME/TAHSİLAT (TÜM ZAMANLAR) ve Son Ödeme/Tahsilat Tarihi
            # SADECE MANUEL TAHSİLAT/ÖDEME VE PEŞİN FATURALARIN TAHSİLAT/ÖDEMELERİNİ DAHİL EDİYORUZ.
            # AÇIK HESAP FATURALARI BU TOPLAMA DAHİL EDİLMEYECEKTİR.
            son_odeme_tarihi = "-"
            if cari_tip == self.CARI_TIP_MUSTERI:
                query_tahsilat = """
                    SELECT SUM(ch.tutar), MAX(ch.tarih) FROM cari_hareketler ch
                    WHERE ch.cari_id = ? AND ch.cari_tip = ? AND ch.islem_tipi = ? -- SADECE TAHSILAT
                """
                self.c.execute(query_tahsilat, (cari_id, self.CARI_TIP_MUSTERI, self.ISLEM_TIP_TAHSILAT))
                
                tahsilat_sonuc = self.c.fetchone()
                odeme_tahsilat_toplam = tahsilat_sonuc[0] or 0.0
                if tahsilat_sonuc[1]:
                    son_odeme_tarihi = datetime.strptime(tahsilat_sonuc[1], '%Y-%m-%d').strftime('%d.%m.%Y')
                
                # Kalan Borç (Net Bakiye) - Bu kısım zaten doğru şekilde filtreliyor
                kalan_borc = self.get_musteri_net_bakiye(cari_id)

            else: # TEDARIKCI
                query_odeme = """
                    SELECT SUM(ch.tutar), MAX(ch.tarih) FROM cari_hareketler ch
                    WHERE ch.cari_id = ? AND ch.cari_tip = ? AND ch.islem_tipi = ? -- SADECE ODEME
                """
                self.c.execute(query_odeme, (cari_id, self.CARI_TIP_TEDARIKCI, self.ISLEM_TIP_ODEME))
                
                odeme_sonuc = self.c.fetchone()
                odeme_tahsilat_toplam = odeme_sonuc[0] or 0.0
                if odeme_sonuc[1]:
                    son_odeme_tarihi = datetime.strptime(odeme_sonuc[1], '%Y-%m-%d').strftime('%d.%m.%Y')

                # Kalan Borç (Net Bakiye) - Bu kısım zaten doğru şekilde filtreliyor
                kalan_borc = self.get_tedarikci_net_bakiye(cari_id)

            # Vadesi Geçmiş Borç (Bu kısım zaten doğru şekilde _get_vade_durumu çağırıyor)
            vadesi_gecmis_borc = 0.0
            vade_durumu = self._get_vade_durumu(cari_id, cari_tip)
            if kalan_borc > 0: # Sadece cari borçlu ise ve vadesi geçmiş borcu varsa
                 vadesi_gecmis_borc = vade_durumu["vadesi_gelmis"]

            data_list.append({
                'id': cari_id,
                'cari_adi': cari_adi,
                'fatura_sayisi': fatura_sayisi,
                'acik_hesap_toplam': acik_hesap_fatura_toplam,
                'odeme_tahsilat_toplam': odeme_tahsilat_toplam,
                'kalan_borc': kalan_borc,
                'vadesi_gecmis_borc': vadesi_gecmis_borc,
                'son_odeme_tarihi': son_odeme_tarihi
            })
        return data_list
        
    def get_cari_count(self, cari_tip, arama_terimi=None, perakende_haric=False):
        """Müşteri veya tedarikçi sayısını arama terimine göre döndürür."""
        if cari_tip == self.CARI_TIP_MUSTERI:
            return self.get_musteri_count(arama_terimi=arama_terimi, perakende_haric=perakende_haric)
        elif cari_tip == self.CARI_TIP_TEDARIKCI:
            return self.get_tedarikci_count(arama_terimi=arama_terimi)
        return 0

    def siparis_listele(self, baslangic_tarih=None, bitis_tarih=None, arama_terimi=None, cari_id_filter=None, durum_filter=None, siparis_tipi_filter=None, limit=None, offset=None):
        """
        Belirtilen kriterlere göre siparişleri listeler.
        Dönüş: (id, siparis_no, tarih, cari_tip, cari_id, toplam_tutar, durum, teslimat_tarihi)
        """
        q = """
            SELECT 
                s.id, s.siparis_no, s.tarih, s.cari_tip, s.cari_id, s.toplam_tutar, s.durum, s.teslimat_tarihi
            FROM siparisler s
            LEFT JOIN musteriler m ON s.cari_id = m.id AND s.cari_tip = 'MUSTERI'
            LEFT JOIN tedarikciler t ON s.cari_id = t.id AND s.cari_tip = 'TEDARIKCI'
            LEFT JOIN siparis_kalemleri sk ON s.id = sk.siparis_id
            LEFT JOIN tbl_stoklar urun ON sk.urun_id = urun.id 
        """
        p = []
        conditions = []

        if baslangic_tarih:
            conditions.append("s.tarih >= ?")
            p.append(baslangic_tarih)
        if bitis_tarih:
            conditions.append("s.tarih <= ?")
            p.append(bitis_tarih)

        if cari_id_filter:
            conditions.append("s.cari_id = ?")
            p.append(cari_id_filter)

        if durum_filter and durum_filter != "TÜMÜ":
            conditions.append("s.durum = ?")
            p.append(durum_filter)
        
        if siparis_tipi_filter and siparis_tipi_filter != "TÜMÜ":
            conditions.append("s.cari_tip = ?") # Sipariş tipi 'cari_tip' sütununda tutuluyor
            p.append(siparis_tipi_filter)

        if arama_terimi:
            term = f"%{arama_terimi}%"
            conditions.append("""(
                                s.siparis_no LIKE ? OR
                                m.ad LIKE ? OR
                                t.ad LIKE ? OR
                                urun.urun_adi LIKE ? OR
                                urun.urun_kodu LIKE ?
                              )""")
            p.extend([term, term, term, term, term])

        if conditions:
            q += " WHERE " + " AND ".join(conditions)

        q += " GROUP BY s.id ORDER BY s.tarih DESC, s.id DESC" # Gruplayarak her siparişi bir kez al

        if limit is not None:
            q += " LIMIT ?"
            p.append(limit)
        if offset is not None:
            q += " OFFSET ?"
            p.append(offset)
        
        self.c.execute(q, p)
        return self.c.fetchall()


    def get_siparis_count(self, baslangic_tarih=None, bitis_tarih=None, arama_terimi=None, cari_id_filter=None, durum_filter=None, siparis_tipi_filter=None):
        """
        Belirtilen kriterlere uyan toplam sipariş sayısını döndürür.
        """
        q = """
            SELECT COUNT(DISTINCT s.id)
            FROM siparisler s
            LEFT JOIN musteriler m ON s.cari_id = m.id AND s.cari_tip = 'MUSTERI'
            LEFT JOIN tedarikciler t ON s.cari_id = t.id AND s.cari_tip = 'TEDARIKCI'
            LEFT JOIN siparis_kalemleri sk ON s.id = sk.siparis_id
            LEFT JOIN tbl_stoklar urun ON sk.urun_id = urun.id 
        """
        p = []
        conditions = []

        if baslangic_tarih:
            conditions.append("s.tarih >= ?")
            p.append(baslangic_tarih)
        if bitis_tarih:
            conditions.append("s.tarih <= ?")
            p.append(bitis_tarih)

        if cari_id_filter:
            conditions.append("s.cari_id = ?")
            p.append(cari_id_filter)

        if durum_filter and durum_filter != "TÜMÜ":
            conditions.append("s.durum = ?")
            p.append(durum_filter)
        
        if siparis_tipi_filter and siparis_tipi_filter != "TÜMÜ":
            conditions.append("s.cari_tip = ?")
            p.append(siparis_tipi_filter)

        if arama_terimi:
            term = f"%{arama_terimi}%"
            conditions.append("""(
                                s.siparis_no LIKE ? OR
                                m.ad LIKE ? OR
                                t.ad LIKE ? OR
                                urun.urun_adi LIKE ? OR
                                urun.urun_kodu LIKE ?
                              )""")
            p.extend([term, term, term, term, term])

        if conditions:
            q += " WHERE " + " AND ".join(conditions)
            
        self.c.execute(q, p)
        return self.c.fetchone()[0]

    def get_next_siparis_no(self, prefix="", length=6): # prefix eklendi (örneğin "MS", "AS")
        """
        Mevcut sipariş kodları arasında en yüksek sayısal değeri bulur ve bir sonraki kodu döndürür.
        Belirtilen uzunluğa kadar baştan sıfırlarla doldurur.
        Opsiyonel olarak bir önek alabilir (örn. 'MS' için Müşteri Siparişi, 'AS' için Alış Siparişi).
        """
        self.c.execute("SELECT siparis_no FROM siparisler WHERE siparis_no LIKE ? || '%'", (prefix,))
        existing_codes = self.c.fetchall()
        
        max_numeric_code = 0
        for code_tuple in existing_codes:
            code = code_tuple[0]
            # Önekten sonraki sayısal kısmı ayır
            numeric_part = code[len(prefix):]
            if numeric_part.isdigit():
                try:
                    numeric_code = int(numeric_part)
                    if numeric_code > max_numeric_code:
                        max_numeric_code = numeric_code
                except ValueError:
                    pass # Sayısal olmayan kısmı atla
        
        next_code = max_numeric_code + 1
        return prefix + str(next_code).zfill(length)

    def siparis_ekle(self, siparis_no, siparis_tipi, cari_id, toplam_tutar, durum, kalemler, siparis_notlari=None, teslimat_tarihi=None, genel_iskonto_tipi='YOK', genel_iskonto_degeri=0.0):
        try:
            self.conn.execute("BEGIN TRANSACTION")
            self.c.execute("SELECT id FROM siparisler WHERE siparis_no = ?", (siparis_no,))
            if self.c.fetchone():
                self.conn.rollback(); return False, f"'{siparis_no}' sipariş numarası zaten mevcut."

            cari_tip_for_db = self.CARI_TIP_MUSTERI if siparis_tipi == self.SIPARIS_TIP_SATIS else self.CARI_TIP_TEDARIKCI
            
            toplam_kdv_haric_kalemler, toplam_kdv_dahil_kalemler = 0.0, 0.0
            kalemler_for_insert = []
            
            for item in kalemler:
                urun_id, miktar, birim_fiyat_haric, kdv_orani, alis_fiyati, isk1, isk2 = item
                iskontolu_bf_haric = self.safe_float(birim_fiyat_haric) * (1 - self.safe_float(isk1) / 100) * (1 - self.safe_float(isk2) / 100)
                kalem_tkh = iskontolu_bf_haric * self.safe_float(miktar)
                kdv_tutari = kalem_tkh * (self.safe_float(kdv_orani) / 100)
                kalem_tkd = kalem_tkh + kdv_tutari
                toplam_kdv_haric_kalemler += kalem_tkh
                toplam_kdv_dahil_kalemler += kalem_tkd
                satis_fiyati = birim_fiyat_haric * (1 + self.safe_float(kdv_orani) / 100)
                
                kalemler_for_insert.append((urun_id, miktar, birim_fiyat_haric, kdv_orani, kdv_tutari, kalem_tkh, kalem_tkd, alis_fiyati, satis_fiyati, isk1, isk2))

            uygulanan_genel_iskonto = genel_iskonto_degeri if genel_iskonto_tipi == self.ISKONTO_TIP_TUTAR else toplam_kdv_haric_kalemler * (genel_iskonto_degeri / 100)
            final_total_tutar = toplam_kdv_dahil_kalemler - uygulanan_genel_iskonto

            current_time, olusturan_id, tarih_str = self.get_current_datetime_str(), self._get_current_user_id(), datetime.now().strftime('%Y-%m-%d')

            self.c.execute("INSERT INTO siparisler (siparis_no, tarih, cari_tip, cari_id, toplam_tutar, durum, siparis_notlari, teslimat_tarihi, genel_iskonto_tipi, genel_iskonto_degeri, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                            (siparis_no, tarih_str, cari_tip_for_db, cari_id, final_total_tutar, durum, siparis_notlari, teslimat_tarihi, genel_iskonto_tipi, genel_iskonto_degeri, current_time, olusturan_id))
            siparis_id = self.c.lastrowid

            for urun_id, miktar, bf, kdv, ktutar, tkh, tkd, af, sf, i1, i2 in kalemler_for_insert:
                 self.c.execute("INSERT INTO siparis_kalemleri (siparis_id, urun_id, miktar, birim_fiyat, kdv_orani, kdv_tutari, kalem_toplam_kdv_haric, kalem_toplam_kdv_dahil, alis_fiyati_siparis_aninda, satis_fiyati_siparis_aninda, iskonto_yuzde_1, iskonto_yuzde_2, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                  (siparis_id, urun_id, miktar, bf, kdv, ktutar, tkh, tkd, af, sf, i1, i2, current_time, olusturan_id))

            self.conn.commit()
            return True, f"'{siparis_no}' numaralı sipariş başarıyla oluşturuldu."
        except Exception as e:
            self.conn.rollback()
            return False, f"Sipariş oluşturulurken hata: {e}\n{traceback.format_exc()}"
        
    def siparis_guncelle(self, siparis_id, yeni_siparis_no, yeni_siparis_tipi, yeni_cari_id, yeni_toplam_tutar, yeni_durum, yeni_kalemler, yeni_siparis_notlari=None, yeni_teslimat_tarihi=None, genel_iskonto_tipi='YOK', genel_iskonto_degeri=0.0):
        try:
            self.conn.execute("BEGIN TRANSACTION")
            self.c.execute("DELETE FROM siparis_kalemleri WHERE siparis_id = ?", (siparis_id,))

            toplam_kdv_haric_kalemler, toplam_kdv_dahil_kalemler = 0.0, 0.0
            kalemler_for_insert = []
            for item in yeni_kalemler:
                urun_id, miktar, birim_fiyat_haric, kdv_orani, alis_fiyati, isk1, isk2 = item
                iskontolu_bf_haric = self.safe_float(birim_fiyat_haric) * (1 - self.safe_float(isk1) / 100) * (1 - self.safe_float(isk2) / 100)
                kalem_tkh = iskontolu_bf_haric * self.safe_float(miktar)
                kdv_tutari = kalem_tkh * (self.safe_float(kdv_orani) / 100)
                kalem_tkd = kalem_tkh + kdv_tutari
                toplam_kdv_haric_kalemler += kalem_tkh
                toplam_kdv_dahil_kalemler += kalem_tkd
                satis_fiyati = birim_fiyat_haric * (1 + self.safe_float(kdv_orani) / 100)
                
                kalemler_for_insert.append((urun_id, miktar, birim_fiyat_haric, kdv_orani, kdv_tutari, kalem_tkh, kalem_tkd, alis_fiyati, satis_fiyati, isk1, isk2))

            uygulanan_genel_iskonto = genel_iskonto_degeri if genel_iskonto_tipi == self.ISKONTO_TIP_TUTAR else toplam_kdv_haric_kalemler * (genel_iskonto_degeri / 100)
            final_total_tutar = toplam_kdv_dahil_kalemler - uygulanan_genel_iskonto

            current_time, guncelleyen_id, tarih_str = self.get_current_datetime_str(), self._get_current_user_id(), datetime.now().strftime('%Y-%m-%d')
            cari_tip_for_db = self.CARI_TIP_MUSTERI if yeni_siparis_tipi == self.SIPARIS_TIP_SATIS else self.CARI_TIP_TEDARIKCI

            self.c.execute("UPDATE siparisler SET siparis_no=?, tarih=?, cari_tip=?, cari_id=?, toplam_tutar=?, durum=?, siparis_notlari=?, teslimat_tarihi=?, genel_iskonto_tipi=?, genel_iskonto_degeri=?, son_guncelleme_tarihi_saat=?, son_guncelleyen_kullanici_id=? WHERE id=?",
                            (yeni_siparis_no, tarih_str, cari_tip_for_db, yeni_cari_id, final_total_tutar, yeni_durum, yeni_siparis_notlari, yeni_teslimat_tarihi, genel_iskonto_tipi, genel_iskonto_degeri, current_time, guncelleyen_id, siparis_id))

            for urun_id, miktar, bf, kdv, ktutar, tkh, tkd, af, sf, i1, i2 in kalemler_for_insert:
                self.c.execute("INSERT INTO siparis_kalemleri (siparis_id, urun_id, miktar, birim_fiyat, kdv_orani, kdv_tutari, kalem_toplam_kdv_haric, kalem_toplam_kdv_dahil, alis_fiyati_siparis_aninda, satis_fiyati_siparis_aninda, iskonto_yuzde_1, iskonto_yuzde_2, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                 (siparis_id, urun_id, miktar, bf, kdv, ktutar, tkh, tkd, af, sf, i1, i2, current_time, guncelleyen_id))

            self.conn.commit()
            return True, f"Sipariş '{yeni_siparis_no}' başarıyla güncellendi."
        except Exception as e:
            self.conn.rollback()
            return False, f"Sipariş güncellenirken hata: {e}\n{traceback.format_exc()}"
                        
    def get_siparis_by_id(self, siparis_id):
        """
        Belirli bir siparişin tüm ana bilgilerini döndürür.
        """
        self.c.execute("SELECT id, siparis_no, tarih, cari_tip, cari_id, toplam_tutar, durum, fatura_id, olusturma_tarihi_saat, olusturan_kullanici_id, son_guncelleme_tarihi_saat, son_guncelleyen_kullanici_id, siparis_notlari, onay_durumu, teslimat_tarihi, genel_iskonto_tipi, genel_iskonto_degeri FROM siparisler WHERE id=?", (siparis_id,))
        return self.c.fetchone()


    def siparis_sil(self, siparis_id):
        """
        Belirli bir siparişi ve ilişkili sipariş kalemlerini siler.
        Eğer sipariş bir faturaya dönüştürüldüyse silinmesine izin vermez.
        Dönüş: (bool, mesaj)
        """
        try:
            # Sipariş bilgilerini al
            siparis_bilgisi = self.get_siparis_by_id(siparis_id)
            if not siparis_bilgisi:
                return False, "Silinecek sipariş bulunamadı."
            
            # Eğer sipariş bir faturaya dönüştürüldüyse silinemez
            # siparis_bilgisi[7] -> fatura_id sütunu (eğer doluysa, yani bir fatura ID'si varsa)
            if siparis_bilgisi[7] is not None:
                # Fatura numarasını da alıp mesajda gösterebiliriz.
                # fatura_id_ref_db = siparis_bilgisi[7]
                # fatura_info = self.fatura_getir_by_id(fatura_id_ref_db)
                # fatura_no = fatura_info[1] if fatura_info else "Bilinmiyor"
                return False, f"Bu sipariş bir faturaya dönüştürülmüştür. Lütfen önce ilgili faturayı silin."

            self.conn.execute("BEGIN TRANSACTION")

            # Sipariş kalemlerini sil
            self.c.execute("DELETE FROM siparis_kalemleri WHERE siparis_id=?", (siparis_id,))
            
            # Ana sipariş kaydını sil
            self.c.execute("DELETE FROM siparisler WHERE id=?", (siparis_id,))
            
            self.conn.commit()
            return True, f"Sipariş '{siparis_bilgisi[1]}' başarıyla silindi."
        except Exception as e:
            self.conn.rollback()
            return False, f"Sipariş silinirken bir hata oluştu: {e}\n{traceback.format_exc()}"


    def get_siparis_kalemleri(self, siparis_id):
        """
        Belirli bir siparişin kalemlerini döndürür.
        """
        self.c.execute("SELECT sk.id, sk.siparis_id, sk.urun_id, sk.miktar, sk.birim_fiyat, sk.kdv_orani, sk.kdv_tutari, sk.kalem_toplam_kdv_haric, sk.kalem_toplam_kdv_dahil, sk.alis_fiyati_siparis_aninda, sk.satis_fiyati_siparis_aninda, sk.iskonto_yuzde_1, sk.iskonto_yuzde_2, sk.olusturma_tarihi_saat, sk.olusturan_kullanici_id, sk.son_guncelleme_tarihi_saat, sk.son_guncelleyen_kullanici_id FROM siparis_kalemleri sk JOIN tbl_stoklar s ON sk.urun_id=s.id WHERE sk.siparis_id=?", (siparis_id,))
        return self.c.fetchall()

    def get_musteri_sayisi(self, arama_terimi="", perakende_haric=False):
        """
        Belirtilen filtre kriterlerine göre toplam müşteri sayısını döner.
        """
        query = "SELECT COUNT(id) FROM musteriler"
        params = []
        
        conditions = []

        if perakende_haric:
            conditions.append("kod != ?")
            params.append(self.PERAKENDE_MUSTERI_KODU)
        
        if arama_terimi:
            # Arama terimini 'ad', 'soyad', 'kod', 'telefon' gibi ilgili sütunlarda arayabilirsiniz.
            # '%{}%' kullanarak LIKE operatörü ile arama yaparız.
            search_query = " (ad LIKE ? OR soyad LIKE ? OR kod LIKE ? OR telefon LIKE ?) "
            conditions.append(search_query)
            search_param = f"%{arama_terimi}%"
            params.extend([search_param, search_param, search_param, search_param])

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        self.c.execute(query, params)
        return self.c.fetchone()[0] # İlk sütunu (COUNT değeri) döner

    def get_musteri_count(self, arama_terimi=None, perakende_haric=False):
        query = "SELECT COUNT(id) FROM musteriler"
        params = []
        conditions = []

        if perakende_haric and self.perakende_musteri_id is not None:
            conditions.append("id != ?")
            params.append(self.PERAKENDE_MUSTERI_KODU) # ID değil, KOD kullanıyoruz
            
        if arama_terimi:
            normalized_term = normalize_turkish_chars(arama_terimi)
            term_wildcard = f"%{normalized_term}%"

            turkish_normalize_sql = lambda col: f"""
                LOWER(
                    REPLACE(
                        REPLACE(
                            REPLACE(
                                REPLACE(
                                    REPLACE(
                                        REPLACE(
                                            REPLACE(
                                                REPLACE(
                                                    REPLACE(
                                                        REPLACE(
                                                            REPLACE(
                                                                REPLACE(
                                                                    REPLACE(
                                                                        REPLACE({col}, 'Ş', 'S'),
                                                                    'İ', 'I'),
                                                                'Ç', 'C'),
                                                            'Ğ', 'G'),
                                                        'Ö', 'O'),
                                                    'Ü', 'U'),
                                                'ş', 's'),
                                            'ı', 'i'),
                                        'ç', 'c'),
                                    'ğ', 'g'),
                                'ö', 'o'),
                            'ü', 'u'),
                        'I', 'i'), -- Büyük I'yi küçük i'ye çevir
                    'İ', 'i')  -- Büyük İ'yi küçük i'ye çevir
                )
            """
            search_clauses = [
                f"{turkish_normalize_sql('kod')} LIKE ?",
                f"{turkish_normalize_sql('ad')} LIKE ?",
                f"{turkish_normalize_sql('soyad')} LIKE ?",
                f"{turkish_normalize_sql('telefon')} LIKE ?",
                f"{turkish_normalize_sql('adres')} LIKE ?",
                f"{turkish_normalize_sql('vergi_dairesi')} LIKE ?",
                f"{turkish_normalize_sql('vergi_no')} LIKE ?"
            ]
            
            conditions.append(f"({ ' OR '.join(search_clauses) })")
            params.extend([term_wildcard] * len(search_clauses))
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
            
        self.c.execute(query, params)
        return self.c.fetchone()[0]
                                
    def get_stok_count(self, arama_terimi=None, kategori_id_filter=None, marka_id_filter=None, urun_grubu_id_filter=None, urun_birimi_id_filter=None, ulke_id_filter=None): # Yeni filtre parametreleri
        query = "SELECT COUNT(s.id) FROM tbl_stoklar s"
        params = []
        conditions = []

        if arama_terimi:
            conditions.append("(s.urun_kodu LIKE ? OR s.urun_adi LIKE ?)")
            term = f"%{arama_terimi}%"
            params.extend([term, term])
        
        if kategori_id_filter is not None:
            conditions.append("s.kategori_id = ?")
            params.append(kategori_id_filter)

        if marka_id_filter is not None:
            conditions.append("s.marka_id = ?")
            params.append(marka_id_filter)

        if urun_grubu_id_filter is not None:
            conditions.append("s.urun_grubu_id = ?")
            params.append(urun_grubu_id_filter)

        if urun_birimi_id_filter is not None:
            conditions.append("s.urun_birimi_id = ?")
            params.append(urun_birimi_id_filter)

        if ulke_id_filter is not None:
            conditions.append("s.ulke_id = ?")
            params.append(ulke_id_filter)

        if conditions:
            query += " WHERE " + " AND ".join(conditions)
            
        self.c.execute(query, params)
        return self.c.fetchone()[0]

    def get_kategoriler_for_combobox(self):
        self.c.execute("SELECT id, kategori_adi FROM urun_kategorileri ORDER BY kategori_adi ASC")
        return {row['kategori_adi']: row['id'] for row in self.c.fetchall()}

    def get_markalar_for_combobox(self):
        self.c.execute("SELECT id, marka_adi FROM urun_markalari ORDER BY marka_adi ASC")
        return {row['marka_adi']: row['id'] for row in self.c.fetchall()}

    def get_urun_gruplari_for_combobox(self):
        self.c.execute("SELECT id, grup_adi FROM urun_gruplari ORDER BY grup_adi ASC")
        return {row['grup_adi']: row['id'] for row in self.c.fetchall()}

    def get_urun_birimleri_for_combobox(self):
        self.c.execute("SELECT id, birim_adi FROM urun_birimleri ORDER BY birim_adi ASC")
        return {row['birim_adi']: row['id'] for row in self.c.fetchall()}

    def get_ulkeler_for_combobox(self):
        self.c.execute("SELECT id, ulke_adi FROM urun_ulkeleri ORDER BY ulke_adi ASC")
        return {row['ulke_adi']: row['id'] for row in self.c.fetchall()}

    def _fatura_hareketlerini_geri_al(self, fatura_ana_bilgileri, fatura_kalemleri_tuple_listesi):
        try:
            fatura_tipi = fatura_ana_bilgileri['tip']
            fatura_id = fatura_ana_bilgileri['id']
            odeme_turu = fatura_ana_bilgileri['odeme_turu']
            kasa_banka_id = fatura_ana_bilgileri['kasa_banka_id']
            toplam_kdv_dahil = fatura_ana_bilgileri['toplam_kdv_dahil']

            logging.info(f"_fatura_hareketlerini_geri_al çağrıldı. Fatura ID: {fatura_id}, Tip: {fatura_tipi}, Toplam KDV Dahil: {toplam_kdv_dahil}")

            # Stok hareketlerini geri al (bu kısım zaten doğru)
            for kalem in fatura_kalemleri_tuple_listesi:
                urun_id = kalem[0] # urun_id
                miktar = kalem[1]   # miktar

                miktar_degisimi = 0
                if fatura_tipi == self.FATURA_TIP_SATIS:
                    miktar_degisimi = miktar
                elif fatura_tipi == self.FATURA_TIP_ALIS:
                    miktar_degisimi = -miktar
                elif fatura_tipi == self.FATURA_TIP_SATIS_IADE:
                    miktar_degisimi = -miktar
                elif fatura_tipi == self.FATURA_TIP_ALIS_IADE:
                    miktar_degisimi = miktar

                self.c.execute("UPDATE tbl_stoklar SET stok_miktari = stok_miktari + ? WHERE id=?", (miktar_degisimi, urun_id))
                logging.info(f"Stok geri alındı: Ürün ID {urun_id}, Miktar Değişimi {miktar_degisimi}")


            # Kasa/Banka hareketlerini geri al
            if odeme_turu in self.pesin_odeme_turleri and kasa_banka_id:
                logging.info(f"Kasa/Banka hareketi geri alınıyor. Kasa/Banka ID: {kasa_banka_id}, Tutar: {toplam_kdv_dahil}, Fatura Tipi: {fatura_tipi}")
                if fatura_tipi == self.FATURA_TIP_SATIS:
                    self.kasa_banka_bakiye_guncelle(kasa_banka_id, toplam_kdv_dahil, artir=False)
                    logging.info(f"Kasa/Banka bakiyesi satış faturası geri alımı için azaltıldı: {toplam_kdv_dahil}")
                elif fatura_tipi == self.FATURA_TIP_ALIS:
                    self.kasa_banka_bakiye_guncelle(kasa_banka_id, toplam_kdv_dahil, artir=True)
                    logging.info(f"Kasa/Banka bakiyesi alış faturası geri alımı için artırıldı: {toplam_kdv_dahil}")
                elif fatura_tipi == self.FATURA_TIP_SATIS_IADE:
                    self.kasa_banka_bakiye_guncelle(kasa_banka_id, toplam_kdv_dahil, artir=True)
                    logging.info(f"Kasa/Banka bakiyesi satış iade faturası geri alımı için artırıldı: {toplam_kdv_dahil}")
                elif fatura_tipi == self.FATURA_TIP_ALIS_IADE:
                    self.kasa_banka_bakiye_guncelle(kasa_banka_id, toplam_kdv_dahil, artir=False)
                    logging.info(f"Kasa/Banka bakiyesi alış iade faturası geri alımı için azaltıldı: {toplam_kdv_dahil}")

            return True, "Hareketler başarıyla geri alındı."
        except Exception as e:
            logging.error(f"Fatura hareketlerini geri alma sırasında beklenmeyen hata: {e}\n{traceback.format_exc()}")
            return False, f"Fatura hareketleri geri alınırken beklenmeyen bir hata oluştu: {e}"

    def _fatura_hareketlerini_kaydet(self, fatura_ana_bilgileri, fatura_kalemleri_tuple_listesi):
        try:
            fatura_id = fatura_ana_bilgileri['id']
            fatura_tipi = fatura_ana_bilgileri['tip']
            odeme_turu = fatura_ana_bilgileri['odeme_turu']
            kasa_banka_id = fatura_ana_bilgileri['kasa_banka_id']
            toplam_kdv_dahil = fatura_ana_bilgileri['toplam_kdv_dahil']

            logging.info(f"_fatura_hareketlerini_kaydet çağrıldı. Fatura ID: {fatura_id}, Tip: {fatura_tipi}, Toplam KDV Dahil: {toplam_kdv_dahil}")

            # Stok hareketlerini kaydet (bu kısım zaten doğru)
            for kalem in fatura_kalemleri_tuple_listesi:
                urun_id = kalem[0] 
                miktar = kalem[1]   

                miktar_degisimi = 0
                if fatura_tipi == self.FATURA_TIP_SATIS: 
                    miktar_degisimi = -miktar
                elif fatura_tipi == self.FATURA_TIP_ALIS: 
                    miktar_degisimi = miktar
                elif fatura_tipi == self.FATURA_TIP_SATIS_IADE: 
                    miktar_degisimi = miktar
                elif fatura_tipi == self.FATURA_TIP_ALIS_IADE: 
                    miktar_degisimi = -miktar

                self.c.execute("UPDATE tbl_stoklar SET stok_miktari = stok_miktari + ? WHERE id=?", (miktar_degisimi, urun_id))
                logging.info(f"Stok kaydedildi: Ürün ID {urun_id}, Miktar Değişimi {miktar_degisimi}")


            # Kasa/Banka hareketlerini kaydet
            if odeme_turu in self.pesin_odeme_turleri and kasa_banka_id:
                logging.info(f"Kasa/Banka hareketi kaydediliyor. Kasa/Banka ID: {kasa_banka_id}, Tutar: {toplam_kdv_dahil}, Fatura Tipi: {fatura_tipi}")
                if fatura_tipi == self.FATURA_TIP_SATIS: 
                    self.kasa_banka_bakiye_guncelle(kasa_banka_id, toplam_kdv_dahil, artir=True)
                    logging.info(f"Kasa/Banka bakiyesi satış faturası için artırıldı: {toplam_kdv_dahil}")
                elif fatura_tipi == self.FATURA_TIP_ALIS: 
                    self.kasa_banka_bakiye_guncelle(kasa_banka_id, toplam_kdv_dahil, artir=False)
                    logging.info(f"Kasa/Banka bakiyesi alış faturası için azaltıldı: {toplam_kdv_dahil}")
                elif fatura_tipi == self.FATURA_TIP_SATIS_IADE: 
                    self.kasa_banka_bakiye_guncelle(kasa_banka_id, toplam_kdv_dahil, artir=False)
                    logging.info(f"Kasa/Banka bakiyesi satış iade faturası için azaltıldı: {toplam_kdv_dahil}")
                elif fatura_tipi == self.FATURA_TIP_ALIS_IADE: 
                    self.kasa_banka_bakiye_guncelle(kasa_banka_id, toplam_kdv_dahil, artir=True)
                    logging.info(f"Kasa/Banka bakiyesi alış iade faturası için artırıldı: {toplam_kdv_dahil}")

            return True, "Hareketler başarıyla kaydedildi."
        except Exception as e:
            logging.error(f"Fatura hareketlerini kaydetme sırasında beklenmeyen hata: {e}\n{traceback.format_exc()}")
            return False, f"Fatura hareketleri kaydedilirken beklenmeyen bir hata oluştu: {e}"

    def _format_currency(self, value):
        """Sayısal değeri Türkçe para birimi formatına dönüştürür."""
        if value is None: return "0,00 TL"
        try:
            # locale.format_string kullanımı için locale ayarının yapılmış olması gerekir.
            # yardimcilar.py'deki setup_locale() çağrılıyor olmalı.
            return locale.format_string("%.2f", self.safe_float(value), grouping=True) + " TL"
        except TypeError:
            return f"{self.safe_float(value):.2f} TL" # Hata durumunda basit formatlama
        except Exception:
            return f"{self.safe_float(value):.2f} TL" # Diğer hatalar için
                        
    def siparis_faturaya_donustur(self, siparis_id, olusturan_kullanici_id, odeme_turu_secilen, kasa_banka_id_secilen, vade_tarihi_secilen):
        """
        Belirtilen siparişi bir faturaya dönüştürür. Tüm mantık artık bu servistedir.
        """
        try:
            self.db.conn.execute("BEGIN TRANSACTION")
            siparis_ana = self.db.get_siparis_by_id(siparis_id)
            if not siparis_ana:
                return False, "Dönüştürülecek sipariş bulunamadı."
            
            if siparis_ana['fatura_id']:
                return False, f"Bu sipariş zaten bir faturaya dönüştürülmüş."
            if siparis_ana['durum'] == 'İPTAL EDİLDİ':
                return False, "İptal edilmiş bir sipariş faturaya dönüştürülemez."

            siparis_kalemleri = self.db.get_siparis_kalemleri(siparis_id)
            if not siparis_kalemleri:
                return False, "Sipariş kalemleri bulunamadı. Fatura oluşturulamıyor."

            fatura_tipi = 'SATIŞ' if siparis_ana['cari_tip'] == 'MUSTERI' else 'ALIŞ'
            fatura_no = self.db.son_fatura_no_getir(fatura_tipi)

            kalemler_for_fatura = []
            for sk in siparis_kalemleri:
                # fatura_olustur metodunun beklediği format:
                # (urun_id, miktar, birim_fiyat, kdv_orani, alis_fiyati, isk1, isk2, isk_tip, isk_deger)
                kalemler_for_fatura.append((
                    sk['urun_id'], sk['miktar'], sk['birim_fiyat'], sk['kdv_orani'], 
                    sk['alis_fiyati_siparis_aninda'], sk['iskonto_yuzde_1'], sk['iskonto_yuzde_2'], 
                    "YOK", 0.0
                ))

            success_fatura, message_fatura_id = self.fatura_olustur(
                fatura_no, fatura_tipi, siparis_ana['cari_id'], kalemler_for_fatura,
                odeme_turu=odeme_turu_secilen,
                kasa_banka_id=kasa_banka_id_secilen,
                fatura_notlari=f"Sipariş No: {siparis_ana['siparis_no']} ile oluşturulmuştur. {siparis_ana['siparis_notlari'] or ''}".strip(),
                vade_tarihi=vade_tarihi_secilen,
                genel_iskonto_tipi=siparis_ana['genel_iskonto_tipi'],
                genel_iskonto_degeri=siparis_ana['genel_iskonto_degeri'],
                manage_transaction=False
            )

            if success_fatura:
                yeni_fatura_id = message_fatura_id
                current_time = self.db.get_current_datetime_str()
                self.db.c.execute("UPDATE siparisler SET durum=?, fatura_id=?, son_guncelleme_tarihi_saat=?, son_guncelleyen_kullanici_id=? WHERE id=?",
                                  ('TAMAMLANDI', yeni_fatura_id, current_time, olusturan_kullanici_id, siparis_id))
                self.db.conn.commit()
                return True, f"Sipariş '{siparis_ana['siparis_no']}' başarıyla '{fatura_no}' nolu faturaya dönüştürüldü."
            else:
                self.db.conn.rollback()
                return False, f"Sipariş faturaya dönüştürülürken hata oluştu: {message_fatura_id}"

        except Exception as e:
            if self.db.conn: self.db.conn.rollback()
            return False, f"Sipariş faturaya dönüştürülürken beklenmeyen bir hata oluştu: {e}\n{traceback.format_exc()}"
                                        
    def get_siparisler_by_cari(self, cari_tip, cari_id):
        """
        Belirli bir müşteriye veya tedarikçiye ait tüm siparişleri,
        ilişkili fatura ve kullanıcı bilgileriyle birlikte getirir.
        """
        try:
            query = """
                SELECT 
                    s.id, s.siparis_no, s.tarih, s.teslimat_tarihi,
                    s.toplam_tutar, s.durum, s.siparis_notlari,
                    f.fatura_no AS iliskili_fatura_no,
                    u.kullanici_adi AS olusturan_kullanici
                FROM siparisler s
                LEFT JOIN faturalar f ON s.fatura_id = f.id
                LEFT JOIN kullanicilar u ON s.olusturan_kullanici_id = u.id
                WHERE s.cari_tip = ? AND s.cari_id = ?
                ORDER BY s.tarih DESC, s.id DESC
            """
            self.c.execute(query, (cari_tip, cari_id))
            return self.c.fetchall()
        except Exception as e:
            logging.error(f"get_siparisler_by_cari hatası: {e}\n{traceback.format_exc()}")
            return []
                
    def _get_or_create_id(self, table_name, name_column, name_value, manage_transaction=True): # manage_transaction parametresi eklendi
        if name_value is None or str(name_value).strip() == '':
            return None

        try:
            self.c.execute(f"SELECT id FROM {table_name} WHERE {name_column}=?", (name_value,))
            result = self.c.fetchone()
            if result:
                return result[0]
            else:
                if manage_transaction: # Sadece kendi transaction'ını yönetecekse başlat
                    self.conn.execute("BEGIN TRANSACTION")

                current_time = self.get_current_datetime_str()
                olusturan_id = self.app.current_user[0] if self.app and self.app.current_user else None 
                self.c.execute(f"INSERT INTO {table_name} ({name_column}, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?)",
                               (name_value, current_time, olusturan_id))
                new_id = self.c.lastrowid

                if manage_transaction: # Sadece kendi transaction'ını yönetecekse commit yap
                    self.conn.commit() 
                return new_id
        except sqlite3.IntegrityError as e: # Hata objesini yakala
            if manage_transaction: self.conn.rollback() 
            # Yeniden sorgula. Eğer bu bir UNIQUE hatasıysa, kayıt zaten eklenmiş demektir.
            self.c.execute(f"SELECT id FROM {table_name} WHERE {name_column}=?", (name_value,))
            result = self.c.fetchone()
            if result:
                return result[0]
            else:
                # Bu durum, IntegrityError'ın farklı bir tipi (FK ihlali vb.) veya yarış durumu olabilir.
                # Ya da kayıt eklenirken aynı anda başka bir işlem tarafından eklendiği için
                # tekrar SELECT yapıldığında bulunamaması durumu.
                print(f"KRİTİK HATA (IntegrityError): '{table_name}' tablosunda '{name_value}' için benzersizlik/bütünlük hatası: {e}. Kayıt tekrar bulunamadı.")
                return None
        except Exception as e:
            if manage_transaction: self.conn.rollback() 
            print(f"UYARI: '{table_name}' tablosunda '{name_value}' için ID bulunamadı veya oluşturulamadı: {e}\n{traceback.format_exc()}") # traceback ekle
            return None
            
    def fatura_listele_urun_ad_dahil(self, tip=None, baslangic_tarih=None, bitis_tarih=None, arama_terimi=None, cari_id_filter=None, odeme_turu_filter=None, kasa_banka_id_filter=None, limit=None, offset=None):
        
        # Bu metod, iade faturalarını da içerecek şekilde güncellendi.
        # Sorgu, farklı tablolardan (musteriler, tedarikciler, kullanicilar vb.) veri almak için JOIN'ler kullanır.
        
        perakende_id_param = self.perakende_musteri_id if self.perakende_musteri_id is not None else -999

        q = """SELECT f.id, f.fatura_no, f.tarih, f.tip,
                       CASE
                           WHEN f.cari_id = ? AND f.tip = 'SATIŞ' THEN IFNULL(f.misafir_adi, 'Perakende Satış')
                           WHEN f.tip IN ('SATIŞ', 'SATIŞ İADE') THEN mus.ad
                           WHEN f.tip IN ('ALIŞ', 'ALIŞ İADE', 'DEVİR_GİRİŞ') THEN ted.ad
                           ELSE 'Bilinmeyen Cari'
                       END AS cari_adi,
                       f.toplam_kdv_dahil, 
                       f.odeme_turu,
                       kb.hesap_adi AS kasa_banka_adi,
                       f.vade_tarihi,
                       f.genel_iskonto_degeri,
                       olusturan.kullanici_adi AS olusturan_kul_adi,
                       guncelleyen.kullanici_adi AS guncelleyen_kul_adi
               FROM faturalar f
               LEFT JOIN musteriler mus ON f.cari_id = mus.id
               LEFT JOIN tedarikciler ted ON f.cari_id = ted.id
               LEFT JOIN kasalar_bankalar kb ON f.kasa_banka_id = kb.id
               LEFT JOIN kullanicilar olusturan ON f.olusturan_kullanici_id = olusturan.id
               LEFT JOIN kullanicilar guncelleyen ON f.son_guncelleyen_kullanici_id = guncelleyen.id
               """
        p = [perakende_id_param] 
        conditions = []

        if tip:
            if isinstance(tip, list):
                placeholders = ','.join(['?'] * len(tip))
                conditions.append(f"f.tip IN ({placeholders})")
                p.extend(tip)
            elif tip != "TÜMÜ":
                conditions.append("f.tip=?")
                p.append(tip)

        if baslangic_tarih:
            conditions.append("f.tarih>=?")
            p.append(baslangic_tarih)
        if bitis_tarih:
            conditions.append("f.tarih<=?")
            p.append(bitis_tarih)

        if cari_id_filter is not None:
            conditions.append("f.cari_id = ?")
            p.append(int(cari_id_filter))

        if odeme_turu_filter is not None and odeme_turu_filter != "TÜMÜ":
            conditions.append("f.odeme_turu = ?")
            p.append(odeme_turu_filter)

        if kasa_banka_id_filter is not None:
            conditions.append("f.kasa_banka_id = ?")
            p.append(kasa_banka_id_filter)

        if arama_terimi:
            term = f"%{normalize_turkish_chars(arama_terimi)}%"
            # Arama sorgusunu daha basit ve okunabilir hale getirelim
            # NOT: Bu arama performansı büyük veritabanlarında yavaş olabilir.
            # Daha iyi performans için FTS5 gibi SQLite eklentileri gerekir.
            search_conditions = [
                "f.fatura_no LIKE ?",
                "f.misafir_adi LIKE ?",
                "mus.ad LIKE ?",
                "ted.ad LIKE ?",
                "kb.hesap_adi LIKE ?",
                """EXISTS (
                    SELECT 1 FROM fatura_kalemleri fk 
                    JOIN tbl_stoklar s ON fk.urun_id = s.id 
                    WHERE fk.fatura_id = f.id AND (s.urun_adi LIKE ? OR s.urun_kodu LIKE ?)
                )"""
            ]
            conditions.append(f"({' OR '.join(search_conditions)})")
            # Her bir LIKE için parametre ekle
            p.extend([term, term, term, term, term, term, term])

        if conditions:
            q += " WHERE " + " AND ".join(conditions)

        q += " ORDER BY f.tarih DESC, f.id DESC"

        if limit is not None:
            q += " LIMIT ?"
            p.append(limit)
        if offset is not None:
            q += " OFFSET ?"
            p.append(offset)
        
        self.c.execute(q, p)
        return self.c.fetchall()

    def get_fatura_count(self, tip=None, baslangic_tarih=None, bitis_tarih=None, arama_terimi=None, cari_id_filter=None, odeme_turu_filter=None, kasa_banka_id_filter=None):
        perakende_id_param = self.perakende_musteri_id if self.perakende_musteri_id is not None else -999

        q = """SELECT COUNT(DISTINCT f.id)
               FROM faturalar f
               LEFT JOIN musteriler mus ON f.cari_id = mus.id
               LEFT JOIN tedarikciler ted ON f.cari_id = ted.id
               LEFT JOIN kasalar_bankalar kb ON f.kasa_banka_id = kb.id
               LEFT JOIN fatura_kalemleri fk ON f.id = fk.fatura_id
               LEFT JOIN tbl_stoklar s ON fk.urun_id = s.id
               """

        p = []
        conditions = []

        conditions.append("""
            (
                (f.tip = 'SATIŞ' AND f.cari_id = ? AND IFNULL(f.misafir_adi, '') IS NOT NULL) OR 
                (f.tip = 'SATIŞ' AND f.cari_id != ?) OR 
                (f.tip = 'ALIŞ') OR
                (f.tip = 'SATIŞ İADE') OR 
                (f.tip = 'ALIŞ İADE')
            )
        """)
        p.append(perakende_id_param)
        p.append(perakende_id_param)
        
        # Düzeltme burada: tip parametresi artık bir liste veya string olabilir.
        if tip:
            if isinstance(tip, list): # Eğer tip bir liste ise
                placeholders = ','.join(['?'] * len(tip))
                conditions.append(f"f.tip IN ({placeholders})")
                p.extend(tip)
            elif tip != "TÜMÜ": # Tek bir string ise ve "TÜMÜ" değilse
                conditions.append("f.tip=?")
                p.append(tip)

        if baslangic_tarih:
            conditions.append("f.tarih>=?")
            p.append(baslangic_tarih)
        if bitis_tarih:
            conditions.append("f.tarih<=?")
            p.append(bitis_tarih)

        if cari_id_filter is not None:
            conditions.append("f.cari_id = ?")
            p.append(int(cari_id_filter))

        if odeme_turu_filter is not None:
            conditions.append("f.odeme_turu = ?")
            p.append(odeme_turu_filter)

        if kasa_banka_id_filter is not None:
            conditions.append("f.kasa_banka_id = ?")
            p.append(kasa_banka_id_filter)

        if arama_terimi:
            term = f"%{normalize_turkish_chars(arama_terimi)}%"
            conditions.append("""(
                                f.fatura_no LIKE ? OR
                                (f.tip IN ('SATIŞ', 'SATIŞ İADE') AND IFNULL(f.misafir_adi, 'Perakende Satış') LIKE ?) OR
                                (f.tip IN ('SATIŞ', 'SATIŞ İADE') AND REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LOWER(mus.ad), 'ş', 's'), 'ı', 'i'), 'ç', 'c'), 'ğ', 'g'), 'ö', 'o'), 'ü', 'u') LIKE ?) OR
                                (f.tip IN ('ALIŞ', 'ALIŞ İADE') AND REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LOWER(ted.ad), 'ş', 's'), 'ı', 'i'), 'ç', 'c'), 'ğ', 'g'), 'ö', 'o'), 'ü', 'u') LIKE ?) OR
                                REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LOWER(kb.hesap_adi), 'ş', 's'), 'ı', 'i'), 'ç', 'c'), 'ğ', 'g'), 'ö', 'o'), 'ü', 'u') LIKE ? OR
                                EXISTS (SELECT 1 FROM fatura_kalemleri fk JOIN tbl_stoklar s ON fk.urun_id = s.id WHERE fk.fatura_id = f.id AND (REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LOWER(s.urun_adi), 'ş', 's'), 'ı', 'i'), 'ç', 'c'), 'ğ', 'g'), 'ö', 'o'), 'ü', 'u') LIKE ? OR REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(LOWER(s.urun_kodu), 'ş', 's'), 'ı', 'i'), 'ç', 'c'), 'ğ', 'g'), 'ö', 'o'), 'ü', 'u') LIKE ?))
                              )""")
            p.extend([term, term, term, term, term, term, term])

        if conditions:
            q += " WHERE " + " AND ".join(conditions)

        print(f"DEBUG SQL Query (get_fatura_count): {q}")
        print(f"DEBUG SQL Params (get_fatura_count): {p}")

        self.c.execute(q, p)
        return self.c.fetchone()[0]
            
    def fatura_getir_by_id(self, fatura_id):
        self.c.execute("SELECT id,fatura_no,tarih,tip,cari_id,toplam_kdv_haric,toplam_kdv_dahil,odeme_turu,misafir_adi,kasa_banka_id,olusturma_tarihi_saat,olusturan_kullanici_id,son_guncelleme_tarihi_saat,son_guncelleyen_kullanici_id, fatura_notlari, vade_tarihi, genel_iskonto_tipi, genel_iskonto_degeri, orijinal_fatura_id FROM faturalar WHERE id=?", (fatura_id,))
        return self.c.fetchone()
    
    def fatura_detay_al(self, fatura_id):
        self.c.execute("SELECT s.urun_kodu, s.urun_adi, fk.miktar, fk.birim_fiyat, fk.kdv_orani, fk.kdv_tutari, fk.kalem_toplam_kdv_haric, fk.kalem_toplam_kdv_dahil, s.id as urun_id, fk.alis_fiyati_fatura_aninda, fk.kdv_orani_fatura_aninda, fk.iskonto_yuzde_1, fk.iskonto_yuzde_2, fk.iskonto_tipi, fk.iskonto_degeri FROM fatura_kalemleri fk JOIN tbl_stoklar s ON fk.urun_id=s.id WHERE fk.fatura_id=?",(fatura_id,))
        return self.c.fetchall()

    def gelir_siniflandirma_ekle(self, siniflandirma_adi):
        if not siniflandirma_adi:
            return False, "Gelir sınıflandırma adı boş olamaz."
        try:
            self.conn.execute("BEGIN TRANSACTION")
            current_time = self.get_current_datetime_str()
            olusturan_id = self._get_current_user_id() # Değişiklik
            self.c.execute("INSERT INTO gelir_siniflandirmalari (siniflandirma_adi, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?)",
                           (siniflandirma_adi, current_time, olusturan_id))
            self.conn.commit()
            return True, f"'{siniflandirma_adi}' gelir sınıflandırması başarıyla eklendi."
        except sqlite3.IntegrityError:
            self.conn.rollback()
            return False, "Bu gelir sınıflandırma adı zaten mevcut."
        except Exception as e:
            self.conn.rollback()
            logging.error(f"Gelir sınıflandırma eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Gelir sınıflandırma eklenirken beklenmeyen hata: {e}"

    def gelir_siniflandirma_listele(self):
        self.c.execute("SELECT id, siniflandirma_adi FROM gelir_siniflandirmalari ORDER BY siniflandirma_adi ASC")
        return self.c.fetchall()

    def gelir_siniflandirma_guncelle(self, siniflandirma_id, yeni_siniflandirma_adi):
        if not yeni_siniflandirma_adi:
            return False, "Gelir sınıflandırma adı boş olamaz."
        try:
            self.conn.execute("BEGIN TRANSACTION")
            current_time = self.get_current_datetime_str()
            guncelleyen_id = self.app.current_user[0] if self.app and self.app.current_user else None
            self.c.execute("UPDATE gelir_siniflandirmalari SET siniflandirma_adi=?, son_guncelleme_tarihi_saat=?, son_guncelleyen_kullanici_id=? WHERE id=?",
                           (yeni_siniflandirma_adi, current_time, guncelleyen_id, siniflandirma_id))
            self.conn.commit()
            if self.c.rowcount > 0:
                return True, "Gelir sınıflandırması başarıyla güncellendi."
            else:
                return False, "Gelir sınıflandırması bulunamadı veya bir değişiklik yapılmadı."
        except sqlite3.IntegrityError:
            self.conn.rollback()
            return False, "Bu gelir sınıflandırma adı zaten mevcut."
        except Exception as e:
            self.conn.rollback()
            logging.error(f"Gelir sınıflandırma güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Gelir sınıflandırma güncellenirken beklenmeyen hata: {e}"

    def gelir_siniflandirma_sil(self, siniflandirma_id):
        try:
            self.conn.execute("BEGIN TRANSACTION")
            # Bu sınıflandırmaya bağlı gelir kaydı var mı kontrol et
            self.c.execute("SELECT COUNT(*) FROM gelir_gider WHERE gelir_siniflandirma_id=?", (siniflandirma_id,))
            if self.c.fetchone()[0] > 0:
                self.conn.rollback()
                return False, "Bu gelir sınıflandırmasına bağlı gelir kayıtları bulunmaktadır. Lütfen önce kayıtları güncelleyin veya silin."
            self.c.execute("DELETE FROM gelir_siniflandirmalari WHERE id=?", (siniflandirma_id,))
            self.conn.commit()
            if self.c.rowcount > 0:
                return True, "Gelir sınıflandırması başarıyla silindi."
            else:
                return False, "Gelir sınıflandırması bulunamadı veya silinemedi."
        except Exception as e:
            self.conn.rollback()
            logging.error(f"Gelir sınıflandırma silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Gelir sınıflandırma silinirken beklenmeyen hata: {e}"

    # --- Gider Sınıflandırma Yönetimi Metotları ---
    def gider_siniflandirma_ekle(self, siniflandirma_adi):
        if not siniflandirma_adi:
            return False, "Gider sınıflandırma adı boş olamaz."
        try:
            self.conn.execute("BEGIN TRANSACTION")
            current_time = self.get_current_datetime_str()
            olusturan_id = self._get_current_user_id() # Değişiklik
            self.c.execute("INSERT INTO gider_siniflandirmalari (siniflandirma_adi, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?)",
                           (siniflandirma_adi, current_time, olusturan_id))
            self.conn.commit()
            return True, f"'{siniflandirma_adi}' gider sınıflandırması başarıyla eklendi."
        except sqlite3.IntegrityError:
            self.conn.rollback()
            return False, "Bu gider sınıflandırma adı zaten mevcut."
        except Exception as e:
            self.conn.rollback()
            logging.error(f"Gider sınıflandırma eklenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Gider sınıflandırma eklenirken beklenmeyen hata: {e}"

    def gider_siniflandirma_listele(self):
        self.c.execute("SELECT id, siniflandirma_adi FROM gider_siniflandirmalari ORDER BY siniflandirma_adi ASC")
        return self.c.fetchall()

    def gider_siniflandirma_guncelle(self, siniflandirma_id, yeni_siniflandirma_adi):
        if not yeni_siniflandirma_adi:
            return False, "Gider sınıflandırma adı boş olamaz."
        try:
            self.conn.execute("BEGIN TRANSACTION")
            current_time = self.get_current_datetime_str()
            guncelleyen_id = self.app.current_user[0] if self.app and self.app.current_user else None
            self.c.execute("UPDATE gider_siniflandirmalari SET siniflandirma_adi=?, son_guncelleme_tarihi_saat=?, son_guncelleyen_kullanici_id=? WHERE id=?",
                           (yeni_siniflandirma_adi, current_time, guncelleyen_id, siniflandirma_id))
            self.conn.commit()
            if self.c.rowcount > 0:
                return True, "Gider sınıflandırması başarıyla güncellendi."
            else:
                return False, "Gider sınıflandırması bulunamadı veya bir değişiklik yapılmadı."
        except sqlite3.IntegrityError:
            self.conn.rollback()
            return False, "Bu gider sınıflandırma adı zaten mevcut."
        except Exception as e:
            self.conn.rollback()
            logging.error(f"Gider sınıflandırma güncellenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Gider sınıflandırma güncellenirken beklenmeyen hata: {e}"

    def gider_siniflandirma_sil(self, siniflandirma_id):
        try:
            self.conn.execute("BEGIN TRANSACTION")
            # Bu sınıflandırmaya bağlı gider kaydı var mı kontrol et
            self.c.execute("SELECT COUNT(*) FROM gelir_gider WHERE gider_siniflandirma_id=?", (siniflandirma_id,))
            if self.c.fetchone()[0] > 0:
                self.conn.rollback()
                return False, "Bu gider sınıflandırmasına bağlı gider kayıtları bulunmaktadır. Lütfen önce kayıtları güncelleyin veya silin."
            self.c.execute("DELETE FROM gider_siniflandirmalari WHERE id=?", (siniflandirma_id,))
            self.conn.commit()
            if self.c.rowcount > 0:
                return True, "Gider sınıflandırması başarıyla silindi."
            else:
                return False, "Gider sınıflandırması bulunamadı veya silinemedi."
        except Exception as e:
            self.conn.rollback()
            logging.error(f"Gider sınıflandırma silinirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Gider sınıflandırma silinirken beklenmeyen hata: {e}"

    # --- Yardımcı Metot: Sınıflandırma ID'sini ada göre getir (combobox'lar için) ---
    def get_gelir_siniflandirmalari_for_combobox(self):
        self.c.execute("SELECT id, siniflandirma_adi FROM gelir_siniflandirmalari ORDER BY siniflandirma_adi ASC")
        return {row['siniflandirma_adi']: row['id'] for row in self.c.fetchall()}

    def get_gider_siniflandirmalari_for_combobox(self):
        self.c.execute("SELECT id, siniflandirma_adi FROM gider_siniflandirmalari ORDER BY siniflandirma_adi ASC")
        return {row['siniflandirma_adi']: row['id'] for row in self.c.fetchall()}

    def gelir_gider_ekle(self, tarih, tip, tutar, aciklama, kasa_banka_id=None, gelir_siniflandirma_id=None, gider_siniflandirma_id=None):
        if not (tarih and tip and aciklama):
            return False, "Tarih, Tip ve Açıklama alanları zorunludur."
        
        tutar_f = 0.0
        try:
            tutar_f = float(str(tutar).replace(',','.'))
            if tutar_f <= 0:
                return False, "Tutar pozitif bir sayı olmalıdır."
            datetime.strptime(tarih, '%Y-%m-%d') # Tarih formatını doğrula
        except ValueError:
            return False, "Tutar sayısal ve geçerli bir değer olmalı.\nTarih formatı (YYYY-AA-GG) şeklinde olmalıdır (örn: 2024-12-31)."
        
        try:
            self.conn.execute("BEGIN TRANSACTION") # Transaction başlatıldı
            current_time = self.get_current_datetime_str()
            olusturan_id = self.app.current_user[0] if self.app and self.app.current_user else None
            
            # Düzeltilen kısım: 'kaynak' için '?' ve parametre listesine 'MANUEL' eklendi.
            self.c.execute("INSERT INTO gelir_gider (tarih, tip, tutar, aciklama, kaynak, kasa_banka_id, gelir_siniflandirma_id, gider_siniflandirma_id, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?,?,?,?,?,?,?,?)",
                           (tarih, tip, tutar_f, aciklama, 'MANUEL', kasa_banka_id, gelir_siniflandirma_id, gider_siniflandirma_id, current_time, olusturan_id))
            
            gg_id = self.c.lastrowid
            
            if kasa_banka_id: # Kasa/Banka seçiliyse bakiyeyi güncelle
                is_bakiye_artir_gg = True if tip == 'GELİR' else False
                if not self.kasa_banka_bakiye_guncelle(kasa_banka_id, tutar_f, artir=is_bakiye_artir_gg):
                    self.conn.rollback() # Bakiye güncellenemezse rollback
                    return False, f"Manuel gelir/gider eklenirken bir hata oluştu: {e}\n\nDetaylar:\n{traceback.format_exc()}" # traceback eklendi

            self.conn.commit() # Tüm işlemler başarılıysa commit et
            return True, f"Manuel {tip.lower()} kaydı başarıyla eklendi." # Mesaj eklendi
        except Exception as e:
            self.conn.rollback() # Hata durumunda rollback
            error_details = traceback.format_exc()
            return False, f"Manuel gelir/gider eklenirken bir hata oluştu: {e}\n\nDetaylar:\n{error_details}"

    def gelir_gider_listele(self, baslangic_tarih=None, bitis_tarih=None, tip_filtre=None, aciklama_filtre=None, limit=None, offset=None): 
        q="SELECT gg.id, gg.tarih, gg.tip, gg.tutar, gg.aciklama, gg.kaynak, gg.kaynak_id, kb.hesap_adi as kasa_banka_adi FROM gelir_gider gg LEFT JOIN kasalar_bankalar kb ON gg.kasa_banka_id = kb.id"
        p=[]; cond=[]
        if baslangic_tarih: cond.append("gg.tarih>=?"); p.append(baslangic_tarih)
        if bitis_tarih: cond.append("gg.tarih<=?"); p.append(bitis_tarih)
        if tip_filtre and tip_filtre!="TÜMÜ": cond.append("gg.tip=?"); p.append(tip_filtre)
        if aciklama_filtre: cond.append("(gg.aciklama LIKE ? OR kb.hesap_adi LIKE ?)"); p.extend([f"%{aciklama_filtre}%", f"%{aciklama_filtre}%"])
        if cond: q+=" WHERE "+" AND ".join(cond)
        q+=" ORDER BY gg.tarih DESC,gg.id DESC"
        
        # Sayfalama için LIMIT ve OFFSET ekle
        if limit is not None:
            q += " LIMIT ?"
            p.append(limit)
        if offset is not None:
            q += " OFFSET ?"
            p.append(offset)
        
        self.c.execute(q,p)
        return self.c.fetchall()
    
    def get_gelir_gider_count(self, baslangic_tarih=None, bitis_tarih=None, tip_filtre=None, aciklama_filtre=None):
        q = "SELECT COUNT(gg.id) FROM gelir_gider gg LEFT JOIN kasalar_bankalar kb ON gg.kasa_banka_id = kb.id"
        p = []; cond = []
        if baslangic_tarih: cond.append("gg.tarih>=?"); p.append(baslangic_tarih)
        if bitis_tarih: cond.append("gg.tarih<=?"); p.append(bitis_tarih)
        if tip_filtre and tip_filtre != "TÜMÜ": cond.append("gg.tip=?"); p.append(tip_filtre)
        if aciklama_filtre: cond.append("(gg.aciklama LIKE ? OR kb.hesap_adi LIKE ?)"); p.extend([f"%{aciklama_filtre}%", f"%{aciklama_filtre}%"])
        if cond: q += " WHERE " + " AND ".join(cond)
        self.c.execute(q, p)
        return self.c.fetchone()[0]  

    def gelir_gider_sil(self, gg_id):
        try:
            self.conn.execute("BEGIN TRANSACTION")
            # Sadece manuel eklenenler silinebilir
            self.c.execute("SELECT kaynak, tutar, tip, kasa_banka_id FROM gelir_gider WHERE id=?",(gg_id,))
            kaynak_bilgisi = self.c.fetchone()
            if not kaynak_bilgisi:
                self.conn.rollback()
                return False, "Silinecek gelir/gider kaydı bulunamadı."
            kaynak, tutar_gg, tip_gg, kasa_banka_id_gg = kaynak_bilgisi
            if kaynak !='MANUEL':
                self.conn.rollback()
                return False, "Sadece manuel olarak eklenmiş gelir/gider kayıtları silinebilir.\nOtomatik oluşan kayıtlar (Fatura, Tahsilat, Ödeme vb.) ilgili modüllerden yönetilmelidir."

            if kasa_banka_id_gg: # Kasa/Banka seçiliyse bakiyeyi düzelt
                is_ters_bakiye_artir_gg = False # Varsayılan: azalt
                if tip_gg == 'GELİR':
                    is_ters_bakiye_artir_gg = False # Gelir silinirse kasa bakiyesi azalır (artır=False)
                elif tip_gg == 'GİDER':
                    is_ters_bakiye_artir_gg = True # Gider silinirse kasa bakiyesi artar (artır=True)

                if not self.kasa_banka_bakiye_guncelle(kasa_banka_id_gg, tutar_gg, artir=is_ters_bakiye_artir_gg):
                    self.conn.rollback()
                    return False, "Manuel gelir/gider silinirken kasa/banka bakiyesi düzeltilemedi."
            
            self.c.execute("DELETE FROM gelir_gider WHERE id=?",(gg_id,))
            self.conn.commit()
            # DÜZELTME BAŞLANGICI: Başarı durumu ve mesaj döndürüyoruz
            if self.c.rowcount > 0:
                return True, "Gelir/Gider kaydı başarıyla silindi."
            else:
                return False, "Gelir/Gider kaydı bulunamadı veya silinemedi."
            
        except Exception as e:
            self.conn.rollback()
            return False, f"Gelir/Gider silme hatası: {e}"

    def tahsilat_ekle(self, musteri_id, tarih, tutar, odeme_sekli, aciklama, kasa_banka_id=None):
        if str(musteri_id) == str(self.perakende_musteri_id):
            return False, "Perakende müşterisinden manuel tahsilat yapılamaz.\nPerakende satışlar zaten peşin kabul edilir."
        if not (musteri_id and tarih and odeme_sekli and aciklama and kasa_banka_id): 
            return False, "Müşteri, Tarih, Ödeme Şekli, Açıklama ve İşlem Kasa/Banka zorunludur."
        try:
            tutar_f = float(str(tutar).replace(",","."))
            if tutar_f <= 0:
                return False, "Tutar pozitif bir sayı olmalıdır."
            datetime.strptime(tarih, '%Y-%m-%d')
        except ValueError:
            return False, "Tutar sayısal olmalı, Tarih (%Y-%m-%d) formatında olmalıdır."
    
        try:
            self.conn.execute("BEGIN TRANSACTION")
            musteri_bilgi = self.musteri_getir_by_id(musteri_id)
            musteri_adi = musteri_bilgi[2] if musteri_bilgi else f"ID: {musteri_id}"

            ch_aciklama = f"Tahsilat: {aciklama} (Ödeme: {odeme_sekli}) - Müşteri: {musteri_adi}"
            current_time = self.get_current_datetime_str()
            olusturan_id = self.app.current_user[0] if self.app and self.app.current_user else None
            self.c.execute("INSERT INTO cari_hareketler (tarih, cari_tip, cari_id, islem_tipi, tutar, aciklama, referans_tip, kasa_banka_id,olusturma_tarihi_saat,olusturan_kullanici_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                           (tarih, 'MUSTERI', musteri_id, 'TAHSILAT', tutar_f, ch_aciklama, "TAHSILAT", kasa_banka_id, current_time, olusturan_id))
            tahsilat_ref_id = self.c.lastrowid

            gelir_aciklama = f"Müşteri Tahsilatı: {musteri_adi} - {aciklama} (Ödeme: {odeme_sekli})"
            current_time_gg = self.get_current_datetime_str()
            olusturan_id_gg = self.app.current_user[0] if self.app and self.app.current_user else None
            self.c.execute("INSERT INTO gelir_gider (tarih, tip, tutar, aciklama, kaynak, kaynak_id, kasa_banka_id,olusturma_tarihi_saat,olusturan_kullanici_id) VALUES (?,?,?,?,?,?,?,?,?)",
                           (tarih, 'GELİR', tutar_f, gelir_aciklama, 'TAHSILAT', tahsilat_ref_id, kasa_banka_id, current_time_gg, olusturan_id_gg))            

            if not self.kasa_banka_bakiye_guncelle(kasa_banka_id, tutar_f, artir=True):
                self.conn.rollback()
                return False, "Tahsilat kaydedilirken kasa/banka bakiyesi güncellenemedi."
        
            self.conn.commit()
            # DÜZELTME BURADA: Başarılı durumda (True, "mesaj") formatında dönüyoruz.
            return True, f"Tahsilat (ID: {tahsilat_ref_id}) başarıyla kaydedildi."
        except Exception as e:
            self.conn.rollback()
            return False, f"Tahsilat eklenirken hata: {e}\n{traceback.format_exc()}"

    def odeme_ekle(self, tedarikci_id, tarih, tutar, odeme_sekli, aciklama, kasa_banka_id):
        if not (tedarikci_id and tarih and odeme_sekli and aciklama and kasa_banka_id):
            return False, "Tedarikçi, Tarih, Ödeme Şekli, Açıklama ve İşlem Kasa/Banka zorunludur."
        try:
            tutar_f = float(str(tutar).replace(",", "."))
            if tutar_f <= 0:
                return False, "Tutar pozitif bir sayı olmalıdır."
            datetime.strptime(tarih, '%Y-%m-%d')
        except ValueError:
            return False, "Tutar sayısal olmalı, Tarih (YYYY-AA-GG) formatında olmalıdır."

        try:
            self.conn.execute("BEGIN TRANSACTION")
            tedarikci_bilgi = self.tedarikci_getir_by_id(tedarikci_id)
            tedarikci_adi = tedarikci_bilgi[2] if tedarikci_bilgi else f"ID: {tedarikci_id}"

            ch_aciklama = f"Ödeme: {aciklama} (Ödeme: {odeme_sekli}) - Tedarikçi: {tedarikci_adi}"
            current_time = self.get_current_datetime_str()
            olusturan_id = self.app.current_user[0] if self.app and self.app.current_user else None
            
            self.c.execute("INSERT INTO cari_hareketler (tarih, cari_tip, cari_id, islem_tipi, tutar, aciklama, referans_id, referans_tip, kasa_banka_id, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                           (tarih, 'TEDARIKCI', tedarikci_id, 'ODEME', tutar_f, ch_aciklama, None, "ODEME", kasa_banka_id, current_time, olusturan_id))
            odeme_ref_id = self.c.lastrowid

            gider_aciklama = f"Tedarikçi Ödemesi: {tedarikci_adi} - {aciklama} (Ödeme: {odeme_sekli})"
            current_time_gg = self.get_current_datetime_str()
            olusturan_id_gg = self.app.current_user[0] if self.app and self.app.current_user else None
            self.c.execute("INSERT INTO gelir_gider (tarih, tip, tutar, aciklama, kaynak, kaynak_id, kasa_banka_id,olusturma_tarihi_saat,olusturan_kullanici_id) VALUES (?,?,?,?,?,?,?,?,?)",
                           (tarih, 'GİDER', tutar_f, gider_aciklama, 'ODEME', odeme_ref_id, kasa_banka_id, current_time_gg, olusturan_id_gg))

            if not self.kasa_banka_bakiye_guncelle(kasa_banka_id, tutar_f, artir=False):
                self.conn.rollback()
                return False, "Ödeme kaydedilirken kasa/banka bakiyesi güncellenemedi."

            self.conn.commit()
            # DÜZELTME BURADA: Başarılı durumda (True, "mesaj") formatında dönüyoruz.
            return True, f"Ödeme (ID: {odeme_ref_id}) başarıyla kaydedildi."
        except Exception as e:
            self.conn.rollback()
            return False, f"Ödeme eklenirken hata: {e}\n{traceback.format_exc()}"

    def tahsilat_odeme_sil(self, cari_hareket_id):
        referans_tipi = "" 
        try:
            self.conn.execute("BEGIN TRANSACTION") # Transaction başlatıldı
            # Silinecek cari hareketin detaylarını al
            self.c.execute("SELECT referans_tip, referans_id, tutar, kasa_banka_id, cari_tip FROM cari_hareketler WHERE id=?", (cari_hareket_id,))
            hareket_bilgisi = self.c.fetchone()
            if not hareket_bilgisi:
                self.conn.rollback() # Kayıt bulunamazsa rollback
                return False, "Silinecek kayıt bulunamadı."
            
            referans_tipi, referans_id, tutar_silinecek_ch, kasa_banka_id_ch, cari_tip_ch = hareket_bilgisi
    
            # Sadece nakit akışı yaratan işlemler için kasa/banka bakiyesini düzelt
            if referans_tipi in ['TAHSILAT', 'ODEME', 'FATURA_SATIS_PESIN', 'FATURA_ALIS_PESIN'] and kasa_banka_id_ch is not None:
                is_ters_bakiye_artir_ch = False # Varsayılan: azalt
                if referans_tipi == 'TAHSILAT' or referans_tipi == 'FATURA_SATIS_PESIN':
                    # Tahsilat silinirse (kasaya para girmişti) kasadan para AZALMALI (artir=False)
                    is_ters_bakiye_artir_ch = False 
                elif referans_tipi == 'ODEME' or referans_tipi == 'FATURA_ALIS_PESIN':
                    # Ödeme silinirse (kasadan para çıkmıştı) kasaya para ARTMALI (artir=True)
                    is_ters_bakiye_artir_ch = True 
        
                if not self.kasa_banka_bakiye_guncelle(kasa_banka_id_ch, tutar_silinecek_ch, artir=is_ters_bakiye_artir_ch):
                    self.conn.rollback() # Bakiye düzeltme başarısız olursa rollback
                    return False, "İşlem silinirken bakiye düzeltilemedi."

            # Eğer bu hareket peşin bir faturadan otomatik oluşmuşsa (FATURA_SATIS_PESIN veya FATURA_ALIS_PESIN)
            # Hem cari hareketi hem de ilişkili gelir/gider kaydını sil. Faturayı "AÇIK HESAP" durumuna getir.
            if referans_tipi in ['FATURA_SATIS_PESIN', 'FATURA_ALIS_PESIN']:
                if referans_id: # Fatura ID'si varsa
                    self.c.execute("UPDATE faturalar SET odeme_turu = 'AÇIK HESAP', kasa_banka_id = NULL WHERE id = ?", (referans_id,))
                self.c.execute("DELETE FROM cari_hareketler WHERE id=?", (cari_hareket_id,))
                # İlişkili gelir/gider kaydını sil (FATURA_SATIS_PESIN ve FATURA_ALIS_PESIN'den gelen gelir/giderler)
                self.c.execute("DELETE FROM gelir_gider WHERE kaynak='FATURA' AND kaynak_id = ?", (referans_id,)) 
    
            # Eğer bu hareket manuel bir tahsilat veya ödeme ise (TAHSILAT veya ODEME)
            # Hem cari hareketi hem de ilişkili gelir/gider kaydını sil
            elif referans_tipi in ['TAHSILAT', 'ODEME']:
                self.c.execute("DELETE FROM gelir_gider WHERE kaynak=? AND kaynak_id=? AND kasa_banka_id=?", (referans_tipi, cari_hareket_id, kasa_banka_id_ch))
                self.c.execute("DELETE FROM cari_hareketler WHERE id=?", (cari_hareket_id,))
            elif referans_tipi == 'VERESIYE_BORC_MANUEL': # Manuel veresiye borç silme
                # Bu sadece cari hareketler tablosunda bir kayıttır, gelir_gider tablosunda karşılığı yoktur
                # ve kasa/banka bakiyesi üzerinde etkisi yoktur.
                self.c.execute("DELETE FROM cari_hareketler WHERE id=?", (cari_hareket_id,))
            elif referans_tipi == 'FATURA': # Bu durum `arayuz.py`'deki `secili_islemi_sil` metodunda ele alınıyor olmalı.
                # Burada sadece mesaj gösterip `rollback` yapacağız, asıl silme `fatura_sil` metodu üzerinden olmalı.
                self.conn.rollback() # Mevcut işlemi geri al
                return False, "Bu işlem tipi (FATURA) buradan doğrudan silinemez. Lütfen fatura modülünden fatura silme işlemini kullanın."
            else: # Diğer işlem tipleri buradan silinemez
                self.conn.rollback() # Bilinmeyen tipse rollback
                return False, "Bu işlem tipi buradan silinemez veya tanımlanamadı.\nFatura kaynaklı borç/alacak hareketleri fatura silinerek veya güncellenerek yönetilir."
            
            self.conn.commit() # Tüm işlemler başarılıysa commit et       
            return True, f"{referans_tipi.capitalize().replace('_',' ')} başarıyla silindi."
        except Exception as e:
            self.conn.rollback() # Hata durumunda rollback
            error_details = traceback.format_exc()
            return False, f"{referans_tipi.capitalize() if referans_tipi else 'İşlem'} silinirken hata: {e}\n\nDetaylar:\n{error_details}"
        


    def veresiye_borc_ekle(self, cari_id, cari_tip, tarih, tutar, aciklama):
        if not all([cari_id, cari_tip, tarih, tutar, aciklama]):
                return False, "Lütfen tüm zorunlu (*) alanları (Cari, Tarih, Tutar, Açıklama) doldurun."

        try:
            tutar_f = float(str(tutar).replace(',', '.'))
            if tutar_f <= 0:
                return False, "Tutar pozitif bir sayı olmalıdır."
            datetime.strptime(tarih, '%Y-%m-%d')
        except ValueError:
            return False, "Tutar sayısal bir değer olmalı veya Tarih formatı (YYYY-AA-GG) olmalıdır."

        try:
            self.conn.execute("BEGIN TRANSACTION")
            current_time = self.get_current_datetime_str()
            olusturan_id = self.app.current_user[0] if self.app and self.app.current_user else None

            if cari_tip == 'MUSTERI':
                islem_tipi_ch = 'ALACAK'
                cari_bilgi = self.musteri_getir_by_id(cari_id)
                ch_aciklama = f"Manuel Veresiye Borç (Alacak): {aciklama} - Müşteri: {cari_bilgi[2] if cari_bilgi else 'Bilinmiyor'}"
            elif cari_tip == 'TEDARIKCI':
                islem_tipi_ch = 'BORC'
                cari_bilgi = self.tedarikci_getir_by_id(cari_id)
                ch_aciklama = f"Manuel Veresiye Borç (Borç): {aciklama} - Tedarikçi: {cari_bilgi[2] if cari_bilgi else 'Bilinmiyor'}"
            else:
                self.conn.rollback()
                return False, "Geçersiz cari tipi. Veresiye borç eklenemiyor."

            self.c.execute("""
                INSERT INTO cari_hareketler
                (tarih, cari_tip, cari_id, islem_tipi, tutar, aciklama, referans_id, referans_tip, kasa_banka_id, olusturma_tarihi_saat, olusturan_kullanici_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (tarih, cari_tip, cari_id, islem_tipi_ch, tutar_f, ch_aciklama, None, "VERESIYE_BORC_MANUEL", None, current_time, olusturan_id))

            self.conn.commit()
            # DÜZELTME BURADA: Başarılı durumda (True, "mesaj") formatında dönüyoruz.
            return True, "Veresiye borç başarıyla eklendi."
        except Exception as e:
            self.conn.rollback() 
            return False, f"Veresiye borç eklenirken bir hata oluştu: {e}"
    
    def get_kar_zarar_verileri(self, baslangic_tarih=None, bitis_tarih=None):
        """
        Belirtilen tarih aralığındaki toplam gelir ve toplam gideri hesaplar.
        """
        toplam_gelir = 0.0
        toplam_gider = 0.0

        query = "SELECT tip, tutar FROM gelir_gider"
        params = []
        conditions = []

        if baslangic_tarih:
            conditions.append("tarih >= ?")
            params.append(baslangic_tarih)
        if bitis_tarih:
            conditions.append("tarih <= ?")
            params.append(bitis_tarih)
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
            
        self.c.execute(query, params)
        results = self.c.fetchall()

        for tip, tutar in results:
            if tip == 'GELİR':
                toplam_gelir += tutar
            elif tip == 'GİDER':
                toplam_gider += tutar
        
        return toplam_gelir, toplam_gider        
    def get_overdue_receivables(self):
        """
        Vadesi geçmiş alacakları (müşteri borçlarını) getirir.
        """
        self.c.execute("""
            SELECT 
                f.id AS fatura_id,
                f.fatura_no,
                f.tarih AS fatura_tarihi,
                f.vade_tarihi,
                (f.toplam_kdv_dahil - COALESCE(SUM(ch.tutar), 0)) AS kalan_tutar,
                m.ad || ' ' || m.soyad AS musteri_adi,
                m.kod AS musteri_kodu, -- <<<< BURAYI DİKKATLİCE DÜZELTİN: 'm.musteri_kodu' yerine 'm.kod' >>>>>
                m.telefon AS musteri_telefon
            FROM faturalar f
            JOIN musteriler m ON f.cari_id = m.id
            LEFT JOIN cari_hareketler ch ON ch.referans_id = f.id 
                                        AND ch.referans_tip = 'FATURA' 
                                        AND ch.cari_tip = 'MUSTERI' 
                                        AND ch.islem_tipi IN ('TAHSILAT', 'IADE_FATURASI') -- İade faturalarını da tahsilat gibi düşmeli
            WHERE f.tip = 'SATIŞ' 
            AND f.vade_tarihi IS NOT NULL 
            AND f.vade_tarihi < CURRENT_DATE 
            GROUP BY f.id
            HAVING kalan_tutar > 0
            ORDER BY f.vade_tarihi ASC
        """)
        return self.c.fetchall()

    def get_overdue_payables(self, current_date_str=None):
        """
        Vadesi geçmiş tedarikçi borçlarını (net borçları) döndürür.
        current_date_str: 'YYYY-AA-GG' formatında raporun alınacağı tarih. Yoksa bugünün tarihi kullanılır.
        Dönüş: [(cari_id, tedarikci_adi, net_borc, vadesi_gecen_gun)]
        """
        if current_date_str:
            current_date = datetime.strptime(current_date_str, '%Y-%m-%d')
        else:
            current_date = datetime.now()

        # Tedarikçilerin net bakiyelerini al (sadece bizim borçlu olduğumuz tedarikçiler)
        # SQL sorgusu güncellendi: islem_tipi = 'BORC' ve 'ODEME' olarak netleştirildi.
        self.c.execute("""
            SELECT 
                c.cari_id,
                t.ad AS tedarikci_adi,
                SUM(CASE WHEN c.islem_tipi = 'BORC' THEN c.tutar ELSE 0 END) - 
                SUM(CASE WHEN c.islem_tipi = 'ODEME' THEN c.tutar ELSE 0 END) AS net_borc,
                MAX(c.tarih) AS son_islem_tarihi -- Vade hesaplaması için son işlem tarihi
            FROM cari_hareketler c
            JOIN tedarikciler t ON c.cari_id = t.id
            WHERE c.cari_tip = 'TEDARIKCI'
            GROUP BY c.cari_id, t.ad
            HAVING net_borc > 0
        """)
        
        results = self.c.fetchall()
        overdue_payables = []

        for cari_id, tedarikci_adi, net_borc, son_islem_tarihi_str in results:
            try:
                # Tarih formatı 'YYYY-MM-DD' olarak varsayılıyor
                son_islem_tarihi = datetime.strptime(son_islem_tarihi_str, '%Y-%m-%d')
                vadesi_gecen_gun = (current_date - son_islem_tarihi).days
                if vadesi_gecen_gun > 0: # Sadece vadesi geçmiş olanları al
                    overdue_payables.append((cari_id, tedarikci_adi, net_borc, vadesi_gecen_gun))
            except ValueError:
                # Tarih formatı hatası olursa bu tedarikçiyi atla veya farklı ele al
                print(f"Uyarı: Tedarikçi ID {cari_id} için son işlem tarihi formatı hatalı: {son_islem_tarihi_str}")
                pass
        
        # En vadesi geçmiş olanları öne almak için sırala
        overdue_payables.sort(key=lambda x: x[3], reverse=True)
        return overdue_payables

    def _get_cari_devir_bakiye(self, cari_tip, cari_id, baslangic_tarih_str):
        devir_bakiye = 0.0
        try:
            # SADECE AÇIK HESAP VE MANUEL HAREKETLER DAHİL EDİLECEK.
            query = """
                SELECT ch.islem_tipi, ch.tutar, ch.referans_tip, f.odeme_turu
                FROM cari_hareketler ch
                LEFT JOIN faturalar f ON ch.referans_id = f.id
                WHERE ch.cari_tip = ? AND ch.cari_id = ? AND ch.tarih < ?
            """
            self.c.execute(query, (cari_tip, cari_id, baslangic_tarih_str))
            hareketler_oncesi = self.c.fetchall()

            for hareket in hareketler_oncesi:
                tutar = hareket['tutar'] or 0.0
                referans_tip = hareket['referans_tip']
                odeme_turu = hareket['odeme_turu'] # Faturadan gelen ödeme türü (None olabilir)

                if cari_tip == self.CARI_TIP_MUSTERI:
                    # Alacak artırıcı hareketler (Müşterinin bize borcu - sadece açık hesap faturalar veya manuel veresiye)
                    if (referans_tip == self.KAYNAK_TIP_FATURA and odeme_turu == self.ODEME_TURU_ACIK_HESAP) or \
                       referans_tip == self.KAYNAK_TIP_VERESIYE_BORC_MANUEL:
                        devir_bakiye += tutar
                    # Borç azaltıcı hareketler (Müşteriden tahsilat veya Satış İadesi - sadece açık hesaplar veya manuel tahsilat)
                    elif referans_tip == self.KAYNAK_TIP_TAHSILAT or \
                         (referans_tip == self.KAYNAK_TIP_IADE_FATURA and odeme_turu == self.ODEME_TURU_ACIK_HESAP):
                        devir_bakiye -= tutar
                elif cari_tip == self.CARI_TIP_TEDARIKCI:
                    # Borç artırıcı hareketler (Bizim tedarikçiye borcumuz - sadece açık hesap faturalar veya manuel veresiye)
                    if (referans_tip == self.KAYNAK_TIP_FATURA and odeme_turu == self.ODEME_TURU_ACIK_HESAP) or \
                       referans_tip == self.KAYNAK_TIP_VERESIYE_BORC_MANUEL:
                        devir_bakiye += tutar
                    # Borç azaltıcı hareketler (Tedarikçiye ödeme veya Alış İadesi - sadece açık hesaplar veya manuel ödeme)
                    elif referans_tip == self.KAYNAK_TIP_ODEME or \
                         (referans_tip == self.KAYNAK_TIP_IADE_FATURA and odeme_turu == self.ODEME_TURU_ACIK_HESAP):
                        devir_bakiye -= tutar
            
            logging.debug(f"Devir bakiye hesaplandı: {devir_bakiye}")
            return devir_bakiye
        except Exception as e:
            logging.error(f"_get_cari_devir_bakiye hatası: {e}\n{traceback.format_exc()}")
            return 0.0
                        
    def cari_hesap_ekstresi_al(self, cari_id, cari_tip, baslangic_tarihi=None, bitis_tarihi=None):
        """
        Belirli bir cariye ait hesap ekstresini (tüm cari hareketleri) getirir ve devreden bakiyeyi hesaplar.
        Dönüş: (ekstre_hareketleri_listesi, devreden_bakiye_float, success_bool, message_str)
        """
        try:
            # Devreden bakiyeyi al
            devreden_bakiye = self._get_cari_devir_bakiye(cari_tip, cari_id, baslangic_tarihi)

            query = """
                SELECT
                    ch.id,
                    ch.tarih,
                    STRFTIME('%H:%M:%S', ch.olusturma_tarihi_saat) AS islem_saati,
                    ch.islem_tipi,
                    ch.tutar,
                    ch.aciklama,
                    ch.referans_tip,
                    ch.referans_id,
                    COALESCE(f.fatura_no, '') AS fatura_no,
                    COALESCE(f.odeme_turu, '') AS odeme_turu,
                    COALESCE(f.vade_tarihi, '') AS vade_tarihi,
                    COALESCE(f.tip, '') AS fatura_tipi -- faturalar.tip
                FROM
                    cari_hareketler ch
                LEFT JOIN
                    faturalar f ON ch.referans_id = f.id 
                               AND ch.referans_tip IN (?, ?, ?, ?, ?) -- Referans tiplerini genişletiyoruz
                WHERE
                    ch.cari_id = ? AND ch.cari_tip = ?
            """
            params = [
                self.KAYNAK_TIP_FATURA, 
                self.KAYNAK_TIP_IADE_FATURA, 
                self.KAYNAK_TIP_FATURA_SATIS_PESIN, # Peşin satış faturalarını dahil et
                self.KAYNAK_TIP_FATURA_ALIS_PESIN,  # Peşin alış faturalarını dahil et
                self.KAYNAK_TIP_VERESIYE_BORC_MANUEL, # Veresiye borçlar da burada JOIN'e dahil edilebilir, mantığına göre.
                                                    # Ancak genellikle sadece faturalar için JOIN yaparız.
                                                    # Şimdilik sadece peşin faturaları ekliyorum.
                cari_id, 
                cari_tip
            ]

            if baslangic_tarihi:
                query += " AND ch.tarih >= ?"
                params.append(baslangic_tarihi)
            if bitis_tarihi:
                query += " AND ch.tarih <= ?"
                params.append(bitis_tarihi)
            
            query += " ORDER BY ch.tarih ASC, ch.olusturma_tarihi_saat ASC"

            self.c.execute(query, params)
            ekstre_hareketleri_raw = self.c.fetchall()

            ekstre_sonuclari_islenmis = []
            for hareket in ekstre_hareketleri_raw:
                hareket_bilgisi = {
                    'id': hareket['id'],
                    'tarih': hareket['tarih'],
                    'islem_saati': hareket['islem_saati'],
                    'islem_tipi': hareket['islem_tipi'],
                    'tutar': self.safe_float(hareket['tutar']),
                    'aciklama': hareket['aciklama'],
                    'referans_tip': hareket['referans_tip'],
                    'referans_id': hareket['referans_id'],
                    'fatura_no': hareket['fatura_no'],
                    'odeme_turu': hareket['odeme_turu'],
                    'vade_tarihi': hareket['vade_tarihi'],
                    'fatura_tipi': hareket['fatura_tipi'] # Fatura tipi de eklendi
                }
                ekstre_sonuclari_islenmis.append(hareket_bilgisi)
            
            return ekstre_sonuclari_islenmis, devreden_bakiye, True, "Ekstre başarıyla alındı."

        except Exception as e:
            logging.error(f"Cari hesap ekstresi alınırken hata: {e}\n{traceback.format_exc()}")
            return [], 0.0, False, f"Cari hesap ekstresi alınırken hata oluştu: {e}"
                            
    def get_cari_ozet_bilgileri(self, cari_id, cari_tip):
        ozet = {
            "donem_basi_bakiye": 0.0,
            "donem_sonu_bakiye": 0.0,
            "donem_toplam_borc_hareketi": 0.0,
            "donem_toplam_alacak_hareketi": 0.0,
            "toplam_tahsilat": 0.0, # Sadece tahsilat
            "toplam_odeme": 0.0,     # Sadece ödeme
            "vadesi_gelmis_borc_alacak": 0.0,
            "vadesi_gelecek_borc_alacak": 0.0
        }
        try:
            # Dönem başı bakiyeyi almak için (bugünden önceki tüm hareketler)
            ozet["donem_basi_bakiye"] = self.get_musteri_net_bakiye(cari_id) if cari_tip == self.CARI_TIP_MUSTERI else self.get_tedarikci_net_bakiye(cari_id)

            baslangic_tarih = "1900-01-01" 
            bitis_tarih = datetime.now().strftime('%Y-%m-%d')

            if cari_tip == 'MUSTERI':
                # Toplam Alacak Hareketi: Müşterinin borcunu artıran hareketler (Açık Fatura, Manuel Veresiye Borç)
                self.c.execute("""
                    SELECT SUM(ch.tutar) FROM cari_hareketler ch
                    LEFT JOIN faturalar f ON ch.referans_id = f.id
                    WHERE ch.cari_id = ? AND ch.cari_tip = ? AND ch.tarih BETWEEN ? AND ? AND (
                        ch.referans_tip = ? OR -- VERESIYE_BORC_MANUEL
                        (ch.referans_tip = ? AND f.odeme_turu = ?) -- FATURA ve AÇIK HESAP
                    )
                """, (cari_id, self.CARI_TIP_MUSTERI, baslangic_tarih, bitis_tarih, self.KAYNAK_TIP_VERESIYE_BORC_MANUEL, self.KAYNAK_TIP_FATURA, self.ODEME_TURU_ACIK_HESAP))
                ozet["donem_toplam_alacak_hareketi"] = self.c.fetchone()[0] or 0.0
                
                # Toplam Borç Hareketi: Müşterinin borcunu azaltan hareketler (Manuel Tahsilat, Açık Hesap İade Faturası)
                self.c.execute("""
                    SELECT SUM(ch.tutar) FROM cari_hareketler ch
                    LEFT JOIN faturalar f ON ch.referans_id = f.id
                    WHERE ch.cari_id = ? AND ch.cari_tip = ? AND ch.tarih BETWEEN ? AND ? AND (
                        ch.referans_tip = ? OR -- TAHSILAT
                        (ch.referans_tip = ? AND f.odeme_turu = ?) -- İADE_FATURA ve AÇIK HESAP
                    )
                """, (cari_id, self.CARI_TIP_MUSTERI, baslangic_tarih, bitis_tarih, self.KAYNAK_TIP_TAHSILAT, self.KAYNAK_TIP_IADE_FATURA, self.ODEME_TURU_ACIK_HESAP))
                ozet["donem_toplam_borc_hareketi"] = self.c.fetchone()[0] or 0.0

                # Toplam Tahsilat: SADECE fiili tahsilat hareketlerini toplamalı
                # Referans tipi TAHSILAT veya FATURA_SATIS_PESIN olan cari hareketleri topla.
                # NOT: Bu sorgu içinde odeme_turu='AÇIK HESAP' faturaları hariç tutulmalı
                self.c.execute("""
                    SELECT SUM(tutar) FROM cari_hareketler
                    WHERE cari_id = ? AND cari_tip = ? AND (
                        referans_tip = ? OR -- Manuel tahsilatlar
                        referans_tip = ? -- Peşin satış faturalarından gelen tahsilatlar
                    ) AND tarih BETWEEN ? AND ?
                """, (cari_id, self.CARI_TIP_MUSTERI, self.KAYNAK_TIP_TAHSILAT, self.KAYNAK_TIP_FATURA_SATIS_PESIN, baslangic_tarih, bitis_tarih))
                ozet["toplam_tahsilat"] = self.c.fetchone()[0] or 0.0
                
                current_net_bakiye = self.get_musteri_net_bakiye(cari_id)
                ozet["net_bakiye"] = current_net_bakiye
                ozet["donem_sonu_bakiye"] = current_net_bakiye

                vade_durumu = self._get_vade_durumu(cari_id, cari_tip)
                ozet["vadesi_gelmis_borc_alacak"] = vade_durumu["vadesi_gelmis"]
                ozet["vadesi_gelecek_borc_alacak"] = vade_durumu["vadesi_gelecek"]


            elif cari_tip == 'TEDARIKCI':
                # Toplam Borç Hareketi: Tedarikçiye borcumuzu artıran hareketler (Açık Fatura, Manuel Veresiye Borç)
                self.c.execute("""
                    SELECT SUM(ch.tutar) FROM cari_hareketler ch
                    LEFT JOIN faturalar f ON ch.referans_id = f.id
                    WHERE ch.cari_id = ? AND ch.cari_tip = ? AND ch.tarih BETWEEN ? AND ? AND (
                        ch.referans_tip = ? OR -- VERESIYE_BORC_MANUEL
                        (ch.referans_tip = ? AND f.odeme_turu = ?) -- FATURA ve AÇIK HESAP
                    )
                """, (cari_id, self.CARI_TIP_TEDARIKCI, baslangic_tarih, bitis_tarih, self.KAYNAK_TIP_VERESIYE_BORC_MANUEL, self.KAYNAK_TIP_FATURA, self.ODEME_TURU_ACIK_HESAP))
                ozet["donem_toplam_borc_hareketi"] = self.c.fetchone()[0] or 0.0

                # Toplam Alacak Hareketi: Tedarikçiye borcumuzu azaltan hareketler (Manuel Ödeme, Açık Hesap Alış İade Faturası)
                self.c.execute("""
                    SELECT SUM(ch.tutar) FROM cari_hareketler ch
                    LEFT JOIN faturalar f ON ch.referans_id = f.id
                    WHERE ch.cari_id = ? AND ch.cari_tip = ? AND ch.tarih BETWEEN ? AND ? AND (
                        ch.referans_tip = ? OR -- ODEME
                        (ch.referans_tip = ? AND f.odeme_turu = ?) -- İADE_FATURA ve AÇIK HESAP
                    )
                """, (cari_id, self.CARI_TIP_TEDARIKCI, baslangic_tarih, bitis_tarih, self.KAYNAK_TIP_ODEME, self.KAYNAK_TIP_IADE_FATURA, self.ODEME_TURU_ACIK_HESAP))
                ozet["donem_toplam_alacak_hareketi"] = self.c.fetchone()[0] or 0.0

                # Toplam Ödeme: SADECE fiili ödeme hareketlerini toplamalı
                # Referans tipi ODEME veya FATURA_ALIS_PESIN olan cari hareketleri topla.
                self.c.execute("""
                    SELECT SUM(tutar) FROM cari_hareketler
                    WHERE cari_id = ? AND cari_tip = ? AND (
                        referans_tip = ? OR -- Manuel ödemeler
                        referans_tip = ? -- Peşin alış faturalarından gelen ödemeler
                    ) AND tarih BETWEEN ? AND ?
                """, (cari_id, self.CARI_TIP_TEDARIKCI, self.KAYNAK_TIP_ODEME, self.KAYNAK_TIP_FATURA_ALIS_PESIN, baslangic_tarih, bitis_tarih))
                ozet["toplam_odeme"] = self.c.fetchone()[0] or 0.0

                current_net_bakiye = self.get_tedarikci_net_bakiye(cari_id)
                ozet["net_bakiye"] = current_net_bakiye
                ozet["donem_sonu_bakiye"] = current_net_bakiye

                vade_durumu = self._get_vade_durumu(cari_id, cari_tip)
                ozet["vadesi_gelmis_borc_alacak"] = vade_durumu["vadesi_gelmis"]
                ozet["vadesi_gelecek_borc_alacak"] = vade_durumu["vadesi_gelecek"]
            
            return ozet
        except Exception as e:
            logging.error(f"get_cari_ozet_bilgileri hatası: {e}\n{traceback.format_exc()}")
            return ozet
                        
    def _fatura_finansal_etki_olustur(self, fatura_id, fatura_no, fatura_tarihi, fatura_tipi, cari_id, nihai_toplam_kdv_dahil, odeme_turu, kasa_banka_id, misafir_adi):
        try:
            current_time = self.get_current_datetime_str()
            olusturan_id = self._get_current_user_id()

            cari_adi = ""
            # <<< DÜZELTME BAŞLANGICI: 'cari_tip_for_db' adında yeni bir değişken tanımlıyoruz >>>
            cari_tip_for_db = ""
            if fatura_tipi in (self.FATURA_TIP_SATIS, self.FATURA_TIP_SATIS_IADE):
                cari_data = self.musteri_getir_by_id(cari_id)
                cari_adi = cari_data['ad'] if cari_data else "Bilinmeyen Müşteri"
                cari_tip_for_db = self.CARI_TIP_MUSTERI # Müşteri ise 'MUSTERI'
            elif fatura_tipi in (self.FATURA_TIP_ALIS, self.FATURA_TIP_ALIS_IADE, self.FATURA_TIP_DEVIR_GIRIS):
                cari_data = self.tedarikci_getir_by_id(cari_id)
                cari_adi = cari_data['ad'] if cari_data else "Bilinmeyen Tedarikçi"
                cari_tip_for_db = self.CARI_TIP_TEDARIKCI # Tedarikçi ise 'TEDARIKCI'
            # <<< DÜZELTME BİTİŞİ >>>
            
            if misafir_adi and fatura_tipi == self.FATURA_TIP_SATIS:
                cari_adi = f"{cari_adi} (Misafir: {misafir_adi})"

            # A. Cari Hareket Kaydı
            cari_islem_tipi = ""
            cari_referans_tipi = self.KAYNAK_TIP_FATURA # Varsayılan olarak fatura

            if odeme_turu in self.pesin_odeme_turleri:
                if fatura_tipi == self.FATURA_TIP_SATIS:
                    cari_islem_tipi = self.ISLEM_TIP_TAHSILAT
                    cari_referans_tipi = self.KAYNAK_TIP_FATURA_SATIS_PESIN
                elif fatura_tipi == self.FATURA_TIP_ALIS:
                    cari_islem_tipi = self.ISLEM_TIP_ODEME
                    cari_referans_tipi = self.KAYNAK_TIP_FATURA_ALIS_PESIN
                elif fatura_tipi == self.FATURA_TIP_SATIS_IADE:
                    cari_islem_tipi = self.ISLEM_TIP_BORC
                    cari_referans_tipi = self.KAYNAK_TIP_IADE_FATURA
                elif fatura_tipi == self.FATURA_TIP_ALIS_IADE:
                    cari_islem_tipi = self.ISLEM_TIP_ALACAK
                    cari_referans_tipi = self.KAYNAK_TIP_IADE_FATURA
                elif fatura_tipi == self.FATURA_TIP_DEVIR_GIRIS:
                    return True, "Devir girişi için cari hareket oluşturulmadı."
                
            elif odeme_turu == self.ODEME_TURU_ACIK_HESAP:
                if fatura_tipi == self.FATURA_TIP_SATIS or fatura_tipi == self.FATURA_TIP_DEVIR_GIRIS:
                    cari_islem_tipi = self.ISLEM_TIP_ALACAK
                elif fatura_tipi == self.FATURA_TIP_ALIS:
                    cari_islem_tipi = self.ISLEM_TIP_BORC
                elif fatura_tipi == self.FATURA_TIP_SATIS_IADE:
                    cari_islem_tipi = self.ISLEM_TIP_BORC
                    cari_referans_tipi = self.KAYNAK_TIP_IADE_FATURA
                elif fatura_tipi == self.FATURA_TIP_ALIS_IADE:
                    cari_islem_tipi = self.ISLEM_TIP_ALACAK
                    cari_referans_tipi = self.KAYNAK_TIP_IADE_FATURA

            ch_aciklama = f"Fatura No: {fatura_no} ({fatura_tipi} Faturası) - Cari: {cari_adi}"

            self.c.execute("""
                INSERT INTO cari_hareketler (
                    tarih, cari_tip, cari_id, islem_tipi, tutar, aciklama,
                    referans_id, referans_tip, kasa_banka_id,
                    olusturma_tarihi_saat, olusturan_kullanici_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                fatura_tarihi, cari_tip_for_db, cari_id, cari_islem_tipi, nihai_toplam_kdv_dahil, ch_aciklama, # <<< BURADA cari_tip_for_db KULLANILDI >>>
                fatura_id, cari_referans_tipi, kasa_banka_id,
                current_time, olusturan_id
            ))
            
            # B. Gelir/Gider Kaydı
            if odeme_turu in self.pesin_odeme_turleri:
                gg_tipi = ""
                if fatura_tipi == self.FATURA_TIP_SATIS or fatura_tipi == self.FATURA_TIP_ALIS_IADE:
                    gg_tipi = self.ISLEM_TIP_GELIR
                elif fatura_tipi == self.FATURA_TIP_ALIS or fatura_tipi == self.FATURA_TIP_SATIS_IADE:
                    gg_tipi = self.ISLEM_TIP_GIDER
                
                gg_aciklama = f"{fatura_tipi} Faturası: {fatura_no} - {cari_adi}"
                
                self.c.execute("""
                    INSERT INTO gelir_gider (
                        tarih, tip, tutar, aciklama, kaynak, kaynak_id, kasa_banka_id,
                        olusturma_tarihi_saat, olusturan_kullanici_id
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    fatura_tarihi, gg_tipi, nihai_toplam_kdv_dahil, gg_aciklama,
                    self.KAYNAK_TIP_FATURA, fatura_id, kasa_banka_id,
                    current_time, olusturan_id
                ))
            
            return True, "Finansal etkiler başarıyla oluşturuldu."
        except Exception as e:
            logging.error(f"Fatura finansal etkileri oluşturulurken hata: {e}\n{traceback.format_exc()}")
            raise # Hatanın yukarıya yayılmasını sağlar
        
    def _fatura_finansal_etki_geri_al(self, fatura_id):
        """
        Belirli bir fatura ID'sine ait tüm cari hareketleri ve gelir/gider kayıtlarını siler.
        Bu metod kendi transaction'ını yönetmez, çağrıldığı transaction'ın bir parçasıdır.
        """
        try:
            # Fatura ile ilişkili tüm cari hareketleri sil
            self.c.execute("""
                DELETE FROM cari_hareketler 
                WHERE referans_id = ? AND referans_tip IN (?, ?, ?, ?)
            """, (fatura_id, self.KAYNAK_TIP_FATURA, self.KAYNAK_TIP_IADE_FATURA, 
                  self.KAYNAK_TIP_FATURA_SATIS_PESIN, self.KAYNAK_TIP_FATURA_ALIS_PESIN))
            
            # Fatura ile ilişkili tüm gelir/gider kayıtlarını sil (sadece kaynak FATURA olanlar)
            self.c.execute("DELETE FROM gelir_gider WHERE kaynak = ? AND kaynak_id = ?", (self.KAYNAK_TIP_FATURA, fatura_id))
            
            return True, "Finansal etkiler başarıyla geri alındı."
        except Exception as e:
            logging.error(f"Fatura finansal etkileri geri alınırken hata: {e}\n{traceback.format_exc()}")
            raise # Hatanın yukarıya yayılmasını sağlar

    def _get_vade_durumu(self, cari_id, cari_tip):
        """
        Bir cari hesabın vadesi gelmiş ve vadesi gelecek borç/alacaklarını hesaplar.
        Sadece Açık Hesap faturalarını dikkate alır.
        """
        vade_durumu = {"vadesi_gelmis": 0.0, "vadesi_gelecek": 0.0}
        
        # Sadece açık hesap faturalarını ve vadesi olanları çek
        query = """
            SELECT SUM(toplam_kdv_dahil) AS tutar, vade_tarihi, tip
            FROM faturalar
            WHERE cari_id = ? AND odeme_turu = ? AND vade_tarihi IS NOT NULL
            GROUP BY vade_tarihi, tip
        """
        
        self.c.execute(query, (cari_id, self.ODEME_TURU_ACIK_HESAP))
        results = self.c.fetchall()

        bugunun_tarihi = date.today()

        for row in results:
            tutar = row['tutar'] or 0.0
            try:
                vade_tarihi = datetime.strptime(row['vade_tarihi'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                continue # Geçersiz tarih formatı veya boş tarihleri atla

            # Fatura tipine göre tutarın yönünü belirle
            # Müşteri alacakları (Satış faturaları) ve Tedarikçi borçları (Alış faturaları)
            # İade faturaları da ters etki yapar.
            if (cari_tip == self.CARI_TIP_MUSTERI and row['tip'] == self.FATURA_TIP_SATIS) or \
               (cari_tip == self.CARI_TIP_TEDARIKCI and row['tip'] == self.FATURA_TIP_ALIS):
                
                if vade_tarihi < bugunun_tarihi:
                    vade_durumu["vadesi_gelmis"] += tutar
                else:
                    vade_durumu["vadesi_gelecek"] += tutar
            
            # İade faturaları ters etki yapar (müşteri borcu azaltır, tedarikçi alacağı azaltır)
            elif (cari_tip == self.CARI_TIP_MUSTERI and row['tip'] == self.FATURA_TIP_SATIS_IADE) or \
                 (cari_tip == self.CARI_TIP_TEDARIKCI and row['tip'] == self.FATURA_TIP_ALIS_IADE):
                
                # İade faturaları için vade mantığı daha karmaşık olabilir.
                # Genellikle iade faturasının vadesi hemen gelir.
                # Ancak burada orijinal fatura ile aynı vade mantığını sürdürmüş olduk.
                if vade_tarihi < bugunun_tarihi:
                    vade_durumu["vadesi_gelmis"] -= tutar # Vadesi geçmiş iade, borcu azaltır
                else:
                    vade_durumu["vadesi_gelecek"] -= tutar # Vadesi gelecek iade, borcu azaltır

        return vade_durumu

    def stok_envanterini_yeniden_hesapla(self):
        """
        Tüm ürünlerin stok miktarlarını, stok_hareketleri tablosundaki TÜM giriş ve çıkışları
        dikkate alarak sıfırdan yeniden hesaplar. Bu, tüm tutarsızlıkları düzeltir.
        """
        try:
            self.conn.execute("BEGIN TRANSACTION")

            # 1. Adım: Stok hareketleri olan tüm benzersiz ürün ID'lerini al.
            self.c.execute("SELECT DISTINCT urun_id FROM stok_hareketleri")
            urun_idler = self.c.fetchall()
            
            if not urun_idler:
                self.conn.commit()
                return True, "Hesaplanacak stok hareketi bulunamadı."

            # 2. Adım: Her bir ürün için net stok farkını hesapla.
            for urun in urun_idler:
                urun_id = urun['urun_id']
                
                # Tüm girişleri topla
                self.c.execute("""
                    SELECT SUM(miktar) FROM stok_hareketleri 
                    WHERE urun_id = ? AND (
                        islem_tipi LIKE '%Giriş%' OR 
                        islem_tipi LIKE '%Alış%' OR 
                        islem_tipi LIKE '%Fazlası%' OR
                        islem_tipi LIKE '%İptal%')
                """, (urun_id,))
                toplam_giris = self.c.fetchone()[0] or 0.0

                # Tüm çıkışları topla
                self.c.execute("""
                    SELECT SUM(miktar) FROM stok_hareketleri 
                    WHERE urun_id = ? AND (
                        islem_tipi LIKE '%Çıkış%' OR 
                        islem_tipi LIKE '%Satış%' OR 
                        islem_tipi LIKE '%Eksiği%' OR
                        islem_tipi LIKE '%Zayiat%')
                """, (urun_id,))
                toplam_cikis = self.c.fetchone()[0] or 0.0
                
                # Nihai doğru stoku hesapla
                dogru_stok = toplam_giris - toplam_cikis

                # 3. Adım: tbl_stoklar tablosundaki stok miktarını bu doğru rakamla güncelle.
                self.c.execute("UPDATE tbl_stoklar SET stok_miktari = ? WHERE id = ?", (dogru_stok, urun_id))

            self.conn.commit()
            return True, f"{len(urun_idler)} adet ürünün stoku, tüm hareketlere göre başarıyla yeniden hesaplandı."

        except Exception as e:
            self.conn.rollback()
            error_details = traceback.format_exc()
            logging.error(f"Stok envanteri yeniden hesaplanırken hata: {e}\nDetay: {error_details}")
            return False, "Stoklar yeniden hesaplanırken beklenmeyen bir hata oluştu."

    def get_toplam_musteri_sayisi(self): # Eğer böyle bir metodunuz varsa
        # self.connect() # Bu satırı sildiğinizden emin olun
        self.c.execute("SELECT COUNT(id) FROM musteriler WHERE kod != ?", (self.PERAKENDE_MUSTERI_KODU,)) # musteri_kodu yerine KOD
        return self.c.fetchone()[0]

    def get_toplam_tedarikci_sayisi(self):
        self.c.execute("SELECT COUNT(id) FROM tedarikciler")
        sonuc = self.c.fetchone()
        return sonuc[0] if sonuc else 0

    def get_toplam_stok_cesidi_sayisi(self):
        self.c.execute("SELECT COUNT(id) FROM tbl_stoklar")
        sonuc = self.c.fetchone()
        return sonuc[0] if sonuc else 0
    def get_nakit_akis_verileri(self, baslangic_tarih=None, bitis_tarih=None, limit=None, offset=None):
        """
        Belirtilen tarih aralığındaki kasa ve banka hareketlerini (nakit akışını) getirir.
        GELİR (kasaya/bankaya giren) ve GİDER (kasadan/bankadan çıkan) hareketleri dahil eder.
        """
        query = """
            SELECT 
                gg.tarih,
                gg.tip, -- GELİR/GİDER
                gg.tutar,
                gg.aciklama,
                kb.hesap_adi, -- Kasa/Banka Adı
                kb.tip, -- KASA/BANKA
                gg.kaynak, -- FATURA, TAHSILAT, ODEME, MANUEL
                gg.kaynak_id -- İlişkili ID (fatura_id, cari_hareket_id)
            FROM gelir_gider gg
            JOIN kasalar_bankalar kb ON gg.kasa_banka_id = kb.id
            WHERE gg.kasa_banka_id IS NOT NULL -- Sadece kasa/banka ile ilişkili hareketler
        """
        params = []
        conditions = []

        if baslangic_tarih:
            conditions.append("gg.tarih >= ?")
            params.append(baslangic_tarih)
        if bitis_tarih:
            conditions.append("gg.tarih <= ?")
            params.append(bitis_tarih)
        
        if conditions:
            query += " AND " + " AND ".join(conditions)
            
        query += " ORDER BY gg.tarih ASC, gg.id ASC"
        
        # Sayfalama için LIMIT ve OFFSET ekle
        if limit is not None:
            query += " LIMIT ?"
            params.append(limit)
        if offset is not None:
            query += " OFFSET ?"
            params.append(offset)
        
        self.c.execute(query, params)
        return self.c.fetchall()
    
    def get_nakit_akis_count(self, baslangic_tarih=None, bitis_tarih=None): # Yeni metot: toplam nakit akışı sayısını alır
        query = """
            SELECT COUNT(gg.id)
            FROM gelir_gider gg
            JOIN kasalar_bankalar kb ON gg.kasa_banka_id = kb.id
            WHERE gg.kasa_banka_id IS NOT NULL
        """
        params = []
        conditions = []

        if baslangic_tarih:
            conditions.append("gg.tarih >= ?")
            params.append(baslangic_tarih)
        if bitis_tarih:
            conditions.append("gg.tarih <= ?")
            params.append(bitis_tarih)
        
        if conditions:
            query += " AND " + " AND ".join(conditions)
            
        self.c.execute(query, params)
        return self.c.fetchone()[0]    

    def get_kasa_banka_toplam_bakiye(self, hesap_id):
        """Belirli bir kasa/banka hesabının toplam bakiyesini döndürür."""
        self.c.execute("SELECT bakiye FROM kasalar_bankalar WHERE id = ?", (hesap_id,))
        result = self.c.fetchone()
        return result[0] if result else 0.0

    def get_tum_kasa_banka_bakiyeleri(self):
        """Tüm kasa ve banka hesaplarının güncel bakiyelerini döndürür."""
        self.c.execute("SELECT id, hesap_adi, bakiye, tip FROM kasalar_bankalar ORDER BY tip DESC, hesap_adi ASC")
        return self.c.fetchall()
    
    # --- Excel ve PDF İşlemleri ---
    def stok_raporu_excel_olustur(self, dosya_yolu):
        try:
            stok_listesi = self.stok_listele() # Tüm stokları al
            if not stok_listesi:
                return False

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stok_Raporu"

            # Başlıklar
            headers = ["Ürün Kodu", "Ürün Adı", "Mevcut Stok", "Alış Fiyatı (TL)", "Satış Fiyatı (TL)", "KDV Oranı (%)"]
            ws.append(headers)

            # Başlık Stili
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for col_idx, header_text in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header_text) + 5
            # Verileri yazdır
            for stok_item in stok_listesi:
                # stok_item: (id, urun_kodu, urun_adi, stok_miktari, alis_fiyati, satis_fiyati, kdv_orani)
                urun_kodu = stok_item[1]
                urun_adi = stok_item[2]
                stok_miktari = stok_item[3]
                alis_fiyati = stok_item[4]
                satis_fiyati = stok_item[5]
                kdv_orani = stok_item[6]
                ws.append([urun_kodu, urun_adi, stok_miktari, alis_fiyati, satis_fiyati, kdv_orani])
            
            # Para birimi ve sayı formatları (isteğe bağlı)
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                if row[2].value is not None: row[2].number_format = '#,##0.00' # Stok miktarı
                if row[3].value is not None: row[3].number_format = '#,##0.00₺' # Alış fiyatı
                if row[4].value is not None: row[4].number_format = '#,##0.00₺' # Satış fiyatı
                if row[5].value is not None: row[5].number_format = '0"%"' # KDV

            wb.save(dosya_yolu)
            return True
        except Exception as e:
            traceback.print_exc()
            return False

    def mevcut_stok_verilerini_excel_disa_aktar(self, dosya_yolu):
        return self.stok_raporu_excel_olustur(dosya_yolu)

    def _get_cari_bakiye_snapshot(self, cari_id, cari_tip, tarih_str):
        """
        Belirli bir tarihteki cari bakiyesini hesaplar.
        tarih_str: 'YYYY-MM-DD' formatında tarih.
        Dönüş: {'onceki_bakiye': float, 'bugun_odenen': float, 'kalan_borc': float}
        """
        onceki_bakiye = 0.0
        bugun_odenen = 0.0
        bugun_tahsil_edilen = 0.0 # Müşteri için
        
        try:
            # Önceki Bakiye: Belirtilen tarihten önceki tüm hareketlerin toplamı
            query_onceki = """
                SELECT islem_tipi, tutar FROM cari_hareketler
                WHERE cari_id = ? AND cari_tip = ? AND tarih < ?
            """
            self.c.execute(query_onceki, (cari_id, cari_tip, tarih_str))
            onceki_hareketler = self.c.fetchall()

            for h in onceki_hareketler:
                if cari_tip == 'MUSTERI':
                    if h['islem_tipi'] == 'ALACAK': onceki_bakiye += h['tutar'] # Müşterinin bize borcu
                    elif h['islem_tipi'] == 'TAHSILAT': onceki_bakiye -= h['tutar'] # Müşteriden tahsilat
                elif cari_tip == 'TEDARIKCI':
                    if h['islem_tipi'] == 'BORC': onceki_bakiye += h['tutar'] # Bizim tedarikçiye borcumuz
                    elif h['islem_tipi'] == 'ODEME': onceki_bakiye -= h['tutar'] # Tedarikçiye ödeme

            # Bugün Yapılan Ödeme/Tahsilat (Fatura tarihi ile aynı güne ait cari hareketler)
            query_bugun = """
                SELECT islem_tipi, tutar FROM cari_hareketler
                WHERE cari_id = ? AND cari_tip = ? AND tarih = ?
            """
            self.c.execute(query_bugun, (cari_id, cari_tip, tarih_str))
            bugun_hareketler = self.c.fetchall()

            for h in bugun_hareketler:
                if cari_tip == 'MUSTERI':
                    if h['islem_tipi'] == 'TAHSILAT': bugun_tahsil_edilen += h['tutar']
                elif cari_tip == 'TEDARIKCI':
                    if h['islem_tipi'] == 'ODEME': bugun_odenen += h['tutar']

            # Kalan Borç: Önceki bakiye + (bugünkü ALACAK/BORÇ hareketleri) - (bugünkü TAHSİLAT/ÖDEME hareketleri)
            # Bu, faturanın kendisi haricindeki gün içi hareketleri hesaba katar.
            kalan_borc = onceki_bakiye
            for h in bugun_hareketler:
                if cari_tip == 'MUSTERI':
                    if h['islem_tipi'] == 'ALACAK': kalan_borc += h['tutar']
                    elif h['islem_tipi'] == 'TAHSILAT': kalan_borc -= h['tutar']
                elif cari_tip == 'TEDARIKCI':
                    if h['islem_tipi'] == 'BORC': kalan_borc += h['tutar']
                    elif h['islem_tipi'] == 'ODEME': kalan_borc -= h['tutar']

            return {
                'onceki_bakiye': onceki_bakiye,
                'bugun_odenen': bugun_tahsil_edilen if cari_tip == 'MUSTERI' else bugun_odenen,
                'kalan_borc': kalan_borc
            }
        except Exception as e:
            logging.error(f"Cari bakiye anlık görüntü hatası: {e}", exc_info=True)
            return {
                'onceki_bakiye': 0.0,
                'bugun_odenen': 0.0,
                'kalan_borc': 0.0
            }

    def _get_cari_bakiye_snapshot(self, cari_id, cari_tip, tarih_str):
        """
        Belirli bir tarihteki cari bakiyesini hesaplar.
        tarih_str: 'YYYY-MM-DD' formatında tarih.
        Dönüş: {'onceki_bakiye': float, 'bugun_odenen': float, 'kalan_borc': float}
        """
        onceki_bakiye = 0.0
        bugun_odenen = 0.0
        bugun_tahsil_edilen = 0.0 # Müşteri için
        
        try:
            # Önceki Bakiye: Belirtilen tarihten önceki tüm hareketlerin toplamı
            query_onceki = """
                SELECT islem_tipi, tutar FROM cari_hareketler
                WHERE cari_id = ? AND cari_tip = ? AND tarih < ?
            """
            self.c.execute(query_onceki, (cari_id, cari_tip, tarih_str))
            onceki_hareketler = self.c.fetchall()

            for h in onceki_hareketler:
                if cari_tip == 'MUSTERI':
                    if h['islem_tipi'] == 'ALACAK': onceki_bakiye += h['tutar'] # Müşterinin bize borcu
                    elif h['islem_tipi'] == 'TAHSILAT': onceki_bakiye -= h['tutar'] # Müşteriden tahsilat
                elif cari_tip == 'TEDARIKCI':
                    if h['islem_tipi'] == 'BORC': onceki_bakiye += h['tutar'] # Bizim tedarikçiye borcumuz
                    elif h['islem_tipi'] == 'ODEME': onceki_bakiye -= h['tutar'] # Tedarikçiye ödeme

            # Bugün Yapılan Ödeme/Tahsilat (Fatura tarihi ile aynı güne ait cari hareketler)
            query_bugun = """
                SELECT islem_tipi, tutar FROM cari_hareketler
                WHERE cari_id = ? AND cari_tip = ? AND tarih = ?
            """
            self.c.execute(query_bugun, (cari_id, cari_tip, tarih_str))
            bugun_hareketler = self.c.fetchall()

            for h in bugun_hareketler:
                if cari_tip == 'MUSTERI':
                    if h['islem_tipi'] == 'TAHSILAT': bugun_tahsil_edilen += h['tutar']
                elif cari_tip == 'TEDARIKCI':
                    if h['islem_tipi'] == 'ODEME': bugun_odenen += h['tutar']

            # Kalan Borç: Önceki bakiye + (bugünkü ALACAK/BORÇ hareketleri) - (bugünkü TAHSİLAT/ÖDEME hareketleri)
            # Bu, faturanın kendisi haricindeki gün içi hareketleri hesaba katar.
            kalan_borc = onceki_bakiye
            for h in bugun_hareketler:
                if cari_tip == 'MUSTERI':
                    if h['islem_tipi'] == 'ALACAK': kalan_borc += h['tutar']
                    elif h['islem_tipi'] == 'TAHSILAT': kalan_borc -= h['tutar']
                elif cari_tip == 'TEDARIKCI':
                    if h['islem_tipi'] == 'BORC': kalan_borc += h['tutar']
                    elif h['islem_tipi'] == 'ODEME': kalan_borc -= h['tutar']

            return {
                'onceki_bakiye': onceki_bakiye,
                'bugun_odenen': bugun_tahsil_edilen if cari_tip == 'MUSTERI' else bugun_odenen,
                'kalan_borc': kalan_borc
            }
        except Exception as e:
            logging.error(f"Cari bakiye anlık görüntü hatası: {e}", exc_info=True)
            return {
                'onceki_bakiye': 0.0,
                'bugun_odenen': 0.0,
                'kalan_borc': 0.0
            }

    def cari_ekstresi_pdf_olustur(self, cari_tip, cari_id, baslangic_tarih_str, bitis_tarih_str, dosya_yolu):
        conn_thread = None # DEĞİŞİKLİK BURADA BAŞLIYOR
        try:
            conn_thread = sqlite3.connect(self.db_name, detect_types=sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES)
            conn_thread.row_factory = sqlite3.Row
            cursor_thread = conn_thread.cursor() # Yeni cursor

            # Sayfa kenar boşlukları ve temel boyutlar
            page_margin_x = 2.0 * cm
            page_margin_y = 2.0 * cm

            # 1. Cari bilgilerini çek
            cari_bilgi = None
            if cari_tip == 'MUSTERI':
                cursor_thread.execute("SELECT ad, musteri_kodu FROM musteriler WHERE id=?", (cari_id,)) # self.c yerine cursor_thread
                cari_bilgi = cursor_thread.fetchone()
                cari_adi = cari_bilgi['ad'] if cari_bilgi else "Bilinmiyor"
                cari_kodu = cari_bilgi['musteri_kodu'] if cari_bilgi else ""
            else: # TEDARIKCI
                cursor_thread.execute("SELECT ad, tedarikci_kodu FROM tedarikciler WHERE id=?", (cari_id,)) # self.c yerine cursor_thread
                cari_bilgi = cursor_thread.fetchone()
                cari_adi = cari_bilgi['ad'] if cari_bilgi else "Bilinmiyor"
                cari_kodu = cari_bilgi['tedarikci_kodu'] if cari_bilgi else ""

            rapor_baslik = f"Cari Hesap Ekstresi: {cari_adi} (Kod: {cari_kodu})"

            # 2. Ekstre verilerini çek
            # Devreden bakiye için de cursor_thread kullanın
            cursor_thread.execute("""
                SELECT islem_tipi, tutar FROM cari_hareketler
                WHERE cari_tip = ? AND cari_id = ? AND tarih < ?
            """, (cari_tip, cari_id, baslangic_tarih_str))
            hareketler_oncesi = cursor_thread.fetchall() # self.c yerine cursor_thread

            devreden_bakiye = 0.0
            for hareket in hareketler_oncesi:
                tutar = hareket['tutar'] or 0.0
                islem_tipi = hareket['islem_tipi']

                if cari_tip == 'MUSTERI':
                    if islem_tipi in ['BORC', 'ALACAK']:
                        devreden_bakiye += tutar
                    elif islem_tipi in ['TAHSILAT', 'ODEME']:
                        devreden_bakiye -= tutar
                elif cari_tip == 'TEDARIKCI':
                    if islem_tipi in ['BORC', 'ALACAK']:
                        devreden_bakiye += tutar
                    elif islem_tipi in ['TAHSILAT', 'ODEME']:
                        devreden_bakiye -= tutar

            # Ana hareketler listesi için cursor_thread kullanın
            query_hareketler = """
                SELECT
                    ch.id, ch.tarih, STRFTIME('%H:%M:%S', ch.olusturma_tarihi_saat) AS islem_saati,
                    ch.islem_tipi, ch.tutar, ch.aciklama, ch.referans_id, ch.referans_tip,
                    f.fatura_no, f.odeme_turu
                FROM cari_hareketler ch
                LEFT JOIN faturalar f ON ch.referans_id = f.id AND ch.referans_tip LIKE 'FATURA%'
                WHERE ch.cari_tip = ? AND ch.cari_id = ? AND ch.tarih >= ? AND ch.tarih <= ?
                ORDER BY ch.tarih ASC, ch.olusturma_tarihi_saat ASC
            """
            params_hareketler = [cari_tip, cari_id, baslangic_tarih_str, bitis_tarih_str]
            cursor_thread.execute(query_hareketler, params_hareketler) # self.c yerine cursor_thread
            hareketler_listesi = cursor_thread.fetchall()

            if not hareketler_listesi and devreden_bakiye == 0:
                return False, "PDF oluşturulacak cari ekstre verisi bulunamadı."

            # 3. PDF Dokümanı oluştur (SimpleDocTemplate)
            doc = SimpleDocTemplate(dosya_yolu, pagesize=A4,
                                    rightMargin=page_margin_x, leftMargin=page_margin_x,
                                    topMargin=page_margin_y, bottomMargin=page_margin_y)

            # Stil yönetimi
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='TitleStyle', fontName=TURKISH_FONT_BOLD, fontSize=14, alignment=1))
            styles.add(ParagraphStyle(name='SubtitleStyle', fontName=TURKISH_FONT_NORMAL, fontSize=10, alignment=1))
            styles.add(ParagraphStyle(name='HeaderDetail', fontName=TURKISH_FONT_NORMAL, fontSize=8, alignment=0))
            styles.add(ParagraphStyle(name='TableHeading', fontName=TURKISH_FONT_BOLD, fontSize=7, alignment=1))
            styles.add(ParagraphStyle(name='TableNormal', fontName=TURKISH_FONT_NORMAL, fontSize=7, alignment=0))
            styles.add(ParagraphStyle(name='TableRight', fontName=TURKISH_FONT_NORMAL, fontSize=7, alignment=2))
            styles.add(ParagraphStyle(name='TableBold', fontName=TURKISH_FONT_BOLD, fontSize=7, alignment=0))

            # Story (PDF içeriği) oluştur
            story = []

            # Ana Başlıklar
            story.append(Paragraph("Cari Hesap Ekstresi", styles['TitleStyle']))
            story.append(Paragraph(rapor_baslik, styles['SubtitleStyle']))
            story.append(Paragraph(f"Tarih Aralığı: {baslangic_tarih_str} - {bitis_tarih_str}", styles['SubtitleStyle']))
            story.append(Spacer(0, 0.5*cm))

            # Şirket Bilgileri (Sol üstte)
            sirket_bilgileri_pdf_local = self.sirket_bilgilerini_yukle() # self.sirket_bilgilerini_yukle
            story.append(Paragraph(f"Şirket Adı: {sirket_bilgileri_pdf_local.get('sirket_adi', 'Şirket Adı')}", styles['HeaderDetail']))
            story.append(Paragraph(f"Adres: {sirket_bilgileri_pdf_local.get('sirket_adresi', '')}", styles['HeaderDetail']))
            story.append(Spacer(0, 0.5*cm))

            # Tablo Başlıkları
            table_headers = [
                Paragraph("ID", styles['TableHeading']), Paragraph("Tarih", styles['TableHeading']), Paragraph("Saat", styles['TableHeading']),
                Paragraph("İşlem Tipi", styles['TableHeading']), Paragraph("Referans", styles['TableHeading']), Paragraph("Ödeme Türü", styles['TableHeading']),
                Paragraph("Açıklama/Detay", styles['TableHeading']), Paragraph("Borç", styles['TableHeading']), Paragraph("Alacak", styles['TableHeading']),
                Paragraph("Bakiye", styles['TableHeading'])
            ]

            col_widths = [0.8*cm, 1.8*cm, 1.2*cm, 2.0*cm, 2.0*cm, 2.0*cm, 5.0*cm, 2.0*cm, 2.0*cm, 2.0*cm]
            table_data_rows = []

            # Devir bakiyesi satırı
            table_data_rows.append([
                Paragraph("", styles['TableNormal']), Paragraph("", styles['TableNormal']), Paragraph("", styles['TableNormal']),
                Paragraph("DEVİR", styles['TableBold']), Paragraph("", styles['TableNormal']), Paragraph("", styles['TableNormal']),
                Paragraph("", styles['TableNormal']), Paragraph("", styles['TableNormal']), Paragraph(self._format_currency(devreden_bakiye), styles['TableRight']),
                Paragraph(self._format_currency(devreden_bakiye), styles['TableRight'])
            ])

            bakiye_current = devreden_bakiye
            for hareket in hareketler_listesi:
                tarih_formatted = datetime.strptime(str(hareket['tarih']), '%Y-%m-%d').strftime('%d.%m.%Y')
                islem_saati = hareket['islem_saati'] if hareket['islem_saati'] else ''
                odeme_turu = hareket['odeme_turu'] if hareket['odeme_turu'] else ''
                aciklama = hareket['aciklama'] if hareket['aciklama'] else ''
                referans_display = hareket['fatura_no'] if hareket['referans_tip'] == 'FATURA' else (hareket['referans_tip'] if hareket['referans_tip'] else '')

                borc_val, alacak_val = "", ""
                if cari_tip == 'MUSTERI':
                    if hareket['islem_tipi'] == 'ALACAK' or hareket['referans_tip'] == 'FATURA' or hareket['referans_tip'] == 'VERESIYE_BORC_MANUEL':
                        alacak_val = self._format_currency(hareket['tutar'])
                        bakiye_current += hareket['tutar']
                    elif hareket['islem_tipi'] == 'TAHSILAT' or hareket['referans_tip'] == 'FATURA_SATIS_PESIN':
                        borc_val = self._format_currency(hareket['tutar'])
                        bakiye_current -= hareket['tutar']
                elif cari_tip == 'TEDARIKCI':
                    if hareket['islem_tipi'] == 'BORC' or hareket['referans_tip'] == 'FATURA' or hareket['referans_tip'] == 'VERESIYE_BORC_MANUEL':
                        alacak_val = self._format_currency(hareket['tutar'])
                        bakiye_current += hareket['tutar']
                    elif hareket['islem_tipi'] == 'ODEME' or hareket['referans_tip'] == 'FATURA_ALIS_PESIN':
                        borc_val = self._format_currency(hareket['tutar'])
                        bakiye_current -= hareket['tutar']

                table_data_rows.append([
                    Paragraph(str(hareket['id']), styles['TableNormal']), Paragraph(tarih_formatted, styles['TableNormal']), Paragraph(islem_saati, styles['TableNormal']),
                    Paragraph(hareket['islem_tipi'], styles['TableNormal']), Paragraph(referans_display, styles['TableNormal']), Paragraph(odeme_turu, styles['TableNormal']),
                    Paragraph(aciklama, styles['TableNormal']), Paragraph(borc_val, styles['TableRight']), Paragraph(alacak_val, styles['TableRight']),
                    Paragraph(self._format_currency(bakiye_current), styles['TableRight'])
                ])

            table = Table([table_headers] + table_data_rows, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#D0D0D0")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), TURKISH_FONT_BOLD),
                ('FONTSIZE', (0,0), (-1,0), 7),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('TOPPADDING', (0,0), (-1,0), 6),

                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),

                ('BACKGROUND', (0,1), (-1,1), colors.HexColor("#EFEFEF")),
                ('FONTNAME', (0,1), (-1,1), TURKISH_FONT_BOLD),

                ('LEFTPADDING', (0,0), (-1,-1), 2),
                ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ]))

            final_bakiye_row = Table([[
                Paragraph(f"Son Bakiye:", styles['TableBold']),
                Paragraph(self._format_currency(bakiye_current), styles['TableRight'])
            ]], [sum(col_widths) - 2*cm, 2*cm])

            def handle_page_footer(canvas, doc):
                canvas.saveState()
                canvas.setFont(TURKISH_FONT_NORMAL, 7)
                canvas.drawString(page_margin_x, page_margin_y / 2, "Bu ekstre On Muhasebe Programı ile oluşturulmuştur.")
                canvas.drawCentredString(A4[0] / 2, page_margin_y / 2, f"Sayfa {doc.page}")
                canvas.restoreState()

            story.append(table)
            story.append(Spacer(0, 0.5*cm))
            story.append(final_bakiye_row)

            doc.build(story, onFirstPage=handle_page_footer, onLaterPages=handle_page_footer)

            return True, f"Cari Ekstresi PDF olarak kaydedildi: {dosya_yolu}"

        except Exception as e:
            logging.error(f"Cari Ekstresi PDF oluşturulurken hata: {e}", exc_info=True)
            return False, f"Cari Ekstresi PDF oluşturulurken bir hata oluştu: {e}"
        finally: 
            if conn_thread:
                conn_thread.close()


    def fatura_pdf_olustur(self, fatura_id, dosya_yolu):
        """
        Belirtilen fatura ID'sine göre modern ve detaylı bir PDF fatura çıktısı oluşturur.
        İskonto detayları, cari bakiye, şirket bilgileri, imza alanı, sayfa numarası içerir.
        """
        try:
            fatura_ana = self.fatura_getir_by_id(fatura_id)
            if not fatura_ana:
                return False, "PDF oluşturulacak fatura bilgileri bulunamadı."


            _id, f_no, tarih_db, tip, c_id, toplam_kdv_haric_fatura_ana, toplam_kdv_dahil_fatura_ana, odeme_turu_db, misafir_adi_db, kb_id_db, \
            olusturma_tarihi_saat, olusturan_kullanici_id, son_guncelleme_tarihi_saat, \
            son_guncelleyen_kullanici_id, fatura_notlari_db, vade_tarihi_db, genel_iskonto_tipi_db, genel_iskonto_degeri_db = fatura_ana

            fatura_kalemleri_db = self.fatura_detay_al(fatura_id)

            # --- Şirket Bilgilerini Çekme ---
            sirket_bilgileri = self.sirket_bilgilerini_yukle()

            # --- Cari Bilgilerini Çekme ve Hazırlama ---
            is_perakende_satis_pdf = (str(c_id) == str(self.perakende_musteri_id) and tip == 'SATIŞ')
            cari_bilgileri_pdf = {}
            cari_tip_for_db_query = 'MUSTERI' if tip == 'SATIŞ' else 'TEDARIKCI'

            if is_perakende_satis_pdf:
                cari_adi_pdf = "Perakende Satış Müşterisi"
                if misafir_adi_db: cari_adi_pdf += f" (Misafir: {misafir_adi_db})"
                cari_bilgileri_pdf = {
                    "ad": cari_adi_pdf, "adres": "N/A", "telefon": "N/A",
                    "vergi_dairesi": "N/A", "vergi_no": "N/A"
                }
            else:
                if tip == 'SATIŞ': cari_db = self.musteri_getir_by_id(c_id)
                else: cari_db = self.tedarikci_getir_by_id(c_id)
                if cari_db:
                    cari_bilgileri_pdf = {
                        "ad": cari_db['ad'], "adres": cari_db['adres'] or "N/A", "telefon": cari_db['telefon'] or "N/A",
                        "vergi_dairesi": cari_db['vergi_dairesi'] or "N/A", "vergi_no": cari_db['vergi_no'] or "N/A"
                    }
                else:
                    cari_bilgileri_pdf = {"ad": "Bilinmeyen Cari", "adres": "N/A", "telefon": "N/A", "vergi_dairesi": "N/A", "vergi_no": "N/A"}

            # --- Kasa/Banka Bilgilerini Çekme ---
            kb_bilgi = None
            if kb_id_db:
                kb_bilgi = self.kasa_banka_getir_by_id(kb_id_db)

            # --- Cari Bakiye Bilgilerini Çekme (Anlık Görüntü) ---
            # Faturanın kesildiği tarihteki bakiye durumunu almak için
            cari_bakiye_snapshot = self._get_cari_bakiye_snapshot(c_id, cari_tip_for_db_query, tarih_db)


            # --- Kullanıcı Adlarını Çekme ---
            kullanicilar_map = {k['id']: k['kullanici_adi'] for k in self.kullanici_listele()}
            olusturan_adi = kullanicilar_map.get(olusturan_kullanici_id, "Bilinmiyor")
            son_guncelleyen_adi = kullanicilar_map.get(son_guncelleyen_kullanici_id, "Bilinmiyor")


            # --- PDF Oluşturma Başlangıcı ---
            c = rp_canvas.Canvas(dosya_yolu, pagesize=A4)
            width, height = A4 # Sayfa boyutları
            page_margin_x = 30 # Yatay kenar boşluğu
            page_margin_top = 50 # Üst kenar boşluğu
            page_margin_bottom_min = 100 # Alt kenar boşluğu (toplamlar ve dipnotlar için minimum)

            # Stil sözlüğünü güncelle
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='TurkishNormal', fontName=TURKISH_FONT_NORMAL, fontSize=9, leading=11))
            styles.add(ParagraphStyle(name='TurkishBold', fontName=TURKISH_FONT_BOLD, fontSize=9, leading=11))
            styles.add(ParagraphStyle(name='TurkishNormalSmall', fontName=TURKISH_FONT_NORMAL, fontSize=7, leading=9))
            styles.add(ParagraphStyle(name='TurkishBoldSmall', fontName=TURKISH_FONT_BOLD, fontSize=7, leading=9))
            styles.add(ParagraphStyle(name='TitleStyle', fontName=TURKISH_FONT_BOLD, fontSize=14, alignment=1)) # CENTER
            styles.add(ParagraphStyle(name='HeaderDetail', fontName=TURKISH_FONT_NORMAL, fontSize=10, alignment=2)) # RIGHT
            styles.add(ParagraphStyle(name='Footer', fontName=TURKISH_FONT_NORMAL, fontSize=7, alignment=1)) # CENTER

            # Ortak TableStyle objesi tanımlaması - ARKA PLAN RENGİ VE SATIR STİLLERİ DÜZELTİLDİ
            kalem_table_style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#F0F0F0")), # Açık gri arka plan
                ('TEXTCOLOR', (0,0), (-1,0), colors.black), # Başlık metin rengi siyah
                ('ALIGN', (0,0), (-1,0), 'CENTER'), # Başlık hizalaması
                ('FONTNAME', (0,0), (-1,0), TURKISH_FONT_BOLD),
                ('FONTSIZE', (0,0), (-1,0), 7),
                ('BOTTOMPADDING', (0,0), (-1,0), 3), # Başlık altında boşluğu azalttık
                ('TOPPADDING', (0,0), (-1,0), 3),    # Başlık üstünde boşluğu azalttık
                ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor("#D0D0D0")), # Başlık altında ince çizgi

                # Veri satırları
                ('BACKGROUND', (0,1), (-1,-1), colors.white), # Tüm veri satırları beyaz arka plan
                ('TEXTCOLOR', (0,1), (-1,-1), colors.black), # Veri metin rengi siyah
                ('ALIGN', (0,1), (0,-1), 'CENTER'), # Sıra no ortala
                ('ALIGN', (1,1), (1,-1), 'LEFT'), # Ürünler sola yaslı
                ('ALIGN', (2,1), (-1,-1), 'RIGHT'), # Diğer sayısal değerler sağa yaslı
                ('FONTNAME', (0,1), (-1,-1), TURKISH_FONT_NORMAL),
                ('FONTSIZE', (0,1), (-1,-1), 7),
                ('BOTTOMPADDING', (0,1), (-1,-1), 2), # Veri satırları altında boşluğu azalttık
                ('TOPPADDING', (0,1), (-1,-1), 2),    # Veri satırları üstünde boşluğu azalttık
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F0F0F0")]), # Zebra deseni

                ('GRID', (0,0), (-1,-1), 0, colors.white), # Grid çizgilerini kaldır
            ])

            # PDF'e çizmeyi kolaylaştırmak için ana döngü fonksiyonu
            # table_header_row: Tablo başlık satırı (Paragraph objeleri içeren liste)
            # table_data_rows_on_page: Bu sayfada çizilecek kalem veri satırları (Paragraph objeleri içeren listeler)
            def draw_page_content(canvas_obj, current_page_num, total_page_num, header_start_y, table_header_row, table_data_rows_on_page):
                canvas_obj.setFillColor(colors.black)

                # Sol Üst: Şirket Bilgileri
                textobject = canvas_obj.beginText(page_margin_x, header_start_y)
                textobject.setFont(TURKISH_FONT_BOLD, 12)
                textobject.textLine(sirket_bilgileri.get("sirket_adi", "ŞİRKET ADI"))
                textobject.setFont(TURKISH_FONT_NORMAL, 8)
                s_adres_parcalari = sirket_bilgileri.get("sirket_adresi", "").split(',')
                for parca in s_adres_parcalari:
                    textobject.textLine(parca.strip())
                textobject.textLine(f"Tel: {sirket_bilgileri.get('sirket_telefonu', '')}")
                textobject.textLine(f"Email: {sirket_bilgileri.get('sirket_email', '')}")
                textobject.textLine(f"VD: {sirket_bilgileri.get('sirket_vergi_dairesi', '')} / VN: {sirket_bilgileri.get('sirket_vergi_no', '')}")
                canvas_obj.drawText(textobject)

                # Sağ Üst: Fatura Bilgileri
                canvas_obj.setFont(TURKISH_FONT_BOLD, 10) # Örnekteki gibi daha küçük font
                canvas_obj.drawRightString(width - page_margin_x, height - 50, "Satış Kodu") # Örnekteki "Satış Kodu"
                canvas_obj.setFont(TURKISH_FONT_NORMAL, 10)
                canvas_obj.drawRightString(width - page_margin_x, height - 65, f_no) # Fatura No (Örnekteki "2507020015-CT")
                canvas_obj.setFont(TURKISH_FONT_BOLD, 10) # Örnekteki gibi bold isim
                canvas_obj.drawRightString(width - page_margin_x, height - 80, cari_bilgileri_pdf["ad"].split('(')[0].strip()) # Sadece isim kısmı
                canvas_obj.setFont(TURKISH_FONT_NORMAL, 9)
                try: formatted_tarih = datetime.strptime(tarih_db, '%Y-%m-%d').strftime('%d/%m/%Y %H:%M:%S') # Saat de eklendi
                except: formatted_tarih = tarih_db
                canvas_obj.drawRightString(width - page_margin_x, height - 95, f"Tarih: {formatted_tarih}")
                canvas_obj.drawRightString(width - page_margin_x, height - 110, f"Ödeme Tipi: {odeme_turu_db if odeme_turu_db else '-'}")

                # İmza alanının hemen altında bir yerde oluşturulma bilgisi
                canvas_obj.setFont(TURKISH_FONT_NORMAL, 7)
                canvas_obj.drawString(width - page_margin_x - 535, height - 130, f"Oluşturan: {olusturan_adi}")
                if son_guncelleyen_adi and son_guncelleyen_adi != olusturan_adi:
                    canvas_obj.drawString(width - page_margin_x - 535, height - 140, f"Güncelleyen: {son_guncelleyen_adi}")

                # Ortalanmış Logo
                logo_path = sirket_bilgileri.get("sirket_logo_yolu", "")
                if logo_path and os.path.exists(logo_path):
                    try:
                        img = Image.open(logo_path)
                        aspect_ratio = img.width / img.height
                        desired_height_logo = 40
                        desired_width_logo = desired_height_logo * aspect_ratio
                        x_pos_logo = (width / 2) - (desired_width_logo / 2)
                        y_pos_logo = height - 70 # Örnekteki gibi daha aşağıda ve ortalanmış
                        canvas_obj.drawImage(logo_path, x_pos_logo, y_pos_logo - desired_height_logo / 2,
                                             width=desired_width_logo, height=desired_height_logo, preserveAspectRatio=True)
                    except Exception as e_logo:
                        logging.warning(f"Logo çizim hatası: {e_logo}")

                # --- Cari Bilgileri Alanı (Örnekteki gibi kutu dışı) ---
                # Fatura bilgisinin hemen altından başlayacak, örnekte "Sayın" yok
                current_y_for_blocks = height - 150 # Ortak Y koordinatı
                canvas_obj.setFont(TURKISH_FONT_BOLD, 9) # Örnekteki gibi daha küçük font
                # Cari Bilgileri kısmının başlığı yok gibi, direkt "Müşteri Bilgileri" altına alınmış.
                # Şimdilik, örnekteki gibi sadece bilgileri yerleştirelim.

                # Cari bilgi kutusunu kaldırdık, sadece bir çizgiyle ayıracağız.
                canvas_obj.line(page_margin_x, current_y_for_blocks, width - page_margin_x, current_y_for_blocks) 
                current_y_for_blocks -= 15 # Çizgi sonrası boşluk

                current_table_y_pos_in_draw = current_y_for_blocks # Kalemler tablosunun başlangıç Y konumu

                # Tablo başlık satırını çiz
                header_table = Table([table_header_row], colWidths=kalem_col_widths)
                header_table.setStyle(kalem_table_style) # Ortak TableStyle objesini kullanıyoruz
                header_table.wrapOn(canvas_obj, width - 2 * page_margin_x, height)
                header_table.drawOn(canvas_obj, page_margin_x, current_table_y_pos_in_draw)
                current_table_y_pos_in_draw -= header_table._height # Başlık sonrası boşluk

                # Kalemleri çiz
                for row_data in table_data_rows_on_page:
                    row_obj = Table([row_data], colWidths=kalem_col_widths)
                    row_obj.setStyle(kalem_table_style) # Ortak kalem stilini uygula
                    row_obj.wrapOn(canvas_obj, width - 2 * page_margin_x, height)

                    if current_table_y_pos_in_draw - row_obj._height < page_margin_bottom_min:
                        return False, current_table_y_pos_in_draw

                    row_obj.drawOn(canvas_obj, page_margin_x, current_table_y_pos_in_draw - row_obj._height)
                    current_table_y_pos_in_draw -= row_obj._height

                return True, current_table_y_pos_in_draw # Başarılı ve kalan Y pozisyonunu döndür

            # --- Ana PDF Oluşturma ve Sayfalama Döngüsü ---
            c = rp_canvas.Canvas(dosya_yolu, pagesize=A4)

            kalem_data_for_table = [] # Sadece veri satırları (başlık hariç)

            # Yeni sütun genişlikleri, örnek PDF'e göre ayarlandı.
            # Metinlerin sığması için Paragraf objelerini kullanıyoruz, bu yüzden wrapOn metodu işe yarayacak.
            kalem_col_widths = [
                0.6 * cm,   # Sıra
                7.8 * cm,   # Ürünler (Ürün Kodu + Adı) - Geniş tutuldu
                1.0* cm,   # Miktar
                2.0 * cm,   # Birim Fiyat (TL 75,00 gibi)
                1.0 * cm,   # KDV%
                1.35 * cm,   # İsk.1(%)
                1.35 * cm,   # İsk.2(%)
                1.5 * cm,   # İsk Tutarı
                2.6 * cm    # Tutar (KDV Dahil) - En sağda
            ]
            # Toplam genişlik: 0.8+7.0+1.5+2.0+1.0+1.0+1.0+2.0+2.5 = 18.8 cm. Bu A4'e rahat sığar.


            # Fatura kalemlerini topla ve formatla
            sira_no = 1
            for kalem in fatura_kalemleri_db:
                miktar_gosterim = f"{kalem['miktar']:.2f}".rstrip('0').rstrip('.')
                
                # İskontolu Birim Fiyat (KDV Dahil) Hesapla
                iskontolu_birim_fiyat_kdv_dahil = (kalem['kalem_toplam_kdv_dahil'] / kalem['miktar']) if kalem['miktar'] != 0 else 0.0

                # Uygulanan Kalem İskonto Tutarı (KDV Dahil) Hesapla
                original_birim_fiyat_kdv_dahil_kalem = kalem['birim_fiyat'] * (1 + kalem['kdv_orani'] / 100) # orijinal kdv hariç * (1+kdv)
                uygulanan_kalem_iskonto_tutari = (original_birim_fiyat_kdv_dahil_kalem - iskontolu_birim_fiyat_kdv_dahil) * kalem['miktar']
                
                # "Ürünler" sütununa ürün kodu ve adını birlikte alıyoruz (örnekteki gibi)
                urunler_text = f"{kalem['urun_adi']}\n({kalem['urun_kodu']})"


                kalem_data_for_table.append([
                    Paragraph(str(sira_no), styles['TurkishNormalSmall']),
                    Paragraph(urunler_text, styles['TurkishNormalSmall']), # Ürünler sütunu
                    Paragraph(miktar_gosterim, styles['TurkishNormalSmall']),
                    Paragraph(self._format_currency(iskontolu_birim_fiyat_kdv_dahil), styles['TurkishNormalSmall']),
                    Paragraph(f"{kalem['kdv_orani']:.0f}%", styles['TurkishNormalSmall']),
                    Paragraph(f"{kalem['iskonto_yuzde_1']:.2f}".rstrip('0').rstrip('.') if kalem['iskonto_yuzde_1'] is not None else "0", styles['TurkishNormalSmall']),
                    Paragraph(f"{kalem['iskonto_yuzde_2']:.2f}".rstrip('0').rstrip('.') if kalem['iskonto_yuzde_2'] is not None else "0", styles['TurkishNormalSmall']),
                    Paragraph(self._format_currency(uygulanan_kalem_iskonto_tutari), styles['TurkishNormalSmall']),
                    Paragraph(self._format_currency(kalem['kalem_toplam_kdv_dahil']), styles['TurkishNormalSmall'])
                ])
                sira_no += 1

            # Tablo başlığı (sütunlar "Ürün Kodu" ve "Ürün Adı" yerine "Ürünler" olacak şekilde güncellendi)
            kalem_table_header = [
                Paragraph("#", styles['TurkishBoldSmall']),
                Paragraph("Ürünler", styles['TurkishBoldSmall']),
                Paragraph("Mik", styles['TurkishBoldSmall']),
                Paragraph("B.Fiyat", styles['TurkishBoldSmall']),
                Paragraph("KDV%", styles['TurkishBoldSmall']),
                Paragraph("İsk.1(%)", styles['TurkishBoldSmall']),
                Paragraph("İsk.2(%)", styles['TurkishBoldSmall']),
                Paragraph("İsk Tutarı", styles['TurkishBoldSmall']),
                Paragraph("Tutar (KDV Dahil)", styles['TurkishBoldSmall'])
            ]

            total_data_rows_count = len(kalem_data_for_table)

            # Sayfa hesaplamaları
            approx_row_height = 13 # Tek bir veri satırının yaklaşık yüksekliği (point)
            available_table_height_for_rows_per_page = height - page_margin_top - 180 - page_margin_bottom_min # Üst boşluklar, cari bilgiler ve alt boşluklar hariç
            max_rows_per_page = int(available_table_height_for_rows_per_page / approx_row_height)

            if max_rows_per_page <= 0: # En az 1 satır sığmalı
                max_rows_per_page = 1

            total_pages = (total_data_rows_count + max_rows_per_page -1) // max_rows_per_page
            if total_pages == 0 and total_data_rows_count == 0:
                total_pages = 1 # Hiç kalem olmasa bile 1 sayfa olmalı


            current_row_idx_for_drawing = 0
            page_count = 0
            
            # Sayfa döngüsü
            while current_row_idx_for_drawing <= total_data_rows_count:
                page_count += 1

                if page_count > 1: # İlk sayfa dışındaki tüm sayfalarda yeni sayfa başlat
                    c.showPage()

                start_data_idx_for_page = current_row_idx_for_drawing
                end_data_idx_for_page = min(start_data_idx_for_page + max_rows_per_page, total_data_rows_count)
                
                rows_to_draw_on_this_page = kalem_data_for_table[start_data_idx_for_page : end_data_idx_for_page]

                # draw_page_content'ı çağır
                # last_y_pos_on_page değişkeni artık kalemlerin altındaki boşluğun kontrolünde kullanılacak.
                success_drawing, last_y_pos_on_page = draw_page_content(
                    c,
                    page_count,
                    total_pages,
                    height - page_margin_top, # Header'ın başlangıç Y konumu
                    kalem_table_header, # Başlık satırını gönder
                    rows_to_draw_on_this_page # Bu sayfadaki kalem veri satırlarını gönder
                )

                if not success_drawing:
                    if page_count == 1 and len(kalem_data_for_table) == 0:
                        c.setFont(TURKISH_FONT_NORMAL, 9)
                        c.drawCentredString(width/2, height/2, "Bu faturada herhangi bir kalem bulunmamaktadır.")
                    current_row_idx_for_drawing = total_data_rows_count + 1
                    break
                elif total_data_rows_count > 0 and len(rows_to_draw_on_this_page) == 0 and page_count > 1:
                    break
                elif total_data_rows_count == 0 and page_count == 1:
                    break

                current_row_idx_for_drawing += len(rows_to_draw_on_this_page)

                if current_row_idx_for_drawing >= total_data_rows_count and total_data_rows_count > 0:
                    break
                elif total_data_rows_count == 0 and page_count == 1:
                    break

            # --- Alt Bilgiler (Footer) Alanı - Sadece son sayfada çizilir ---
            # Hesaplanan genel iskonto tutarı (faturanın kendi değeri)
            genel_iskonto_uygulanan_tutari_fatura = 0.0
            if genel_iskonto_tipi_db == 'YUZDE' and genel_iskonto_degeri_db > 0:
                genel_iskonto_uygulanan_tutari_fatura = toplam_kdv_haric_fatura_ana * (genel_iskonto_degeri_db / 100)
            elif genel_iskonto_tipi_db == 'TUTAR' and genel_iskonto_degeri_db > 0:
                genel_iskonto_uygulanan_tutari_fatura = genel_iskonto_degeri_db

            # Toplamlar
            total_block_x_label = page_margin_x + 300
            total_block_x_value = width - page_margin_x

            c.setFont(TURKISH_FONT_NORMAL, 9)
            c.drawString(total_block_x_label, 85, "Toplam KDV Hariç:")
            c.drawString(total_block_x_label, 70, "Toplam KDV:")

            if genel_iskonto_uygulanan_tutari_fatura > 0:
                c.drawString(total_block_x_label, 55, "Genel İskonto:")
            
            c.setFont(TURKISH_FONT_BOLD, 10)
            c.drawString(total_block_x_label, 40, "GENEL TOPLAM:")


            c.setFont(TURKISH_FONT_NORMAL, 9)
            c.drawRightString(total_block_x_value, 85, self._format_currency(toplam_kdv_haric_fatura_ana))
            c.drawRightString(total_block_x_value, 70, self._format_currency(toplam_kdv_dahil_fatura_ana - toplam_kdv_haric_fatura_ana))
            if genel_iskonto_uygulanan_tutari_fatura > 0:
                c.drawRightString(total_block_x_value, 55, self._format_currency(genel_iskonto_uygulanan_tutari_fatura))
            
            c.setFont(TURKISH_FONT_BOLD, 10)
            c.drawRightString(total_block_x_value, 40, self._format_currency(toplam_kdv_dahil_fatura_ana))


            # Müşteri Bilgileri (Daha yukarıya ve sade)
            left_bottom_y_start_musteri = 85 # Y koordinatını yukarı çektik
            c.setFont(TURKISH_FONT_BOLD, 9)
            c.drawString(page_margin_x, left_bottom_y_start_musteri, "Müşteri Bilgileri")
            c.setFont(TURKISH_FONT_NORMAL, 8)
            c.drawString(page_margin_x, left_bottom_y_start_musteri - 12, f"Müşteri: {cari_bilgileri_pdf['ad']}")
            c.drawString(page_margin_x, left_bottom_y_start_musteri - 24, f"Önceki bakiye: {self._format_currency(cari_bakiye_snapshot['onceki_bakiye'])}")
            c.drawString(page_margin_x, left_bottom_y_start_musteri - 36, f"Bugün yapılan {'ödeme' if cari_tip_for_db_query == 'TEDARIKCI' else 'tahsilat'}: {self._format_currency(cari_bakiye_snapshot['bugun_odenen'])}")
            c.drawString(page_margin_x, left_bottom_y_start_musteri - 48, f"Kalan borç: {self._format_currency(cari_bakiye_snapshot['kalan_borc'])}")
            
            # Notlar / Açıklamalar (Müşteri bilgilerinin altına veya yanına konumlandırılabilir)
            if fatura_notlari_db:
                c.setFont(TURKISH_FONT_BOLD, 9)
                c.drawString(page_margin_x + 200, left_bottom_y_start_musteri, "Açıklamalar") # Yana kaydırdım
                notes_paragraph = Paragraph(fatura_notlari_db, styles['TurkishNormalSmall'])
                notes_paragraph.wrapOn(c, (width / 2) - page_margin_x - 20, 50)
                notes_paragraph.drawOn(c, page_margin_x + 200, left_bottom_y_start_musteri - 12 - notes_paragraph._height)

            # Banka Bilgileri (Açıklamaların altına veya yanına)
            if sirket_bilgileri.get('sirket_banka_adi') and sirket_bilgileri.get('sirket_iban'):
                c.setFont(TURKISH_FONT_BOLD, 9)
                c.drawString(page_margin_x + 400, left_bottom_y_start_musteri, "Banka Bilgileri") # En sağa hizala
                bank_info_text = f"Banka Adı: {sirket_bilgileri.get('sirket_banka_adi', 'N/A')}\nIBAN: {sirket_bilgileri.get('sirket_iban', 'N/A')}"
                bank_info_paragraph = Paragraph(bank_info_text, styles['TurkishNormalSmall'])
                bank_info_paragraph.wrapOn(c, (width / 2) - page_margin_x, 30)
                bank_info_paragraph.drawOn(c, page_margin_x + 400, left_bottom_y_start_musteri - 12 - bank_info_paragraph._height)


            # "Tanzim Eden" ve çizgisi kaldırıldı.

            # Dipnot
            c.setFont(TURKISH_FONT_NORMAL, 7)
            c.drawCentredString(width/2, 20, "Bu fatura Çınar Yapı Ön Muhasebe Programı ile oluşturulmuştur. Elektronik ortamda hazırlanmıştır.")


            c.save()
            return True, f"Fatura başarıyla PDF olarak kaydedildi: {dosya_yolu}"

        except Exception as e:
            logging.error(f"PDF oluşturulurken beklenmeyen bir hata oluştu: {e}\nDetaylar:\n{traceback.format_exc()}")
            return False, f"PDF oluşturulurken beklenmeyen bir hata oluştu: {e}\nDetaylar için log dosyasına bakınız."

    def tarihsel_satis_raporu_verilerini_al(self, baslangic_tarih_str, bitis_tarih_str):
        """Belirtilen tarih aralığındaki satış faturalarını ve kalemlerini getirir."""
        query = """
            SELECT 
                f.fatura_no, f.tarih, 
                CASE 
                    WHEN f.cari_id = ? THEN IFNULL(f.misafir_adi, 'Perakende Satış')
                    ELSE m.ad 
                END AS musteri_adi, 
                s.urun_kodu, s.urun_adi, 
                fk.miktar, fk.birim_fiyat, fk.kdv_orani, fk.kdv_tutari, 
                fk.kalem_toplam_kdv_haric, fk.kalem_toplam_kdv_dahil
            FROM faturalar f
            JOIN fatura_kalemleri fk ON f.id = fk.fatura_id
            JOIN tbl_stoklar s ON fk.urun_id = s.id
            LEFT JOIN musteriler m ON f.cari_id = m.id AND f.cari_id != ? -- Perakende olmayanlar için join
            WHERE f.tip = 'SATIŞ' AND f.tarih BETWEEN ? AND ?
            ORDER BY f.tarih, f.fatura_no, s.urun_adi
        """
        perakende_id_param = self.perakende_musteri_id if self.perakende_musteri_id is not None else -999
        params = (perakende_id_param, perakende_id_param, baslangic_tarih_str, bitis_tarih_str)
        self.c.execute(query, params)
        return self.c.fetchall()

    def tarihsel_satis_raporu_excel_olustur(self, rapor_verileri, dosya_yolu, bas_t, bit_t):
        try:
            if not rapor_verileri:
                return False, "Belirtilen tarih aralığında raporlanacak satış verisi bulunamadı."

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Satis_Raporu_{bas_t}_to_{bit_t}"

            headers = [
                "Fatura No", "Tarih", "Müşteri Adı", "Ürün Kodu", "Ürün Adı",
                "Miktar", "Birim Fiyat (TL)", "KDV Oranı (%)", "KDV Tutarı (TL)",
                "Toplam (KDV Hariç TL)", "Toplam (KDV Dahil TL)"
            ]
            ws.append(headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Koyu Mavi
            for col_idx, header_text in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header_text) + 2, 15)

            for row_data in rapor_verileri:
                ws.append(list(row_data)) # Gelen veri zaten tuple listesi

            # Sayısal sütunlara formatlama
            for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                # Miktar (F sütunu, index 5)
                if row_cells[5].value is not None: row_cells[5].number_format = '#,##0.00'
                # Birim Fiyat (G sütunu, index 6)
                if row_cells[6].value is not None: row_cells[6].number_format = '#,##0.00₺'
                # KDV Oranı (H sütunu, index 7)
                if row_cells[7].value is not None: row_cells[7].number_format = '0"%"'
                # KDV Tutarı (I sütunu, index 8)
                if row_cells[8].value is not None: row_cells[8].number_format = '#,##0.00₺'
                # Toplam KDV Hariç (J sütunu, index 9)
                if row_cells[9].value is not None: row_cells[9].number_format = '#,##0.00₺'
                # Toplam KDV Dahil (K sütunu, index 10)
                if row_cells[10].value is not None: row_cells[10].number_format = '#,##0.00₺'
            
            ws.freeze_panes = 'A2' # Başlık satırını dondur

            wb.save(dosya_yolu)
            return True
        except Exception as e:
            traceback.print_exc()
            return False, f"Satış raporu Excel'e aktarılırken bir hata oluştu:\n{e}"

    def tarihsel_satis_raporu_pdf_olustur(self, rapor_verileri, dosya_yolu, bas_t, bit_t):
        try:
            if not rapor_verileri:
                return False, "Belirtilen tarih aralığında raporlanacak satış verisi bulunamadı."

            c = rp_canvas.Canvas(dosya_yolu, pagesize=landscape(A4)) # Yatay A4
            width, height = landscape(A4)

            styles = getSampleStyleSheet()
            styleN = styles['Normal']
            styleN.fontName = TURKISH_FONT_NORMAL
            styleN.fontSize = 7 # Daha küçük font
            styleH = styles['Normal']
            styleH.fontName = TURKISH_FONT_BOLD
            styleH.fontSize = 7
            styleH.alignment = 1 # TA_CENTER
            styleRight = styles['Normal']
            styleRight.fontName = TURKISH_FONT_NORMAL
            styleRight.fontSize = 7
            styleRight.alignment = 2 # TA_RIGHT

            # Başlık
            c.setFont(TURKISH_FONT_BOLD, 14)
            c.drawCentredString(width/2, height - 40, f"Satış Raporu ({bas_t} - {bit_t})")
            c.setFont(TURKISH_FONT_NORMAL, 9)
            c.drawCentredString(width/2, height - 55, self.sirket_bilgileri.get("sirket_adi", ""))
            y_pos = height - 80

            data = [
                [Paragraph(h, styleH) for h in ["F.No", "Tarih", "Müşteri", "Ü.Kodu", "Ürün Adı", "Mik.", "B.Fyt", "KDV%", "KDV Tut.", "Tutar(Har.)", "Tutar(Dah.)"]]
            ]
            
            genel_toplam_kdv_haric_rapor = 0
            genel_toplam_kdv_tutar_rapor = 0
            genel_toplam_kdv_dahil_rapor = 0

            for item in rapor_verileri:
                # item: (f_no, tarih, musteri, u_kodu, u_adi, mik, bfyt, kdvo, kdvt, tkh, tkd)
                tarih_f = datetime.strptime(item[1], '%Y-%m-%d').strftime('%d.%m.%y') if item[1] else '-'
                miktar_f = f"{item[5]:.2f}".rstrip('0').rstrip('.') if isinstance(item[5], float) else str(item[5])
                data.append([
                    Paragraph(str(item[0]), styleN), Paragraph(tarih_f, styleN), Paragraph(str(item[2])[:25], styleN), # Müşteri adı kısaltılabilir
                    Paragraph(str(item[3]), styleN), Paragraph(str(item[4])[:30], styleN), # Ürün adı kısaltılabilir
                    Paragraph(miktar_f, styleRight), Paragraph(self._format_currency(item[6]), styleRight),
                    Paragraph(f"{item[7]:.0f}%", styleRight), Paragraph(self._format_currency(item[8]), styleRight),
                    Paragraph(self._format_currency(item[9]), styleRight), Paragraph(self._format_currency(item[10]), styleRight)
                ])
                genel_toplam_kdv_haric_rapor += item[9]
                genel_toplam_kdv_tutar_rapor += item[8]
                genel_toplam_kdv_dahil_rapor += item[10]

            col_widths = [1.8*cm, 1.5*cm, 4*cm, 2*cm, 5*cm, 1.2*cm, 2*cm, 1.2*cm, 2*cm, 2.3*cm, 2.3*cm]
            
            # Sayfa başına satır sayısı tahmini (deneme yanılma ile ayarlanabilir)
            rows_per_page = 25 
            num_pages = (len(data) -1 + rows_per_page - 1) // rows_per_page # Başlık hariç
            if num_pages == 0: num_pages = 1


            for page_num in range(num_pages):
                start_row = page_num * rows_per_page + (1 if page_num > 0 else 0) # İlk sayfada başlık var
                end_row = min((page_num + 1) * rows_per_page, len(data) -1 )
                
                page_data = [data[0]] + data[start_row+1 : end_row+1] # Başlığı her sayfaya ekle, sonra o sayfanın verilerini
                if not page_data[1:]: # Eğer veri kalmadıysa (sadece başlık)
                    if page_num > 0 : break # İlk sayfa değilse ve veri yoksa bitir
                    elif not data[1:]: # İlk sayfa ve hiç veri yoksa (sadece başlık)
                         page_data.append(["Veri Yok"]*len(col_widths)) # Boş satır ekle


                table = Table(page_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E78")),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('ALIGN', (2,1), (2,-1), 'LEFT'), # Müşteri Adı
                    ('ALIGN', (4,1), (4,-1), 'LEFT'), # Ürün Adı
                    ('ALIGN', (5,1), (-1,-1), 'RIGHT'), # Sayısal
                    ('FONTNAME', (0,0), (-1,-1), TURKISH_FONT_NORMAL),
                    ('FONTNAME', (0,0), (-1,0), TURKISH_FONT_BOLD),
                    ('FONTSIZE', (0,0), (-1,-1), 7),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                
                table.wrapOn(c, width - 80, height - 100)
                table_h = table._height
                table.drawOn(c, 40, y_pos - table_h)
                
                # Sayfa Numarası
                c.setFont(TURKISH_FONT_NORMAL, 8)
                c.drawRightString(width - 40, 30, f"Sayfa {page_num + 1} / {num_pages}")
                
                if page_num < num_pages - 1:
                    c.showPage()
                    c.setFont(TURKISH_FONT_BOLD, 14) # Yeni sayfada başlık fontunu ayarla
                    c.drawCentredString(width/2, height - 40, f"Satış Raporu ({bas_t} - {bit_t}) - Devam")
                    c.setFont(TURKISH_FONT_NORMAL, 9)
                    c.drawCentredString(width/2, height - 55, self.sirket_bilgileri.get("sirket_adi", ""))
                    y_pos = height - 80
            
            # Rapor Sonu Toplamları (Son Sayfaya)
            y_pos_summary = y_pos - table_h - 20 # Son tablonun altına
            if y_pos_summary < 80 : # Eğer yer kalmadıysa yeni sayfa açıp oraya yaz
                c.showPage()
                c.setFont(TURKISH_FONT_BOLD, 14)
                c.drawCentredString(width/2, height - 40, f"Satış Raporu ({bas_t} - {bit_t}) - Toplamlar")
                y_pos_summary = height - 70
                c.setFont(TURKISH_FONT_NORMAL, 8) # Sayfa numarasını da ekleyelim
                c.drawRightString(width - 40, 30, f"Sayfa {num_pages} / {num_pages}")


            c.setFont(TURKISH_FONT_BOLD, 9)
            c.drawRightString(width - 50, y_pos_summary, f"Genel Toplam (KDV Hariç): {self._format_currency(genel_toplam_kdv_haric_rapor)}")
            y_pos_summary -= 15
            c.drawRightString(width - 50, y_pos_summary, f"Genel Toplam KDV: {self._format_currency(genel_toplam_kdv_tutar_rapor)}")
            y_pos_summary -= 15
            c.setFont(TURKISH_FONT_BOLD, 10)
            c.drawRightString(width - 50, y_pos_summary, f"Genel Toplam (KDV Dahil): {self._format_currency(genel_toplam_kdv_dahil_rapor)}")

            c.save()
        except Exception as e:
            return False, f"Satış raporu PDF oluşturulurken hata: {e}"

    def __del__(self):
        if self.conn:
            try:
                self.conn.close()
            except Exception as e:
                print(f"Veritabanı bağlantısı kapatılırken hata: {e}")

    def kasa_banka_ekle(self, hesap_adi, hesap_no, bakiye, para_birimi, tip, acilis_tarihi=None, banka_adi=None, sube_adi=None, varsayilan_odeme_turu=None):
        if not (hesap_adi and tip):
            return False, "Hesap Adı ve Tip zorunludur."
        try:
            self.conn.execute("BEGIN TRANSACTION")
            current_time = self.get_current_datetime_str()
            olusturan_id = self.app.current_user[0] if self.app and self.app.current_user else None
            bakiye_f = float(str(bakiye).replace(',', '.')) if bakiye else 0.0

            self.c.execute("INSERT INTO kasalar_bankalar (hesap_adi, hesap_no, bakiye, para_birimi, tip, acilis_tarihi, banka_adi, sube_adi, varsayilan_odeme_turu, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                           (hesap_adi, hesap_no, bakiye_f, para_birimi, tip, acilis_tarihi, banka_adi, sube_adi, varsayilan_odeme_turu, current_time, olusturan_id))
            yeni_hesap_id = self.c.lastrowid
            self.conn.commit()
            return True, f"'{hesap_adi}' adlı hesap başarıyla eklendi. (ID: {yeni_hesap_id})"
        except sqlite3.IntegrityError as e:
            self.conn.rollback()
            if "UNIQUE constraint failed: kasalar_bankalar.hesap_adi" in str(e):
                return False, "Bu hesap adı zaten mevcut."
            elif "UNIQUE constraint failed: kasalar_bankalar.varsayilan_odeme_turu" in str(e):
                return False, f"'{varsayilan_odeme_turu}' ödeme türü zaten başka bir kasa/banka hesabına atanmış. Bir ödeme türü sadece bir hesaba atanabilir."
            else:
                return False, f"Veritabanı bütünlüğü hatası: {e}"
        except ValueError:
            self.conn.rollback()
            return False, "Bakiye sayısal olmalıdır."
        except Exception as e:
            self.conn.rollback()
            return False, f"Kasa/Banka ekleme sırasında hata: {e}"

    def kasa_banka_guncelle(self, hesap_id, hesap_adi, hesap_no, bakiye, para_birimi, tip, acilis_tarihi=None, banka_adi=None, sube_adi=None, varsayilan_odeme_turu=None):
        if not (hesap_adi and tip):
            return False, "Hesap Adı ve Tip zorunludur."
        try:
            self.conn.execute("BEGIN TRANSACTION")
            bakiye_f = float(str(bakiye).replace(',', '.')) if bakiye else 0.0
            current_time = self.get_current_datetime_str()
            guncelleyen_id = self.app.current_user[0] if self.app and self.app.current_user else None
            self.c.execute("UPDATE kasalar_bankalar SET hesap_adi=?, hesap_no=?, bakiye=?, para_birimi=?, tip=?, acilis_tarihi=?, banka_adi=?, sube_adi=?, varsayilan_odeme_turu=?, son_guncelleme_tarihi_saat=?, son_guncelleyen_kullanici_id=? WHERE id=?",
                           (hesap_adi, hesap_no, bakiye_f, para_birimi, tip, acilis_tarihi, banka_adi, sube_adi, varsayilan_odeme_turu, current_time, guncelleyen_id, hesap_id))
            self.conn.commit()
            if self.c.rowcount > 0:
                return True, "Hesap başarıyla güncellendi."
            else:
                return False, "Hesap bulunamadı veya bir değişiklik yapılmadı."
        except sqlite3.IntegrityError as e:
            self.conn.rollback()
        except ValueError:
            self.conn.rollback()
            return False, "Bakiye sayısal olmalıdır."
        except Exception as e:
            self.conn.rollback()
            return False, f"Kasa/Banka güncelleme sırasında hata: {e}"


    def kasa_banka_listesi_al(self, tip_filtre=None, arama_terimi=None):
        query = "SELECT id, hesap_adi, hesap_no, bakiye, para_birimi, tip, acilis_tarihi, banka_adi, sube_adi, varsayilan_odeme_turu FROM kasalar_bankalar"
        params = []; conditions = []
        if tip_filtre and tip_filtre != "TÜMÜ": conditions.append("tip = ?"); params.append(tip_filtre)
        if arama_terimi: term = f"%{arama_terimi}%"; conditions.append("(hesap_adi LIKE ? OR hesap_no LIKE ? OR banka_adi LIKE ?)"); params.extend([term, term, term])
        if conditions: query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY tip, hesap_adi ASC"; self.c.execute(query, params); return self.c.fetchall()

    def kasa_banka_getir_by_id(self, hesap_id):
        self.c.execute("SELECT id, hesap_adi, hesap_no, bakiye, para_birimi, tip, acilis_tarihi, banka_adi, sube_adi, varsayilan_odeme_turu, olusturma_tarihi_saat, olusturan_kullanici_id, son_guncelleme_tarihi_saat, son_guncelleyen_kullanici_id FROM kasalar_bankalar WHERE id=?", (hesap_id,))
        return self.c.fetchone()

    def kasa_banka_sil(self, hesap_id):
        if self.default_nakit_kasa_id is not None and str(hesap_id) == str(self.default_nakit_kasa_id):
            return False, "MERKEZİ NAKİT kasa hesabı silinemez. Sadece adı güncellenebilir."

        try:
            self.conn.execute("BEGIN TRANSACTION")
            for table_name in ['faturalar', 'cari_hareketler', 'gelir_gider']:
                self.c.execute(f"SELECT COUNT(*) FROM {table_name} WHERE kasa_banka_id=?", (hesap_id,))
                if self.c.fetchone()[0] > 0:
                    self.conn.rollback()
                    return False, f"Bu kasa/banka hesabı '{table_name}' tablosunda kullanılıyor.\nİlişkili kayıtları silmeden bu hesabı silemezsiniz."
            
            self.c.execute("DELETE FROM kasalar_bankalar WHERE id=?", (hesap_id,))
            
            self.conn.commit()
            if self.c.rowcount > 0:
                return True, "Kasa/Banka hesabı başarıyla silindi."
            else:
                return False, "Kasa/Banka hesabı bulunamadı veya silinemedi."
        except Exception as e:
            self.conn.rollback()
            return False, f"Kasa/Banka silme sırasında hata: {e}"

    def kasa_banka_bakiye_guncelle(self, kasa_banka_id, islem_tutari, artir=True):
        """
        Belirli bir kasa/banka hesabının bakiyesini günceller.
        Bu metot kendi başına bir transaction başlatmaz/kapatmaz,
        çağrıldığı transaction'ın bir parçası olarak çalışır.
        """
        if kasa_banka_id is None: 
            logging.warning("kasa_banka_bakiye_guncelle: Kasa/Banka ID boş geldi, işlem yapılmadı.")
            return True 
        try:
            self.c.execute("SELECT bakiye FROM kasalar_bankalar WHERE id=?", (kasa_banka_id,))
            mevcut_bakiye_tuple = self.c.fetchone()
            if not mevcut_bakiye_tuple:
                logging.error(f"kasa_banka_bakiye_guncelle: Hesap ID {kasa_banka_id} bulunamadı. Bakiye güncellenemedi.")
                return False 

            mevcut_bakiye = mevcut_bakiye_tuple[0]
            yeni_bakiye = mevcut_bakiye + islem_tutari if artir else mevcut_bakiye - islem_tutari

            logging.info(f"Kasa/Banka Bakiye Güncelleme: ID={kasa_banka_id}, Mevcut Bakiye={mevcut_bakiye}, İşlem Tutarı={islem_tutari}, Artır={artir}, Yeni Bakiye (Hesaplanan)={yeni_bakiye}")

            self.c.execute("UPDATE kasalar_bankalar SET bakiye=? WHERE id=?", (yeni_bakiye, kasa_banka_id))

            # Güncelleme sonrası kontrol (isteğe bağlı, hata ayıklama için)
            self.c.execute("SELECT bakiye FROM kasalar_bankalar WHERE id=?", (kasa_banka_id,))
            guncel_bakiye = self.c.fetchone()[0]
            logging.info(f"Kasa/Banka Bakiye Güncelleme: ID={kasa_banka_id}, Veritabanındaki Güncel Bakiye={guncel_bakiye}")

            return True
        except Exception as e:
            logging.error(f"Kasa/banka bakiyesi güncellenirken hata: Hesap ID {kasa_banka_id}, Tutar: {islem_tutari}, Artır: {artir}. Hata: {e}\n{traceback.format_exc()}")
            return False 
            
    def clear_stok_data(self):
        try:
            self.conn.execute("BEGIN TRANSACTION")
            self.c.execute("DELETE FROM fatura_kalemleri")
            self.c.execute("DELETE FROM siparis_kalemleri")
            self.c.execute("DELETE FROM teklif_kalemleri")
            self.c.execute("DELETE FROM tbl_stoklar")
            self.conn.commit()
            return True, "Tüm stok verileri başarıyla temizlendi." # Düzeltme: Mesaj eklendi
        except Exception as e:
            self.conn.rollback()
            return False, f"Stok verileri temizlenirken hata oluştu: {e}" # Düzeltme: Mesaj eklendi

    def clear_musteri_data(self):
        """Perakende müşteri hariç tüm müşteri verilerini ve ilişkili hareketleri temizler."""
        try:
            self.conn.execute("BEGIN TRANSACTION")
            
            # Perakende müşterinin ID'sini al
            perakende_id = self.perakende_musteri_id
            
            # Müşterilere ait faturaları, cari hareketleri, siparişleri, teklifleri sil
            self.c.execute("DELETE FROM fatura_kalemleri WHERE fatura_id IN (SELECT id FROM faturalar WHERE cari_id IN (SELECT id FROM musteriler WHERE id != ?))", (perakende_id,))
            self.c.execute("DELETE FROM faturalar WHERE cari_id IN (SELECT id FROM musteriler WHERE id != ?)", (perakende_id,))
            self.c.execute("DELETE FROM cari_hareketler WHERE cari_id IN (SELECT id FROM musteriler WHERE id != ?) AND cari_tip='MUSTERI'", (perakende_id,))
            self.c.execute("DELETE FROM siparis_kalemleri WHERE siparis_id IN (SELECT id FROM siparisler WHERE cari_id IN (SELECT id FROM musteriler WHERE id != ?) AND cari_tip='MUSTERI')", (perakende_id,))
            self.c.execute("DELETE FROM siparisler WHERE cari_id IN (SELECT id FROM musteriler WHERE id != ?) AND cari_tip='MUSTERI'", (perakende_id,))
            self.c.execute("DELETE FROM teklif_kalemleri WHERE teklif_id IN (SELECT id FROM teklifler WHERE musteri_id IN (SELECT id FROM musteriler WHERE id != ?))", (perakende_id,))
            self.c.execute("DELETE FROM teklifler WHERE musteri_id IN (SELECT id FROM musteriler WHERE id != ?)", (perakende_id,))

            # Perakende müşteri hariç tüm müşterileri sil
            self.c.execute("DELETE FROM musteriler WHERE id != ?", (perakende_id,))
            self.conn.commit()
            return True, "Tüm müşteri verileri (perakende hariç) başarıyla temizlendi." # Düzeltme: Mesaj eklendi
        except Exception as e:
            self.conn.rollback()
            return False, f"Müşteri verileri temizlenirken hata oluştu: {e}" # Düzeltme: Mesaj eklendi
        
    def clear_tedarikci_data(self):
        """Tüm tedarikçi verilerini ve ilişkili hareketleri temizler."""
        try:
            self.conn.execute("BEGIN TRANSACTION")
            # Tedarikçilere ait faturaları, cari hareketleri, siparişleri sil
            self.c.execute("DELETE FROM fatura_kalemleri WHERE fatura_id IN (SELECT id FROM faturalar WHERE cari_id IN (SELECT id FROM tedarikciler))")
            self.c.execute("DELETE FROM faturalar WHERE cari_id IN (SELECT id FROM tedarikciler)")
            self.c.execute("DELETE FROM cari_hareketler WHERE cari_id IN (SELECT id FROM tedarikciler) AND cari_tip='TEDARIKCI'")
            self.c.execute("DELETE FROM siparis_kalemleri WHERE siparis_id IN (SELECT id FROM siparisler WHERE cari_id IN (SELECT id FROM tedarikciler) AND cari_tip='TEDARIKCI')")
            self.c.execute("DELETE FROM siparisler WHERE cari_id IN (SELECT id FROM tedarikciler) AND cari_tip='TEDARIKCI'")
            
            # Tedarikçi tablosunu temizle
            self.c.execute("DELETE FROM tedarikciler")       
            self.conn.commit()
            return True, "Tüm tedarikçi verileri başarıyla temizlendi." # Düzeltme: Mesaj eklendi
        except Exception as e:
            self.conn.rollback()
            return False, f"Tedarikçi verileri temizlenirken hata oluştu: {e}" # Düzeltme: Mesaj eklendi

    def clear_kasa_banka_data(self):
        """Tüm kasa/banka verilerini ve ilişkili hareketleri temizler."""
        try:
            self.conn.execute("BEGIN TRANSACTION")
            # İlişkili faturaları, gelir/gider ve cari hareketleri sil
            
            # Faturalardaki kasa_banka_id'yi NULL yap
            self.c.execute("UPDATE faturalar SET kasa_banka_id = NULL WHERE kasa_banka_id IS NOT NULL")
            # Cari hareketlerdeki kasa_banka_id'yi NULL yap
            self.c.execute("UPDATE cari_hareketler SET kasa_banka_id = NULL WHERE kasa_banka_id IS NOT NULL")
            # Gelir/giderlerdeki kasa_banka_id'yi NULL yap
            self.c.execute("UPDATE gelir_gider SET kasa_banka_id = NULL WHERE kasa_banka_id IS NOT NULL")

            # Kasa/Banka tablosunu temizle
            self.c.execute("DELETE FROM kasalar_bankalar")
            self.conn.commit()
            return True, "Tüm kasa/banka verileri başarıyla temizlendi." # Düzeltme: Mesaj eklendi
        except Exception as e:
            self.conn.rollback()
            return False, f"Kasa/Banka verileri temizlenirken hata oluştu: {e}" # Düzeltme: Mesaj eklendi

    def clear_all_transaction_data(self):
        """Tüm işlem (fatura, gelir/gider, cari hareket, sipariş, teklif) verilerini temizler.
        Stok, müşteri, tedarikçi, kasa/banka, kullanıcı ve şirket ayarları korunur."""
        try:
            self.conn.execute("BEGIN TRANSACTION")
            self.c.execute("DELETE FROM fatura_kalemleri")
            self.c.execute("DELETE FROM faturalar")
            self.c.execute("DELETE FROM gelir_gider")
            self.c.execute("DELETE FROM cari_hareketler")
            self.c.execute("DELETE FROM siparis_kalemleri")
            self.c.execute("DELETE FROM siparisler")
            self.c.execute("DELETE FROM teklif_kalemleri")
            self.c.execute("DELETE FROM teklifler")

            # Stok miktarlarını sıfırla (ürünleri silmeden)
            self.c.execute("UPDATE tbl_stoklar SET stok_miktari = 0.0")
            
            # Kasa/Banka bakiyelerini sıfırla (hesapları silmeden)
            self.c.execute("UPDATE kasalar_bankalar SET bakiye = 0.0")
            self.conn.commit()

            return True, "Tüm işlem verileri başarıyla temizlendi. Stok ve kasa/banka bakiyeleri sıfırlandı." # Düzeltme: Mesaj eklendi
        except Exception as e:
            self.conn.rollback()
            return False, f"Tüm işlem verileri temizlenirken hata oluştu: {e}" # Düzeltme: Mesaj eklendi

    def clear_all_data(self):
        """Kullanıcılar ve şirket ayarları hariç tüm veritabanı tablolarını temizler."""
        try:
            self.conn.execute("BEGIN TRANSACTION")
            
            # Tüm tabloları sil (kullanicilar ve sirket_ayarlari hariç)
            # Bağımlı tabloları önce sil
            self.c.execute("DELETE FROM fatura_kalemleri")
            self.c.execute("DELETE FROM gelir_gider")
            self.c.execute("DELETE FROM cari_hareketler")
            self.c.execute("DELETE FROM siparis_kalemleri")
            self.c.execute("DELETE FROM teklif_kalemleri")
            
            # Sonra ana tabloları sil
            self.c.execute("DELETE FROM faturalar")
            self.c.execute("DELETE FROM siparisler")
            self.c.execute("DELETE FROM teklifler")
            self.c.execute("DELETE FROM tbl_stoklar")
            self.c.execute("DELETE FROM tedarikciler")
            self.c.execute("DELETE FROM kasalar_bankalar")
            
            # Müşteriler tablosunu temizle, perakende müşteriyi yeniden oluşturmak için
            self.c.execute("DELETE FROM musteriler")
            self.conn.commit()
            
            return True, "Tüm veritabanı başarıyla temizlendi (kullanıcılar ve şirket ayarları hariç)." # Düzeltme: Mesaj eklendi
        except Exception as e:
            self.conn.rollback()
            return False, f"Tüm veriler temizlenirken hata oluştu: {e}" 

    def optimize_database(self):
        """
        SQLite veritabanını sıkıştırır (VACUUM komutu).
        Bu işlem, veritabanı boyutunu küçültür ve performansı artırabilir.
        """
        try:
            self.conn.execute("VACUUM;")
            self.conn.commit()
            return True, "Veritabanı başarıyla optimize edildi."
        except Exception as e:
            return False, f"Veritabanı optimize edilirken hata oluştu: {e}"


    def _ensure_default_kasa(self):
        try:
            self.c.execute("SELECT id FROM kasalar_bankalar WHERE hesap_adi=? AND varsayilan_odeme_turu=?",
                           ("MERKEZİ NAKİT", "NAKİT"))
            result = self.c.fetchone()
            if result:
                self.default_nakit_kasa_id = result[0]
                return True, "MERKEZİ NAKİT kasa hesabı bulundu."
            else:
                olusturan_id = 1 # Genellikle 'admin' kullanıcısının ID'si
                current_time = self.get_current_datetime_str()
                self.c.execute("INSERT INTO kasalar_bankalar (hesap_adi, hesap_no, bakiye, para_birimi, tip, acilis_tarihi, banka_adi, sube_adi, varsayilan_odeme_turu, olusturma_tarihi_saat, olusturan_kullanici_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                               ("MERKEZİ NAKİT", None, 0.0, "TL", "KASA", datetime.now().strftime('%Y-%m-%d'), None, None, "NAKİT", current_time, olusturan_id))
                self.default_nakit_kasa_id = self.c.lastrowid
                self.conn.commit()
                return True, "MERKEZİ NAKİT kasa hesabı başarıyla oluşturuldu."
        except sqlite3.IntegrityError as e:
            if "UNIQUE constraint failed: kasalar_bankalar.varsayilan_odeme_turu" in str(e):
                return False, "NAKİT ödeme türü zaten başka bir kasa/banka hesabına atanmış. Lütfen kontrol edin."
            else:
                return False, f"Veritabanı bütünlüğü hatası: {e}"
        except Exception as e:
            return False, f"MERKEZİ NAKİT kasa hesabı kontrol edilirken/oluşturulurken kritik hata: {e}\n{traceback.format_exc()}"

    def safe_float(self, value):
        """String veya sayısal değeri güvenli bir şekilde float'a dönüştürür."""
        try:
            if isinstance(value, (int, float)):
                return float(value)
            # Türkçe virgüllü sayıları noktaya çevirerek dönüştür
            return float(str(value).replace('.', '').replace(',', '.'))
        except (ValueError, TypeError):
            return 0.0
        
    def gecmis_hatali_kayitlari_temizle(self):
        """
        Veritabanını tarar ve artık 'faturalar' tablosunda var olmayan faturalara ait
        'cari_hareketler' ve 'gelir_gider' kayıtlarını (hayalet kayıtları) siler.
        Bu, geçmişte yapılan hatalı silme işlemlerini düzeltmek için tek seferlik bir işlemdir.
        """
        # <<< DEĞİŞİKLİK BURADA BAŞLIYOR >>>
        try:
            self.conn.execute("BEGIN TRANSACTION")
            
            # Adım 1: Artık var olmayan faturalara ait cari hareketleri bul ve sil
            # LEFT JOIN tekniği, faturalar tablosunda karşılığı olmayan kayıtları (f.id IS NULL) bulmada daha güvenilirdir.
            self.c.execute("""
                DELETE FROM cari_hareketler 
                WHERE id IN (
                    SELECT ch.id 
                    FROM cari_hareketler ch
                    LEFT JOIN faturalar f ON ch.referans_id = f.id
                    WHERE (ch.referans_tip LIKE '%FATURA%' OR ch.referans_tip LIKE '%IADE_FATURA%') AND f.id IS NULL
                )
            """)
            cari_hareket_sayisi = self.c.rowcount
            if cari_hareket_sayisi > 0:
                logging.info(f"{cari_hareket_sayisi} adet hayalet cari hareketi silindi.")

            # Adım 2: Artık var olmayan faturalara ait gelir/gider hareketlerini bul ve sil
            self.c.execute("""
                DELETE FROM gelir_gider
                WHERE id IN (
                    SELECT gg.id
                    FROM gelir_gider gg
                    LEFT JOIN faturalar f ON gg.kaynak_id = f.id
                    WHERE (gg.kaynak LIKE '%FATURA%' OR gg.kaynak LIKE '%IADE_FATURA%') AND f.id IS NULL
                )
            """)
            gg_hareket_sayisi = self.c.rowcount
            if gg_hareket_sayisi > 0:
                logging.info(f"{gg_hareket_sayisi} adet hayalet gelir/gider hareketi silindi.")

            self.conn.commit()
            toplam_silinen = cari_hareket_sayisi + gg_hareket_sayisi
            if toplam_silinen > 0:
                return True, f"Temizlik tamamlandı. Toplam {toplam_silinen} adet geçmiş hatalı kayıt silindi."
            else:
                return True, "Geçmişe dönük hatalı bir kayıt bulunamadı. Veritabanınız temiz görünüyor."

        except Exception as e:
            self.conn.rollback()
            logging.error(f"Geçmiş hatalı kayıtlar temizlenirken hata: {e}\n{traceback.format_exc()}")
            return False, f"Temizleme sırasında bir hata oluştu: {e}"

